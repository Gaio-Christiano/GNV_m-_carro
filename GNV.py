# =============================================================================
# ARQUIVO.....: GNV.py
# AUTOR.......: Christiano Gaio
# OBJETIVO....: Calculadora de GNV
#
# DESCRIÇÃO
# ----------
# Este programa calcula aproximadamente a quantidade de GNV armazenada em um
# cilindro utilizando a Lei dos Gases Ideais.
#
# OBSERVAÇÃO IMPORTANTE
# ---------------------
# O GNV é um gás REAL.
# Portanto, em pressões elevadas (200 a 220 bar) existe um erro quando usamos
# apenas a Lei dos Gases Ideais.
#
# Futuramente será possível implementar:
#
#   • Fator de Compressibilidade (Z)
#   • Norma AGA8
#   • GERG-2008
#
# Entretanto, para estudo da física e programação este programa é excelente.
#
# Todo o código possui comentários para facilitar o aprendizado.
# =============================================================================






# =============================================================================
# PARTE 01 - IMPORTAÇÃO DAS BIBLIOTECAS
# =============================================================================

# Biblioteca matemática

import math
import csv
import os
import sqlite3
import tkinter as tk
import pandas as pd
from tkinter import ttk

# =============================================================================
# PARTE 253
# IMPORT DATETIME
# =============================================================================

from datetime import datetime
from tkinter import messagebox
from tkinter import filedialog

# =============================================================================
# PARTE 246
# IMPORT FPDF
# =============================================================================
from fpdf import FPDF



# =============================================================================
# PARTE 02 - CONSTANTES FÍSICAS
# =============================================================================

# Pressão atmosférica padrão ao nível do mar (bar)
PRESSAO_ATMOSFERICA_PADRAO = 1.01325

# Temperatura padrão (Kelvin)
TEMPERATURA_PADRAO = 273.15

# Constante universal dos gases
#
# Unidade:
#
# J/(mol.K)
#
R = 8.314462618

# Massa molar aproximada do GNV
#
# O gás natural muda de composição conforme a distribuidora.
#
# Foi adotado um valor médio.
#

# -------------------------------------------------------------------------
# Estes valores mudam conforme a composição do gás fornecido pela
# distribuidora.
#
# O usuário poderá alterá-los na execução do programa.
# -------------------------------------------------------------------------

MASSA_MOLAR_GNV = None      # kg/mol

# Densidade média do GNV nas CNTP
#
# Este valor varia mensalmente.
#
DENSIDADE_GNV = None           # kg/m³

# Fator de compressibilidade
#
# Z = 1,000  -> gás ideal
#
# Para GNV normalmente varia aproximadamente entre
#
# 0,82 e 0,95
#
Z = None






# =============================================================================
# PARTE 03 - FUNÇÃO
# Calcula a pressão atmosférica em função da altitude.
# =============================================================================

def calcular_pressao_atmosferica(altitude):

    """
    Calcula a pressão atmosférica.

    Entrada

        altitude em metros

    Saída

        pressão em bar

    Fórmula aproximada válida para pequenas altitudes.
    """

    # Equação barométrica simplificada

    pressao = PRESSAO_ATMOSFERICA_PADRAO * \
              (1 - 2.25577e-5 * altitude) ** 5.25588

    return pressao





# =============================================================================
# PARTE 04 - FUNÇÃO
# Converte Celsius para Kelvin
# =============================================================================

def celsius_para_kelvin(temperatura):

    return temperatura + 273.15

# =============================================================================
# FUNÇÃO
# Converte litros para metros cúbicos
# =============================================================================

def litros_para_m3(litros):

    return litros / 1000.0

# =============================================================================
# FUNÇÃO
# Calcula pressão absoluta
# =============================================================================

def calcular_pressao_absoluta(pressao_manometrica,
                              pressao_atmosferica):

    """
    A pressão mostrada no manômetro é relativa.

    Para utilizar a Lei dos Gases devemos usar pressão absoluta.

    Pabs = Pman + Patm
    """

    return pressao_manometrica + pressao_atmosferica

# =============================================================================
# FUNÇÃO
# Calcula quantidade de mols
# =============================================================================

def calcular_mols(volume_m3,
                  pressao_bar,
                  temperatura_kelvin,
                  fator_z):

    """
    Calcula a quantidade de matéria.

    Equação

    PV = ZnRT

    Quanto menor o fator Z,
    maior será a quantidade de gás para
    a mesma pressão.
    """

    pressao_pa = pressao_bar * 100000

    mols = (
        pressao_pa *
        volume_m3
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return mols

# =============================================================================
# FUNÇÃO
# Calcula massa do gás
# =============================================================================

def calcular_massa(mols):

    """
    Massa = mol × Massa molar
    """

    return mols * MASSA_MOLAR_GNV



# =============================================================================
# FUNÇÃO
# Calcula volume equivalente em condições normais
# =============================================================================

def calcular_volume_equivalente(massa):

    """
    m³ = massa / densidade
    """

    return massa / DENSIDADE_GNV




# =============================================================================
# FUNÇÃO PARA CALCULAR O ABASTECIMENTO
# =============================================================================

def calcular_abastecimento(
    volume_abastecido,
    preco_m3,
    odometro_anterior,
    odometro_atual,
    massa_gnv
):
    """
    Calcula todas as informações referentes ao abastecimento.

    Parâmetros
    ----------
    volume_abastecido : float
        Volume abastecido em m³.

    preco_m3 : float
        Preço do GNV por metro cúbico.

    odometro_anterior : float
        Quilometragem anterior.

    odometro_atual : float
        Quilometragem atual.

    massa_gnv : float
        Massa estimada do GNV.

    Retorno
    -------
    dict
        Dicionário contendo todos os resultados.
    """

    valor_total = volume_abastecido * preco_m3

    distancia = odometro_atual - odometro_anterior

    if distancia < 0:

        distancia = 0

    if volume_abastecido > 0:

        rendimento = distancia / volume_abastecido

    else:

        rendimento = 0

    if distancia > 0:

        custo_km = valor_total / distancia

    else:

        custo_km = 0

    if massa_gnv > 0:

        custo_kg = valor_total / massa_gnv

    else:

        custo_kg = 0

    custo_100km = custo_km * 100

    return {

        "valor_total": valor_total,

        "distancia": distancia,

        "rendimento": rendimento,

        "custo_km": custo_km,

        "custo_100km": custo_100km,

        "custo_kg": custo_kg

    }




# =============================================================================
# PARTE 21
# FUNÇÃO PARA CALCULAR O VALOR ENERGÉTICO DO GNV
# =============================================================================

def calcular_energia_gnv(
    volume_m3,
    pcs=39.5
):
    """
    Calcula a energia contida no GNV.

    volume_m3
        Volume do gás em metros cúbicos.

    pcs
        Poder Calorífico Superior
        MJ/m³

    Retorna

        energia_mj
        energia_kwh
    """

    energia_mj = volume_m3 * pcs

    energia_kwh = energia_mj / 3.6

    return {

        "MJ": energia_mj,

        "kWh": energia_kwh

    }


# =============================================================================
# PARTE 22
# FUNÇÃO PARA CALCULAR A MASSA DO GNV
# =============================================================================

def calcular_massa_gnv(
    volume_m3,
    densidade=0.717
):
    """
    Calcula a massa do GNV.

    densidade padrão em CNTP.

    Retorna a massa em quilogramas.
    """

    return volume_m3 * densidade


# =============================================================================
# PARTE 25
# CÁLCULO DA DENSIDADE REAL DO GNV
# =============================================================================

def calcular_densidade_real(

    massa,

    volume

):
    """
    Calcula a densidade real do gás.

    densidade = massa / volume

    Retorna kg/m³.
    """

    if volume <= 0:

        return 0.0

    return massa / volume

# =============================================================================
# PARTE 26
# PESO DO GNV
# =============================================================================

def calcular_peso(

    massa

):
    """
    Calcula o peso do gás.

    Retorna Newton.
    """

    G = 9.80665

    return massa * G


# =============================================================================
# PARTE 27
# CÁLCULO DA ENERGIA PELO PESO DO GNV
# =============================================================================

def calcular_energia_por_massa(
    massa_kg,
    pci=50.0
):
    """
    Calcula a energia disponível no GNV.

    Parâmetros
    ----------
    massa_kg : float
        Massa do GNV em quilogramas.

    pci : float
        Poder calorífico inferior em MJ/kg.
        Valor padrão aproximado para GNV.

    Retorno
    -------
    dict contendo:

        energia_mj
        energia_kwh
    """

    energia_mj = massa_kg * pci

    energia_kwh = energia_mj / 3.6

    return {

        "energia_mj": energia_mj,

        "energia_kwh": energia_kwh

    }

# =============================================================================
# PARTE 28
# CÁLCULO DO CONSUMO DO VEÍCULO
# =============================================================================

def calcular_consumo(

    distancia_km,

    volume_m3,

    massa_kg

):
    """
    Calcula diversos índices de consumo.

    Retorna um dicionário contendo:

        km_por_m3

        m3_por_100km

        km_por_kg

        kg_por_100km
    """

    resultado = {}

    if volume_m3 > 0:

        resultado["km_por_m3"] = (
            distancia_km /
            volume_m3
        )

        resultado["m3_por_100km"] = (
            volume_m3 /
            distancia_km
        ) * 100 if distancia_km > 0 else 0

    else:

        resultado["km_por_m3"] = 0

        resultado["m3_por_100km"] = 0

    if massa_kg > 0:

        resultado["km_por_kg"] = (
            distancia_km /
            massa_kg
        )

        resultado["kg_por_100km"] = (
            massa_kg /
            distancia_km
        ) * 100 if distancia_km > 0 else 0

    else:

        resultado["km_por_kg"] = 0

        resultado["kg_por_100km"] = 0

    return resultado


# =============================================================================
# PARTE 29
# CÁLCULO DA MASSA DO GNV PELA DENSIDADE
# =============================================================================

def calcular_massa_por_densidade(

    volume_m3,

    densidade

):
    """
    Calcula a massa do GNV.

    massa = volume × densidade

    volume_m3
        Volume em metros cúbicos

    densidade
        kg/m³

    Retorna
        Massa em quilogramas.
    """

    return volume_m3 * densidade


# =============================================================================
# PARTE 30
# DENSIDADE APARENTE
# =============================================================================

def calcular_densidade_aparente(

    massa,

    volume

):
    """
    Calcula a densidade aparente.

    densidade = massa / volume
    """

    if volume <= 0:

        return 0

    return massa / volume


# =============================================================================
# PARTE 31
# ENERGIA ESPECÍFICA
# =============================================================================

def calcular_energia_especifica(

    energia_kwh,

    massa

):
    """
    Calcula a energia específica.

    kWh/kg
    """

    if massa <= 0:

        return 0

    return energia_kwh / massa

# =============================================================================
# PARTE 32
# CÁLCULO DO VOLUME LIVRE DO CILINDRO
# =============================================================================

def calcular_volume_livre(

    volume_total_litros,

    percentual_ocupacao

):
    """
    Calcula o volume livre existente
    no cilindro.

    percentual_ocupacao

        0 a 100 %
    """

    if percentual_ocupacao < 0:

        percentual_ocupacao = 0

    if percentual_ocupacao > 100:

        percentual_ocupacao = 100

    volume_ocupado = (

        volume_total_litros

        *

        percentual_ocupacao

        /

        100

    )

    volume_livre = (

        volume_total_litros

        -

        volume_ocupado

    )

    return volume_livre


# =============================================================================
# PARTE 33
# PERCENTUAL DE ENCHIMENTO
# =============================================================================

def calcular_percentual_enchimento(

    pressao_bar,

    pressao_maxima=220

):
    """
    Retorna o percentual de enchimento
    do cilindro baseado na pressão.

    A pressão máxima pode ser alterada.
    """

    if pressao_maxima <= 0:

        return 0

    percentual = (

        pressao_bar

        /

        pressao_maxima

    ) * 100

    if percentual < 0:

        percentual = 0

    if percentual > 100:

        percentual = 100

    return percentual


# =============================================================================
# PARTE 34
# PRESSÃO RESTANTE
# =============================================================================

def calcular_pressao_restante(

    pressao_atual,

    pressao_minima=5

):
    """
    Calcula a pressão realmente
    utilizável do cilindro.
    """

    restante = (

        pressao_atual

        -

        pressao_minima

    )

    if restante < 0:

        restante = 0

    return restante


# =============================================================================
# PARTE 35
# CÁLCULO DA MASSA CONSUMIDA
# =============================================================================

def calcular_massa_consumida(

    massa_inicial,

    massa_final

):
    """
    Calcula a massa consumida.

    Retorna sempre um valor positivo.
    """

    massa = (

        massa_inicial

        -

        massa_final

    )

    if massa < 0:

        massa = 0

    return massa



# =============================================================================
# PARTE 36
# CÁLCULO DO VOLUME CONSUMIDO
# =============================================================================

def calcular_volume_consumido(

    volume_inicial,

    volume_final

):
    """
    Calcula o volume efetivamente consumido.
    """

    volume = (

        volume_inicial

        -

        volume_final

    )

    if volume < 0:

        volume = 0

    return volume


# =============================================================================
# PARTE 37
# PERCENTUAL CONSUMIDO
# =============================================================================

def calcular_percentual_consumido(

    valor_inicial,

    valor_final

):
    """
    Calcula o percentual consumido.

    Exemplo

    Inicial = 16 m³

    Final = 4 m³

    Consumo = 75%
    """

    if valor_inicial <= 0:

        return 0

    percentual = (

        (

            valor_inicial

            -

            valor_final

        )

        /

        valor_inicial

    ) * 100

    if percentual < 0:

        percentual = 0

    if percentual > 100:

        percentual = 100

    return percentual


# =============================================================================
# PARTE 38
# EFICIÊNCIA DO ABASTECIMENTO
# =============================================================================

def calcular_eficiencia(

    volume_teorico,

    volume_real

):
    """
    Calcula a eficiência do abastecimento.

    Quanto mais próximo de 100%,
    mais próximo do cálculo teórico.
    """

    if volume_teorico <= 0:

        return 0

    eficiencia = (

        volume_real

        /

        volume_teorico

    ) * 100

    return eficiencia


# =============================================================================
# PARTE 39
# PRESSÃO DE ENCHIMENTO
# =============================================================================

def calcular_pressao_enchimento(

    pressao_inicial,

    pressao_final

):
    """
    Calcula quanto a pressão aumentou
    durante o abastecimento.

    Retorna o aumento em bar.
    """

    aumento = (

        pressao_final

        -

        pressao_inicial

    )

    if aumento < 0:

        aumento = 0

    return aumento



# =============================================================================
# PARTE 40
# TAXA DE COMPRESSÃO
# =============================================================================

def calcular_taxa_compressao(

    pressao_absoluta,

    pressao_atmosferica

):
    """
    Calcula a taxa de compressão.

    Exemplo

    201 bar absolutos

    1,01325 bar atmosféricos

    Resultado ≈ 198
    """

    if pressao_atmosferica <= 0:

        return 0

    return (

        pressao_absoluta

        /

        pressao_atmosferica

    )



# =============================================================================
# PARTE 41
# ENERGIA POR METRO CÚBICO
# =============================================================================

def calcular_energia_m3(

    energia_total,

    volume_m3

):
    """
    Calcula a energia existente
    em cada metro cúbico.
    """

    if volume_m3 <= 0:

        return 0

    return (

        energia_total

        /

        volume_m3

    )


# =============================================================================
# PARTE 42
# FATOR DE UTILIZAÇÃO
# =============================================================================

def calcular_fator_utilizacao(

    volume_utilizado,

    volume_total

):
    """
    Calcula o percentual realmente
    utilizado do cilindro.
    """

    if volume_total <= 0:

        return 0

    fator = (

        volume_utilizado

        /

        volume_total

    ) * 100

    if fator < 0:

        fator = 0

    if fator > 100:

        fator = 100

    return fator


# =============================================================================
# PARTE 43
# BIBLIOTECA DE PROPRIEDADES DO GNV
# =============================================================================

def obter_propriedades_gnv():

    """
    Retorna um dicionário contendo as
    propriedades médias do GNV.

    Todos os valores poderão ser alterados
    futuramente conforme a composição do gás.
    """

    propriedades = {

        "massa_molar": 0.01812,      # kg/mol

        "densidade": 0.717,          # kg/m³ CNTP

        "pcs": 39.50,                # MJ/m³

        "pci": 35.80,                # MJ/m³

        "cp": 2.20,                  # kJ/(kg.K)

        "cv": 1.70,                  # kJ/(kg.K)

        "k": 1.294,                  # Cp/Cv

        "R_especifico": 518.3,       # J/(kg.K)

        "temperatura_critica": 190.6,# Kelvin

        "pressao_critica": 45.99     # bar

    }

    return propriedades


# =============================================================================
# PARTE 44
# MOSTRAR AS PROPRIEDADES
# =============================================================================

def mostrar_propriedades_gnv():

    propriedades = obter_propriedades_gnv()

    print()

    print("=" * 75)

    print("PROPRIEDADES DO GNV")

    print("=" * 75)

    print()

    for chave, valor in propriedades.items():

        print(f"{chave:<25} : {valor}")

    print()

    print("=" * 75)

    print()


# =============================================================================
# PARTE 45
# OBTER UMA PROPRIEDADE
# =============================================================================

def propriedade(nome):

    """
    Retorna uma propriedade específica.

    Exemplo

    pcs = propriedade("pcs")
    """

    propriedades = obter_propriedades_gnv()

    return propriedades.get(nome)


# =============================================================================
# PARTE 46
# ENERGIA UTILIZANDO A BIBLIOTECA
# =============================================================================

def calcular_energia_volume(

    volume_m3

):
    """
    Calcula automaticamente a energia
    utilizando o PCS armazenado na biblioteca.
    """

    pcs = propriedade(

        "pcs"

    )

    energia = (

        volume_m3

        *

        pcs

    )

    return energia



# =============================================================================
# PARTE 47
# COMPOSIÇÃO DO GNV
# =============================================================================

def obter_composicao_padrao():

    """
    Retorna uma composição típica do GNV.

    Os valores são em porcentagem (%).
    """

    composicao = {

        "CH4": 89.00,

        "C2H6": 6.00,

        "C3H8": 2.00,

        "CO2": 1.50,

        "N2": 1.00,

        "O2": 0.30,

        "Outros": 0.20

    }

    return composicao




# =============================================================================
# PARTE 48
# MOSTRAR COMPOSIÇÃO DO GNV
# =============================================================================

def mostrar_composicao():

    composicao = obter_composicao_padrao()

    print()

    print("=" * 75)

    print("COMPOSIÇÃO DO GNV")

    print("=" * 75)

    print()

    for gas, percentual in composicao.items():

        print(f"{gas:<10} : {percentual:6.2f} %")

    print()

    print("=" * 75)

    print()



# =============================================================================
# PARTE 49
# MASSA MOLAR DA MISTURA
# =============================================================================

def calcular_massa_molar_mistura():

    """
    Calcula a massa molar média
    da composição do GNV.
    """

    massas = {

        "CH4":16.043,

        "C2H6":30.070,

        "C3H8":44.097,

        "CO2":44.010,

        "N2":28.014,

        "O2":31.999,

        "Outros":20.000

    }

    composicao = obter_composicao_padrao()

    massa = 0

    for gas in composicao:

        massa += (

            composicao[gas]

            *

            massas[gas]

        )

    massa /= 100

    return massa / 1000


# =============================================================================
# PARTE 50
# RESUMO DAS PROPRIEDADES DO GNV
# =============================================================================

def resumo_gnv():

    print()

    print("=" * 75)

    print("RESUMO DO GNV")

    print("=" * 75)

    print()

    print(
        f"Massa molar média : "
        f"{calcular_massa_molar_mistura():.6f} kg/mol"
    )

    print(
        f"Densidade........ : "
        f"{propriedade('densidade'):.3f} kg/m³"
    )

    print(
        f"PCS............... : "
        f"{propriedade('pcs'):.2f} MJ/m³"
    )

    print(
        f"PCI............... : "
        f"{propriedade('pci'):.2f} MJ/m³"
    )

    print()

    print("=" * 75)

    print()



=============================================================================
PARTE 51
PROPRIEDADES REDUZIDAS
=============================================================================

# =============================================================================
# PARTE 51
# TEMPERATURA REDUZIDA
# =============================================================================

def calcular_temperatura_reduzida(

    temperatura_kelvin,

    temperatura_critica

):
    """
    Calcula a temperatura reduzida.

    Tr = T / Tc
    """

    if temperatura_critica <= 0:

        return 0

    return (

        temperatura_kelvin

        /

        temperatura_critica

    )


# =============================================================================
# PARTE 52
# PRESSÃO REDUZIDA
# =============================================================================

def calcular_pressao_reduzida(

    pressao_bar,

    pressao_critica

):
    """
    Calcula a pressão reduzida.

    Pr = P / Pc
    """

    if pressao_critica <= 0:

        return 0

    return (

        pressao_bar

        /

        pressao_critica

    )



# =============================================================================
# PARTE 53
# CONSTANTE ESPECÍFICA DO GNV
# =============================================================================

def calcular_R_especifico(

    massa_molar

):
    """
    Calcula a constante específica.

    R = 8,314462618 / M
    """

    R_UNIVERSAL = 8.314462618

    if massa_molar <= 0:

        return 0

    return (

        R_UNIVERSAL

        /

        massa_molar

    )



# =============================================================================
# PARTE 54
# MASSA ESPECÍFICA
# =============================================================================

def calcular_massa_especifica(

    massa,

    volume

):
    """
    Calcula a massa específica.

    kg/m³
    """

    if volume <= 0:

        return 0

    return (

        massa

        /

        volume

    )


# =============================================================================
# PARTE 55
# FATOR Z APROXIMADO
# =============================================================================

def calcular_Z_aproximado(

    pressao_reduzida,

    temperatura_reduzida

):
    """
    Calcula uma aproximação do fator Z.

    Esta função será substituída futuramente
    pela Equação de Peng-Robinson ou
    Dranchuk-Abou-Kassem.

    Retorna um fator Z entre aproximadamente
    0,65 e 1,10.
    """

    Z = (

        1

        +

        0.08

        *

        pressao_reduzida

        /

        temperatura_reduzida

    )

    if Z < 0.65:

        Z = 0.65

    if Z > 1.10:

        Z = 1.10

    return Z


# =============================================================================
# PARTE 56
# CÁLCULO AUTOMÁTICO DE Tr E Pr
# =============================================================================

def calcular_propriedades_reduzidas(

    pressao_absoluta,

    temperatura_kelvin

):

    propriedades = obter_propriedades_gnv()

    Tr = calcular_temperatura_reduzida(

        temperatura_kelvin,

        propriedades["temperatura_critica"]

    )

    Pr = calcular_pressao_reduzida(

        pressao_absoluta,

        propriedades["pressao_critica"]

    )

    return {

        "Tr": Tr,

        "Pr": Pr

    }



# =============================================================================
# PARTE 57
# CÁLCULO COMPLETO DO FATOR Z
# =============================================================================

def calcular_Z(

    pressao_absoluta,

    temperatura_kelvin

):

    reduzidas = calcular_propriedades_reduzidas(

        pressao_absoluta,

        temperatura_kelvin

    )

    Z = calcular_Z_aproximado(

        reduzidas["Pr"],

        reduzidas["Tr"]

    )

    return {

        "Z": Z,

        "Pr": reduzidas["Pr"],

        "Tr": reduzidas["Tr"]

    }


# =============================================================================
# PARTE 83
# CÁLCULO DA MASSA DO GÁS REAL
# =============================================================================

def calcular_massa_gas_real(
    pressao_absoluta_bar,
    volume_m3,
    temperatura_kelvin,
    fator_z,
    massa_molar
):
    """
    Calcula a massa do GNV utilizando
    a equação dos gases reais.

    Entradas
    --------
    pressao_absoluta_bar : Pressão absoluta (bar)
    volume_m3            : Volume (m³)
    temperatura_kelvin   : Temperatura (K)
    fator_z              : Fator de compressibilidade
    massa_molar          : Massa molar (kg/mol)

    Retorna
    -------
    Massa em kg
    """

    R = 8.314462618

    pressao_pa = pressao_absoluta_bar * 100000

    massa = (
        pressao_pa *
        volume_m3 *
        massa_molar
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return massa


# =============================================================================
# PARTE 84
# CÁLCULO DO NÚMERO DE MOLS
# =============================================================================

def calcular_numero_mols(
    massa_kg,
    massa_molar
):
    """
    Calcula a quantidade de matéria (mol).

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    massa_molar : Massa molar (kg/mol)

    Retorno
    -------
    Quantidade de mols.
    """

    if massa_molar <= 0:

        return 0.0

    numero_mols = (

        massa_kg

        /

        massa_molar

    )

    return numero_mols


# =============================================================================
# PARTE 85
# CÁLCULO DA DENSIDADE DO GÁS REAL
# =============================================================================

def calcular_densidade_gas_real(
    pressao_absoluta_bar,
    temperatura_kelvin,
    fator_z,
    massa_molar
):
    """
    Calcula a densidade do GNV utilizando
    a equação dos gases reais.

    Entradas
    --------
    pressao_absoluta_bar : Pressão absoluta (bar)

    temperatura_kelvin : Temperatura (K)

    fator_z : Fator de compressibilidade

    massa_molar : Massa molar (kg/mol)

    Retorna
    -------
    Densidade em kg/m³
    """

    R = 8.314462618

    pressao_pa = pressao_absoluta_bar * 100000

    if (
        temperatura_kelvin <= 0
        or fator_z <= 0
        or massa_molar <= 0
    ):

        return 0.0

    densidade = (
        pressao_pa *
        massa_molar
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return densidade


# =============================================================================
# PARTE 86
# CÁLCULO DO VOLUME ESPECÍFICO DO GÁS REAL
# =============================================================================

def calcular_volume_especifico_gas_real(
    massa_kg,
    volume_m3
):
    """
    Calcula o volume específico do GNV.

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    volume_m3 : Volume ocupado pelo gás (m³)

    Retorna
    -------
    Volume específico em m³/kg
    """

    if massa_kg <= 0:

        return 0.0

    volume_especifico = (

        volume_m3

        /

        massa_kg

    )

    return volume_especifico



# =============================================================================
# PARTE 87
# CÁLCULO DO VOLUME REAL DO GNV
# =============================================================================

def calcular_volume_real_gnv(
    massa_kg,
    densidade_kg_m3
):
    """
    Calcula o volume realmente ocupado
    pelo GNV.

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    densidade_kg_m3 : Densidade (kg/m³)

    Retorna
    -------
    Volume em m³
    """

    if densidade_kg_m3 <= 0:

        return 0.0

    volume = (

        massa_kg

        /

        densidade_kg_m3

    )

    return volume




# =============================================================================
# PARTE 88
# CALCULA A QUANTIDADE DE GNV NO CILINDRO
# =============================================================================

def calcular_quantidade_gnv(
    volume_cilindro_litros,
    pressao_bar,
    temperatura_c,
    altitude_m,
    fator_z,
    massa_molar
):
    """
    Calcula automaticamente diversas propriedades
    do GNV armazenado no cilindro.

    Retorna um dicionário contendo todos os resultados.
    """

    # ---------------------------------------------------------
    # Conversões
    # ---------------------------------------------------------

    volume_m3 = volume_cilindro_litros / 1000.0

    temperatura_k = temperatura_c + 273.15

    pressao_atm = calcular_pressao_atmosferica(
        altitude_m
    )

    pressao_abs = pressao_bar + pressao_atm

    # ---------------------------------------------------------
    # Massa
    # ---------------------------------------------------------

    massa = calcular_massa_gas_real(
        pressao_abs,
        volume_m3,
        temperatura_k,
        fator_z,
        massa_molar
    )

    # ---------------------------------------------------------
    # Densidade
    # ---------------------------------------------------------

    densidade = calcular_densidade_gas_real(
        pressao_abs,
        temperatura_k,
        fator_z,
        massa_molar
    )

    # ---------------------------------------------------------
    # Número de mols
    # ---------------------------------------------------------

    mols = calcular_numero_mols(
        massa,
        massa_molar
    )

    # ---------------------------------------------------------
    # Volume específico
    # ---------------------------------------------------------

    volume_especifico = calcular_volume_especifico_gas_real(
        massa,
        volume_m3
    )

    # ---------------------------------------------------------
    # Volume real
    # ---------------------------------------------------------

    volume_real = calcular_volume_real_gnv(
        massa,
        densidade
    )

    # ---------------------------------------------------------
    # Retorno
    # ---------------------------------------------------------

    return {

        "volume_cilindro_l": volume_cilindro_litros,

        "volume_cilindro_m3": volume_m3,

        "temperatura_c": temperatura_c,

        "temperatura_k": temperatura_k,

        "pressao_bar": pressao_bar,

        "pressao_atm": pressao_atm,

        "pressao_absoluta": pressao_abs,

        "massa": massa,

        "densidade": densidade,

        "mols": mols,

        "volume_especifico": volume_especifico,

        "volume_real": volume_real

    }


# =============================================================================
# PARTE 89
# COMPARAÇÃO DO ABASTECIMENTO
# =============================================================================

def comparar_abastecimento(
    volume_informado_m3,
    volume_calculado_m3,
    massa_informada_kg=0.0,
    massa_calculada_kg=0.0
):
    """
    Compara os valores informados pelo posto
    com os valores calculados pelo programa.

    Retorna um dicionário contendo as diferenças.
    """

    diferenca_volume = (
        volume_informado_m3 -
        volume_calculado_m3
    )

    if volume_calculado_m3 > 0:

        erro_volume = (
            diferenca_volume /
            volume_calculado_m3
        ) * 100

    else:

        erro_volume = 0.0

    diferenca_massa = (
        massa_informada_kg -
        massa_calculada_kg
    )

    if massa_calculada_kg > 0:

        erro_massa = (
            diferenca_massa /
            massa_calculada_kg
        ) * 100

    else:

        erro_massa = 0.0

    return {

        "volume_informado": volume_informado_m3,

        "volume_calculado": volume_calculado_m3,

        "diferenca_volume": diferenca_volume,

        "erro_volume_percentual": erro_volume,

        "massa_informada": massa_informada_kg,

        "massa_calculada": massa_calculada_kg,

        "diferenca_massa": diferenca_massa,

        "erro_massa_percentual": erro_massa

    }


# =============================================================================
# PARTE 90
# CÁLCULO DA ENERGIA TOTAL DO GNV
# =============================================================================

def calcular_energia_gnv(
    massa_kg,
    pcs_mj_kg=50.02,
    pci_mj_kg=45.00
):
    """
    Calcula a energia contida no GNV.

    Entradas
    --------
    massa_kg : Massa do GNV (kg)

    pcs_mj_kg : Poder Calorífico Superior (MJ/kg)

    pci_mj_kg : Poder Calorífico Inferior (MJ/kg)

    Retorna
    -------
    Dicionário contendo a energia.
    """

    energia_pcs = massa_kg * pcs_mj_kg

    energia_pci = massa_kg * pci_mj_kg

    energia_pcs_kwh = energia_pcs / 3.6

    energia_pci_kwh = energia_pci / 3.6

    return {

        "PCS_MJ": energia_pcs,

        "PCI_MJ": energia_pci,

        "PCS_kWh": energia_pcs_kwh,

        "PCI_kWh": energia_pci_kwh

    }



# =============================================================================
# PARTE 91
# CÁLCULO COMPLETO DO CONSUMO DO VEÍCULO
# =============================================================================

def calcular_consumo_veiculo(
    distancia_km,
    volume_m3,
    massa_kg,
    valor_abastecimento,
    energia_mj
):
    """
    Calcula diversos indicadores de consumo.

    Entradas
    --------
    distancia_km : Distância percorrida (km)

    volume_m3 : Volume abastecido (m³)

    massa_kg : Massa abastecida (kg)

    valor_abastecimento : Valor pago (R$)

    energia_mj : Energia do abastecimento (MJ)

    Retorna
    -------
    Dicionário contendo os indicadores.
    """

    if distancia_km <= 0:

        return None

    if volume_m3 > 0:

        km_m3 = distancia_km / volume_m3

    else:

        km_m3 = 0.0

    if massa_kg > 0:

        km_kg = distancia_km / massa_kg

    else:

        km_kg = 0.0

    if energia_mj > 0:

        km_mj = distancia_km / energia_mj

        mj_km = energia_mj / distancia_km

    else:

        km_mj = 0.0

        mj_km = 0.0

    custo_km = valor_abastecimento / distancia_km

    custo_m3 = 0.0

    if volume_m3 > 0:

        custo_m3 = valor_abastecimento / volume_m3

    custo_kg = 0.0

    if massa_kg > 0:

        custo_kg = valor_abastecimento / massa_kg

    custo_mj = 0.0

    if energia_mj > 0:

        custo_mj = valor_abastecimento / energia_mj

    return {

        "km_m3": km_m3,

        "km_kg": km_kg,

        "km_MJ": km_mj,

        "MJ_km": mj_km,

        "custo_km": custo_km,

        "custo_m3": custo_m3,

        "custo_kg": custo_kg,

        "custo_MJ": custo_mj

    }


# =============================================================================
# PARTE 92
# CLASSE ABASTECIMENTO
# =============================================================================

class Abastecimento:

    """
    Armazena todas as informações de um abastecimento.
    """

    def __init__(

        self,

        data,

        posto,

        cidade,

        odometro,

        volume_m3,

        preco_m3,

        temperatura,

        pressao,

        altitude

    ):

        self.data = data

        self.posto = posto

        self.cidade = cidade

        self.odometro = odometro

        self.volume_m3 = volume_m3

        self.preco_m3 = preco_m3

        self.temperatura = temperatura

        self.pressao = pressao

        self.altitude = altitude

        self.valor_total = (

            volume_m3 *

            preco_m3

        )


# =============================================================================
# PARTE 93
# RESUMO DO ABASTECIMENTO
# =============================================================================

    def resumo(self):

        print()

        print("=" * 75)

        print("DADOS DO ABASTECIMENTO")

        print("=" * 75)

        print()

        print(f"Data...............: {self.data}")

        print(f"Posto..............: {self.posto}")

        print(f"Cidade.............: {self.cidade}")

        print(f"Odômetro...........: {self.odometro:.1f} km")

        print(f"Volume.............: {self.volume_m3:.3f} m³")

        print(f"Preço por m³.......: R$ {self.preco_m3:.3f}")

        print(f"Valor Total........: R$ {self.valor_total:.2f}")

        print(f"Temperatura........: {self.temperatura:.1f} °C")

        print(f"Pressão............: {self.pressao:.1f} bar")

        print(f"Altitude...........: {self.altitude:.1f} m")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 94
# EXPORTAÇÃO PARA DICIONÁRIO
# =============================================================================

    def to_dict(self):

        return {

            "data": self.data,

            "posto": self.posto,

            "cidade": self.cidade,

            "odometro": self.odometro,

            "volume_m3": self.volume_m3,

            "preco_m3": self.preco_m3,

            "valor_total": self.valor_total,

            "temperatura": self.temperatura,

            "pressao": self.pressao,

            "altitude": self.altitude

        }


# =============================================================================
# PARTE 95
# HISTÓRICO DE ABASTECIMENTOS
# =============================================================================

class HistoricoAbastecimentos:

    """
    Armazena vários abastecimentos.
    """

    def __init__(self):

        self.abastecimentos = []


# =============================================================================
# PARTE 96
# ADICIONAR ABASTECIMENTO
# =============================================================================

    def adicionar(

        self,

        abastecimento

    ):

        self.abastecimentos.append(

            abastecimento

        )




# =============================================================================
# PARTE 97
# TOTAL DE ABASTECIMENTOS
# =============================================================================

    def quantidade(self):

        return len(

            self.abastecimentos

        )


# =============================================================================
# PARTE 98
# LISTAR ABASTECIMENTOS
# =============================================================================

    def listar(self):

        print()

        print("=" * 75)

        print("HISTÓRICO DE ABASTECIMENTOS")

        print("=" * 75)

        print()

        for indice, abastecimento in enumerate(

            self.abastecimentos,

            start=1

        ):

            print(

                f"{indice:03d} - "

                f"{abastecimento.data} - "

                f"{abastecimento.posto} - "

                f"{abastecimento.volume_m3:.3f} m³"

            )

        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 99
# PROCURAR ABASTECIMENTO POR DATA
# =============================================================================

    def procurar_data(
        self,
        data
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.data == data:

                resultados.append(

                    abastecimento

                )

        return resultados


# =============================================================================
# PARTE 100
# PROCURAR ABASTECIMENTO POR POSTO
# =============================================================================

    def procurar_posto(
        self,
        posto
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.posto.lower() == posto.lower():

                resultados.append(

                    abastecimento

                )

        return resultados



# =============================================================================
# PARTE 101
# PROCURAR ABASTECIMENTO POR CIDADE
# =============================================================================

    def procurar_cidade(
        self,
        cidade
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.cidade.lower() == cidade.lower():

                resultados.append(

                    abastecimento

                )

        return resultados



# =============================================================================
# PARTE 102
# VALOR TOTAL GASTO
# =============================================================================

    def valor_total(self):

        total = 0.0

        for abastecimento in self.abastecimentos:

            total += abastecimento.valor_total

        return total


# =============================================================================
# PARTE 103
# VOLUME TOTAL ABASTECIDO
# =============================================================================

    def volume_total(self):

        total = 0.0

        for abastecimento in self.abastecimentos:

            total += abastecimento.volume_m3

        return total


# =============================================================================
# PARTE 104
# MÉDIA DO PREÇO DO m³
# =============================================================================

    def media_preco_m3(self):

        if len(self.abastecimentos) == 0:

            return 0.0

        soma = 0.0

        for abastecimento in self.abastecimentos:

            soma += abastecimento.preco_m3

        return soma / len(self.abastecimentos)


# =============================================================================
# PARTE 105
# MÉDIA DO VOLUME ABASTECIDO
# =============================================================================

    def media_volume(self):

        if len(self.abastecimentos) == 0:

            return 0.0

        return self.volume_total() / len(self.abastecimentos)


# =============================================================================
# PARTE 106
# MAIOR ABASTECIMENTO
# =============================================================================

    def maior_abastecimento(self):

        if len(self.abastecimentos) == 0:

            return None

        maior = self.abastecimentos[0]

        for abastecimento in self.abastecimentos:

            if abastecimento.volume_m3 > maior.volume_m3:

                maior = abastecimento

        return maior


# =============================================================================
# PARTE 107
# MENOR ABASTECIMENTO
# =============================================================================

    def menor_abastecimento(self):

        if len(self.abastecimentos) == 0:

            return None

        menor = self.abastecimentos[0]

        for abastecimento in self.abastecimentos:

            if abastecimento.volume_m3 < menor.volume_m3:

                menor = abastecimento

        return menor


# =============================================================================
# PARTE 108
# RELATÓRIO GERAL
# =============================================================================

    def relatorio(self):

        print()

        print("=" * 75)

        print("RELATÓRIO DO HISTÓRICO")

        print("=" * 75)

        print()

        print(f"Quantidade...............: {self.quantidade()}")

        print(f"Volume Total.............: {self.volume_total():.3f} m³")

        print(f"Valor Total..............: R$ {self.valor_total():.2f}")

        print(f"Média do Volume..........: {self.media_volume():.3f} m³")

        print(f"Média do Preço...........: R$ {self.media_preco_m3():.3f}")

        maior = self.maior_abastecimento()

        if maior is not None:

            print()

            print("Maior abastecimento")

            print("-------------------")

            print(f"Data.....................: {maior.data}")

            print(f"Posto....................: {maior.posto}")

            print(f"Volume...................: {maior.volume_m3:.3f} m³")

        menor = self.menor_abastecimento()

        if menor is not None:

            print()

            print("Menor abastecimento")

            print("-------------------")

            print(f"Data.....................: {menor.data}")

            print(f"Posto....................: {menor.posto}")

            print(f"Volume...................: {menor.volume_m3:.3f} m³")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 109
# EXPORTAR HISTÓRICO
# =============================================================================

    def exportar_lista(self):

        """
        Retorna todos os abastecimentos
        em formato de lista de dicionários.
        """

        lista = []

        for abastecimento in self.abastecimentos:

            lista.append(

                abastecimento.to_dict()

            )

        return lista


# =============================================================================
# PARTE 110
# LIMPAR HISTÓRICO
# =============================================================================

    def limpar(self):

        """
        Remove todos os abastecimentos
        da memória.
        """

        self.abastecimentos.clear()


# =============================================================================
# PARTE 111
# ÚLTIMO ABASTECIMENTO
# =============================================================================

    def ultimo(self):

        """
        Retorna o último abastecimento.
        """

        if len(self.abastecimentos) == 0:

            return None

        return self.abastecimentos[-1]


# =============================================================================
# PARTE 112
# PRIMEIRO ABASTECIMENTO
# =============================================================================

    def primeiro(self):

        """
        Retorna o primeiro abastecimento.
        """

        if len(self.abastecimentos) == 0:

            return None

        return self.abastecimentos[0]



# =============================================================================
# PARTE 113
# TOTAL DE QUILÔMETROS
# =============================================================================

    def quilometragem_total(self):

        """
        Calcula a quilometragem entre
        o primeiro e o último abastecimento.
        """

        if len(self.abastecimentos) < 2:

            return 0.0

        primeiro = self.primeiro()

        ultimo = self.ultimo()

        return (

            ultimo.odometro

            -

            primeiro.odometro

        )



# =============================================================================
# PARTE 114
# BANCO DE DADOS SQLITE
# =============================================================================

class BancoGNV:

    """
    Classe responsável pelo banco SQLite.
    """

    def __init__(

        self,

        nome_banco="gnv.db"

    ):

        self.nome_banco = nome_banco

        self.conexao = None

        self.cursor = None



# =============================================================================
# PARTE 115
# CONECTAR AO SQLITE
# =============================================================================

    def conectar(self):

        self.conexao = sqlite3.connect(

            self.nome_banco

        )

        self.cursor = self.conexao.cursor()



# =============================================================================
# PARTE 116
# FECHAR SQLITE
# =============================================================================

    def fechar(self):

        if self.conexao:

            self.conexao.close()



# =============================================================================
# PARTE 117
# CRIAR TABELA
# =============================================================================

    def criar_tabela(self):

        self.cursor.execute("""

        CREATE TABLE IF NOT EXISTS abastecimentos(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            data TEXT,

            posto TEXT,

            cidade TEXT,

            odometro REAL,

            volume REAL,

            preco REAL,

            valor REAL,

            temperatura REAL,

            pressao REAL,

            altitude REAL

        )

        """)

        self.conexao.commit()



# =============================================================================
# PARTE 118
# SALVAR ABASTECIMENTO
# =============================================================================

    def salvar_abastecimento(
        self,
        abastecimento
    ):

        self.cursor.execute(

            """

            INSERT INTO abastecimentos(

                data,

                posto,

                cidade,

                odometro,

                volume,

                preco,

                valor,

                temperatura,

                pressao,

                altitude

            )

            VALUES(

                ?,?,?,?,?,?,?,?,?,?

            )

            """,

            (

                abastecimento.data,

                abastecimento.posto,

                abastecimento.cidade,

                abastecimento.odometro,

                abastecimento.volume_m3,

                abastecimento.preco_m3,

                abastecimento.valor_total,

                abastecimento.temperatura,

                abastecimento.pressao,

                abastecimento.altitude

            )

        )

        self.conexao.commit()



# =============================================================================
# PARTE 119
# LISTAR TODOS OS ABASTECIMENTOS
# =============================================================================

    def listar_abastecimentos(self):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            ORDER BY id

            """

        )

        return self.cursor.fetchall()


# =============================================================================
# PARTE 120
# EXCLUIR ABASTECIMENTO
# =============================================================================

    def excluir_abastecimento(
        self,
        id_abastecimento
    ):

        self.cursor.execute(

            """

            DELETE FROM abastecimentos

            WHERE id = ?

            """,

            (

                id_abastecimento,

            )

        )

        self.conexao.commit()



# =============================================================================
# PARTE 121
# PESQUISAR ABASTECIMENTO
# =============================================================================

    def buscar_por_id(
        self,
        id_abastecimento
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE id = ?

            """,

            (

                id_abastecimento,

            )

        )

        return self.cursor.fetchone()


# =============================================================================
# PARTE 122
# ATUALIZAR ABASTECIMENTO
# =============================================================================

    def atualizar_abastecimento(
        self,
        id_abastecimento,
        abastecimento
    ):

        self.cursor.execute(

            """

            UPDATE abastecimentos

            SET

                data = ?,

                posto = ?,

                cidade = ?,

                odometro = ?,

                volume = ?,

                preco = ?,

                valor = ?,

                temperatura = ?,

                pressao = ?,

                altitude = ?

            WHERE id = ?

            """,

            (

                abastecimento.data,

                abastecimento.posto,

                abastecimento.cidade,

                abastecimento.odometro,

                abastecimento.volume_m3,

                abastecimento.preco_m3,

                abastecimento.valor_total,

                abastecimento.temperatura,

                abastecimento.pressao,

                abastecimento.altitude,

                id_abastecimento

            )

        )

        self.conexao.commit()


# =============================================================================
# PARTE 123
# PESQUISAR POR POSTO
# =============================================================================

    def buscar_por_posto(
        self,
        posto
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE posto LIKE ?

            ORDER BY data

            """,

            (

                "%" + posto + "%",

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 124
# PESQUISAR POR CIDADE
# =============================================================================

    def buscar_por_cidade(
        self,
        cidade
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE cidade LIKE ?

            ORDER BY data

            """,

            (

                "%" + cidade + "%",

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 125
# CONTAR ABASTECIMENTOS
# =============================================================================

    def contar_registros(self):

        self.cursor.execute(

            """

            SELECT COUNT(*)

            FROM abastecimentos

            """

        )

        return self.cursor.fetchone()[0]



# =============================================================================
# PARTE 126
# BUSCAR ENTRE DUAS DATAS
# =============================================================================

    def buscar_periodo(
        self,
        data_inicial,
        data_final
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE data BETWEEN ? AND ?

            ORDER BY data

            """,

            (

                data_inicial,

                data_final

            )

        )

        return self.cursor.fetchall()


# =============================================================================
# PARTE 127
# BUSCAR POR ODÔMETRO
# =============================================================================

    def buscar_odometro(
        self,
        km_inicial,
        km_final
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE odometro BETWEEN ? AND ?

            ORDER BY odometro

            """,

            (

                km_inicial,

                km_final

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 128
# CRIAR ÍNDICES
# =============================================================================

    def criar_indices(self):

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_data

            ON abastecimentos(data)

            """

        )

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_posto

            ON abastecimentos(posto)

            """

        )

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_odometro

            ON abastecimentos(odometro)

            """

        )

        self.conexao.commit()



# =============================================================================
# PARTE 129
# APAGAR TODOS OS REGISTROS
# =============================================================================

    def apagar_todos(self):

        self.cursor.execute(

            """

            DELETE FROM abastecimentos

            """

        )

        self.conexao.commit()


# =============================================================================
# PARTE 130
# TOTAL GASTO
# =============================================================================

    def total_gasto(self):

        self.cursor.execute(

            """

            SELECT SUM(valor)

            FROM abastecimentos

            """

        )

        resultado = self.cursor.fetchone()

        if resultado[0] is None:

            return 0.0

        return resultado[0]



# =============================================================================
# PARTE 401
# CLASSE PDF
# =============================================================================

class RelatorioPDF(FPDF):

    def header(self):

        self.set_font(

            "Arial",

            "B",

            14

        )

        self.cell(

            0,

            10,

            "RELATÓRIO DO SISTEMA GNV",

            0,

            1,

            "C"

        )

        self.ln(

            5

        )


# =============================================================================
# PARTE 402
# RODAPÉ
# =============================================================================

    def footer(self):

        self.set_y(

            -15

        )

        self.set_font(

            "Arial",

            "I",

            8

        )

        self.cell(

            0,

            10,

            f"Página {self.page_no()}",

            0,

            0,

            "C"

        )


# =============================================================================
# PARTE 403
# TÍTULO
# =============================================================================

    def titulo(

        self,

        texto

    ):

        self.set_font(

            "Arial",

            "B",

            12

        )

        self.cell(

            0,

            8,

            texto,

            0,

            1

        )



# =============================================================================
# PARTE 404
# LINHA
# =============================================================================

    def linha(

        self,

        titulo,

        valor

    ):

        self.set_font(

            "Arial",

            "",

            10

        )

        self.cell(

            70,

            7,

            titulo

        )

        self.cell(

            0,

            7,

            str(

                valor

            ),

            ln=True

        )


# =============================================================================
# PARTE 405
# FIM DA CLASSE PDF
# =============================================================================








# =============================================================================
# PARTE 131
# INTERFACE GRÁFICA
# =============================================================================

class InterfaceGNV:

    """
    Interface gráfica principal do sistema.
    """

    def __init__(

        self,

        janela

    ):

        self.janela = janela

        self.janela.title(

            "Sistema de Cálculo de GNV"

        )

        self.janela.geometry(

            "1400x850"

        )

        self.janela.minsize(

            1200,

            700

        )




# =============================================================================
# PARTE 132
# NOTEBOOK
# =============================================================================

        self.notebook = ttk.Notebook(

            self.janela

        )

        self.notebook.pack(

            fill="both",

            expand=True
        )



# =============================================================================
# PARTE 133
# ABAS
# =============================================================================

        self.aba_calculos = ttk.Frame(

            self.notebook

        )

        self.aba_abastecimentos = ttk.Frame(

            self.notebook

        )

        self.aba_historico = ttk.Frame(

            self.notebook

        )

        self.aba_sqlite = ttk.Frame(

            self.notebook

        )

        self.aba_excel = ttk.Frame(

            self.notebook

        )

        self.aba_graficos = ttk.Frame(

            self.notebook

        )

        self.aba_configuracoes = ttk.Frame(

            self.notebook

        )



# =============================================================================
# PARTE 254
# ABA ESTATÍSTICAS
# =============================================================================

        self.aba_estatisticas = ttk.Frame(

            self.notebook

        )

        self.notebook.add(

            self.aba_estatisticas,

            text="Estatísticas"

        )


# =============================================================================
# PARTE 255
# FRAME ESTATÍSTICAS
# =============================================================================

        self.frame_estatisticas = ttk.LabelFrame(

            self.aba_estatisticas,

            text="Resumo Geral"

        )

        self.frame_estatisticas.pack(

            fill="both",

            expand=True,

            padx=10,

            pady=10

        )


# =============================================================================
# PARTE 256
# TÍTULO
# =============================================================================

        self.label_estatisticas = ttk.Label(

            self.frame_estatisticas,

            text="Nenhuma estatística calculada.",

            font=(

                "Segoe UI",

                10

            )

        )

        self.label_estatisticas.pack(

            pady=20

        )



# =============================================================================
# PARTE 257
# BOTÃO ATUALIZAR
# =============================================================================

        self.botao_estatisticas = ttk.Button(

            self.frame_estatisticas,

            text="Atualizar Estatísticas",

            command=self.atualizar_estatisticas

        )

        self.botao_estatisticas.pack(

            pady=10

        )


# =============================================================================
# PARTE 258
# ATUALIZAR ESTATÍSTICAS
# =============================================================================

    def atualizar_estatisticas(self):

        registros = self.banco.listar_abastecimentos()

        total = len(

            registros

        )

        self.label_estatisticas.configure(

            text=f"Total de abastecimentos: {total}"

        )


# =============================================================================
# PARTE 271
# VARIÁVEIS
# =============================================================================

        total_gasto = 0.0

        total_volume = 0.0

        total_preco = 0.0

        maior_volume = 0.0

        menor_volume = None

        postos = set()

        cidades = set()


# =============================================================================
# PARTE 272
# PERCORRE REGISTROS
# =============================================================================

        for registro in registros:


# =============================================================================
# PARTE 276
# CÁLCULO DAS MÉDIAS
# =============================================================================

        if registros:

            media_preco = total_preco / len(registros)

            media_volume = total_volume / len(registros)

        else:

            media_preco = 0

            media_volume = 0

# =============================================================================
# PARTE 277
# TOTAL GASTO
# =============================================================================

        self.lbl_total_gasto.configure(

            text=f"Total Gasto: R$ {total_gasto:,.2f}"

        )


# =============================================================================
# PARTE 278
# TOTAL M3
# =============================================================================

        self.lbl_total_m3.configure(

            text=f"Total Abastecido: {total_volume:.2f} m³"

        )



# =============================================================================
# PARTE 279
# MÉDIA PREÇO
# =============================================================================

        self.lbl_media_preco.configure(

            text=f"Preço Médio: R$ {media_preco:.2f}"

        )


# =============================================================================
# PARTE 280
# MÉDIA VOLUME
# =============================================================================

        self.lbl_media_volume.configure(

            text=f"Volume Médio: {media_volume:.2f} m³"

        )


# =============================================================================
# PARTE 281
# MELHOR ABASTECIMENTO
# =============================================================================

        self.lbl_melhor.configure(

            text=f"Maior abastecimento: {maior_volume:.2f} m³"

        )


# =============================================================================
# PARTE 282
# MENOR ABASTECIMENTO
# =============================================================================

        if menor_volume is None:

            menor_volume = 0

        self.lbl_menor.configure(

            text=f"Menor abastecimento: {menor_volume:.2f} m³"

        )


# =============================================================================
# PARTE 283
# POSTOS
# =============================================================================

        self.lbl_postos.configure(

            text=f"Postos cadastrados: {len(postos)}"

        )



# =============================================================================
# PARTE 284
# CIDADES
# =============================================================================

        self.lbl_cidades.configure(

            text=f"Cidades cadastradas: {len(cidades)}"

        )



# =============================================================================
# PARTE 285
# PIOR ABASTECIMENTO
# =============================================================================

        self.lbl_pior.configure(

            text="Pior abastecimento: em desenvolvimento"

        )


# =============================================================================
# PARTE 286
# MAIOR VALOR
# =============================================================================

        maior_valor = 0

        for registro in registros:

            valor = float(

                registro[7]

            )

            if valor > maior_valor:

                maior_valor = valor



# =============================================================================
# PARTE 287
# MENOR VALOR
# =============================================================================

        if registros:

            menor_valor = float(

                registros[0][7]

            )

            for registro in registros:

                valor = float(

                    registro[7]

                )

                if valor < menor_valor:

                    menor_valor = valor

        else:

            menor_valor = 0


# =============================================================================
# PARTE 288
# LABEL MAIOR VALOR
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"Maior valor pago : R$ {maior_valor:.2f}\n"

        )



# =============================================================================
# PARTE 289
# LABEL MENOR VALOR
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"Menor valor pago : R$ {menor_valor:.2f}\n"

        )


# =============================================================================
# PARTE 290
# FINALIZA TEXTO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\nEstatísticas atualizadas com sucesso."

        )


# =============================================================================
# PARTE 291
# LABEL MAIOR VALOR
# =============================================================================

        self.lbl_maior.configure(

            text=f"Maior valor pago: R$ {maior_valor:.2f}"

        )


# =============================================================================
# PARTE 292
# LABEL MENOR VALOR
# =============================================================================

        self.lbl_pior.configure(

            text=f"Menor valor pago: R$ {menor_valor:.2f}"

        )


# =============================================================================
# PARTE 293
# QUANTIDADE DE POSTOS
# =============================================================================

        quantidade_postos = len(

            postos

        )


# =============================================================================
# PARTE 294
# QUANTIDADE DE CIDADES
# =============================================================================

        quantidade_cidades = len(

            cidades

        )


# =============================================================================
# PARTE 295
# ATUALIZA LABELS
# =============================================================================

        self.lbl_postos.configure(

            text=f"Postos cadastrados: {quantidade_postos}"

        )

        self.lbl_cidades.configure(

            text=f"Cidades cadastradas: {quantidade_cidades}"

        )


# =============================================================================
# PARTE 296
# RANKING DE POSTOS
# =============================================================================

        ranking_postos = {}

        for registro in registros:

            posto = registro[2]

            ranking_postos[posto] = ranking_postos.get(

                posto,

                0

            ) + 1


# =============================================================================
# PARTE 297
# POSTO MAIS UTILIZADO
# =============================================================================

        if ranking_postos:

            posto_favorito = max(

                ranking_postos,

                key=ranking_postos.get

            )

        else:

            posto_favorito = "-"


# =============================================================================
# PARTE 298
# RANKING DE CIDADES
# =============================================================================

        ranking_cidades = {}

        for registro in registros:

            cidade = registro[3]

            ranking_cidades[cidade] = ranking_cidades.get(

                cidade,

                0

            ) + 1


# =============================================================================
# PARTE 299
# CIDADE MAIS UTILIZADA
# =============================================================================

        if ranking_cidades:

            cidade_favorita = max(

                ranking_cidades,

                key=ranking_cidades.get

            )

        else:

            cidade_favorita = "-"


# =============================================================================
# PARTE 300
# EXIBIR RANKING
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nPosto mais utilizado : {posto_favorito}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nCidade mais utilizada : {cidade_favorita}"

        )


# =============================================================================
# PARTE 301
# TOTAL POR POSTO
# =============================================================================

        total_por_posto = {}

        for registro in registros:

            posto = registro[2]

            valor = float(

                registro[7]

            )

            total_por_posto[posto] = total_por_posto.get(

                posto,

                0

            ) + valor




# =============================================================================
# PARTE 302
# POSTO COM MAIOR GASTO
# =============================================================================

        if total_por_posto:

            posto_maior_gasto = max(

                total_por_posto,

                key=total_por_posto.get

            )

        else:

            posto_maior_gasto = "-"



# =============================================================================
# PARTE 303
# TOTAL POR CIDADE
# =============================================================================

        total_por_cidade = {}

        for registro in registros:

            cidade = registro[3]

            valor = float(

                registro[7]

            )

            total_por_cidade[cidade] = total_por_cidade.get(

                cidade,

                0

            ) + valor



# =============================================================================
# PARTE 304
# CIDADE COM MAIOR GASTO
# =============================================================================

        if total_por_cidade:

            cidade_maior_gasto = max(

                total_por_cidade,

                key=total_por_cidade.get

            )

        else:

            cidade_maior_gasto = "-"





# =============================================================================
# PARTE 305
# RESUMO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior gasto por posto : {posto_maior_gasto}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior gasto por cidade : {cidade_maior_gasto}"

        )


# =============================================================================
# PARTE 306
# MAIOR VOLUME POR POSTO
# =============================================================================

        volume_por_posto = {}

        for registro in registros:

            posto = registro[2]

            volume = float(

                registro[5]

            )

            volume_por_posto[posto] = volume_por_posto.get(

                posto,

                0

            ) + volume



# =============================================================================
# PARTE 307
# POSTO MAIOR VOLUME
# =============================================================================

        if volume_por_posto:

            posto_maior_volume = max(

                volume_por_posto,

                key=volume_por_posto.get

            )

        else:

            posto_maior_volume = "-"



# =============================================================================
# PARTE 308
# MÉDIA DE VALOR
# =============================================================================

        if registros:

            media_valor = total_gasto / len(

                registros

            )

        else:

            media_valor = 0


# =============================================================================
# PARTE 309
# MÉDIA GERAL
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nValor médio por abastecimento : "

            f"R$ {media_valor:.2f}"

        )



# =============================================================================
# PARTE 310
# POSTO MAIOR VOLUME
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nPosto com maior volume : "

            f"{posto_maior_volume}"

        )


# =============================================================================
# PARTE 311
# ANO MAIS ANTIGO
# =============================================================================

        anos = []

        for registro in registros:

            try:

                ano = int(

                    str(registro[1])[-4:]

                )

                anos.append(

                    ano

                )

            except:

                pass


# =============================================================================
# PARTE 312
# ANO MAIS RECENTE
# =============================================================================

        if anos:

            primeiro_ano = min(

                anos

            )

            ultimo_ano = max(

                anos

            )

        else:

            primeiro_ano = "-"

            ultimo_ano = "-"


# =============================================================================
# PARTE 313
# PERÍODO DOS DADOS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nPeríodo: {primeiro_ano} até {ultimo_ano}"

        )


# =============================================================================
# PARTE 314
# MÉDIA POR POSTO
# =============================================================================

        if quantidade_postos:

            media_posto = len(

                registros

            ) / quantidade_postos

        else:

            media_posto = 0



# =============================================================================
# PARTE 315
# EXIBE MÉDIA
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nAbastecimentos por posto: "

            f"{media_posto:.2f}"

        )


# =============================================================================
# PARTE 316
# MÉDIA POR CIDADE
# =============================================================================

        if quantidade_cidades:

            media_cidade = len(

                registros

            ) / quantidade_cidades

        else:

            media_cidade = 0


# =============================================================================
# PARTE 317
# EXIBE MÉDIA CIDADE
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nAbastecimentos por cidade: "

            f"{media_cidade:.2f}"

        )



# =============================================================================
# PARTE 318
# VOLUME MÉDIO POR POSTO
# =============================================================================

        if quantidade_postos:

            volume_medio_posto = total_volume / quantidade_postos

        else:

            volume_medio_posto = 0



# =============================================================================
# PARTE 319
# EXIBE VOLUME MÉDIO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nVolume médio por posto: "

            f"{volume_medio_posto:.2f} m³"

        )



# =============================================================================
# PARTE 320
# FINALIZA RELATÓRIO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n\n========================================"

        )

        self.texto_estatisticas.insert(

            tk.END,

            "\nFim das estatísticas."

        )


# =============================================================================
# PARTE 321
# LIMPAR TEXTO
# =============================================================================

        self.texto_estatisticas.see(

            tk.END

        )


# =============================================================================
# PARTE 322
# STATUS
# =============================================================================

        self.status.set(

            "Estatísticas atualizadas."

        )



# =============================================================================
# PARTE 323
# REFRESH
# =============================================================================

        self.root.update_idletasks()



# =============================================================================
# PARTE 324
# LOG
# =============================================================================

        print(

            "Estatísticas atualizadas."

        )


# =============================================================================
# PARTE 325
# FIM
# =============================================================================

        return


# =============================================================================
# PARTE 326
# ABASTECIMENTOS POR ANO
# =============================================================================

        abastecimentos_ano = {}

        for registro in registros:

            try:

                ano = str(

                    registro[1]

                )[-4:]

                abastecimentos_ano[ano] = abastecimentos_ano.get(

                    ano,

                    0

                ) + 1

            except:

                pass


# =============================================================================
# PARTE 327
# MOSTRA ANOS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n\nAbastecimentos por ano:\n"

        )

        for ano in sorted(

            abastecimentos_ano.keys()

        ):

            self.texto_estatisticas.insert(

                tk.END,

                f"{ano}: {abastecimentos_ano[ano]}\n"

            )



# =============================================================================
# PARTE 328
# ANO COM MAIS ABASTECIMENTOS
# =============================================================================

        if abastecimentos_ano:

            ano_recorde = max(

                abastecimentos_ano,

                key=abastecimentos_ano.get

            )

        else:

            ano_recorde = "-"



# =============================================================================
# PARTE 329
# MOSTRAR RECORDE
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nAno com mais abastecimentos: {ano_recorde}"

        )


# =============================================================================
# PARTE 330
# ENCERRA ESTATÍSTICAS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n\n=========================================\n"

        )


# =============================================================================
# PARTE 331
# ABASTECIMENTOS POR MÊS
# =============================================================================

        abastecimentos_mes = {}

        for registro in registros:

            try:

                data = str(

                    registro[1]

                )

                mes = data[3:5]

                abastecimentos_mes[mes] = abastecimentos_mes.get(

                    mes,

                    0

                ) + 1

            except:

                pass


# =============================================================================
# PARTE 332
# MOSTRAR POR MÊS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\nAbastecimentos por mês:\n"

        )

        for mes in sorted(

            abastecimentos_mes.keys()

        ):

            self.texto_estatisticas.insert(

                tk.END,

                f"Mês {mes}: {abastecimentos_mes[mes]}\n"

            )


# =============================================================================
# PARTE 333
# KM RODADOS
# =============================================================================

        km_total = 0

        if len(registros) >= 2:

            km_inicial = float(

                registros[0][4]

            )

            km_final = float(

                registros[-1][4]

            )

            km_total = km_final - km_inicial


# =============================================================================
# PARTE 334
# MOSTRAR KM
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nQuilometragem percorrida: "

            f"{km_total:.0f} km"

        )

# =============================================================================
# PARTE 335
# MÉDIA POR KM
# =============================================================================

        if km_total > 0:

            media_km = total_volume / km_total

        else:

            media_km = 0

        self.texto_estatisticas.insert(

            tk.END,

            f"\nConsumo médio: "

            f"{media_km:.4f} m³/km"

        )



# =============================================================================
# PARTE 336
# MAIOR KM
# =============================================================================

        maior_km = 0

        for registro in registros:

            km = float(

                registro[4]

            )

            if km > maior_km:

                maior_km = km



# =============================================================================
# PARTE 337
# MENOR KM
# =============================================================================

        if registros:

            menor_km = float(

                registros[0][4]

            )

            for registro in registros:

                km = float(

                    registro[4]

                )

                if km < menor_km:

                    menor_km = km

        else:

            menor_km = 0




# =============================================================================
# PARTE 338
# EXIBE MAIOR KM
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior hodômetro: {maior_km:.0f} km"

        )



# =============================================================================
# PARTE 339
# EXIBE MENOR KM
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor hodômetro: {menor_km:.0f} km"

        )



# =============================================================================
# PARTE 340
# TOTAL DE POSTOS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nQuantidade de postos: "

            f"{len(postos)}"

        )



# =============================================================================
# PARTE 341
# TOTAL DE CIDADES
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nQuantidade de cidades: "

            f"{len(cidades)}"

        )



# =============================================================================
# PARTE 342
# MÉDIA POR POSTO
# =============================================================================

        if len(postos) > 0:

            media_gasto_posto = total_gasto / len(postos)

        else:

            media_gasto_posto = 0



# =============================================================================
# PARTE 343
# EXIBE MÉDIA POSTO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nGasto médio por posto: "

            f"R$ {media_gasto_posto:.2f}"

        )



# =============================================================================
# PARTE 344
# MÉDIA POR CIDADE
# =============================================================================

        if len(cidades) > 0:

            media_gasto_cidade = total_gasto / len(cidades)

        else:

            media_gasto_cidade = 0


# =============================================================================
# PARTE 345
# EXIBE MÉDIA CIDADE
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nGasto médio por cidade: "

            f"R$ {media_gasto_cidade:.2f}"

        )


# =============================================================================
# PARTE 346
# MÉDIA KM
# =============================================================================

        if len(registros) > 0:

            media_km_abastecimento = km_total / len(registros)

        else:

            media_km_abastecimento = 0


# =============================================================================
# PARTE 347
# EXIBE MÉDIA KM
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nKM médio por abastecimento: "

            f"{media_km_abastecimento:.2f} km"

        )


# =============================================================================
# PARTE 348
# GASTO POR KM
# =============================================================================

        if km_total > 0:

            gasto_km = total_gasto / km_total

        else:

            gasto_km = 0


# =============================================================================
# PARTE 349
# EXIBE GASTO KM
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nGasto médio por km: "

            f"R$ {gasto_km:.4f}"

        )


# =============================================================================
# PARTE 350
# SEPARADOR
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n--------------------------------------------"

        )


# =============================================================================
# PARTE 351
# TOTAL DE DIAS
# =============================================================================

        datas = set()

        for registro in registros:

            datas.add(

                registro[1]

            )



# =============================================================================
# PARTE 352
# MOSTRAR DIAS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nDias com abastecimento: "

            f"{len(datas)}"

        )


# =============================================================================
# PARTE 353
# MÉDIA POR DIA
# =============================================================================

        if len(datas) > 0:

            media_dia = len(

                registros

            ) / len(datas)

        else:

            media_dia = 0


# =============================================================================
# PARTE 354
# EXIBE MÉDIA DIA
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMédia por dia: "

            f"{media_dia:.2f}"

        )


# =============================================================================
# PARTE 355
# FIM DO BLOCO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n============================================\n"

        )



# =============================================================================
# PARTE 356
# TOTAL DE MESES
# =============================================================================

        meses = set()

        for registro in registros:

            try:

                mes = str(

                    registro[1]

                )[3:5]

                meses.add(

                    mes

                )

            except:

                pass




# =============================================================================
# PARTE 357
# MOSTRA MESES
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMeses registrados: "

            f"{len(meses)}"

        )



# =============================================================================
# PARTE 358
# MÉDIA MENSAL
# =============================================================================

        if len(meses) > 0:

            media_mes = len(

                registros

            ) / len(meses)

        else:

            media_mes = 0



# =============================================================================
# PARTE 359
# EXIBE MÉDIA MENSAL
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMédia mensal: "

            f"{media_mes:.2f} abastecimentos"

        )



# =============================================================================
# PARTE 360
# FINAL DO RELATÓRIO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\nRelatório concluído."

        )



# =============================================================================
# PARTE 361
# TOTAL DE ANOS
# =============================================================================

        total_anos = len(

            abastecimentos_ano

        )




# =============================================================================
# PARTE 362
# EXIBE ANOS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nAnos registrados: "

            f"{total_anos}"

        )



# =============================================================================
# PARTE 363
# MÉDIA ANUAL
# =============================================================================

        if total_anos > 0:

            media_anual = len(

                registros

            ) / total_anos

        else:

            media_anual = 0



# =============================================================================
# PARTE 364
# EXIBE MÉDIA ANUAL
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMédia anual: "

            f"{media_anual:.2f} abastecimentos"

        )



# =============================================================================
# PARTE 365
# ENCERRAMENTO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n============================================"

        )


# =============================================================================
# PARTE 366
# MAIOR PREÇO DO M³
# =============================================================================

        maior_preco = 0

        for registro in registros:

            preco = float(

                registro[6]

            )

            if preco > maior_preco:

                maior_preco = preco


# =============================================================================
# PARTE 367
# MENOR PREÇO DO M³
# =============================================================================

        if registros:

            menor_preco = float(

                registros[0][6]

            )

            for registro in registros:

                preco = float(

                    registro[6]

                )

                if preco < menor_preco:

                    menor_preco = preco

        else:

            menor_preco = 0


# =============================================================================
# PARTE 368
# EXIBE MAIOR PREÇO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior preço do m³: "

            f"R$ {maior_preco:.3f}"

        )



# =============================================================================
# PARTE 369
# EXIBE MENOR PREÇO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor preço do m³: "

            f"R$ {menor_preco:.3f}"

        )




# =============================================================================
# PARTE 370
# DIFERENÇA DE PREÇO
# =============================================================================

        diferenca_preco = maior_preco - menor_preco

        self.texto_estatisticas.insert(

            tk.END,

            f"\nDiferença de preço: "

            f"R$ {diferenca_preco:.3f}"

        )



# =============================================================================
# PARTE 371
# SOMA DAS PRESSÕES
# =============================================================================

        soma_pressao = 0

        for registro in registros:

            soma_pressao += float(

                registro[8]

            )



# =============================================================================
# PARTE 372
# MÉDIA DA PRESSÃO
# =============================================================================

        if registros:

            media_pressao = (

                soma_pressao /

                len(registros)

            )

        else:

            media_pressao = 0



# =============================================================================
# PARTE 373
# EXIBE PRESSÃO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nPressão média: "

            f"{media_pressao:.3f} bar"

        )




# =============================================================================
# PARTE 374
# SOMA DAS TEMPERATURAS
# =============================================================================

        soma_temperatura = 0

        for registro in registros:

            soma_temperatura += float(

                registro[9]

            )




# =============================================================================
# PARTE 375
# MÉDIA TEMPERATURA
# =============================================================================

        if registros:

            media_temperatura = (

                soma_temperatura /

                len(registros)

            )

        else:

            media_temperatura = 0

        self.texto_estatisticas.insert(

            tk.END,

            f"\nTemperatura média: "

            f"{media_temperatura:.2f} °C"

        )



# =============================================================================
# PARTE 376
# MAIOR PRESSÃO
# =============================================================================

        maior_pressao = 0

        for registro in registros:

            pressao = float(

                registro[8]

            )

            if pressao > maior_pressao:

                maior_pressao = pressao


# =============================================================================
# PARTE 377
# MENOR PRESSÃO
# =============================================================================

        if registros:

            menor_pressao = float(

                registros[0][8]

            )

            for registro in registros:

                pressao = float(

                    registro[8]

                )

                if pressao < menor_pressao:

                    menor_pressao = pressao

        else:

            menor_pressao = 0



# =============================================================================
# PARTE 378
# EXIBE PRESSÕES
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior pressão: "

            f"{maior_pressao:.3f} bar"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor pressão: "

            f"{menor_pressao:.3f} bar"

        )


# =============================================================================
# PARTE 379
# MAIOR TEMPERATURA
# =============================================================================

        maior_temperatura = max(

            (

                float(r[9])

                for r in registros

            ),

            default=0

        )



# =============================================================================
# PARTE 380
# MENOR TEMPERATURA
# =============================================================================

        menor_temperatura = min(

            (

                float(r[9])

                for r in registros

            ),

            default=0

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior temperatura: "

            f"{maior_temperatura:.2f} °C"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor temperatura: "

            f"{menor_temperatura:.2f} °C"

        )



# =============================================================================
# PARTE 381
# MAIOR ALTITUDE
# =============================================================================

        maior_altitude = max(

            (

                float(r[10])

                for r in registros

            ),

            default=0

        )


# =============================================================================
# PARTE 382
# MENOR ALTITUDE
# =============================================================================

        menor_altitude = min(

            (

                float(r[10])

                for r in registros

            ),

            default=0

        )



# =============================================================================
# PARTE 383
# MÉDIA ALTITUDE
# =============================================================================

        if registros:

            media_altitude = sum(

                float(r[10])

                for r in registros

            ) / len(

                registros

            )

        else:

            media_altitude = 0



# =============================================================================
# PARTE 384
# EXIBE ALTITUDE
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior altitude: "

            f"{maior_altitude:.2f} m"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor altitude: "

            f"{menor_altitude:.2f} m"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nAltitude média: "

            f"{media_altitude:.2f} m"

        )



# =============================================================================
# PARTE 385
# FIM DO BLOCO ALTITUDE
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n--------------------------------------------"

        )


# =============================================================================
# PARTE 386
# MAIOR VALOR POR M³
# =============================================================================

        maior_valor_m3 = max(

            (

                float(r[6])

                for r in registros

            ),

            default=0

        )



# =============================================================================
# PARTE 387
# MENOR VALOR POR M³
# =============================================================================

        menor_valor_m3 = min(

            (

                float(r[6])

                for r in registros

            ),

            default=0

        )


# =============================================================================
# PARTE 388
# MÉDIA DO PREÇO DO M³
# =============================================================================

        if registros:

            media_valor_m3 = sum(

                float(r[6])

                for r in registros

            ) / len(

                registros

            )

        else:

            media_valor_m3 = 0


# =============================================================================
# PARTE 389
# EXIBIR PREÇOS
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior preço do m³: "

            f"R$ {maior_valor_m3:.3f}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor preço do m³: "

            f"R$ {menor_valor_m3:.3f}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nPreço médio do m³: "

            f"R$ {media_valor_m3:.3f}"

        )


# =============================================================================
# PARTE 390
# SEPARADOR
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n============================================"

        )

# =============================================================================
# PARTE 391
# MAIOR VALOR TOTAL
# =============================================================================

        maior_valor_total = max(

            (

                float(r[7])

                for r in registros

            ),

            default=0

        )


# =============================================================================
# PARTE 392
# MENOR VALOR TOTAL
# =============================================================================

        menor_valor_total = min(

            (

                float(r[7])

                for r in registros

            ),

            default=0

        )


# =============================================================================
# PARTE 393
# DIFERENÇA
# =============================================================================

        diferenca = (

            maior_valor_total -

            menor_valor_total

        )


# =============================================================================
# PARTE 394
# EXIBIR VALORES
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMaior valor pago: R$ {maior_valor_total:.2f}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nMenor valor pago: R$ {menor_valor_total:.2f}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nDiferença: R$ {diferenca:.2f}"

        )


# =============================================================================
# PARTE 395
# SEPARADOR
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n--------------------------------------------"

        )


# =============================================================================
# PARTE 396
# PRIMEIRO ABASTECIMENTO
# =============================================================================

        if registros:

            primeira_data = registros[0][1]

        else:

            primeira_data = "-"



# =============================================================================
# PARTE 397
# ÚLTIMO ABASTECIMENTO
# =============================================================================

        if registros:

            ultima_data = registros[-1][1]

        else:

            ultima_data = "-"


# =============================================================================
# PARTE 398
# EXIBIR PERÍODO
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            f"\nPrimeiro abastecimento: {primeira_data}"

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"\nÚltimo abastecimento: {ultima_data}"

        )


# =============================================================================
# PARTE 399
# MENSAGEM FINAL
# =============================================================================

        self.texto_estatisticas.insert(

            tk.END,

            "\n\nRelatório estatístico concluído com sucesso."

        )


# =============================================================================
# PARTE 400
# FIM DO MÓDULO ESTATÍSTICAS
# =============================================================================

        return










# =============================================================================
# PARTE 273
# SOMATÓRIOS
# =============================================================================

            volume = float(

                registro[5]

            )

            valor = float(

                registro[7]

            )

            preco = float(

                registro[6]

            )

            total_volume += volume

            total_gasto += valor

            total_preco += preco


# =============================================================================
# PARTE 274
# MAIOR E MENOR
# =============================================================================

            if volume > maior_volume:

                maior_volume = volume

            if menor_volume is None:

                menor_volume = volume

            elif volume < menor_volume:

                menor_volume = volume


# =============================================================================
# PARTE 275
# POSTOS E CIDADES
# =============================================================================

            postos.add(

                registro[2]

            )

            cidades.add(

                registro[3]

            )


# =============================================================================
# PARTE 134
# ADICIONAR ABAS
# =============================================================================

        self.notebook.add(

            self.aba_calculos,

            text="Cálculos"

        )

        self.notebook.add(

            self.aba_abastecimentos,

            text="Abastecimentos"

        )

        self.notebook.add(

            self.aba_historico,

            text="Histórico"

        )

        self.notebook.add(

            self.aba_sqlite,

            text="Banco SQLite"

        )

        self.notebook.add(

            self.aba_excel,

            text="Excel"

        )

        self.notebook.add(

            self.aba_graficos,

            text="Gráficos"

        )

        self.notebook.add(

            self.aba_configuracoes,

            text="Configurações"

        )



# =============================================================================
# PARTE 177
# INICIALIZA O BANCO DE DADOS
# =============================================================================

        self.banco = BancoGNV()

        self.banco.conectar()

        self.banco.criar_tabela()

        self.banco.criar_indices()


# =============================================================================
# PARTE 178
# CARREGA O HISTÓRICO
# =============================================================================

        self.registros = self.banco.listar_abastecimentos()


# =============================================================================
# PARTE 179
# TOTAL DE REGISTROS
# =============================================================================

        self.total_registros = self.banco.contar_registros()


# =============================================================================
# PARTE 180
# STATUS INICIAL
# =============================================================================

        if hasattr(self, "status"):

            self.status.set(

                f"Sistema iniciado - {self.total_registros} abastecimentos cadastrados."

            )


# =============================================================================
# PARTE 182
# FRAME DO HISTÓRICO
# =============================================================================

        self.frame_historico = ttk.Frame(

            self.aba_historico,

            padding=10

        )

        self.frame_historico.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 255
# FRAME ESTATÍSTICAS
# =============================================================================

        self.frame_estatisticas = ttk.Frame(

            self.aba_estatisticas,

            padding=10

        )

        self.frame_estatisticas.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 256
# LABEL ESTATÍSTICAS
# =============================================================================

        self.label_estatisticas = ttk.Label(

            self.frame_estatisticas,

            text="Resumo Geral",

            font=(

                "Segoe UI",

                12,

                "bold"

            )

        )

        self.label_estatisticas.pack(

            pady=10

        )



# =============================================================================
# PARTE 257
# TEXTO ESTATÍSTICAS
# =============================================================================

        self.texto_estatisticas = tk.Text(

            self.frame_estatisticas,

            width=90,

            height=25

        )

        self.texto_estatisticas.pack(

            fill="both",

            expand=True

        )



# =============================================================================
# PARTE 258
# BOTÃO ATUALIZAR
# =============================================================================

        self.botao_atualizar_estatisticas = ttk.Button(

            self.frame_estatisticas,

            text="Atualizar",

            command=self.atualizar_estatisticas

        )

        self.botao_atualizar_estatisticas.pack(

            pady=10

        )


# =============================================================================
# PARTE 260
# SEPARADOR
# =============================================================================

        ttk.Separator(

            self.frame_estatisticas,

            orient="horizontal"

        ).pack(

            fill="x",

            pady=10

        )


# =============================================================================
# PARTE 261
# TOTAL GASTO
# =============================================================================

        self.lbl_total_gasto = ttk.Label(

            self.frame_estatisticas,

            text="Total Gasto: R$ 0,00",

            font=(

                "Segoe UI",

                10,

                "bold"

            )

        )

        self.lbl_total_gasto.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 262
# TOTAL M3
# =============================================================================

        self.lbl_total_m3 = ttk.Label(

            self.frame_estatisticas,

            text="Total Abastecido: 0,00 m³"

        )

        self.lbl_total_m3.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 263
# MÉDIA PREÇO
# =============================================================================

        self.lbl_media_preco = ttk.Label(

            self.frame_estatisticas,

            text="Preço Médio: R$ 0,00"

        )

        self.lbl_media_preco.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 264
# MÉDIA VOLUME
# =============================================================================

        self.lbl_media_volume = ttk.Label(

            self.frame_estatisticas,

            text="Volume Médio: 0,00 m³"

        )

        self.lbl_media_volume.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 265
# MELHOR ABASTECIMENTO
# =============================================================================

        self.lbl_melhor = ttk.Label(

            self.frame_estatisticas,

            text="Melhor abastecimento:"

        )

        self.lbl_melhor.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 266
# PIOR ABASTECIMENTO
# =============================================================================

        self.lbl_pior = ttk.Label(

            self.frame_estatisticas,

            text="Pior abastecimento:"

        )

        self.lbl_pior.pack(

            anchor="w",

            padx=10

        )



# =============================================================================
# PARTE 267
# MAIOR ABASTECIMENTO
# =============================================================================

        self.lbl_maior = ttk.Label(

            self.frame_estatisticas,

            text="Maior abastecimento:"

        )

        self.lbl_maior.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 268
# MENOR ABASTECIMENTO
# =============================================================================

        self.lbl_menor = ttk.Label(

            self.frame_estatisticas,

            text="Menor abastecimento:"

        )

        self.lbl_menor.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 269
# POSTOS
# =============================================================================

        self.lbl_postos = ttk.Label(

            self.frame_estatisticas,

            text="Postos cadastrados: 0"

        )

        self.lbl_postos.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 270
# CIDADES
# =============================================================================

        self.lbl_cidades = ttk.Label(

            self.frame_estatisticas,

            text="Cidades cadastradas: 0"

        )

        self.lbl_cidades.pack(

            anchor="w",

            padx=10

        )








# =============================================================================
# PARTE 183
# TÍTULO
# =============================================================================

        ttk.Label(

            self.frame_historico,

            text="Histórico de Abastecimentos",

            font=("Arial",14,"bold")

        ).pack(

            pady=10

        )

# =============================================================================
# PARTE 192
# CAMPO DE PESQUISA
# =============================================================================

        self.frame_pesquisa = ttk.Frame(

            self.frame_historico

        )

        self.frame_pesquisa.pack(

            fill="x",

            pady=5

        )

        ttk.Label(

            self.frame_pesquisa,

            text="Pesquisar:"

        ).pack(

            side="left"

        )

        self.entry_pesquisa = ttk.Entry(

            self.frame_pesquisa,

            width=40

        )

        self.entry_pesquisa.pack(

            side="left",

            padx=5

        )

# =============================================================================
# PARTE 193
# BOTÃO PESQUISAR
# =============================================================================

        self.botao_pesquisar = ttk.Button(

            self.frame_pesquisa,

            text="Pesquisar",

            command=self.pesquisar_historico

        )

        self.botao_pesquisar.pack(

            side="left",

            padx=5

        )


# =============================================================================
# PARTE 194
# BOTÃO MOSTRAR TODOS
# =============================================================================

        self.botao_todos = ttk.Button(

            self.frame_pesquisa,

            text="Mostrar Todos",

            command=self.atualizar_historico

        )

        self.botao_todos.pack(

            side="left",

            padx=5

        )


# =============================================================================
# PARTE 196
# ENTER PESQUISA
# =============================================================================

        self.entry_pesquisa.bind(

            "<Return>",

            lambda event: self.pesquisar_historico()

        )

# =============================================================================
# PARTE 199
# DUPLO CLIQUE
# =============================================================================

        self.tree.bind(

            "<Double-1>",

            self.abrir_registro

        )

# =============================================================================
# PARTE 201
# MENU CONTEXTUAL
# =============================================================================

        self.menu_tree = tk.Menu(

            self.tree,

            tearoff=False

        )


# =============================================================================
# PARTE 202
# MENU EDITAR
# =============================================================================

        self.menu_tree.add_command(

            label="Editar"

        )


# =============================================================================
# PARTE 209
# MENU EDITAR
# =============================================================================

        self.menu_tree.add_command(

            label="Editar",

            command=self.editar_registro

        )


# =============================================================================
# PARTE 203
# MENU EXCLUIR
# =============================================================================

        self.menu_tree.add_command(

            label="Excluir"

        )


# =============================================================================
# PARTE 210
# MENU EXCLUIR
# =============================================================================

        self.menu_tree.add_command(

            label="Excluir",

            command=self.excluir_registro

        )



# =============================================================================
# PARTE 204
# BOTÃO DIREITO
# =============================================================================

        self.tree.bind(

            "<Button-3>",

            self.menu_contexto

        )


# =============================================================================
# PARTE 222
# BOTÃO EXPORTAR EXCEL
# =============================================================================

        self.botao_excel = ttk.Button(

            self.frame_historico,

            text="Exportar Excel",

            command=self.exportar_excel

        )

        self.botao_excel.pack(

            side="left",

            padx=5,

            pady=5

        )



# =============================================================================
# PARTE 197
# TOTAL DE REGISTROS
# =============================================================================

        self.label_total = ttk.Label(

            self.frame_historico,

            text="Total: 0 registros"

        )

        self.label_total.pack(

            anchor="e",

            pady=5

        )





# =============================================================================
# PARTE 184
# TREEVIEW
# =============================================================================

        self.tree = ttk.Treeview(

            self.frame_historico,

            columns=(

                "data",

                "posto",

                "cidade",

                "volume",

                "valor"

            ),

            show="headings"

        )


# =============================================================================
# PARTE 185
# CABEÇALHOS
# =============================================================================

        self.tree.heading(

            "data",

            text="Data"

        )

        self.tree.heading(

            "posto",

            text="Posto"

        )

        self.tree.heading(

            "cidade",

            text="Cidade"

        )

        self.tree.heading(

            "volume",

            text="Volume"

        )

        self.tree.heading(

            "valor",

            text="Valor"

        )


# =============================================================================
# PARTE 191
# LARGURA DAS COLUNAS
# =============================================================================

        self.tree.column(

            "data",

            width=120

        )

        self.tree.column(

            "posto",

            width=260

        )

        self.tree.column(

            "cidade",

            width=180

        )

        self.tree.column(

            "volume",

            width=100,

            anchor="center"

        )

        self.tree.column(

            "valor",

            width=120,

            anchor="e"

        )




# =============================================================================
# PARTE 186
# EXIBIR TREEVIEW
# =============================================================================

        self.tree.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 187
# SCROLLBAR DO HISTÓRICO
# =============================================================================

        self.scroll_historico = ttk.Scrollbar(

            self.frame_historico,

            orient="vertical",

            command=self.tree.yview

        )

        self.tree.configure(

            yscrollcommand=self.scroll_historico.set

        )

        self.scroll_historico.pack(

            side="right",

            fill="y"

        )



# =============================================================================
# PARTE 190
# CARREGA O HISTÓRICO
# =============================================================================

        self.atualizar_historico()




# =============================================================================
# PARTE 135
# FRAME PRINCIPAL DA ABA CÁLCULOS
# =============================================================================

        self.frame_calculos = ttk.Frame(

            self.aba_calculos,

            padding=10

        )

        self.frame_calculos.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 136
# VOLUME DO CILINDRO
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Volume do cilindro (L):"

        ).grid(

            row=0,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_volume = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_volume.grid(

            row=0,

            column=1,

            padx=5,

            pady=5

        )

# =============================================================================
# PARTE 137
# QUANTIDADE DE CILINDROS
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Quantidade de cilindros:"

        ).grid(

            row=1,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_quantidade = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_quantidade.grid(

            row=1,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_quantidade.insert(

            0,

            "1"

        )


# =============================================================================
# PARTE 138
# PRESSÃO
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Pressão (bar):"

        ).grid(

            row=2,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_pressao = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_pressao.grid(

            row=2,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_pressao.insert(

            0,

            "200"

        )

# =============================================================================
# PARTE 139
# TEMPERATURA
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Temperatura (°C):"

        ).grid(

            row=3,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_temperatura = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_temperatura.grid(

            row=3,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_temperatura.insert(

            0,

            "20"
        )


# =============================================================================
# PARTE 140
# ALTITUDE
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Altitude (m):"

        ).grid(

            row=4,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_altitude = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_altitude.grid(

            row=4,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_altitude.insert(

            0,

            "0"

        )



# =============================================================================
# PARTE 141
# FATOR Z
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Fator Z:"

        ).grid(

            row=5,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_fator_z = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_fator_z.grid(

            row=5,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_fator_z.insert(

            0,

            "0.92"

        )


# =============================================================================
# PARTE 142
# MASSA MOLAR
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Massa molar (kg/mol):"

        ).grid(

            row=6,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_massa_molar = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_massa_molar.grid(

            row=6,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_massa_molar.insert(

            0,

            "0.01604"

        )



# =============================================================================
# PARTE 143
# BOTÃO CALCULAR
# =============================================================================

        self.botao_calcular = ttk.Button(

            self.frame_calculos,

            text="Calcular"

        )

        self.botao_calcular.grid(

            row=7,

            column=0,

            columnspan=2,

            padx=10,

            pady=15,

            sticky="ew"

        )


# =============================================================================
# PARTE 146
# LIGAR BOTÃO AO MÉTODO
# =============================================================================

        self.botao_calcular.configure(

            command=self.executar_calculo

        )


# =============================================================================
# PARTE 150
# BOTÃO LIMPAR
# =============================================================================

        self.botao_limpar = ttk.Button(

            self.frame_calculos,

            text="Limpar",

            command=self.limpar_resultados

        )

        self.botao_limpar.grid(

            row=7,

            column=2,

            padx=10,

            pady=15,

            sticky="ew"

        )



# =============================================================================
# PARTE 152
# BOTÃO LIMPAR CAMPOS
# =============================================================================

        self.botao_limpar_campos = ttk.Button(

            self.frame_calculos,

            text="Limpar Campos",

            command=self.limpar_campos

        )

        self.botao_limpar_campos.grid(

            row=7,

            column=3,

            padx=10,

            pady=15,

            sticky="ew"

        )


# =============================================================================
# PARTE 406
# BOTÃO PDF
# =============================================================================

        self.botao_pdf = ttk.Button(

            self.aba_excel,

            text="Gerar Relatório PDF",

            command=self.exportar_pdf

        )

        self.botao_pdf.pack(

            pady=10

        )






""" # =============================================================================
# PARTE 144
# ÁREA DE RESULTADOS
# =============================================================================

        self.texto_resultados = tk.Text(

            self.frame_calculos,

            width=90,

            height=25,

            font=("Consolas", 10)

        )

        self.texto_resultados.grid(

            row=8,

            column=0,

            columnspan=2,

            padx=10,

            pady=10,

            sticky="nsew"

        )

        self.frame_calculos.grid_rowconfigure(

            8,

            weight=1

        )

        self.frame_calculos.grid_columnconfigure(

            1,

            weight=1

        ) """



# =============================================================================
# PARTE 157
# FRAME ABA ABASTECIMENTOS
# =============================================================================

        self.frame_abastecimentos = ttk.Frame(

            self.aba_abastecimentos,

            padding=10

        )

        self.frame_abastecimentos.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 158
# TÍTULO
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Cadastro de Abastecimentos",

            font=("Arial",14,"bold")

        ).grid(

            row=0,

            column=0,

            columnspan=2,

            pady=15

        )


# =============================================================================
# PARTE 159
# DATA
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Data"

        ).grid(

            row=1,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_data = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_data.grid(

            row=1,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 160
# POSTO
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Posto"

        ).grid(

            row=2,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_posto = ttk.Entry(

            self.frame_abastecimentos,

            width=40

        )

        self.entry_posto.grid(

            row=2,

            column=1,

            padx=5,

            pady=5

        )



# =============================================================================
# PARTE 161
# CIDADE
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Cidade"

        ).grid(

            row=3,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_cidade = ttk.Entry(

            self.frame_abastecimentos,

            width=40

        )

        self.entry_cidade.grid(

            row=3,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 162
# ODÔMETRO
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Odômetro (km)"

        ).grid(

            row=4,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_odometro = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_odometro.grid(

            row=4,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 163
# VOLUME ABASTECIDO
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Volume Abastecido (m³)"

        ).grid(

            row=5,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_volume_abastecido = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_volume_abastecido.grid(

            row=5,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 164
# PREÇO DO m³
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Preço por m³ (R$)"

        ).grid(

            row=6,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_preco_m3 = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_preco_m3.grid(

            row=6,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 165
# TEMPERATURA DO ABASTECIMENTO
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Temperatura (°C)"

        ).grid(

            row=7,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_temp_abastecimento = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_temp_abastecimento.grid(

            row=7,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 166
# PRESSÃO DO ABASTECIMENTO
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Pressão (bar)"

        ).grid(

            row=8,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_pressao_abastecimento = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_pressao_abastecimento.grid(

            row=8,

            column=1,

            padx=5,

            pady=5

        )



# =============================================================================
# PARTE 167
# ALTITUDE
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Altitude (m)"

        ).grid(

            row=9,

            column=0,

            sticky="w",

            padx=5,

            pady=5

        )

        self.entry_altitude_abastecimento = ttk.Entry(

            self.frame_abastecimentos,

            width=20

        )

        self.entry_altitude_abastecimento.grid(

            row=9,

            column=1,

            padx=5,

            pady=5

        )



# =============================================================================
# PARTE 168
# OBSERVAÇÕES
# =============================================================================

        ttk.Label(

            self.frame_abastecimentos,

            text="Observações"

        ).grid(

            row=10,

            column=0,

            sticky="nw",

            padx=5,

            pady=5

        )

        self.texto_observacoes = tk.Text(

            self.frame_abastecimentos,

            width=45,

            height=5

        )

        self.texto_observacoes.grid(

            row=10,

            column=1,

            padx=5,

            pady=5

        )


# =============================================================================
# PARTE 169
# BOTÃO SALVAR
# =============================================================================

        self.botao_salvar = ttk.Button(

            self.frame_abastecimentos,

            text="Salvar Abastecimento",

            command=self.salvar_abastecimento

        )

        self.botao_salvar.grid(

            row=11,

            column=0,

            padx=10,

            pady=15,

            sticky="ew"

        )


# =============================================================================
# PARTE 170
# BOTÃO NOVO
# =============================================================================

        self.botao_novo = ttk.Button(

            self.frame_abastecimentos,

            text="Novo",

            command=self.novo_abastecimento

        )

        self.botao_novo.grid(

            row=11,

            column=1,

            padx=10,

            pady=15,

            sticky="ew"

        )


# =============================================================================
# PARTE 171
# NOVO ABASTECIMENTO
# =============================================================================

    def novo_abastecimento(self):

        self.entry_data.delete(0, tk.END)

        self.entry_posto.delete(0, tk.END)

        self.entry_cidade.delete(0, tk.END)

        self.entry_odometro.delete(0, tk.END)

        self.entry_volume_abastecido.delete(0, tk.END)

        self.entry_preco_m3.delete(0, tk.END)

        self.entry_temp_abastecimento.delete(0, tk.END)

        self.entry_pressao_abastecimento.delete(0, tk.END)

        self.entry_altitude_abastecimento.delete(0, tk.END)

        self.texto_observacoes.delete(

            "1.0",

            tk.END

        )


# =============================================================================
# PARTE 172
# SALVAR ABASTECIMENTO
# =============================================================================

    def salvar_abastecimento(self):

        try:

            abastecimento = Abastecimento(

                data=self.entry_data.get(),

                posto=self.entry_posto.get(),

                cidade=self.entry_cidade.get(),

                odometro=float(self.entry_odometro.get()),

                volume_m3=float(self.entry_volume_abastecido.get()),

                preco_m3=float(self.entry_preco_m3.get()),

                temperatura=float(self.entry_temp_abastecimento.get()),

                pressao=float(self.entry_pressao_abastecimento.get()),

                altitude=float(self.entry_altitude_abastecimento.get())

            )

        except ValueError:

            messagebox.showerror(

                "Erro",

                "Existem campos numéricos inválidos."

            )

            return

# =============================================================================
# PARTE 173
# SALVAR NO SQLITE
# =============================================================================

        self.banco.salvar_abastecimento(

            abastecimento

        )

# =============================================================================
# PARTE 174
# MENSAGEM DE SUCESSO
# =============================================================================

        messagebox.showinfo(

            "Sucesso",

            "Abastecimento gravado com sucesso."

        )


# =============================================================================
# PARTE 175
# LIMPAR FORMULÁRIO
# =============================================================================

        self.novo_abastecimento()



# =============================================================================
# PARTE 176
# ATUALIZAR BARRA DE STATUS
# =============================================================================

        self.status.set(

            "Abastecimento salvo no banco de dados."

        )


# =============================================================================
# PARTE 189
# ATUALIZA O HISTÓRICO
# =============================================================================

        self.atualizar_historico()




# =============================================================================
# PARTE 181
# FECHAR BANCO
# =============================================================================

    def fechar(self):

        if hasattr(

            self,

            "banco"

        ):

            self.banco.fechar()


# =============================================================================
# PARTE 188
# ATUALIZAR HISTÓRICO
# =============================================================================

    def atualizar_historico(self):

        for item in self.tree.get_children():

            self.tree.delete(

                item

            )

        registros = self.banco.listar_abastecimentos()

        for registro in registros:

            self.tree.insert(

                "",

                "end",

                values=(

                    registro[1],

                    registro[2],

                    registro[3],

                    registro[5],

                    registro[7]

                )

            )

# =============================================================================
# PARTE 198
# ATUALIZA TOTAL
# =============================================================================

        self.label_total.configure(

            text=f"Total: {len(registros)} registros"

        )



# =============================================================================
# PARTE 221
# STATUS
# =============================================================================

        self.status.set(

            f"{len(registros)} registros carregados."

        )


# =============================================================================
# PARTE 223
# EXPORTAR EXCEL
# =============================================================================

    def exportar_excel(self):

    try:

        print(

            "Exportando para Excel..."

        )


# =============================================================================
# PARTE 259
# ATUALIZAR ESTATÍSTICAS
# =============================================================================

    def atualizar_estatisticas(self):

        registros = self.banco.listar_abastecimentos()

        self.texto_estatisticas.delete(

            "1.0",

            tk.END

        )

        self.texto_estatisticas.insert(

            tk.END,

            f"Total de abastecimentos : {len(registros)}\n"

        )




# =============================================================================
# PARTE 224
# BUSCA DADOS
# =============================================================================

        registros = self.banco.listar_abastecimentos()

        print(

            f"{len(registros)} registros encontrados."

        )


# =============================================================================
# PARTE 225
# DATAFRAME
# =============================================================================

        df = pd.DataFrame(

            registros

        )


# =============================================================================
# PARTE 226
# COLUNAS
# =============================================================================

        df.columns = [

            "ID",

            "Data",

            "Posto",

            "Cidade",

            "Odometro",

            "Volume",

            "Preco",

            "Valor",

            "Temperatura",

            "Pressao",

            "Altitude"

        ]


# =============================================================================
# PARTE 227
# ESCOLHER ARQUIVO
# =============================================================================

        arquivo = filedialog.asksaveasfilename(

            title="Salvar planilha Excel",

            defaultextension=".xlsx",

            filetypes=[

                ("Planilha Excel","*.xlsx")

            ],

            initialfile="Historico_GNV.xlsx"

        )

        if not arquivo:

            return


# =============================================================================
# PARTE 228
# EXPORTAR DATAFRAME
# =============================================================================

        df.to_excel(

            arquivo,

            index=False

        )


# =============================================================================
# PARTE 229
# MENSAGEM
# =============================================================================

        messagebox.showinfo(

            "Excel",

            "Planilha exportada com sucesso."

        )


# =============================================================================
# PARTE 230
# STATUS
# =============================================================================

        self.status.set(

            "Planilha Excel exportada."

        )


# =============================================================================
# PARTE 232
# AJUSTA LARGURA DAS COLUNAS
# =============================================================================

        with pd.ExcelWriter(

            arquivo,

            engine="openpyxl"

        ) as writer:

            df.to_excel(

                writer,

                index=False,

                sheet_name="Abastecimentos"

            )

            worksheet = writer.sheets["Abastecimentos"]

            for coluna in worksheet.columns:

                tamanho = max(

                    len(str(c.value))

                    if c.value is not None

                    else 0

                    for c in coluna

                )

                worksheet.column_dimensions[

                    coluna[0].column_letter

                ].width = tamanho + 3


# =============================================================================
# PARTE 233
# CONGELA CABEÇALHO
# =============================================================================

            worksheet.freeze_panes = "A2"


# =============================================================================
# PARTE 234
# FILTROS
# =============================================================================

            worksheet.auto_filter.ref = worksheet.dimensions

# =============================================================================
# PARTE 235
# CABEÇALHO
# =============================================================================

            from openpyxl.styles import Font

            for celula in worksheet[1]:

                celula.font = Font(

                    bold=True

                )


# =============================================================================
# PARTE 236
# FINALIZA EXPORTAÇÃO
# =============================================================================

        print()

        print("=" * 70)

        print("PLANILHA EXPORTADA COM SUCESSO")

        print("=" * 70)

        print(arquivo)

        print()


# =============================================================================
# PARTE 407
# EXPORTAR PDF
# =============================================================================

    def exportar_pdf(self):

        pdf = RelatorioPDF()

        pdf.add_page()


# =============================================================================
# PARTE 408
# DATA
# =============================================================================

        pdf.titulo(

            "Informações Gerais"

        )

        pdf.linha(

            "Data:",

            datetime.now().strftime(

                "%d/%m/%Y"

            )

        )

        pdf.linha(

            "Hora:",

            datetime.now().strftime(

                "%H:%M:%S"

            )

        )


# =============================================================================
# PARTE 409
# TOTAL REGISTROS
# =============================================================================

        registros = self.banco.listar_abastecimentos()

        pdf.linha(

            "Abastecimentos:",

            len(

                registros

            )

        )


# =============================================================================
# PARTE 410
# ESPAÇO
# =============================================================================

        pdf.ln(

            5

        )

# =============================================================================
# PARTE 411
# TÍTULO HISTÓRICO
# =============================================================================

        pdf.titulo(

            "Histórico de Abastecimentos"

        )

        pdf.ln(

            3

        )


# =============================================================================
# PARTE 412
# LOOP
# =============================================================================

        for registro in registros:



# =============================================================================
# PARTE 413
# DATA
# =============================================================================

            pdf.linha(

                "Data",

                registro[1]

            )


# =============================================================================
# PARTE 414
# POSTO
# =============================================================================

            pdf.linha(

                "Posto",

                registro[2]

            )


# =============================================================================
# PARTE 415
# CIDADE
# =============================================================================

            pdf.linha(

                "Cidade",

                registro[3]

            )
	
# =============================================================================
# PARTE 416
# HODÔMETRO
# =============================================================================

            pdf.linha(

                "Hodômetro",

                f"{registro[4]} km"

            )


# =============================================================================
# PARTE 417
# VOLUME
# =============================================================================

            pdf.linha(

                "Volume",

                f"{registro[5]} m³"

            )


# =============================================================================
# PARTE 418
# PREÇO DO M³
# =============================================================================

            pdf.linha(

                "Preço do m³",

                f"R$ {registro[6]}"

            )

# =============================================================================
# PARTE 419
# VALOR TOTAL
# =============================================================================

            pdf.linha(

                "Valor Total",

                f"R$ {registro[7]}"

            )


# =============================================================================
# PARTE 420
# PRESSÃO
# =============================================================================

            pdf.linha(

                "Pressão",

                f"{registro[8]} bar"

            )
# =============================================================================
# PARTE 421
# TEMPERATURA
# =============================================================================

            pdf.linha(

                "Temperatura",

                f"{registro[9]} °C"

            )


# =============================================================================
# PARTE 422
# ALTITUDE
# =============================================================================

            pdf.linha(

                "Altitude",

                f"{registro[10]} m"

            )


# =============================================================================
# PARTE 423
# LINHA EM BRANCO
# =============================================================================

            pdf.ln(

                2

            )



# =============================================================================
# PARTE 424
# SEPARADOR
# =============================================================================

            pdf.cell(

                0,

                2,

                "_" * 80,

                ln=True

            )

# =============================================================================
# PARTE 425
# ESPAÇO
# =============================================================================

            pdf.ln(

                3

            )



# =============================================================================
# PARTE 237
# EXPORTAR CSV
# =============================================================================

        arquivo_csv = arquivo.replace(

            ".xlsx",

            ".csv"

        )

        df.to_csv(

            arquivo_csv,

            index=False,

            sep=";",

            decimal=",",

            encoding="utf-8-sig"

        )



# =============================================================================
# PARTE 238
# STATUS CSV
# =============================================================================

        print(

            "CSV exportado:",

            arquivo_csv

        )

        self.status.set(

            "Excel e CSV exportados."

        )


# =============================================================================
# PARTE 239
# MENSAGEM COMPLETA
# =============================================================================

        messagebox.showinfo(

            "Exportação",

            "Arquivos Excel e CSV criados com sucesso."

        )






# =============================================================================
# PARTE 231
# TRATAMENTO DE ERROS
# =============================================================================

        except Exception as erro:

            messagebox.showerror(

                "Erro",

                str(erro)

            )

# =============================================================================
# PARTE 240
# EXPORTAR PDF
# =============================================================================

        arquivo_pdf = arquivo.replace(

            ".xlsx",

            ".pdf"

        )

# =============================================================================
# PARTE 241
# CRIAR PDF
# =============================================================================

        pdf = FPDF(

            orientation="L",

            unit="mm",

            format="A4"

        )

        pdf.add_page()

        pdf.set_font(

            "Arial",

            "B",

            14

        )


# =============================================================================
# PARTE 242
# TÍTULO
# =============================================================================

        pdf.cell(

            0,

            10,

            "RELATORIO DE ABASTECIMENTOS GNV",

            ln=True,

            align="C"

        )

        pdf.ln(5)



# =============================================================================
# PARTE 243
# CABEÇALHO
# =============================================================================

        pdf.set_font(

            "Arial",

            "B",

            8

        )

        colunas = [

            ("Data",25),

            ("Posto",55),

            ("Cidade",40),

            ("Volume",25),

            ("Valor",25)

        ]

        for titulo, largura in colunas:

            pdf.cell(

                largura,

                8,

                titulo,

                border=1,

                align="C"

            )

        pdf.ln()



# =============================================================================
# PARTE 244
# DADOS
# =============================================================================

        pdf.set_font(

            "Arial",

            "",

            8

        )

        for registro in registros:

            pdf.cell(25,8,str(registro[1]),1)

            pdf.cell(55,8,str(registro[2]),1)

            pdf.cell(40,8,str(registro[3]),1)

            pdf.cell(25,8,str(registro[5]),1)

            pdf.cell(25,8,f"R$ {registro[7]:.2f}",1)

            pdf.ln()



# =============================================================================
# PARTE 245
# SALVAR PDF
# =============================================================================

        pdf.output(

            arquivo_pdf

        )

        print(

            "PDF exportado:",

            arquivo_pdf

        )

        self.status.set(

            "Excel, CSV e PDF exportados."

        )





# =============================================================================
# PARTE 247
# MENSAGEM FINAL
# =============================================================================

        messagebox.showinfo(

            "Exportação concluída",

            f"Arquivos gerados:\n\n"

            f"Excel:\n{arquivo}\n\n"

            f"CSV:\n{arquivo_csv}\n\n"

            f"PDF:\n{arquivo_pdf}"

        )



# =============================================================================
# PARTE 248
# LOG
# =============================================================================

        print()

        print("=" * 70)

        print("ARQUIVOS EXPORTADOS")

        print("=" * 70)

        print("Excel :", arquivo)

        print("CSV   :", arquivo_csv)

        print("PDF   :", arquivo_pdf)

        print("=" * 70)

        print()



# =============================================================================
# PARTE 249
# ABRIR PASTA
# =============================================================================

        pasta = os.path.dirname(

            arquivo

        )

        try:

            os.startfile(

                pasta

            )

        except Exception:

            pass



# =============================================================================
# PARTE 251
# TOTAL EXPORTADO
# =============================================================================

        print(

            f"Foram exportados {len(registros)} abastecimentos."

        )


# =============================================================================
# PARTE 252
# DATA E HORA
# =============================================================================

        print(

            datetime.now().strftime(

                "Exportado em %d/%m/%Y às %H:%M:%S"

            )

        )




# =============================================================================
# PARTE 195
# PESQUISAR HISTÓRICO
# =============================================================================

    def pesquisar_historico(self):

        texto = self.entry_pesquisa.get().strip()

        if texto == "":

            self.atualizar_historico()

            return

        for item in self.tree.get_children():

            self.tree.delete(

                item

            )

        registros = self.banco.buscar_por_posto(

            texto

        )

        for registro in registros:

            self.tree.insert(

                "",

                "end",

                values=(

                    registro[1],

                    registro[2],

                    registro[3],

                    registro[5],

                    registro[7]

                )

            )


# =============================================================================
# PARTE 200
# ABRIR REGISTRO
# =============================================================================

    def abrir_registro(

        self,

        event

    ):

        item = self.tree.focus()

        if not item:

            return

        valores = self.tree.item(

            item,

            "values"

        )

        print(

            valores

        )


# =============================================================================
# PARTE 205
# MENU CONTEXTO
# =============================================================================

    def menu_contexto(

        self,

        event

    ):

        self.menu_tree.post(

            event.x_root,

            event.y_root

        )


# =============================================================================
# PARTE 206
# SELECIONAR ITEM
# =============================================================================

        item = self.tree.identify_row(

            event.y

        )

        if item:

            self.tree.selection_set(

                item

            )

# =============================================================================
# PARTE 217
# EXCLUSÃO COMPLETA
# =============================================================================

    def editar_registro(self):

        item = self.tree.focus()

        if not item:

            return

        valores = self.tree.item(

            item,

            "values"

        )

        print()

        print("=" * 70)

        print("REGISTRO SELECIONADO")

        print("=" * 70)

        print(valores)

        print()


    def excluir_registro(self):

        item = self.tree.focus()

        if not item:

            messagebox.showwarning(

                "Histórico",

                "Selecione um registro."

            )

            return

        resposta = messagebox.askyesno(

            "Excluir",

            "Deseja realmente excluir este registro?"

        )

        if not resposta:

            return

        print()

        print("=" * 70)

        print("REGISTRO EXCLUÍDO")

        print("=" * 70)

        print(item)

        print()

        self.atualizar_historico()




""" # =============================================================================
# PARTE 207
# EDITAR REGISTRO
# =============================================================================

    def editar_registro(self):

        print(

            "Editar registro"

        ) """


# =============================================================================
# PARTE 208
# EXCLUIR REGISTRO
# =============================================================================

    def excluir_registro(self):

        print(

            "Excluir registro"

        )

# =============================================================================
# PARTE 212
# EDITAR REGISTRO
# =============================================================================

    def editar_registro(self):

        item = self.tree.focus()

        if not item:

            return

        valores = self.tree.item(

            item,

            "values"

        )

        print(

            "Editar:",

            valores

        )

# =============================================================================
# PARTE 215
# ATUALIZA TREEVIEW
# =============================================================================

        self.recarregar_treeview()


# =============================================================================
# PARTE 213
# EXCLUIR REGISTRO
# =============================================================================

    def excluir_registro(self):

        item = self.tree.focus()

        if not item:

            return

        resposta = messagebox.askyesno(

            "Excluir",

            "Deseja realmente excluir este abastecimento?"

        )

        if not resposta:

            return

        print(

            "Excluir:",

            item

        )

# =============================================================================
# PARTE 216
# ATUALIZA TREEVIEW
# =============================================================================

        self.recarregar_treeview()


# =============================================================================
# PARTE 214
# RECARREGA TREEVIEW
# =============================================================================

    def recarregar_treeview(self):

        self.atualizar_historico()




# =============================================================================
# PARTE 211
# CONFIRMAÇÃO
# =============================================================================

        resposta = messagebox.askyesno(

            "Excluir",

            "Deseja realmente excluir este abastecimento?"

        )

        if not resposta:

            return




# =============================================================================
# PARTE 148
# ÁREA DE RESULTADOS COM SCROLLBAR
# =============================================================================

        self.frame_resultados = ttk.Frame(

            self.frame_calculos

        )

        self.frame_resultados.grid(

            row=8,

            column=0,

            columnspan=2,

            padx=10,

            pady=10,

            sticky="nsew"

        )

        self.scroll_resultados = ttk.Scrollbar(

            self.frame_resultados,

            orient="vertical"

        )

        self.scroll_resultados.pack(

            side="right",

            fill="y"

        )

        self.texto_resultados = tk.Text(

            self.frame_resultados,

            width=90,

            height=25,

            font=("Consolas", 10),

            yscrollcommand=self.scroll_resultados.set

        )

        self.texto_resultados.pack(

            side="left",

            fill="both",

            expand=True

        )

        self.scroll_resultados.config(

            command=self.texto_resultados.yview

        )

        self.frame_calculos.grid_rowconfigure(

            8,

            weight=1

        )

        self.frame_calculos.grid_columnconfigure(

            1,

            weight=1

        )


# =============================================================================
# PARTE 154
# TÍTULO DOS RESULTADOS
# =============================================================================

        self.label_resultados = ttk.Label(

            self.frame_calculos,

            text="Resultados",

            font=("Arial",12,"bold")

        )

        self.label_resultados.grid(

            row=8,

            column=0,

            sticky="w",

            padx=10,

            pady=(10,0)

        )





# =============================================================================
# PARTE 145
# EXECUTAR CÁLCULO
# =============================================================================

    def executar_calculo(self):

        try:

            volume = float(

                self.entry_volume.get()

            )

            quantidade = int(

                self.entry_quantidade.get()

            )

            pressao = float(

                self.entry_pressao.get()

            )

            temperatura = float(

                self.entry_temperatura.get()

            )

            altitude = float(

                self.entry_altitude.get()

            )

            fator_z = float(

                self.entry_fator_z.get()

            )

            massa_molar = float(

                self.entry_massa_molar.get()

            )

        except ValueError:

            self.texto_resultados.delete(

                "1.0",

                tk.END

            )

            self.texto_resultados.insert(

                tk.END,

                "Erro: verifique os valores informados."

            )

            return

        volume_total = (

            volume *

            quantidade

        )

        resultado = calcular_quantidade_gnv(

            volume_total,

            pressao,

            temperatura,

            altitude,

            fator_z,

            massa_molar

        )

                self.texto_resultados.delete(
            "1.0",
            tk.END
        )

        self.texto_resultados.insert(
            tk.END,
            "=" * 70 + "\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "RELATÓRIO DOS CÁLCULOS DE GNV\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "=" * 70 + "\n\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Volume Total..............: {volume_total:.2f} L\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Quantidade de Cilindros...: {quantidade}\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Pressão...................: {pressao:.2f} bar\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Temperatura...............: {temperatura:.2f} °C\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Altitude..................: {altitude:.2f} m\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Fator Z...................: {fator_z:.4f}\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Massa Molar...............: {massa_molar:.5f} kg/mol\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "-" * 70 + "\n"
        )

        for chave, valor in resultado.items():

            if isinstance(valor, float):

                texto = f"{valor:.6f}"

            else:

                texto = str(valor)

            self.texto_resultados.insert(
                tk.END,
                f"{chave:<30} {texto}\n"
            )

        self.texto_resultados.insert(
            tk.END,
            "\n" + "=" * 70 + "\n"
        )

# =============================================================================
# PARTE 156
# ATUALIZA STATUS
# =============================================================================

        self.status.set(

            "Cálculo executado com sucesso."

        )




# =============================================================================
# PARTE 149
# LIMPAR RESULTADOS
# =============================================================================

    def limpar_resultados(self):

        self.texto_resultados.delete(

            "1.0",

            tk.END

        )


# =============================================================================
# PARTE 151
# LIMPAR CAMPOS
# =============================================================================

    def limpar_campos(self):

        self.entry_volume.delete(0, tk.END)

        self.entry_quantidade.delete(0, tk.END)

        self.entry_pressao.delete(0, tk.END)

        self.entry_temperatura.delete(0, tk.END)

        self.entry_altitude.delete(0, tk.END)

        self.entry_fator_z.delete(0, tk.END)

        self.entry_massa_molar.delete(0, tk.END)

        self.entry_quantidade.insert(0, "1")

        self.entry_pressao.insert(0, "200")

        self.entry_temperatura.insert(0, "20")

        self.entry_altitude.insert(0, "0")

        self.entry_fator_z.insert(0, "0.92")

        self.entry_massa_molar.insert(0, "0.01604")

        self.limpar_resultados()








# =============================================================================
# PARTE 58
# CLASSE GNV
# =============================================================================

class GNV:

    """
    Classe principal para armazenar
    todos os dados referentes ao GNV.
    """

    def __init__(

        self,

        volume_m3,

        pressao_bar,

        temperatura_c,

        altitude

    ):

        self.volume_m3 = volume_m3

        self.pressao_bar = pressao_bar

        self.temperatura_c = temperatura_c

        self.altitude = altitude

        self.temperatura_k = (

            temperatura_c

            +

            273.15

        )


# =============================================================================
# PARTE 59
# PRESSÃO ATMOSFÉRICA
# =============================================================================

    def pressao_atmosferica(self):

        return calcular_pressao_atmosferica(

            self.altitude

        )


# =============================================================================
# PARTE 60
# PRESSÃO ABSOLUTA
# =============================================================================

    def pressao_absoluta(self):

        return (

            self.pressao_bar

            +

            self.pressao_atmosferica()

        )


# =============================================================================
# PARTE 61
# FATOR Z
# =============================================================================

    def fator_Z(self):

        resultado = calcular_Z(

            self.pressao_absoluta(),

            self.temperatura_k

        )

        return resultado


# =============================================================================
# PARTE 62
# MASSA DO GNV
# =============================================================================

    def massa(self):

        densidade = propriedade(

            "densidade"

        )

        return calcular_massa_por_densidade(

            self.volume_m3,

            densidade

        )


# =============================================================================
# PARTE 63
# ENERGIA DO GNV
# =============================================================================

    def energia(self):

        pcs = propriedade(

            "pcs"

        )

        energia_mj = (

            self.volume_m3

            *

            pcs

        )

        energia_kwh = (

            energia_mj

            /

            3.6

        )

        return {

            "MJ": energia_mj,

            "kWh": energia_kwh

        }


# =============================================================================
# PARTE 64
# AUTONOMIA
# =============================================================================

    def autonomia(

        self,

        consumo_km_m3

    ):

        return (

            self.volume_m3

            *

            consumo_km_m3

        )


# =============================================================================
# PARTE 65
# DENSIDADE
# =============================================================================

    def densidade(self):

        return propriedade(

            "densidade"

        )


# =============================================================================
# PARTE 66
# MASSA MOLAR
# =============================================================================

    def massa_molar(self):

        return calcular_massa_molar_mistura()


# =============================================================================
# PARTE 67
# RELATÓRIO COMPLETO
# =============================================================================

    def relatorio(self, consumo_km_m3=15.0):
        """
        Exibe um relatório completo do estado do GNV.

        consumo_km_m3:
            Rendimento médio do veículo em km por m³.
        """

        pressao_atm = self.pressao_atmosferica()

        pressao_abs = self.pressao_absoluta()

        fator = self.fator_Z()

        massa = self.massa()

        energia = self.energia()

        autonomia = self.autonomia(consumo_km_m3)

        print()

        print("=" * 75)

        print("RELATÓRIO COMPLETO DO GNV")

        print("=" * 75)

        print(f"Volume................: {self.volume_m3:.3f} m³")

        print(f"Pressão...............: {self.pressao_bar:.2f} bar")

        print(f"Pressão Atmosférica...: {pressao_atm:.4f} bar")

        print(f"Pressão Absoluta......: {pressao_abs:.4f} bar")

        print(f"Temperatura...........: {self.temperatura_c:.2f} °C")

        print(f"Temperatura...........: {self.temperatura_k:.2f} K")

        print(f"Pressão Reduzida......: {fator['Pr']:.4f}")

        print(f"Temperatura Reduzida..: {fator['Tr']:.4f}")

        print(f"Fator Z...............: {fator['Z']:.5f}")

        print(f"Massa................: {massa:.3f} kg")

        print(f"Energia..............: {energia['MJ']:.2f} MJ")

        print(f"Energia..............: {energia['kWh']:.2f} kWh")

        print(f"Autonomia Estimada...: {autonomia:.2f} km")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 68
# DICIONÁRIO COMPLETO
# =============================================================================

    def dados(self, consumo_km_m3=15.0):
        """
        Retorna todos os cálculos em formato
        de dicionário.

        Essa função será utilizada futuramente
        para:

            Excel

            SQLite

            PDF

            API

            Interface Gráfica
        """

        pressao_atm = self.pressao_atmosferica()

        pressao_abs = self.pressao_absoluta()

        fator = self.fator_Z()

        energia = self.energia()

        return {

            "Volume_m3": self.volume_m3,

            "Pressao_bar": self.pressao_bar,

            "Pressao_Atmosferica": pressao_atm,

            "Pressao_Absoluta": pressao_abs,

            "Temperatura_C": self.temperatura_c,

            "Temperatura_K": self.temperatura_k,

            "Pressao_Reduzida": fator["Pr"],

            "Temperatura_Reduzida": fator["Tr"],

            "Fator_Z": fator["Z"],

            "Massa": self.massa(),

            "Energia_MJ": energia["MJ"],

            "Energia_kWh": energia["kWh"],

            "Autonomia": self.autonomia(
                consumo_km_m3
            )

        }


# =============================================================================
# PARTE 69
# MOSTRAR DADOS
# =============================================================================

    def mostrar_dados(self):

        dados = self.dados()

        print()

        print("=" * 75)

        print("DADOS DA CLASSE GNV")

        print("=" * 75)

        print()

        for chave, valor in dados.items():

            print(

                f"{chave:<30}: {valor}"

            )

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 70
# EXPORTAR DADOS
# =============================================================================

    def exportar(self):

        return self.dados()

# =============================================================================
# PARTE 71
# EXPORTAÇÃO JSON
# =============================================================================

    def json(self):

        import json

        return json.dumps(

            self.dados(),

            indent=4,

            ensure_ascii=False

        )


# =============================================================================
# PARTE 72
# SIMULAÇÃO DA PRESSÃO EM OUTRA TEMPERATURA
# =============================================================================

    def simular_pressao(

        self,

        nova_temperatura_c

    ):
        """
        Simula a nova pressão considerando
        que a quantidade de gás permanece
        constante.

        Utiliza uma aproximação baseada
        na Lei dos Gases Ideais.

        Retorna a pressão em bar.
        """

        temperatura1 = self.temperatura_k

        temperatura2 = (

            nova_temperatura_c

            +

            273.15

        )

        pressao_absoluta = (

            self.pressao_absoluta()

            *

            temperatura2

            /

            temperatura1

        )

        pressao_manometrica = (

            pressao_absoluta

            -

            self.pressao_atmosferica()

        )

        return pressao_manometrica



# =============================================================================
# PARTE 73
# RESFRIAMENTO
# =============================================================================

    def simular_resfriamento(

        self,

        temperatura_final

    ):

        return self.simular_pressao(

            temperatura_final

        )


# =============================================================================
# PARTE 74
# AQUECIMENTO
# =============================================================================

    def simular_aquecimento(

        self,

        temperatura_final

    ):

        return self.simular_pressao(

            temperatura_final

        )


# =============================================================================
# PARTE 75
# TABELA DE PRESSÃO x TEMPERATURA
# =============================================================================

    def tabela_temperatura(

        self,

        temperatura_inicial=-20,

        temperatura_final=60,

        passo=2

    ):

        print()

        print("=" * 75)

        print("SIMULAÇÃO DE PRESSÃO")

        print("=" * 75)

        print()

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            pressao = self.simular_pressao(

                temperatura

            )

            print(

                f"{temperatura:6.1f} °C   "

                f"{pressao:8.2f} bar"

            )

            temperatura += passo

        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 76
# SIMULAÇÃO DE ALTITUDE
# =============================================================================

    def simular_altitude(

        self,

        nova_altitude

    ):
        """
        Simula a pressão atmosférica para
        outra altitude.
        """

        return calcular_pressao_atmosferica(

            nova_altitude

        )


# =============================================================================
# PARTE 77
# PRESSÃO ABSOLUTA EM OUTRA ALTITUDE
# =============================================================================

    def pressao_absoluta_altitude(

        self,

        nova_altitude

    ):

        pressao_atm = calcular_pressao_atmosferica(

            nova_altitude

        )

        return (

            self.pressao_bar

            +

            pressao_atm

        )


# =============================================================================
# PARTE 78
# SIMULAÇÃO COMPLETA
# =============================================================================

    def simular(

        self,

        temperatura,

        altitude

    ):

        pressao = self.simular_pressao(

            temperatura

        )

        pressao_atm = calcular_pressao_atmosferica(

            altitude

        )

        pressao_abs = (

            pressao

            +

            pressao_atm

        )

        return {

            "temperatura": temperatura,

            "altitude": altitude,

            "pressao_atmosferica": pressao_atm,

            "pressao_manometrica": pressao,

            "pressao_absoluta": pressao_abs

        }


# =============================================================================
# PARTE 79
# RELATÓRIO DA SIMULAÇÃO
# =============================================================================

    def relatorio_simulacao(

        self,

        temperatura,

        altitude

    ):

        dados = self.simular(

            temperatura,

            altitude

        )

        print()

        print("=" * 75)

        print("SIMULAÇÃO")

        print("=" * 75)

        print()

        print(f"Temperatura...........: {dados['temperatura']:.2f} °C")

        print(f"Altitude..............: {dados['altitude']:.2f} m")

        print(f"Pressão Atmosférica...: {dados['pressao_atmosferica']:.4f} bar")

        print(f"Pressão Manométrica...: {dados['pressao_manometrica']:.4f} bar")

        print(f"Pressão Absoluta......: {dados['pressao_absoluta']:.4f} bar")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 80
# SIMULAÇÃO EM FAIXA
# =============================================================================

    def simular_faixa_temperatura(

        self,

        temperatura_inicial,

        temperatura_final,

        incremento=2

    ):

        resultados = []

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            resultados.append(

                self.simular(

                    temperatura,

                    self.altitude

                )

            )

            temperatura += incremento

        return resultados



# =============================================================================
# PARTE 81
# GERAÇÃO DA TABELA PRESSÃO x TEMPERATURA
# =============================================================================

    def gerar_tabela_pressao_temperatura(

        self,

        temperatura_inicial=-20,

        temperatura_final=60,

        passo_temperatura=2,

        pressao_inicial=0,

        pressao_final=300,

        passo_pressao=5

    ):
        """
        Gera uma tabela contendo todas as
        combinações entre temperatura e pressão.

        Retorna uma lista de dicionários.
        """

        tabela = []

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            pressao = pressao_inicial

            while pressao <= pressao_final:

                """linha = {

                    "Temperatura": temperatura,

                    "Pressao": pressao

                }

                tabela.append(

                    linha

                )"""

# ==============================================================
# CONVERSÕES
# ==============================================================

temperatura_kelvin = (

    temperatura

    +

    273.15

)

pressao_atm = calcular_pressao_atmosferica(

    self.altitude

)

pressao_absoluta = (

    pressao

    +

    pressao_atm

)

# ==============================================================
# FATOR Z
# ==============================================================

resultado_z = calcular_Z(

    pressao_absoluta,

    temperatura_kelvin

)

# ==============================================================
# VOLUME NAS CNTP
# ==============================================================

volume_cntp = calcular_volume_equivalente(

    self.volume_m3,

    pressao_absoluta,

    temperatura_kelvin,

    resultado_z["Z"]

)

# ==============================================================
# MASSA
# ==============================================================

massa = calcular_massa_por_densidade(

    volume_cntp,

    propriedade(

        "densidade"

    )

)

# ==============================================================
# ENERGIA
# ==============================================================

energia = calcular_energia_volume(

    volume_cntp

)


# ==============================================================
# ENERGIA EM kWh
# ==============================================================

energia_kwh = (

    energia

    /

    3.6

)


                # ==============================================================
                # DENSIDADE APARENTE
                # ==============================================================

                if volume_cntp > 0:

                    densidade_aparente = massa / volume_cntp

                else:

                    densidade_aparente = 0.0


# ==============================================================
# DENSIDADE
# ==============================================================

densidade = calcular_massa_especifica(

    massa,

    volume_cntp

)

# ==============================================================
# CONSTANTE ESPECÍFICA
# ==============================================================

R_especifico = calcular_R_especifico(

    calcular_massa_molar_mistura()

)






# ==============================================================
# LINHA DA TABELA
# ==============================================================

linha = {

    "Temperatura_C": temperatura,

    "Temperatura_K": temperatura_kelvin,

    "Pressao_bar": pressao,

    "Pressao_Absoluta": pressao_absoluta,

    "Altitude": self.altitude,

    "Pressao_Atmosferica": pressao_atm,

    "Pr": resultado_z["Pr"],

    "Tr": resultado_z["Tr"],

    "Fator_Z": resultado_z["Z"],

    "Volume_CNTP": volume_cntp,

    "Massa": massa,

    # ==============================================================
    # ENERGIA EM kWh
    # ==============================================================

    energia_kwh = (

        energia

        /

        3.6

    )

# ==============================================================
# LINHA DA TABELA
# ==============================================================

linha = {

    "Temperatura_C": temperatura,

    "Temperatura_K": temperatura_kelvin,

    "Pressao_bar": pressao,

    "Pressao_Absoluta": pressao_absoluta,

    "Altitude": self.altitude,

    "Pressao_Atmosferica": pressao_atm,

    "Pr": resultado_z["Pr"],

    "Tr": resultado_z["Tr"],

    "Fator_Z": resultado_z["Z"],

    "Volume_CNTP": volume_cntp,

    "Massa": massa,

    "Energia_MJ": energia,

                    "Energia_kWh": energia_kwh,

                    "Densidade": densidade,

                    "Densidade_Aparente": densidade_aparente,

                    "R_especifico": R_especifico


}

}

tabela.append(

    linha

)

                pressao += passo_pressao

            temperatura += passo_temperatura

        return tabela


# =============================================================================
# PARTE 82
# MOSTRAR TABELA
# =============================================================================

    def mostrar_tabela(

        self

    ):

        tabela = self.gerar_tabela_pressao_temperatura()

        print()

        print("=" * 75)

        print("TABELA PRESSÃO x TEMPERATURA")

        print("=" * 75)

        print()

        for linha in tabela:

            print(

                f"{linha['Temperatura']:6.1f} °C"

                f"    "

                f"{linha['Pressao']:6.1f} bar"

            )

        print()

        print("=" * 75)

        print()











# =============================================================================
# PARTE 23
# FUNÇÃO PARA CALCULAR AUTONOMIA
# =============================================================================

def calcular_autonomia(
    volume_m3,
    consumo_km_m3
):
    """
    Calcula a autonomia.

    consumo_km_m3

        rendimento do veículo
        em km por metro cúbico.
    """

    return volume_m3 * consumo_km_m3


# =============================================================================
# PARTE 24
# RELATÓRIO RESUMIDO
# =============================================================================

def mostrar_relatorio(
    volume,
    massa,
    energia,
    autonomia
):

    print()

    print("=" * 70)

    print("RELATÓRIO")

    print("=" * 70)

    print()

    print(f"Volume..........: {volume:.3f} m³")

    print(f"Massa...........: {massa:.3f} kg")

    print(f"Energia.........: {energia:.2f} kWh")

    print(f"Autonomia.......: {autonomia:.1f} km")

    print()

    print("=" * 70)





# =============================================================================
# PARTE 16 - FUNÇÃO PARA EXPORTAR O HISTÓRICO PARA EXCEL
# =============================================================================

from openpyxl import Workbook

def exportar_excel(historico, nome_arquivo="Historico_GNV.xlsx"):
    """
    Exporta o histórico de abastecimentos para um arquivo Excel.

    Parâmetros
    ----------
    historico : list
        Lista contendo os registros dos abastecimentos.

    nome_arquivo : str
        Nome do arquivo Excel.
    """

    wb = Workbook()

    ws = wb.active

    ws.title = "Histórico"

    if len(historico) == 0:

        wb.save(nome_arquivo)

        return

    cabecalho = list(historico[0].keys())

    ws.append(cabecalho)

    for registro in historico:

        ws.append(list(registro.values()))

    wb.save(nome_arquivo)

    print()

    print("Arquivo Excel criado com sucesso.")

    print(nome_arquivo)

    print()


# =============================================================================
# FUNÇÃO PARA LER NÚMEROS REAIS
# =============================================================================

def ler_float(
    mensagem,
    valor_padrao=None
):
    """
    Lê um número decimal informado pelo usuário.

    Aceita ponto ou vírgula.

    Se existir um valor padrão e o usuário apenas
    pressionar ENTER, retorna esse valor.
    """

    while True:

        texto = input(
            mensagem
        ).strip()

        if texto == "":

            if valor_padrao is not None:

                return valor_padrao

            print()

            print("Digite um valor.")

            print()

            continue

        texto = texto.replace(",", ".")

        try:

            return float(texto)

        except ValueError:

            print()

            print("Valor inválido.")

            print()



# =============================================================================
# PARTE 06 - TELA INICIAL
# =============================================================================

print("=" * 70)
print("              CALCULADORA DE GNV")
print("=" * 70)

print()

print("Informe os dados abaixo.")

print()

# =============================================================================
# LEITURA DOS DADOS
# =============================================================================

# Volume do cilindro

volume_litros = float(
    input("Volume do cilindro (Litros): ")
)

# Pressão do cilindro

pressao_bar = ler_float(
    "Pressão indicada no manômetro (bar): "
)

# Temperatura

temperatura = ler_float(
    "Temperatura do gás (°C): "
)

# Altitude

altitude = float(
    input("Altitude (metros): ")
)

# =============================================================================
# PARÂMETROS DO GÁS
# =============================================================================

print()

print("=" * 70)
print("PARÂMETROS DO GNV")
print("=" * 70)

print()

print("Se você não souber algum valor, pressione ENTER.")

print()

# -------------------------------------------------------------------------
# Massa molar
# -------------------------------------------------------------------------

texto = input(
    "Massa molar do GNV [0.018 kg/mol]: "
).strip()

if texto == "":

    MASSA_MOLAR_GNV = 0.018

else:

    MASSA_MOLAR_GNV = float(texto)

# -------------------------------------------------------------------------
# Densidade
# -------------------------------------------------------------------------

texto = input(
    "Densidade do GNV [0.78 kg/m³]: "
).strip()

if texto == "":

    DENSIDADE_GNV = 0.78

else:

    DENSIDADE_GNV = float(texto)

# -------------------------------------------------------------------------
# Compressibilidade
# -------------------------------------------------------------------------

# =============================================================================
# CÁLCULO AUTOMÁTICO DO FATOR Z
# =============================================================================

#
# Esta função calcula uma aproximação para o fator de
# compressibilidade do GNV.
#
# Futuramente será substituída pela Equação de Peng-Robinson.
#

def calcular_fator_z(pressao_bar, temperatura_c):

    temperatura_k = temperatura_c + 273.15

    #
    # Valores críticos aproximados do Metano
    #

    Tc = 190.56

    Pc = 45.99

    #
    # Pressão reduzida
    #

    Pr = pressao_bar / Pc

    #
    # Temperatura reduzida
    #

    Tr = temperatura_k / Tc

    #
    # Aproximação inicial
    #

    Z = 1.0

    #
    # Correção empírica.
    #
    # Esta equação foi ajustada apenas para produzir
    # valores próximos aos encontrados em cilindros de GNV.
    #

    Z = 1 - (
        0.065
        *
        (
            Pr /
            (Tr ** 1.20)
        )
    )

    #
    # Limites
    #

    if Z < 0.72:

        Z = 0.72

    if Z > 1.00:

        Z = 1.00

    return Z

# =============================================================================
# FATOR Z
# =============================================================================

Z = calcular_fator_z(
    pressao_bar,
    temperatura
)

print()

print(f"Fator Z calculado automaticamente: {Z:.4f}")
print()

print("=" * 75)
print("Calculando...")
print("=" * 75)
print()


# =============================================================================
# QUANTIDADE DE CILINDROS
# =============================================================================

print()

print("=" * 75)
print("CONFIGURAÇÃO DOS CILINDROS")
print("=" * 75)

while True:

    try:

        quantidade_cilindros = int(
            input("Quantidade de cilindros [1 a 8]: ")
        )

        if 1 <= quantidade_cilindros <= 8:

            break

        else:

            print("Digite um valor entre 1 e 8.")

    except ValueError:

        print("Digite apenas números.")

print()

# =============================================================================
# LISTA DOS CILINDROS
# =============================================================================

cilindros = []

volume_total_litros = 0.0

# =============================================================================
# LEITURA DOS CILINDROS
# =============================================================================

for numero in range(1, quantidade_cilindros + 1):

    while True:

        try:

            volume = ler_float(
    		f"Volume do cilindro {numero} (Litros): "
	    )

            if volume > 0:

                break

            print("O volume deve ser maior que zero.")

        except ValueError:

            print("Digite um número válido.")

    cilindros.append(volume)

    volume_total_litros += volume

print()

print("=" * 75)

print("CILINDROS INFORMADOS")

print("=" * 75)

print()

for indice, volume in enumerate(cilindros, start=1):

    print(f"Cilindro {indice}: {volume:.2f} litros")

print()

print(f"Volume Total: {volume_total_litros:.2f} litros")

print()

# =============================================================================
# SUBSTITUI O VOLUME INFORMADO PELO VOLUME TOTAL
# =============================================================================

volume_litros = volume_total_litros


# =============================================================================
# INÍCIO DOS CÁLCULOS
# =============================================================================


# =============================================================================
# RESUMO DOS DADOS INFORMADOS
# =============================================================================

print()

print("=" * 75)
print("DADOS INFORMADOS")
print("=" * 75)

print()

print(f"Quantidade de cilindros........: {quantidade_cilindros}")

print()

for indice, volume in enumerate(cilindros, start=1):

    print(
        f"Cilindro {indice:02d}...................: "
        f"{volume:7.2f} litros"
    )

print()

print(
    f"Volume Total...................: "
    f"{volume_total_litros:7.2f} litros"
)

print()

print(
    f"Pressão........................: "
    f"{pressao_bar:7.2f} bar"
)

print(
    f"Temperatura....................: "
    f"{temperatura:7.2f} °C"
)

print(
    f"Altitude.......................: "
    f"{altitude:7.2f} metros"
)

print()

print(
    f"Densidade do GNV...............: "
    f"{DENSIDADE_GNV:7.3f} kg/m³"
)

print(
    f"Massa molar....................: "
    f"{MASSA_MOLAR_GNV:7.5f} kg/mol"
)

print(
    f"Fator Z........................: "
    f"{Z:7.4f}"
)

print()

print("=" * 75)

print()


# =============================================================================
# CÁLCULO DA PRESSÃO ATMOSFÉRICA
# =============================================================================

# Calcula automaticamente a pressão atmosférica em função da altitude.
pressao_atmosferica = calcular_pressao_atmosferica(
    altitude
)

# =============================================================================
# CÁLCULO DA PRESSÃO ABSOLUTA
# =============================================================================

# A Lei dos Gases utiliza pressão absoluta.
pressao_absoluta = calcular_pressao_absoluta(
    pressao_bar,
    pressao_atmosferica
)

# =============================================================================
# TEMPERATURA ABSOLUTA
# =============================================================================

temperatura_kelvin = celsius_para_kelvin(
    temperatura
)



# =============================================================================
# SUBSTITUI O VOLUME PELO TOTAL DOS CILINDROS
# =============================================================================

volume_litros = volume_total_litros

# =============================================================================
# VOLUME DO CILINDRO
# =============================================================================

volume_cilindro_m3 = litros_para_m3(
    volume_litros
)



# =============================================================================
# CÁLCULO DO VOLUME EQUIVALENTE NAS CNTP
# =============================================================================
#
# Fórmula:
#
#        P1 x V1      P2 x V2
#       --------- = ----------
#           T1            T2
#
# Rearranjando:
#
#              P1
# V2 = V1 x ------- x T2/T1
#              P2
#
# Considerando fator de compressibilidade:
#
#              P1
# V2 = V1 x ------- x T2/T1 x 1/Z
#              P2
#
# Onde:
#
# V1 = Volume do cilindro
# V2 = Volume equivalente em CNTP
#
# =============================================================================

volume_equivalente = (
    volume_cilindro_m3
    *
    (pressao_absoluta / PRESSAO_ATMOSFERICA_PADRAO)
    *
    (TEMPERATURA_PADRAO / temperatura_kelvin)
    /
    Z
)

# =============================================================================
# CÁLCULO DOS MOLS (APENAS PARA ESTUDO)
# =============================================================================

mols = calcular_mols(
    volume_cilindro_m3,
    pressao_absoluta,
    temperatura_kelvin,
    Z
)

# =============================================================================
# MASSA DO GÁS
# =============================================================================

massa_gnv = calcular_massa(
    mols
)

# =============================================================================
# DENSIDADE DENTRO DO CILINDRO
# =============================================================================

densidade_interna = (
    massa_gnv /
    volume_cilindro_m3
)

# =============================================================================
# VOLUME EM LITROS NAS CNTP
# =============================================================================

volume_equivalente_litros = (
    volume_equivalente * 1000
)

# =============================================================================
# MASSA ESPECÍFICA
# =============================================================================

massa_especifica = (
    massa_gnv /
    volume_equivalente
)

# =============================================================================
# ENERGIA DO GÁS
# =============================================================================

#
# PCS aproximado
#
# Pode ser alterado futuramente conforme informado
# pela distribuidora.
#

PCS = 13.9

energia = (
    massa_gnv *
    PCS
)

# =============================================================================
# PERCENTUAL DE ENCHIMENTO
# =============================================================================

percentual = (
    pressao_bar /
    200
) * 100

if percentual < 0:

    percentual = 0

if percentual > 100:

    percentual = 100

# =============================================================================
# AUTONOMIA
# =============================================================================

#
# Valor apenas ilustrativo.
#
# Depois criaremos uma função onde o usuário poderá
# informar o consumo em km/m³.
#

consumo_medio = 13.5

autonomia = (
    volume_equivalente *
    consumo_medio
)


# =============================================================================
# RELATÓRIO DOS RESULTADOS
# =============================================================================

print()

print("=" * 75)
print("                 RESULTADO DOS CÁLCULOS")
print("=" * 75)

print()

print("---------------- DADOS INFORMADOS ----------------")

print(f"Volume do cilindro .............: {volume_litros:10.2f} litros")

print(f"Pressão do manômetro ...........: {pressao_bar:10.2f} bar")

print(f"Altitude .......................: {altitude:10.2f} metros")

print(f"Temperatura ....................: {temperatura:10.2f} °C")

print()

print("---------------- GÁS INFORMADO ----------------")

print(f"Massa molar ....................: {MASSA_MOLAR_GNV:10.6f} kg/mol")

print(f"Densidade utilizada ............: {DENSIDADE_GNV:10.3f} kg/m³")

print(f"Fator de Compressibilidade (Z)..: {Z:10.3f}")

print()

print("---------------- PRESSÕES ----------------")

print(f"Pressão Atmosférica ............: {pressao_atmosferica:10.5f} bar")

print(f"Pressão Absoluta ...............: {pressao_absoluta:10.5f} bar")

print()

print("---------------- TEMPERATURA ----------------")

print(f"Temperatura em Celsius .........: {temperatura:10.2f} °C")

print(f"Temperatura em Kelvin ..........: {temperatura_kelvin:10.2f} K")

print()

print("---------------- VOLUMES ----------------")

print(f"Volume do cilindro .............: {volume_cilindro_m3:10.6f} m³")

print(f"Volume equivalente .............: {volume_equivalente:10.4f} m³")

print(f"Volume equivalente .............: {volume_equivalente_litros:10.2f} litros")

print()

print("---------------- MASSA ----------------")

print(f"Quantidade de mols .............: {mols:10.2f} mol")

print(f"Massa estimada do GNV ..........: {massa_gnv:10.4f} kg")

print(f"Densidade dentro do cilindro ...: {densidade_interna:10.2f} kg/m³")

print(f"Densidade nas CNTP .............: {massa_especifica:10.3f} kg/m³")

print()

print("---------------- ENERGIA ----------------")

print(f"PCS utilizado ..................: {PCS:10.2f} kWh/kg")

print(f"Energia armazenada .............: {energia:10.2f} kWh")

print()

print("---------------- AUTONOMIA ----------------")

print(f"Consumo médio utilizado ........: {consumo_medio:10.2f} km/m³")

print(f"Autonomia estimada .............: {autonomia:10.2f} km")

print()

print("---------------- ENCHIMENTO ----------------")

print(f"Percentual estimado ............: {percentual:10.2f} %")

print()

print("=" * 75)

# =============================================================================
# INFORMAÇÕES IMPORTANTES
# =============================================================================

print()

print("OBSERVAÇÕES")

print()

print("1 - Este cálculo utiliza a Lei dos Gases Ideais.")

print("2 - O fator Z reduz o erro para altas pressões.")

print("3 - A precisão depende dos dados informados.")

print("4 - A composição do GNV varia conforme a distribuidora.")

print("5 - Os resultados devem ser considerados aproximações.")

print()

print("=" * 75)

# =============================================================================
# PERGUNTA AO USUÁRIO
# =============================================================================

print()

# =============================================================================
# PARTE 06 - MENU PRINCIPAL
# =============================================================================

while True:

    print()
    print("=" * 75)
    print("               MENU PRINCIPAL")
    print("=" * 75)

    print("1 - Novo cálculo")
    print("2 - Alterar parâmetros do GNV")
    print("3 - Mostrar parâmetros atuais")
    print("0 - Sair")

    print()

    opcao = input("Escolha uma opção: ").strip()

    # ----------------------------------------------------------------------
    # NOVO CÁLCULO
    # ----------------------------------------------------------------------

    if opcao == "1":

        print()

        # ==============================================================
        # DADOS DO CILINDRO
        # ==============================================================

        volume_litros = float(
            input("Volume do cilindro (Litros): ")
        )

        pressao_bar = float(
            input("Pressão indicada no manômetro (bar): ")
        )

        temperatura = float(
            input("Temperatura do gás (°C): ")
        )

        altitude = float(
            input("Altitude (metros): ")
        )

        # ==============================================================
        # CÁLCULOS
        # ==============================================================

        pressao_atmosferica = calcular_pressao_atmosferica(
            altitude
        )

        pressao_absoluta = calcular_pressao_absoluta(
            pressao_bar,
            pressao_atmosferica
        )

        temperatura_kelvin = celsius_para_kelvin(
            temperatura
        )

        volume_cilindro_m3 = litros_para_m3(
            volume_litros
        )

        volume_equivalente = (
            volume_cilindro_m3
            *
            (pressao_absoluta / PRESSAO_ATMOSFERICA_PADRAO)
            *
            (TEMPERATURA_PADRAO / temperatura_kelvin)
            /
            Z
        )

        mols = calcular_mols(
            volume_cilindro_m3,
            pressao_absoluta,
            temperatura_kelvin,
            Z
        )

        massa_gnv = calcular_massa(
            mols
        )

        energia = massa_gnv * PCS

        consumo = input(
            "Consumo do veículo (km/m³) [13.5]: "
        ).strip()

        if consumo == "":

            consumo = 13.5

        else:

            consumo = float(consumo)

        autonomia = volume_equivalente * consumo

        # ==============================================================
        # RELATÓRIO
        # ==============================================================

        print()

        print("=" * 75)
        print("RESULTADO")
        print("=" * 75)

        print()

        print(f"Volume equivalente : {volume_equivalente:.3f} m³")

        print(f"Massa estimada     : {massa_gnv:.3f} kg")

        print(f"Energia            : {energia:.2f} kWh")

        print(f"Autonomia          : {autonomia:.1f} km")

        print()

        input("ENTER para continuar...")

    # ----------------------------------------------------------------------
    # ALTERA PARÂMETROS
    # ----------------------------------------------------------------------

    # ==============================================================
    # CÁLCULO MAIS DETALHADO
    # ==============================================================

    print()

    print("=" * 75)
    print("CÁLCULO DETALHADO")
    print("=" * 75)

    print()

    print(f"Pressão Atmosférica.............: {pressao_atmosferica:.5f} bar")

    print(f"Pressão Absoluta.................: {pressao_absoluta:.5f} bar")

    print(f"Temperatura Kelvin...............: {temperatura_kelvin:.2f} K")

    print(f"Volume do Cilindro...............: {volume_cilindro_m3:.6f} m³")

    print()

    # ==============================================================
    # GÁS IDEAL
    # ==============================================================

    volume_ideal = (
        volume_cilindro_m3
        *
        (pressao_absoluta / PRESSAO_ATMOSFERICA_PADRAO)
        *
        (TEMPERATURA_PADRAO / temperatura_kelvin)
    )

    print("USANDO GÁS IDEAL")

    print(f"Volume Equivalente...............: {volume_ideal:.4f} m³")

    print()

    # ==============================================================
    # GÁS REAL
    # ==============================================================

    volume_real = volume_ideal / Z

    print("USANDO FATOR Z")

    print(f"Fator Z..........................: {Z:.3f}")

    print(f"Volume Equivalente...............: {volume_real:.4f} m³")

    print()





# =============================================================================
# PRESSÃO MÁXIMA DO CILINDRO
# =============================================================================

#
# Normalmente os cilindros de GNV são abastecidos até
# aproximadamente 200 bar.
#
# Caso seu posto trabalhe com outra pressão,
# basta alterar este valor.
#

PRESSAO_MAXIMA = 200.0

# =============================================================================
# PERCENTUAL DE ENCHIMENTO
# =============================================================================

percentual_enchimento = (
    pressao_bar /
    PRESSAO_MAXIMA
) * 100

# Evita valores menores que zero

if percentual_enchimento < 0:

    percentual_enchimento = 0

# Evita valores maiores que 100%

if percentual_enchimento > 100:

    percentual_enchimento = 100

# =============================================================================
# GÁS RESTANTE
# =============================================================================

#
# Quanto ainda existe dentro do cilindro.
#

m3_restante = volume_real

kg_restante = massa_cnpt

# =============================================================================
# GÁS CONSUMIDO
# =============================================================================

#
# Considerando que o cilindro cheio possui
# o volume calculado para PRESSAO_MAXIMA.
#

volume_cheio = (
    volume_cilindro_m3
    *
    (
        (PRESSAO_MAXIMA + pressao_atmosferica)
        /
        PRESSAO_ATMOSFERICA_PADRAO
    )
    *
    (
        TEMPERATURA_PADRAO
        /
        temperatura_kelvin
    )
    /
    Z
)

m3_consumido = volume_cheio - m3_restante

if m3_consumido < 0:

    m3_consumido = 0

# =============================================================================
# BARRA DE ENCHIMENTO
# =============================================================================

#
# Barra com 50 posições.
#

barra_total = 50

barra_cheia = int(
    percentual_enchimento / 2
)

barra_vazia = barra_total - barra_cheia

barra = (
    "█" * barra_cheia
    +
    "-" * barra_vazia
)

print()

print("=" * 75)

print("NÍVEL DO CILINDRO")

print("=" * 75)

print()

print(f"[{barra}]")

print()

print(f"Enchimento.............: {percentual_enchimento:.1f}%")

print(f"GNV restante...........: {m3_restante:.3f} m³")

print(f"GNV consumido..........: {m3_consumido:.3f} m³")

print(f"Massa restante.........: {kg_restante:.3f} kg")

print()

print("=" * 75)



# =============================================================================
# DADOS DO ABASTECIMENTO
# =============================================================================

print()

print("DADOS DO ABASTECIMENTO")

print()

# -------------------------------------------------------------------------
# Quantidade abastecida
# -------------------------------------------------------------------------

texto = input(
    "Quantidade abastecida (m³): "
).strip()

if texto == "":

    volume_abastecido = 0.0

else:

    volume_abastecido = float(texto)

# -------------------------------------------------------------------------
# Preço por metro cúbico
# -------------------------------------------------------------------------

texto = input(
    "Preço do GNV (R$/m³): "
).strip()

if texto == "":

    preco_m3 = 0.0

else:

    preco_m3 = float(texto.replace(",", "."))

# -------------------------------------------------------------------------
# Quilometragem anterior
# -------------------------------------------------------------------------

texto = input(
    "Odômetro anterior (km): "
).strip()

if texto == "":

    odometro_anterior = 0.0

else:

    odometro_anterior = float(texto)

# -------------------------------------------------------------------------
# Quilometragem atual
# -------------------------------------------------------------------------

texto = input(
    "Odômetro atual (km): "
).strip()

if texto == "":

    odometro_atual = 0.0

else:

    odometro_atual = float(texto)

# =============================================================================
# CÁLCULOS
# =============================================================================

resultado = calcular_abastecimento(

    volume_abastecido,

    preco_m3,

    odometro_anterior,

    odometro_atual,

    massa_cnpt

)

valor_abastecimento = resultado["valor_total"]

distancia = resultado["distancia"]

rendimento = resultado["rendimento"]

custo_km = resultado["custo_km"]

custo_100km = resultado["custo_100km"]

custo_kg = resultado["custo_kg"]


# =============================================================================
# RESULTADOS
# =============================================================================

print()

print("RESULTADOS DO ABASTECIMENTO")

print()

print(f"Volume abastecido........: {volume_abastecido:.3f} m³")

print(f"Preço do GNV.............: R$ {preco_m3:.3f}")

print(f"Valor pago...............: R$ {valor_abastecimento:.2f}")

print()

print(f"Distância percorrida.....: {distancia:.1f} km")

print(f"Rendimento...............: {rendimento:.2f} km/m³")

print(f"Custo por km.............: R$ {custo_km:.3f}")

print(f"Custo por 100 km.........: R$ {custo_100km:.2f}")

print(f"Custo por kg.............: R$ {custo_kg:.2f}")

print()


# =============================================================================
# HISTÓRICO DOS ABASTECIMENTOS
# =============================================================================

#
# Esta lista armazenará todos os abastecimentos realizados
# durante a execução do programa.
#

if "historico_abastecimentos" not in globals():

    historico_abastecimentos = []

# =============================================================================
# DATA E HORA
# =============================================================================

from datetime import datetime

agora = datetime.now()

# =============================================================================
# REGISTRO
# =============================================================================

registro_abastecimento = {

    "Data": agora.strftime("%d/%m/%Y"),

    "Hora": agora.strftime("%H:%M:%S"),

    "Quantidade Cilindros": quantidade_cilindros,

    "Volume Cilindros (L)": volume_total_litros,

    "Pressão (bar)": pressao_bar,

    "Temperatura (°C)": temperatura,

    "Altitude (m)": altitude,

    "Fator Z": Z,

    "Volume GNV (m³)": volume_real,

    "Massa (kg)": massa_cnpt,

    "Energia (kWh)": energia,

    "Autonomia (km)": autonomia,

    "Abastecido (m³)": volume_abastecido,

    "Preço m³": preco_m3,

    "Valor Pago": valor_abastecimento,

    "Odômetro Inicial": odometro_anterior,

    "Odômetro Final": odometro_atual,

    "Distância": distancia,

    "Rendimento": rendimento,

    "R$/km": custo_km,

    "R$/100km": custo_100km,

    "R$/kg": custo_kg

}

historico_abastecimentos.append(
    registro_abastecimento
)

# =============================================================================
# RESUMO
# =============================================================================

print()

print("Abastecimento gravado com sucesso.")

print()

print(
    f"Total de abastecimentos registrados: "
    f"{len(historico_abastecimentos)}"
)

print()


# =============================================================================
# MOSTRAR HISTÓRICO
# =============================================================================

resposta = input(
    "Deseja visualizar os abastecimentos (S/N)? "
).strip().upper()

if resposta == "S":

    print()

    print("HISTÓRICO DOS ABASTECIMENTOS")

    print()

    for numero, item in enumerate(
            historico_abastecimentos,
            start=1):

        print("-" * 70)

        print(f"Registro {numero}")

        print("-" * 70)

        print(
            f"Data...............: {item['Data']} {item['Hora']}"
        )

        print(
            f"Volume.............: {item['Abastecido (m³)']:.3f} m³"
        )

        print(
            f"Valor Pago.........: R$ {item['Valor Pago']:.2f}"
        )

        print(
            f"Distância..........: {item['Distância']:.1f} km"
        )

        print(
            f"Rendimento.........: {item['Rendimento']:.2f} km/m³"
        )

        print(
            f"Custo por km.......: R$ {item['R$/km']:.3f}"
        )

        print()


# =============================================================================
# SALVAR HISTÓRICO EM CSV
# =============================================================================

#
# Nome do arquivo
#

ARQUIVO_CSV = "Historico_GNV.csv"

#
# Cabeçalho da planilha
#

cabecalho = [

    "Data",

    "Hora",

    "Quantidade Cilindros",

    "Volume Cilindros (L)",

    "Pressão (bar)",

    "Temperatura (°C)",

    "Altitude (m)",

    "Fator Z",

    "Volume GNV (m³)",

    "Massa (kg)",

    "Energia (kWh)",

    "Autonomia (km)",

    "Abastecido (m³)",

    "Preço m³",

    "Valor Pago",

    "Odômetro Inicial",

    "Odômetro Final",

    "Distância",

    "Rendimento",

    "R$/km",

    "R$/100km",

    "R$/kg"

]

#
# Se o arquivo não existir cria um novo
#

arquivo_novo = not os.path.exists(
    ARQUIVO_CSV
)

with open(

    ARQUIVO_CSV,

    "a",

    newline="",

    encoding="utf-8"

) as arquivo:

    escritor = csv.writer(arquivo)

    if arquivo_novo:

        escritor.writerow(cabecalho)

    escritor.writerow([

        registro_abastecimento["Data"],

        registro_abastecimento["Hora"],

        registro_abastecimento["Quantidade Cilindros"],

        registro_abastecimento["Volume Cilindros (L)"],

        registro_abastecimento["Pressão (bar)"],

        registro_abastecimento["Temperatura (°C)"],

        registro_abastecimento["Altitude (m)"],

        registro_abastecimento["Fator Z"],

        registro_abastecimento["Volume GNV (m³)"],

        registro_abastecimento["Massa (kg)"],

        registro_abastecimento["Energia (kWh)"],

        registro_abastecimento["Autonomia (km)"],

        registro_abastecimento["Abastecido (m³)"],

        registro_abastecimento["Preço m³"],

        registro_abastecimento["Valor Pago"],

        registro_abastecimento["Odômetro Inicial"],

        registro_abastecimento["Odômetro Final"],

        registro_abastecimento["Distância"],

        registro_abastecimento["Rendimento"],

        registro_abastecimento["R$/km"],

        registro_abastecimento["R$/100km"],

        registro_abastecimento["R$/kg"]

    ])

print()

print("Histórico salvo em:")

print(os.path.abspath(ARQUIVO_CSV))

print()

resposta = input(
    "Deseja exportar o histórico para Excel (S/N)? "
).strip().upper()

if resposta == "S":

    exportar_excel(
        historico_abastecimentos
    )



# =============================================================================
# PARTE 15 - ABASTECIMENTO DO POSTO
# =============================================================================

print()

print("=" * 75)
print("COMPARAÇÃO COM O ABASTECIMENTO")
print("=" * 75)

print()

print(
    "Informe quanto apareceu na bomba do posto."
)

print(
    "Caso não queira comparar, pressione ENTER."
)

print()

texto = input(
    "Volume abastecido (m³): "
).strip()

# -------------------------------------------------------------------------
# O usuário pode simplesmente pressionar ENTER.
# -------------------------------------------------------------------------

if texto != "":

    volume_bomba = float(texto)

    print()

    # ---------------------------------------------------------------------
    # Diferença absoluta
    # ---------------------------------------------------------------------

    diferenca_m3 = volume_bomba - volume_real

    diferenca_litros = diferenca_m3 * 1000

    # ---------------------------------------------------------------------
    # Diferença percentual
    # ---------------------------------------------------------------------

    if volume_real != 0:

        diferenca_percentual = (
            diferenca_m3 /
            volume_real
        ) * 100

    else:

        diferenca_percentual = 0

    print("=" * 75)

    print("COMPARAÇÃO")

    print("=" * 75)

    print()

    print(
        f"Calculado...............: {volume_real:.3f} m³"
    )

    print(
        f"Bomba...................: {volume_bomba:.3f} m³"
    )

    print(
        f"Diferença...............: {diferenca_m3:.3f} m³"
    )

    print(
        f"Diferença...............: {diferenca_litros:.1f} litros"
    )

    print(
        f"Diferença...............: {diferenca_percentual:.2f}%"
    )

    print()

    # ---------------------------------------------------------------------
    # Avaliação
    # ---------------------------------------------------------------------

    tolerancia = 3.0

    if abs(diferenca_percentual) <= tolerancia:

        print("Resultado dentro da tolerância.")

    else:

        print("ATENÇÃO")

        print(
            "Diferença acima da tolerância."
        )

        print(
            "Verifique pressão, temperatura e fator Z."
        )

    print()

    print("=" * 75)






    # ==============================================================
    # DIFERENÇA ENTRE OS MÉTODOS
    # ==============================================================

    diferenca = volume_real - volume_ideal

    erro = (
        diferenca /
        volume_ideal
    ) * 100

    print("COMPARAÇÃO")

    print(f"Diferença.........................: {diferenca:.4f} m³")

    print(f"Erro..............................: {erro:.2f} %")

    print()

    # ==============================================================
    # MASSA ESTIMADA
    # ==============================================================

    massa_cnpt = volume_real * DENSIDADE_GNV

    print(f"Massa estimada....................: {massa_cnpt:.3f} kg")

    print()

    # ==============================================================
    # AUTONOMIA
    # ==============================================================

    autonomia = volume_real * consumo

    print(f"Autonomia estimada................: {autonomia:.1f} km")

    print()

    print("=" * 75)

    input("ENTER para continuar...")

    elif opcao == "2":

        print()

        texto = input(
            f"Massa molar [{MASSA_MOLAR_GNV}]: "
        ).strip()

        if texto != "":

            MASSA_MOLAR_GNV = float(texto)

        texto = input(
            f"Densidade [{DENSIDADE_GNV}]: "
        ).strip()

        if texto != "":

            DENSIDADE_GNV = float(texto)

        texto = input(
            f"Fator Z [{Z}]: "
        ).strip()

        if texto != "":

            Z = float(texto)

        print()

        print("Parâmetros atualizados.")

    # ----------------------------------------------------------------------
    # MOSTRA PARÂMETROS
    # ----------------------------------------------------------------------

    elif opcao == "3":

        print()

        print("=" * 75)

        print("PARÂMETROS")

        print("=" * 75)

        print()

        print(f"Massa molar : {MASSA_MOLAR_GNV}")

        print(f"Densidade   : {DENSIDADE_GNV}")

        print(f"Fator Z     : {Z}")

        print()

        input("ENTER para voltar ao menu...")

    # ----------------------------------------------------------------------
    # SAIR
    # ----------------------------------------------------------------------

    elif opcao == "0":

        print()

        print("Encerrando...")

        break

    # ----------------------------------------------------------------------
    # OPÇÃO INVÁLIDA
    # ----------------------------------------------------------------------

    else:

        print()

        print("Opção inválida.")

        input("ENTER para continuar...")

# =============================================================================
# HISTÓRICO DOS CÁLCULOS
# =============================================================================

#
# Esta parte cria um histórico simples dos cálculos realizados.
#
# Futuramente esse histórico poderá ser salvo em CSV,
# Excel ou banco de dados.
#

try:
    historico
except NameError:
    historico = []

# =============================================================================
# REGISTRO DO CÁLCULO
# =============================================================================

registro = {

    "Volume Cilindro (L)" : volume_litros,

    "Pressão (bar)" : pressao_bar,

    "Altitude (m)" : altitude,

    "Temperatura (°C)" : temperatura,

    "Temperatura (K)" : temperatura_kelvin,

    "Pressão Atmosférica" : pressao_atmosferica,

    "Pressão Absoluta" : pressao_absoluta,

    "Fator Z" : Z,

    "Volume Equivalente (m³)" : volume_real,

    "Massa (kg)" : massa_cnpt,

    "Energia (kWh)" : energia,

    "Autonomia (km)" : autonomia

}

historico.append(registro)

# =============================================================================
# MENU DE CONSULTA
# =============================================================================

print()

print("=" * 75)

print("HISTÓRICO")

print("=" * 75)

print()

print(f"Foram realizados {len(historico)} cálculo(s).")

print()

mostrar = input(
    "Mostrar histórico (S/N)? "
).strip().upper()

print()

if mostrar == "S":

    print("=" * 75)

    for numero, item in enumerate(historico, start=1):

        print()

        print("-" * 75)

        print(f"CÁLCULO Nº {numero}")

        print("-" * 75)

        print(
            f"Volume do cilindro : {item['Volume Cilindro (L)']:.2f} L"
        )

        print(
            f"Pressão            : {item['Pressão (bar)']:.2f} bar"
        )

        print(
            f"Temperatura        : {item['Temperatura (°C)']:.2f} °C"
        )

        print(
            f"Altitude           : {item['Altitude (m)']:.2f} m"
        )

        print(
            f"Volume equivalente : {item['Volume Equivalente (m³)']:.3f} m³"
        )

        print(
            f"Massa              : {item['Massa (kg)']:.3f} kg"
        )

        print(
            f"Energia            : {item['Energia (kWh)']:.2f} kWh"
        )

        print(
            f"Autonomia          : {item['Autonomia (km)']:.1f} km"
        )

        print()

print()



# =============================================================================
# PARTE 16 - HISTÓRICO DOS CÁLCULOS
# =============================================================================

#
# O histórico ficará armazenado apenas na memória.
#
# Futuramente poderemos gravar em CSV ou SQLite.
#

# Se a lista ainda não existir, cria uma.

if "historico" not in globals():

    historico = []

# Cria um dicionário contendo todas as informações do cálculo.

calculo = {

    "volume_litros": volume_litros,

    "pressao_bar": pressao_bar,

    "temperatura_c": temperatura,

    "temperatura_k": temperatura_kelvin,

    "altitude": altitude,

    "pressao_atmosferica": pressao_atmosferica,

    "pressao_absoluta": pressao_absoluta,

    "fator_z": Z,

    "volume_cilindro_m3": volume_cilindro_m3,

    "volume_equivalente_m3": volume_real,

    "massa_kg": massa_cnpt,

    "energia_kwh": energia,

    "autonomia_km": autonomia

}

# Adiciona o cálculo ao histórico.

historico.append(calculo)

# =============================================================================
# EXIBE UM RESUMO DO CÁLCULO
# =============================================================================

print()

print("=" * 75)

print("RESUMO")

print("=" * 75)

print()

print(f"Volume equivalente : {volume_real:.3f} m³")

print(f"Massa              : {massa_cnpt:.3f} kg")

print(f"Energia            : {energia:.2f} kWh")

print(f"Autonomia          : {autonomia:.1f} km")

print()

print(f"Histórico possui {len(historico)} cálculo(s).")

print()

# =============================================================================
# MOSTRAR HISTÓRICO
# =============================================================================

mostrar = input(
    "Deseja visualizar o histórico (S/N)? "
).strip().upper()

if mostrar == "S":

    print()

    print("=" * 75)

    print("HISTÓRICO DOS CÁLCULOS")

    print("=" * 75)

    print()

    for indice, item in enumerate(historico, start=1):

        print("-" * 75)

        print(f"Cálculo nº {indice}")

        print("-" * 75)

        print(f"Volume cilindro : {item['volume_litros']:.2f} L")

        print(f"Pressão         : {item['pressao_bar']:.2f} bar")

        print(f"Temperatura     : {item['temperatura_c']:.2f} °C")

        print(f"Altitude        : {item['altitude']:.2f} m")

        print(f"Volume em m³    : {item['volume_equivalente_m3']:.3f}")

        print(f"Massa           : {item['massa_kg']:.3f} kg")

        print(f"Energia         : {item['energia_kwh']:.2f} kWh")

        print(f"Autonomia       : {item['autonomia_km']:.1f} km")

        print()



input("ENTER para retornar ao menu...")



