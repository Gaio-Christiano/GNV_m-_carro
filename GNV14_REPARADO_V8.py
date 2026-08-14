# =============================================================================
# ARQUIVO.....: GNV14_REPARADO_V17.py
# AUTOR.......: Christiano Gaio
# OBJETIVO....: Calculadora de GNV
#
# DESCRIÇÃO
# ----------
# Este programa estima a quantidade de GNV em um cilindro usando um modelo
# de gás real simplificado, PV = Z n R T, com Z informado pelo usuário.
#
# OBSERVAÇÃO IMPORTANTE
# ---------------------
# O GNV é um gás REAL.
# Portanto, em pressões elevadas (200 a 220 bar) existe um erro quando usamos
# apenas a Lei dos Gases Ideais.
#
# Limitação importante: Z é fornecido como aproximação. Para maior precisão em
# alta pressão, é necessário calcular Z(P,T,composição) por um modelo validado,
# como AGA8/GERG, e conhecer a temperatura real do gás durante o abastecimento.
#
# Todo o código possui comentários para facilitar o aprendizado.
# =============================================================================






# =============================================================================
# PARTE 01 - IMPORTAÇÃO DAS BIBLIOTECAS
# =============================================================================

# Biblioteca matemática

import math
import csv
import os
import sqlite3
import tkinter as tk
import pandas as pd
import json
from tkinter import ttk

# =============================================================================
# PARTE 253
# IMPORT DATETIME
# =============================================================================

from datetime import datetime
from tkinter import messagebox
from tkinter import filedialog

# =============================================================================
# PARTE 246
# IMPORT FPDF
# =============================================================================
from fpdf import FPDF



# =============================================================================
# PARTE 02 - CONSTANTES FÍSICAS
# =============================================================================

# Pressão atmosférica padrão ao nível do mar (bar)
PRESSAO_ATMOSFERICA_PADRAO = 1.01325

# Temperatura padrão (Kelvin)
TEMPERATURA_PADRAO = 273.15

# Constante universal dos gases
#
# Unidade:
#
# J/(mol.K)
#
R = 8.314462618

# Massa molar aproximada do GNV
#
# O gás natural muda de composição conforme a distribuidora.
#
# Foi adotado um valor médio.
#

# -------------------------------------------------------------------------
# Estes valores mudam conforme a composição do gás fornecido pela
# distribuidora.
#
# O usuário poderá alterá-los na execução do programa.
# -------------------------------------------------------------------------

MASSA_MOLAR_GNV = None      # kg/mol

# Densidade média do GNV nas CNTP
#
# Este valor varia mensalmente.
#
DENSIDADE_GNV = None           # kg/m³

# Fator de compressibilidade
#
# Z = 1,000  -> gás ideal
#
# Para GNV normalmente varia aproximadamente entre
#
# 0,82 e 0,95
#
Z = None

# Idiomas disponíveis na interface.
IDIOMAS_DISPONIVEIS = ("pt-BR", "English", "Español", "Français", "Italiano", "Deutsch")
IDIOMA_TABS = {
    # Ordem oficial das 11 abas do sistema.
    "pt-BR": [
        "Cálculos", "Abastecimentos", "ANP", "Aquecimento / Compressão",
        "Histórico de Abastecimentos", "Banco SQLite", "Exportação / Excel",
        "Gráficos de Abastecimento", "Configurações do Sistema",
        "Fórmulas e Física", "Total de Abastecimentos"
    ],
    "English": [
        "Calculations", "Refuelings", "ANP", "Heating / Compression",
        "Refueling History", "SQLite Database", "Export / Excel",
        "Refueling Charts", "System Settings", "Formulas & Physics",
        "Total Refuelings"
    ],
    "Español": [
        "Cálculos", "Abastecimientos", "ANP", "Calentamiento / Compresión",
        "Historial de Abastecimientos", "Base SQLite", "Exportación / Excel",
        "Gráficos de Abastecimiento", "Configuración del Sistema",
        "Fórmulas y Física", "Total de Abastecimientos"
    ],
    "Français": [
        "Calculs", "Ravitaillements", "ANP", "Chauffage / Compression",
        "Historique des Ravitaillements", "Base SQLite", "Exportation / Excel",
        "Graphiques de Ravitaillement", "Paramètres du Système",
        "Formules et Physique", "Total des Ravitaillements"
    ],
    "Italiano": [
        "Calcoli", "Rifornimenti", "ANP", "Riscaldamento / Compressione",
        "Storico Rifornimenti", "Database SQLite", "Esportazione / Excel",
        "Grafici dei Rifornimenti", "Impostazioni di Sistema",
        "Formule e Fisica", "Totale Rifornimenti"
    ],
    "Deutsch": [
        "Berechnungen", "Tankvorgänge", "ANP", "Erwärmung / Kompression",
        "Tankvorgangsverlauf", "SQLite-Datenbank", "Export / Excel",
        "Tankdiagramme", "Systemeinstellungen", "Formeln & Physik",
        "Gesamte Tankvorgänge"
    ],
}

# =============================================================================
# PARTE 02A - CONDIÇÕES DE REFERÊNCIA DA ANP
# =============================================================================

# A ANP publica volumes de gás natural equivalentes a 20 °C e 1,033 kgf/cm².
# A condição padrão de medição também é expressa como 0,101325 MPa a 20 °C.
# Para os cálculos internos, usamos a forma SI equivalente: 1,01325 bar.
TEMPERATURA_REFERENCIA_ANP_C = 20.0
PRESSAO_REFERENCIA_ANP_KGF_CM2 = 1.033
PRESSAO_REFERENCIA_ANP_BAR = 1.01325

# =============================================================================
# PARTE 02A1 - MODELO DE REFERÊNCIA ANP
# =============================================================================

def calcular_volume_anp_referencia(
    capacidade_cilindro_l,
    pressao_inicial_bar,
    pressao_final_bar,
    temperatura_ambiente_c,
    altitude_m=0.0
):
    """Estima o volume equivalente usando a condição de referência da ANP.

    Este cálculo NÃO é uma reprodução do algoritmo interno do dispenser.
    É uma estimativa baseada na condição de referência publicada pela ANP,
    no volume físico do cilindro e na aproximação de gás ideal (Z=1).

    A temperatura disponível ao usuário é a temperatura ambiente; ela não é
    tratada como medição da temperatura real do gás durante a compressão.
    """
    V = capacidade_cilindro_l / 1000.0
    T = temperatura_ambiente_c + 273.15
    Tref = TEMPERATURA_REFERENCIA_ANP_C + 273.15
    Patm = calcular_pressao_atmosferica(altitude_m)
    Pi = pressao_inicial_bar + Patm
    Pf = pressao_final_bar + Patm

    if V <= 0 or T <= 0 or Pf < Pi:
        raise ValueError("Dados inválidos para o cálculo pela referência ANP.")

    return V * ((Pf - Pi) / PRESSAO_REFERENCIA_ANP_BAR) * (Tref / T)


def calcular_volume_cientifico_gas_real(
    capacidade_cilindro_l,
    pressao_inicial_bar,
    pressao_final_bar,
    temperatura_ambiente_c,
    altitude_m,
    fator_z,
    massa_molar,
    temperatura_referencia_c=20.0,
    pressao_referencia_bar=1.01325
):
    """Modelo físico de gás real usando Z informado pelo usuário.

    Usa PV = Z n R T, pressão absoluta e temperatura absoluta.
    Como a temperatura real do gás no interior do cilindro durante o
    abastecimento não é medida pelo usuário, a temperatura ambiente é
    utilizada como aproximação e isso é explicitado no resultado.

    O modelo é mais completo que a aproximação ideal porque permite Z != 1,
    mas NÃO é AGA8/GERG-2008: para alta precisão metrológica seriam
    necessários composição do gás e propriedades termodinâmicas validadas.
    """
    V = capacidade_cilindro_l / 1000.0
    T = temperatura_ambiente_c + 273.15
    Tref = temperatura_referencia_c + 273.15
    Patm = calcular_pressao_atmosferica(altitude_m)
    Pi = pressao_inicial_bar + Patm
    Pf = pressao_final_bar + Patm

    if V <= 0 or T <= 0 or fator_z <= 0 or massa_molar <= 0:
        raise ValueError("Parâmetros físicos inválidos no modelo de gás real.")
    if Pf < Pi:
        raise ValueError("A pressão final deve ser maior ou igual à pressão inicial.")

    # n = PV/(ZRT)
    n_i = (Pi * 100000.0 * V) / (fator_z * R * T)
    n_f = (Pf * 100000.0 * V) / (fator_z * R * T)
    delta_n = max(0.0, n_f - n_i)
    massa = delta_n * massa_molar

    # Conversão dos mols para a condição de referência, assumindo Zref=1.
    P_ref = pressao_referencia_bar * 100000.0
    volume_ref = delta_n * R * Tref / P_ref

    return {
        "capacidade_m3": V,
        "pressao_atmosferica_bar": Patm,
        "pressao_inicial_absoluta_bar": Pi,
        "pressao_final_absoluta_bar": Pf,
        "temperatura_kelvin": T,
        "mols_iniciais": n_i,
        "mols_finais": n_f,
        "mols_adicionados": delta_n,
        "massa_adicionada_kg": massa,
        "volume_referencia_m3": volume_ref,
        "volume_referencia_litros": volume_ref * 1000.0,
        "fator_z": fator_z,
        "temperatura_referencia_c": temperatura_referencia_c,
        "pressao_referencia_bar": pressao_referencia_bar,
        "temperatura_eh_ambiente": True
    }


# =============================================================================
# PARTE 02A - CONVERSÃO NUMÉRICA PT-BR / INTERNACIONAL
# =============================================================================

def converter_numero(valor):
    """
    Aceita números nos formatos mais comuns no Brasil e internacional.

    Exemplos aceitos:
        24,5
        24.5
        1.234,56
        1,234.56
        1234
    """
    texto = str(valor).strip().replace(" ", "")

    if not texto:
        raise ValueError("Valor numérico vazio.")

    if "," in texto and "." in texto:
        # O último separador é tratado como separador decimal.
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")

    elif "," in texto:
        texto = texto.replace(",", ".")

    return float(texto)


def formatar_numero_br(valor, casas=2):
    """Formata número para exibição brasileira, sem alterar o valor salvo."""
    return f"{float(valor):.{casas}f}".replace(".", ",")


# =============================================================================
# PARTE 02B - VOLUME EQUIVALENTE NA CONDIÇÃO DE REFERÊNCIA
# =============================================================================

def calcular_volume_referencia_m3(
    mols,
    temperatura_referencia_c=20.0,
    pressao_referencia_bar=1.01325,
    fator_z_referencia=1.0
):
    """Converte mols calculados para m³ na condição de referência adotada.

    Referência padrão do aplicativo: 20 °C e 1,01325 bar, com Z=1.
    O resultado é explicitamente chamado de volume equivalente, não de
    volume físico ocupado pelo cilindro.
    """
    temperatura_k = temperatura_referencia_c + 273.15
    pressao_pa = pressao_referencia_bar * 100000.0

    if mols <= 0 or temperatura_k <= 0 or pressao_pa <= 0:
        return 0.0

    return (
        mols * R * temperatura_k * fator_z_referencia
        / pressao_pa
    )


def calcular_volume_equivalente_na_temperatura_m3(mols, temperatura_c, pressao_referencia_bar=1.01325, fator_z_referencia=1.0):
    """Volume equivalente dos mesmos mols na temperatura informada."""
    temperatura_k = temperatura_c + 273.15
    pressao_pa = pressao_referencia_bar * 100000.0
    if mols <= 0 or temperatura_k <= 0 or pressao_pa <= 0:
        return 0.0
    return mols * R * temperatura_k * fator_z_referencia / pressao_pa


def calcular_compressao_ideal_adiabatica(pressao_inicial_manometrica_bar, pressao_final_manometrica_bar, temperatura_inicial_c, altitude_m=0.0, k=1.294):
    """Modelo didático de compressão reversível adiabática; não é o enchimento real do cilindro."""
    if k <= 1.0: raise ValueError("O expoente k deve ser maior que 1.")
    if temperatura_inicial_c <= -273.15: raise ValueError("A temperatura inicial deve ser maior que -273,15 °C.")
    patm=calcular_pressao_atmosferica(altitude_m); p1=pressao_inicial_manometrica_bar+patm; p2=pressao_final_manometrica_bar+patm; t1=temperatura_inicial_c+273.15
    if p1 <= 0 or p2 <= 0 or p2 < p1: raise ValueError("As pressões devem ser positivas e a final maior ou igual à inicial.")
    t2=t1*(p2/p1)**((k-1.0)/k); vr=(p1/p2)**(1.0/k)
    return {"pressao_atmosferica_bar":patm,"pressao_inicial_absoluta_bar":p1,"pressao_final_absoluta_bar":p2,"temperatura_inicial_c":temperatura_inicial_c,"temperatura_final_c":t2-273.15,"temperatura_final_k":t2,"aumento_temperatura_c":t2-t1,"volume_relativo_final":vr,"reducao_volume_percentual":(1-vr)*100,"k":k}


def gerar_pontos_compressao_adiabatica(pressao_inicial_manometrica_bar, pressao_final_manometrica_bar, temperatura_inicial_c, altitude_m=0.0, k=1.294, pontos=40):
    patm=calcular_pressao_atmosferica(altitude_m); p1=pressao_inicial_manometrica_bar+patm; p2=pressao_final_manometrica_bar+patm; t1=temperatura_inicial_c+273.15
    if p1 <= 0 or p2 < p1 or k <= 1: raise ValueError("Parâmetros inválidos para o gráfico de compressão.")
    n=max(2,int(pontos)); dados=[]
    for i in range(n):
        f=i/(n-1); pa=p1+(p2-p1)*f; tk=t1*(pa/p1)**((k-1)/k); vr=(p1/pa)**(1/k); dados.append({"pressao_man_bar":pa-patm,"temperatura_c":tk-273.15,"volume_relativo":vr})
    return dados


# =============================================================================
# PARTE 02C - ANÁLISE FÍSICA DO ABASTECIMENTO
# =============================================================================

def calcular_comparacao_abastecimento(
    capacidade_cilindro_l,
    pressao_inicial_bar,
    pressao_final_bar,
    temperatura_c,
    altitude_m,
    fator_z,
    massa_molar,
    temperatura_referencia_c=20.0,
    pressao_referencia_bar=1.01325
):
    """Compatibilidade: retorna o modelo científico de gás real."""
    resultado = calcular_volume_cientifico_gas_real(
        capacidade_cilindro_l,
        pressao_inicial_bar,
        pressao_final_bar,
        temperatura_c,
        altitude_m,
        fator_z,
        massa_molar,
        temperatura_referencia_c,
        pressao_referencia_bar
    )
    resultado["delta_pressao_bar"] = max(0.0, pressao_final_bar - pressao_inicial_bar)
    resultado["volume_teorico_m3"] = resultado["volume_referencia_m3"]
    resultado["massa_adicionada_kg"] = resultado["massa_adicionada_kg"]
    resultado["temperatura_referencia_c"] = temperatura_referencia_c
    resultado["pressao_referencia_bar"] = pressao_referencia_bar
    return resultado


# =============================================================================
# PARTE 03 - FUNÇÃO
# Calcula a pressão atmosférica em função da altitude.
# =============================================================================

def calcular_pressao_atmosferica(altitude):

    """
    Calcula a pressão atmosférica.

    Entrada

        altitude em metros

    Saída

        pressão em bar

    Fórmula aproximada válida para pequenas altitudes.
    """

    # Equação barométrica simplificada

    pressao = PRESSAO_ATMOSFERICA_PADRAO * \
              (1 - 2.25577e-5 * altitude) ** 5.25588

    return pressao





# =============================================================================
# PARTE 04 - FUNÇÃO
# Converte Celsius para Kelvin
# =============================================================================

def celsius_para_kelvin(temperatura):

    return temperatura + 273.15

# =============================================================================
# FUNÇÃO
# Converte litros para metros cúbicos
# =============================================================================

def litros_para_m3(litros):

    return litros / 1000.0

# =============================================================================
# FUNÇÃO
# Calcula pressão absoluta
# =============================================================================

def calcular_pressao_absoluta(pressao_manometrica,
                              pressao_atmosferica):

    """
    A pressão mostrada no manômetro é relativa.

    Para utilizar a Lei dos Gases devemos usar pressão absoluta.

    Pabs = Pman + Patm
    """

    return pressao_manometrica + pressao_atmosferica

# =============================================================================
# FUNÇÃO
# Calcula quantidade de mols
# =============================================================================

def calcular_mols(volume_m3,
                  pressao_bar,
                  temperatura_kelvin,
                  fator_z):

    """
    Calcula a quantidade de matéria.

    Equação

    PV = ZnRT

    Quanto menor o fator Z,
    maior será a quantidade de gás para
    a mesma pressão.
    """

    pressao_pa = pressao_bar * 100000

    mols = (
        pressao_pa *
        volume_m3
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return mols

# =============================================================================
# FUNÇÃO
# Calcula massa do gás
# =============================================================================

def calcular_massa(mols):

    """
    Massa = mol × Massa molar
    """

    return mols * MASSA_MOLAR_GNV



# =============================================================================
# FUNÇÃO
# Calcula volume equivalente em condições normais
# =============================================================================

def calcular_volume_equivalente(massa):

    """
    m³ = massa / densidade
    """

    return massa / DENSIDADE_GNV




# =============================================================================
# FUNÇÃO PARA CALCULAR O ABASTECIMENTO
# =============================================================================

def calcular_abastecimento(
    volume_abastecido,
    preco_m3,
    odometro_anterior,
    odometro_atual,
    massa_gnv
):
    """
    Calcula todas as informações referentes ao abastecimento.

    Parâmetros
    ----------
    volume_abastecido : float
        Volume abastecido em m³.

    preco_m3 : float
        Preço do GNV por metro cúbico.

    odometro_anterior : float
        Quilometragem anterior.

    odometro_atual : float
        Quilometragem atual.

    massa_gnv : float
        Massa estimada do GNV.

    Retorno
    -------
    dict
        Dicionário contendo todos os resultados.
    """

    valor_total = volume_abastecido * preco_m3

    distancia = odometro_atual - odometro_anterior

    if distancia < 0:

        distancia = 0

    if volume_abastecido > 0:

        rendimento = distancia / volume_abastecido

    else:

        rendimento = 0

    if distancia > 0:

        custo_km = valor_total / distancia

    else:

        custo_km = 0

    if massa_gnv > 0:

        custo_kg = valor_total / massa_gnv

    else:

        custo_kg = 0

    custo_100km = custo_km * 100

    return {

        "valor_total": valor_total,

        "distancia": distancia,

        "rendimento": rendimento,

        "custo_km": custo_km,

        "custo_100km": custo_100km,

        "custo_kg": custo_kg

    }




# =============================================================================
# PARTE 21
# FUNÇÃO PARA CALCULAR O VALOR ENERGÉTICO DO GNV
# =============================================================================

def calcular_energia_gnv(
    volume_m3,
    pcs=39.5
):
    """
    Calcula a energia contida no GNV.

    volume_m3
        Volume do gás em metros cúbicos.

    pcs
        Poder Calorífico Superior
        MJ/m³

    Retorna

        energia_mj
        energia_kwh
    """

    energia_mj = volume_m3 * pcs

    energia_kwh = energia_mj / 3.6

    return {

        "MJ": energia_mj,

        "kWh": energia_kwh

    }


# =============================================================================
# PARTE 22
# FUNÇÃO PARA CALCULAR A MASSA DO GNV
# =============================================================================

def calcular_massa_gnv(
    volume_m3,
    densidade=0.717
):
    """
    Calcula a massa do GNV.

    densidade padrão em CNTP.

    Retorna a massa em quilogramas.
    """

    return volume_m3 * densidade


# =============================================================================
# PARTE 25
# CÁLCULO DA DENSIDADE REAL DO GNV
# =============================================================================

def calcular_densidade_real(

    massa,

    volume

):
    """
    Calcula a densidade real do gás.

    densidade = massa / volume

    Retorna kg/m³.
    """

    if volume <= 0:

        return 0.0

    return massa / volume

# =============================================================================
# PARTE 26
# PESO DO GNV
# =============================================================================

def calcular_peso(

    massa

):
    """
    Calcula o peso do gás.

    Retorna Newton.
    """

    G = 9.80665

    return massa * G


# =============================================================================
# PARTE 27
# CÁLCULO DA ENERGIA PELO PESO DO GNV
# =============================================================================

def calcular_energia_por_massa(
    massa_kg,
    pci=50.0
):
    """
    Calcula a energia disponível no GNV.

    Parâmetros
    ----------
    massa_kg : float
        Massa do GNV em quilogramas.

    pci : float
        Poder calorífico inferior em MJ/kg.
        Valor padrão aproximado para GNV.

    Retorno
    -------
    dict contendo:

        energia_mj
        energia_kwh
    """

    energia_mj = massa_kg * pci

    energia_kwh = energia_mj / 3.6

    return {

        "energia_mj": energia_mj,

        "energia_kwh": energia_kwh

    }

# =============================================================================
# PARTE 28
# CÁLCULO DO CONSUMO DO VEÍCULO
# =============================================================================

def calcular_consumo(

    distancia_km,

    volume_m3,

    massa_kg

):
    """
    Calcula diversos índices de consumo.

    Retorna um dicionário contendo:

        km_por_m3

        m3_por_100km

        km_por_kg

        kg_por_100km
    """

    resultado = {}

    if volume_m3 > 0:

        resultado["km_por_m3"] = (
            distancia_km /
            volume_m3
        )

        resultado["m3_por_100km"] = (
            volume_m3 /
            distancia_km
        ) * 100 if distancia_km > 0 else 0

    else:

        resultado["km_por_m3"] = 0

        resultado["m3_por_100km"] = 0

    if massa_kg > 0:

        resultado["km_por_kg"] = (
            distancia_km /
            massa_kg
        )

        resultado["kg_por_100km"] = (
            massa_kg /
            distancia_km
        ) * 100 if distancia_km > 0 else 0

    else:

        resultado["km_por_kg"] = 0

        resultado["kg_por_100km"] = 0

    return resultado


# =============================================================================
# PARTE 29
# CÁLCULO DA MASSA DO GNV PELA DENSIDADE
# =============================================================================

def calcular_massa_por_densidade(

    volume_m3,

    densidade

):
    """
    Calcula a massa do GNV.

    massa = volume × densidade

    volume_m3
        Volume em metros cúbicos

    densidade
        kg/m³

    Retorna
        Massa em quilogramas.
    """

    return volume_m3 * densidade


# =============================================================================
# PARTE 30
# DENSIDADE APARENTE
# =============================================================================

def calcular_densidade_aparente(

    massa,

    volume

):
    """
    Calcula a densidade aparente.

    densidade = massa / volume
    """

    if volume <= 0:

        return 0

    return massa / volume


# =============================================================================
# PARTE 31
# ENERGIA ESPECÍFICA
# =============================================================================

def calcular_energia_especifica(

    energia_kwh,

    massa

):
    """
    Calcula a energia específica.

    kWh/kg
    """

    if massa <= 0:

        return 0

    return energia_kwh / massa

# =============================================================================
# PARTE 32
# CÁLCULO DO VOLUME LIVRE DO CILINDRO
# =============================================================================

def calcular_volume_livre(

    volume_total_litros,

    percentual_ocupacao

):
    """
    Calcula o volume livre existente
    no cilindro.

    percentual_ocupacao

        0 a 100 %
    """

    if percentual_ocupacao < 0:

        percentual_ocupacao = 0

    if percentual_ocupacao > 100:

        percentual_ocupacao = 100

    volume_ocupado = (

        volume_total_litros

        *

        percentual_ocupacao

        /

        100

    )

    volume_livre = (

        volume_total_litros

        -

        volume_ocupado

    )

    return volume_livre


# =============================================================================
# PARTE 33
# PERCENTUAL DE ENCHIMENTO
# =============================================================================

def calcular_percentual_enchimento(

    pressao_bar,

    pressao_maxima=220

):
    """
    Retorna o percentual de enchimento
    do cilindro baseado na pressão.

    A pressão máxima pode ser alterada.
    """

    if pressao_maxima <= 0:

        return 0

    percentual = (

        pressao_bar

        /

        pressao_maxima

    ) * 100

    if percentual < 0:

        percentual = 0

    if percentual > 100:

        percentual = 100

    return percentual


# =============================================================================
# PARTE 34
# PRESSÃO RESTANTE
# =============================================================================

def calcular_pressao_restante(

    pressao_atual,

    pressao_minima=5

):
    """
    Calcula a pressão realmente
    utilizável do cilindro.
    """

    restante = (

        pressao_atual

        -

        pressao_minima

    )

    if restante < 0:

        restante = 0

    return restante


# =============================================================================
# PARTE 35
# CÁLCULO DA MASSA CONSUMIDA
# =============================================================================

def calcular_massa_consumida(

    massa_inicial,

    massa_final

):
    """
    Calcula a massa consumida.

    Retorna sempre um valor positivo.
    """

    massa = (

        massa_inicial

        -

        massa_final

    )

    if massa < 0:

        massa = 0

    return massa



# =============================================================================
# PARTE 36
# CÁLCULO DO VOLUME CONSUMIDO
# =============================================================================

def calcular_volume_consumido(

    volume_inicial,

    volume_final

):
    """
    Calcula o volume efetivamente consumido.
    """

    volume = (

        volume_inicial

        -

        volume_final

    )

    if volume < 0:

        volume = 0

    return volume


# =============================================================================
# PARTE 37
# PERCENTUAL CONSUMIDO
# =============================================================================

def calcular_percentual_consumido(

    valor_inicial,

    valor_final

):
    """
    Calcula o percentual consumido.

    Exemplo

    Inicial = 16 m³

    Final = 4 m³

    Consumo = 75%
    """

    if valor_inicial <= 0:

        return 0

    percentual = (

        (

            valor_inicial

            -

            valor_final

        )

        /

        valor_inicial

    ) * 100

    if percentual < 0:

        percentual = 0

    if percentual > 100:

        percentual = 100

    return percentual


# =============================================================================
# PARTE 38
# EFICIÊNCIA DO ABASTECIMENTO
# =============================================================================

def calcular_eficiencia(

    volume_teorico,

    volume_real

):
    """
    Calcula a eficiência do abastecimento.

    Quanto mais próximo de 100%,
    mais próximo do cálculo teórico.
    """

    if volume_teorico <= 0:

        return 0

    eficiencia = (

        volume_real

        /

        volume_teorico

    ) * 100

    return eficiencia


# =============================================================================
# PARTE 39
# PRESSÃO DE ENCHIMENTO
# =============================================================================

def calcular_pressao_enchimento(

    pressao_inicial,

    pressao_final

):
    """
    Calcula quanto a pressão aumentou
    durante o abastecimento.

    Retorna o aumento em bar.
    """

    aumento = (

        pressao_final

        -

        pressao_inicial

    )

    if aumento < 0:

        aumento = 0

    return aumento



# =============================================================================
# PARTE 40
# TAXA DE COMPRESSÃO
# =============================================================================

def calcular_taxa_compressao(

    pressao_absoluta,

    pressao_atmosferica

):
    """
    Calcula a taxa de compressão.

    Exemplo

    201 bar absolutos

    1,01325 bar atmosféricos

    Resultado ≈ 198
    """

    if pressao_atmosferica <= 0:

        return 0

    return (

        pressao_absoluta

        /

        pressao_atmosferica

    )



# =============================================================================
# PARTE 41
# ENERGIA POR METRO CÚBICO
# =============================================================================

def calcular_energia_m3(

    energia_total,

    volume_m3

):
    """
    Calcula a energia existente
    em cada metro cúbico.
    """

    if volume_m3 <= 0:

        return 0

    return (

        energia_total

        /

        volume_m3

    )


# =============================================================================
# PARTE 42
# FATOR DE UTILIZAÇÃO
# =============================================================================

def calcular_fator_utilizacao(

    volume_utilizado,

    volume_total

):
    """
    Calcula o percentual realmente
    utilizado do cilindro.
    """

    if volume_total <= 0:

        return 0

    fator = (

        volume_utilizado

        /

        volume_total

    ) * 100

    if fator < 0:

        fator = 0

    if fator > 100:

        fator = 100

    return fator


# =============================================================================
# PARTE 43
# BIBLIOTECA DE PROPRIEDADES DO GNV
# =============================================================================

def obter_propriedades_gnv():

    """
    Retorna um dicionário contendo as
    propriedades médias do GNV.

    Todos os valores poderão ser alterados
    futuramente conforme a composição do gás.
    """

    propriedades = {

        "massa_molar": 0.01812,      # kg/mol

        "densidade": 0.717,          # kg/m³ CNTP

        "pcs": 39.50,                # MJ/m³

        "pci": 35.80,                # MJ/m³

        "cp": 2.20,                  # kJ/(kg.K)

        "cv": 1.70,                  # kJ/(kg.K)

        "k": 1.294,                  # Cp/Cv

        "R_especifico": 518.3,       # J/(kg.K)

        "temperatura_critica": 190.6,# Kelvin

        "pressao_critica": 45.99     # bar

    }

    return propriedades


# =============================================================================
# PARTE 44
# MOSTRAR AS PROPRIEDADES
# =============================================================================

def mostrar_propriedades_gnv():

    propriedades = obter_propriedades_gnv()

    print()

    print("=" * 75)

    print("PROPRIEDADES DO GNV")

    print("=" * 75)

    print()

    for chave, valor in propriedades.items():

        print(f"{chave:<25} : {valor}")

    print()

    print("=" * 75)

    print()


# =============================================================================
# PARTE 45
# OBTER UMA PROPRIEDADE
# =============================================================================

def propriedade(nome):

    """
    Retorna uma propriedade específica.

    Exemplo

    pcs = propriedade("pcs")
    """

    propriedades = obter_propriedades_gnv()

    return propriedades.get(nome)


# =============================================================================
# PARTE 46
# ENERGIA UTILIZANDO A BIBLIOTECA
# =============================================================================

def calcular_energia_volume(

    volume_m3

):
    """
    Calcula automaticamente a energia
    utilizando o PCS armazenado na biblioteca.
    """

    pcs = propriedade(

        "pcs"

    )

    energia = (

        volume_m3

        *

        pcs

    )

    return energia



# =============================================================================
# PARTE 47
# COMPOSIÇÃO DO GNV
# =============================================================================

def obter_composicao_padrao():

    """
    Retorna uma composição típica do GNV.

    Os valores são em porcentagem (%).
    """

    composicao = {

        "CH4": 89.00,

        "C2H6": 6.00,

        "C3H8": 2.00,

        "CO2": 1.50,

        "N2": 1.00,

        "O2": 0.30,

        "Outros": 0.20

    }

    return composicao




# =============================================================================
# PARTE 48
# MOSTRAR COMPOSIÇÃO DO GNV
# =============================================================================

def mostrar_composicao():

    composicao = obter_composicao_padrao()

    print()

    print("=" * 75)

    print("COMPOSIÇÃO DO GNV")

    print("=" * 75)

    print()

    for gas, percentual in composicao.items():

        print(f"{gas:<10} : {percentual:6.2f} %")

    print()

    print("=" * 75)

    print()



# =============================================================================
# PARTE 49
# MASSA MOLAR DA MISTURA
# =============================================================================

def calcular_massa_molar_mistura():

    """
    Calcula a massa molar média
    da composição do GNV.
    """

    massas = {

        "CH4":16.043,

        "C2H6":30.070,

        "C3H8":44.097,

        "CO2":44.010,

        "N2":28.014,

        "O2":31.999,

        "Outros":20.000

    }

    composicao = obter_composicao_padrao()

    massa = 0

    for gas in composicao:

        massa += (

            composicao[gas]

            *

            massas[gas]

        )

    massa /= 100

    return massa / 1000


# =============================================================================
# PARTE 50
# RESUMO DAS PROPRIEDADES DO GNV
# =============================================================================

def resumo_gnv():

    print()

    print("=" * 75)

    print("RESUMO DO GNV")

    print("=" * 75)

    print()

    print(
        f"Massa molar média : "
        f"{calcular_massa_molar_mistura():.6f} kg/mol"
    )

    print(
        f"Densidade........ : "
        f"{propriedade('densidade'):.3f} kg/m³"
    )

    print(
        f"PCS............... : "
        f"{propriedade('pcs'):.2f} MJ/m³"
    )

    print(
        f"PCI............... : "
        f"{propriedade('pci'):.2f} MJ/m³"
    )

    print()

    print("=" * 75)

    print()



# =============================================================================
# PARTE 51
# TEMPERATURA REDUZIDA
# =============================================================================

def calcular_temperatura_reduzida(

    temperatura_kelvin,

    temperatura_critica

):
    """
    Calcula a temperatura reduzida.

    Tr = T / Tc
    """

    if temperatura_critica <= 0:

        return 0

    return (

        temperatura_kelvin

        /

        temperatura_critica

    )


# =============================================================================
# PARTE 52
# PRESSÃO REDUZIDA
# =============================================================================

def calcular_pressao_reduzida(

    pressao_bar,

    pressao_critica

):
    """
    Calcula a pressão reduzida.

    Pr = P / Pc
    """

    if pressao_critica <= 0:

        return 0

    return (

        pressao_bar

        /

        pressao_critica

    )



# =============================================================================
# PARTE 53
# CONSTANTE ESPECÍFICA DO GNV
# =============================================================================

def calcular_R_especifico(

    massa_molar

):
    """
    Calcula a constante específica.

    R = 8,314462618 / M
    """

    R_UNIVERSAL = 8.314462618

    if massa_molar <= 0:

        return 0

    return (

        R_UNIVERSAL

        /

        massa_molar

    )



# =============================================================================
# PARTE 54
# MASSA ESPECÍFICA
# =============================================================================

def calcular_massa_especifica(

    massa,

    volume

):
    """
    Calcula a massa específica.

    kg/m³
    """

    if volume <= 0:

        return 0

    return (

        massa

        /

        volume

    )


# =============================================================================
# PARTE 55
# FATOR Z APROXIMADO
# =============================================================================

def calcular_Z_aproximado(

    pressao_reduzida,

    temperatura_reduzida

):
    """
    Calcula uma aproximação do fator Z.

    Esta função será substituída futuramente
    pela Equação de Peng-Robinson ou
    Dranchuk-Abou-Kassem.

    Retorna um fator Z entre aproximadamente
    0,65 e 1,10.
    """

    Z = (

        1

        +

        0.08

        *

        pressao_reduzida

        /

        temperatura_reduzida

    )

    if Z < 0.65:

        Z = 0.65

    if Z > 1.10:

        Z = 1.10

    return Z


# =============================================================================
# PARTE 56
# CÁLCULO AUTOMÁTICO DE Tr E Pr
# =============================================================================

def calcular_propriedades_reduzidas(

    pressao_absoluta,

    temperatura_kelvin

):

    propriedades = obter_propriedades_gnv()

    Tr = calcular_temperatura_reduzida(

        temperatura_kelvin,

        propriedades["temperatura_critica"]

    )

    Pr = calcular_pressao_reduzida(

        pressao_absoluta,

        propriedades["pressao_critica"]

    )

    return {

        "Tr": Tr,

        "Pr": Pr

    }



# =============================================================================
# PARTE 57
# CÁLCULO COMPLETO DO FATOR Z
# =============================================================================

def calcular_Z(

    pressao_absoluta,

    temperatura_kelvin

):

    reduzidas = calcular_propriedades_reduzidas(

        pressao_absoluta,

        temperatura_kelvin

    )

    Z = calcular_Z_aproximado(

        reduzidas["Pr"],

        reduzidas["Tr"]

    )

    return {

        "Z": Z,

        "Pr": reduzidas["Pr"],

        "Tr": reduzidas["Tr"]

    }


# =============================================================================
# PARTE 83
# CÁLCULO DA MASSA DO GÁS REAL
# =============================================================================

def calcular_massa_gas_real(
    pressao_absoluta_bar,
    volume_m3,
    temperatura_kelvin,
    fator_z,
    massa_molar,
    densidade_informada_kg_m3=None
):
    """
    Calcula a massa do GNV utilizando
    a equação dos gases reais.

    Entradas
    --------
    pressao_absoluta_bar : Pressão absoluta (bar)
    volume_m3            : Volume (m³)
    temperatura_kelvin   : Temperatura (K)
    fator_z              : Fator de compressibilidade
    massa_molar          : Massa molar (kg/mol)

    Retorna
    -------
    Massa em kg
    """

    R = 8.314462618

    pressao_pa = pressao_absoluta_bar * 100000

    massa = (
        pressao_pa *
        volume_m3 *
        massa_molar
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return massa


# =============================================================================
# PARTE 84
# CÁLCULO DO NÚMERO DE MOLS
# =============================================================================

def calcular_numero_mols(
    massa_kg,
    massa_molar
):
    """
    Calcula a quantidade de matéria (mol).

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    massa_molar : Massa molar (kg/mol)

    Retorno
    -------
    Quantidade de mols.
    """

    if massa_molar <= 0:

        return 0.0

    numero_mols = (

        massa_kg

        /

        massa_molar

    )

    return numero_mols


# =============================================================================
# PARTE 85
# CÁLCULO DA DENSIDADE DO GÁS REAL
# =============================================================================

def calcular_densidade_gas_real(
    pressao_absoluta_bar,
    temperatura_kelvin,
    fator_z,
    massa_molar
):
    """
    Calcula a densidade do GNV utilizando
    a equação dos gases reais.

    Entradas
    --------
    pressao_absoluta_bar : Pressão absoluta (bar)

    temperatura_kelvin : Temperatura (K)

    fator_z : Fator de compressibilidade

    massa_molar : Massa molar (kg/mol)

    Retorna
    -------
    Densidade em kg/m³
    """

    R = 8.314462618

    pressao_pa = pressao_absoluta_bar * 100000

    if (
        temperatura_kelvin <= 0
        or fator_z <= 0
        or massa_molar <= 0
    ):

        return 0.0

    densidade = (
        pressao_pa *
        massa_molar
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return densidade


# =============================================================================
# PARTE 86
# CÁLCULO DO VOLUME ESPECÍFICO DO GÁS REAL
# =============================================================================

def calcular_volume_especifico_gas_real(
    massa_kg,
    volume_m3
):
    """
    Calcula o volume específico do GNV.

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    volume_m3 : Volume ocupado pelo gás (m³)

    Retorna
    -------
    Volume específico em m³/kg
    """

    if massa_kg <= 0:

        return 0.0

    volume_especifico = (

        volume_m3

        /

        massa_kg

    )

    return volume_especifico



# =============================================================================
# PARTE 87
# CÁLCULO DO VOLUME REAL DO GNV
# =============================================================================

def calcular_volume_real_gnv(
    massa_kg,
    densidade_kg_m3
):
    """
    Calcula o volume realmente ocupado
    pelo GNV.

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    densidade_kg_m3 : Densidade (kg/m³)

    Retorna
    -------
    Volume em m³
    """

    if densidade_kg_m3 <= 0:

        return 0.0

    volume = (

        massa_kg

        /

        densidade_kg_m3

    )

    return volume




# =============================================================================
# PARTE 88
# CALCULA A QUANTIDADE DE GNV NO CILINDRO
# =============================================================================

def calcular_quantidade_gnv(
    volume_cilindro_litros,
    pressao_bar,
    temperatura_c,
    altitude_m,
    fator_z,
    massa_molar,
    densidade_informada_kg_m3=None
):
    """
    Calcula automaticamente diversas propriedades
    do GNV armazenado no cilindro.

    Retorna um dicionário contendo todos os resultados.
    """

    # ---------------------------------------------------------
    # Conversões
    # ---------------------------------------------------------

    volume_m3 = volume_cilindro_litros / 1000.0

    temperatura_k = temperatura_c + 273.15

    pressao_atm = calcular_pressao_atmosferica(
        altitude_m
    )

    pressao_abs = pressao_bar + pressao_atm

    # ---------------------------------------------------------
    # Massa
    # ---------------------------------------------------------

    massa = calcular_massa_gas_real(
        pressao_abs,
        volume_m3,
        temperatura_k,
        fator_z,
        massa_molar
    )

    # ---------------------------------------------------------
    # Densidade
    # ---------------------------------------------------------

    densidade = calcular_densidade_gas_real(
        pressao_abs,
        temperatura_k,
        fator_z,
        massa_molar
    )

    # ---------------------------------------------------------
    # Número de mols
    # ---------------------------------------------------------

    mols = calcular_numero_mols(
        massa,
        massa_molar
    )

    # ---------------------------------------------------------
    # Volume específico
    # ---------------------------------------------------------

    volume_especifico = calcular_volume_especifico_gas_real(
        massa,
        volume_m3
    )

    # ---------------------------------------------------------
    # Volume real
    # ---------------------------------------------------------

    volume_real = calcular_volume_real_gnv(
        massa,
        densidade
    )

    # ---------------------------------------------------------
    # Retorno
    # ---------------------------------------------------------

    return {

        "volume_cilindro_l": volume_cilindro_litros,

        "volume_cilindro_m3": volume_m3,

        "temperatura_c": temperatura_c,

        "temperatura_k": temperatura_k,

        "pressao_bar": pressao_bar,

        "pressao_atm": pressao_atm,

        "pressao_absoluta": pressao_abs,

        "massa": massa,

        "densidade": densidade,

        "mols": mols,

        "volume_especifico": volume_especifico,

        "volume_real": volume_real,

        "volume_equivalente_m3_temperatura_informada": calcular_volume_equivalente_na_temperatura_m3(mols, temperatura_c, 1.01325, 1.0),

        "volume_equivalente_m3_20c": calcular_volume_referencia_m3(
            mols,
            20.0,
            1.01325,
            1.0
        ),
        # Conversão independente pelo modelo ANP/idealizado (Z=1),
        # usando a variação de pressão absoluta no cilindro.
        "volume_anp_ideal_m3_20c": calcular_volume_anp_referencia(
            volume_cilindro_litros,
            0.0,
            pressao_bar,
            temperatura_c,
            altitude_m
        ),

        "densidade_informada_kg_m3": densidade_informada_kg_m3,

        "massa_referencia_informada_kg": (
            calcular_volume_referencia_m3(
                mols,
                20.0,
                1.01325,
                1.0
            ) * densidade_informada_kg_m3
            if densidade_informada_kg_m3 is not None else None
        ),

        "diferenca_massa_referencia_kg": (
            (
                calcular_volume_referencia_m3(
                    mols,
                    20.0,
                    1.01325,
                    1.0
                ) * densidade_informada_kg_m3
            ) - massa
            if densidade_informada_kg_m3 is not None else None
        ),

        "volume_equivalente_litros_20c": (
            calcular_volume_referencia_m3(
                mols,
                20.0,
                1.01325,
                1.0
            ) * 1000.0
        )

    }


# =============================================================================
# PARTE 89
# COMPARAÇÃO DO ABASTECIMENTO
# =============================================================================

def comparar_abastecimento(
    volume_informado_m3,
    volume_calculado_m3,
    massa_informada_kg=0.0,
    massa_calculada_kg=0.0
):
    """
    Compara os valores informados pelo posto
    com os valores calculados pelo programa.

    Retorna um dicionário contendo as diferenças.
    """

    diferenca_volume = (
        volume_informado_m3 -
        volume_calculado_m3
    )

    if volume_calculado_m3 > 0:

        erro_volume = (
            diferenca_volume /
            volume_calculado_m3
        ) * 100

    else:

        erro_volume = 0.0

    diferenca_massa = (
        massa_informada_kg -
        massa_calculada_kg
    )

    if massa_calculada_kg > 0:

        erro_massa = (
            diferenca_massa /
            massa_calculada_kg
        ) * 100

    else:

        erro_massa = 0.0

    return {

        "volume_informado": volume_informado_m3,

        "volume_calculado": volume_calculado_m3,

        "diferenca_volume": diferenca_volume,

        "erro_volume_percentual": erro_volume,

        "massa_informada": massa_informada_kg,

        "massa_calculada": massa_calculada_kg,

        "diferenca_massa": diferenca_massa,

        "erro_massa_percentual": erro_massa

    }


# =============================================================================
# PARTE 90
# CÁLCULO DA ENERGIA TOTAL DO GNV
# =============================================================================

def calcular_energia_gnv_por_massa(
    massa_kg,
    pcs_mj_kg=50.02,
    pci_mj_kg=45.00
):
    """
    Calcula a energia contida no GNV.

    Entradas
    --------
    massa_kg : Massa do GNV (kg)

    pcs_mj_kg : Poder Calorífico Superior (MJ/kg)

    pci_mj_kg : Poder Calorífico Inferior (MJ/kg)

    Retorna
    -------
    Dicionário contendo a energia.
    """

    energia_pcs = massa_kg * pcs_mj_kg

    energia_pci = massa_kg * pci_mj_kg

    energia_pcs_kwh = energia_pcs / 3.6

    energia_pci_kwh = energia_pci / 3.6

    return {

        "PCS_MJ": energia_pcs,

        "PCI_MJ": energia_pci,

        "PCS_kWh": energia_pcs_kwh,

        "PCI_kWh": energia_pci_kwh

    }



# =============================================================================
# PARTE 91
# CÁLCULO COMPLETO DO CONSUMO DO VEÍCULO
# =============================================================================

def calcular_consumo_veiculo(
    distancia_km,
    volume_m3,
    massa_kg,
    valor_abastecimento,
    energia_mj
):
    """
    Calcula diversos indicadores de consumo.

    Entradas
    --------
    distancia_km : Distância percorrida (km)

    volume_m3 : Volume abastecido (m³)

    massa_kg : Massa abastecida (kg)

    valor_abastecimento : Valor pago (R$)

    energia_mj : Energia do abastecimento (MJ)

    Retorna
    -------
    Dicionário contendo os indicadores.
    """

    if distancia_km <= 0:

        return None

    if volume_m3 > 0:

        km_m3 = distancia_km / volume_m3

    else:

        km_m3 = 0.0

    if massa_kg > 0:

        km_kg = distancia_km / massa_kg

    else:

        km_kg = 0.0

    if energia_mj > 0:

        km_mj = distancia_km / energia_mj

        mj_km = energia_mj / distancia_km

    else:

        km_mj = 0.0

        mj_km = 0.0

    custo_km = valor_abastecimento / distancia_km

    custo_m3 = 0.0

    if volume_m3 > 0:

        custo_m3 = valor_abastecimento / volume_m3

    custo_kg = 0.0

    if massa_kg > 0:

        custo_kg = valor_abastecimento / massa_kg

    custo_mj = 0.0

    if energia_mj > 0:

        custo_mj = valor_abastecimento / energia_mj

    return {

        "km_m3": km_m3,

        "km_kg": km_kg,

        "km_MJ": km_mj,

        "MJ_km": mj_km,

        "custo_km": custo_km,

        "custo_m3": custo_m3,

        "custo_kg": custo_kg,

        "custo_MJ": custo_mj

    }


# =============================================================================
# PARTE 58
# CLASSE GNV
# =============================================================================

class GNV:

    """
    Classe principal para armazenar
    todos os dados referentes ao GNV.
    """

    def __init__(

        self,

        volume_m3,

        pressao_bar,

        temperatura_c,

        altitude

    ):

        self.volume_m3 = volume_m3

        self.pressao_bar = pressao_bar

        self.temperatura_c = temperatura_c

        self.altitude = altitude

        self.temperatura_k = (

            temperatura_c

            +

            273.15

        )


# =============================================================================
# PARTE 59
# PRESSÃO ATMOSFÉRICA
# =============================================================================

    def pressao_atmosferica(self):

        return calcular_pressao_atmosferica(

            self.altitude

        )


# =============================================================================
# PARTE 60
# PRESSÃO ABSOLUTA
# =============================================================================

    def pressao_absoluta(self):

        return (

            self.pressao_bar

            +

            self.pressao_atmosferica()

        )


# =============================================================================
# PARTE 61
# FATOR Z
# =============================================================================

    def fator_Z(self):

        resultado = calcular_Z(

            self.pressao_absoluta(),

            self.temperatura_k

        )

        return resultado


# =============================================================================
# PARTE 62
# MASSA DO GNV
# =============================================================================

    def massa(self):

        densidade = propriedade(

            "densidade"

        )

        return calcular_massa_por_densidade(

            self.volume_m3,

            densidade

        )


# =============================================================================
# PARTE 63
# ENERGIA DO GNV
# =============================================================================

    def energia(self):

        pcs = propriedade(

            "pcs"

        )

        energia_mj = (

            self.volume_m3

            *

            pcs

        )

        energia_kwh = (

            energia_mj

            /

            3.6

        )

        return {

            "MJ": energia_mj,

            "kWh": energia_kwh

        }


# =============================================================================
# PARTE 64
# AUTONOMIA
# =============================================================================

    def autonomia(

        self,

        consumo_km_m3

    ):

        return (

            self.volume_m3

            *

            consumo_km_m3

        )


# =============================================================================
# PARTE 65
# DENSIDADE
# =============================================================================

    def densidade(self):

        return propriedade(

            "densidade"

        )


# =============================================================================
# PARTE 66
# MASSA MOLAR
# =============================================================================

    def massa_molar(self):

        return calcular_massa_molar_mistura()


# =============================================================================
# PARTE 67
# RELATÓRIO COMPLETO
# =============================================================================

    def relatorio(self, consumo_km_m3=15.0):
        """
        Exibe um relatório completo do estado do GNV.

        consumo_km_m3:
            Rendimento médio do veículo em km por m³.
        """

        pressao_atm = self.pressao_atmosferica()

        pressao_abs = self.pressao_absoluta()

        fator = self.fator_Z()

        massa = self.massa()

        energia = self.energia()

        autonomia = self.autonomia(consumo_km_m3)

        print()

        print("=" * 75)

        print("RELATÓRIO COMPLETO DO GNV")

        print("=" * 75)

        print(f"Volume................: {self.volume_m3:.3f} m³")

        print(f"Pressão...............: {self.pressao_bar:.2f} bar")

        print(f"Pressão Atmosférica...: {pressao_atm:.4f} bar")

        print(f"Pressão Absoluta......: {pressao_abs:.4f} bar")

        print(f"Temperatura...........: {self.temperatura_c:.2f} °C")

        print(f"Temperatura...........: {self.temperatura_k:.2f} K")

        print(f"Pressão Reduzida......: {fator['Pr']:.4f}")

        print(f"Temperatura Reduzida..: {fator['Tr']:.4f}")

        print(f"Fator Z...............: {fator['Z']:.5f}")

        print(f"Massa................: {massa:.3f} kg")

        print(f"Energia..............: {energia['MJ']:.2f} MJ")

        print(f"Energia..............: {energia['kWh']:.2f} kWh")

        print(f"Autonomia Estimada...: {autonomia:.2f} km")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 68
# DICIONÁRIO COMPLETO
# =============================================================================

    def dados(self, consumo_km_m3=15.0):
        """
        Retorna todos os cálculos em formato
        de dicionário.

        Essa função será utilizada futuramente
        para:

            Excel

            SQLite

            PDF

            API

            Interface Gráfica
        """

        pressao_atm = self.pressao_atmosferica()

        pressao_abs = self.pressao_absoluta()

        fator = self.fator_Z()

        energia = self.energia()

        return {

            "Volume_m3": self.volume_m3,

            "Pressao_bar": self.pressao_bar,

            "Pressao_Atmosferica": pressao_atm,

            "Pressao_Absoluta": pressao_abs,

            "Temperatura_C": self.temperatura_c,

            "Temperatura_K": self.temperatura_k,

            "Pressao_Reduzida": fator["Pr"],

            "Temperatura_Reduzida": fator["Tr"],

            "Fator_Z": fator["Z"],

            "Massa": self.massa(),

            "Energia_MJ": energia["MJ"],

            "Energia_kWh": energia["kWh"],

            "Autonomia": self.autonomia(
                consumo_km_m3
            )

        }


# =============================================================================
# PARTE 69
# MOSTRAR DADOS
# =============================================================================

    def mostrar_dados(self):

        dados = self.dados()

        print()

        print("=" * 75)

        print("DADOS DA CLASSE GNV")

        print("=" * 75)

        print()

        for chave, valor in dados.items():

            print(

                f"{chave:<30}: {valor}"

            )

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 70
# EXPORTAR DADOS
# =============================================================================

    def exportar(self):

        return self.dados()

# =============================================================================
# PARTE 71
# EXPORTAÇÃO JSON
# =============================================================================

    def json(self):

        import json

        return json.dumps(

            self.dados(),

            indent=4,

            ensure_ascii=False

        )


# =============================================================================
# PARTE 72
# SIMULAÇÃO DA PRESSÃO EM OUTRA TEMPERATURA
# =============================================================================

    def simular_pressao(

        self,

        nova_temperatura_c

    ):
        """
        Simula a nova pressão considerando
        que a quantidade de gás permanece
        constante.

        Utiliza uma aproximação baseada
        na Lei dos Gases Ideais.

        Retorna a pressão em bar.
        """

        temperatura1 = self.temperatura_k

        temperatura2 = (

            nova_temperatura_c

            +

            273.15

        )

        pressao_absoluta = (

            self.pressao_absoluta()

            *

            temperatura2

            /

            temperatura1

        )

        pressao_manometrica = (

            pressao_absoluta

            -

            self.pressao_atmosferica()

        )

        return pressao_manometrica



# =============================================================================
# PARTE 73
# RESFRIAMENTO
# =============================================================================

    def simular_resfriamento(

        self,

        temperatura_final

    ):

        return self.simular_pressao(

            temperatura_final

        )


# =============================================================================
# PARTE 74
# AQUECIMENTO
# =============================================================================

    def simular_aquecimento(

        self,

        temperatura_final

    ):

        return self.simular_pressao(

            temperatura_final

        )


# =============================================================================
# PARTE 75
# TABELA DE PRESSÃO x TEMPERATURA
# =============================================================================

    def tabela_temperatura(

        self,

        temperatura_inicial=-20,

        temperatura_final=60,

        passo=2

    ):

        print()

        print("=" * 75)

        print("SIMULAÇÃO DE PRESSÃO")

        print("=" * 75)

        print()

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            pressao = self.simular_pressao(

                temperatura

            )

            print(

                f"{temperatura:6.1f} °C   "

                f"{pressao:8.2f} bar"

            )

            temperatura += passo

        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 76
# SIMULAÇÃO DE ALTITUDE
# =============================================================================

    def simular_altitude(

        self,

        nova_altitude

    ):
        """
        Simula a pressão atmosférica para
        outra altitude.
        """

        return calcular_pressao_atmosferica(

            nova_altitude

        )


# =============================================================================
# PARTE 77
# PRESSÃO ABSOLUTA EM OUTRA ALTITUDE
# =============================================================================

    def pressao_absoluta_altitude(

        self,

        nova_altitude

    ):

        pressao_atm = calcular_pressao_atmosferica(

            nova_altitude

        )

        return (

            self.pressao_bar

            +

            pressao_atm

        )


# =============================================================================
# PARTE 78
# SIMULAÇÃO COMPLETA
# =============================================================================

    def simular(

        self,

        temperatura,

        altitude

    ):

        pressao = self.simular_pressao(

            temperatura

        )

        pressao_atm = calcular_pressao_atmosferica(

            altitude

        )

        pressao_abs = (

            pressao

            +

            pressao_atm

        )

        return {

            "temperatura": temperatura,

            "altitude": altitude,

            "pressao_atmosferica": pressao_atm,

            "pressao_manometrica": pressao,

            "pressao_absoluta": pressao_abs

        }


# =============================================================================
# PARTE 79
# RELATÓRIO DA SIMULAÇÃO
# =============================================================================

    def relatorio_simulacao(

        self,

        temperatura,

        altitude

    ):

        dados = self.simular(

            temperatura,

            altitude

        )

        print()

        print("=" * 75)

        print("SIMULAÇÃO")

        print("=" * 75)

        print()

        print(f"Temperatura...........: {dados['temperatura']:.2f} °C")

        print(f"Altitude..............: {dados['altitude']:.2f} m")

        print(f"Pressão Atmosférica...: {dados['pressao_atmosferica']:.4f} bar")

        print(f"Pressão Manométrica...: {dados['pressao_manometrica']:.4f} bar")

        print(f"Pressão Absoluta......: {dados['pressao_absoluta']:.4f} bar")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 80
# SIMULAÇÃO EM FAIXA
# =============================================================================

    def simular_faixa_temperatura(

        self,

        temperatura_inicial,

        temperatura_final,

        incremento=2

    ):

        resultados = []

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            resultados.append(

                self.simular(

                    temperatura,

                    self.altitude

                )

            )

            temperatura += incremento

        return resultados



# =============================================================================
# PARTE 81
# GERAÇÃO DA TABELA PRESSÃO x TEMPERATURA
# =============================================================================

    def gerar_tabela_pressao_temperatura(

        self,

        temperatura_inicial=-20,

        temperatura_final=60,

        passo_temperatura=2,

        pressao_inicial=0,

        pressao_final=300,

        passo_pressao=5

    ):
        """
        Gera uma tabela contendo todas as
        combinações entre temperatura e pressão.

        Retorna uma lista de dicionários.
        """

        tabela = []

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            pressao = pressao_inicial

            while pressao <= pressao_final:

                """linha = {

                    "Temperatura": temperatura,

                    "Pressao": pressao

                }

                tabela.append(

                    linha

                )"""


# ==============================================================
# CONVERSÕES
# ==============================================================

                temperatura_kelvin = (
                    temperatura
                    +
                    273.15
                )

                pressao_atm = calcular_pressao_atmosferica(
                    self.altitude
                )

                pressao_absoluta = (
                    pressao
                    +
                    pressao_atm
                )


# ==============================================================
# FATOR Z
# ==============================================================

                resultado_z = calcular_Z(
                    pressao_absoluta,
                    temperatura_kelvin
                )


# ==============================================================
# VOLUME NAS CNTP
# ==============================================================

                # Volume equivalente nas condições de referência (CNTP)
                # usando a equação dos gases reais: P V = Z n R T.
                volume_cntp = (
                    self.volume_m3
                    * pressao_absoluta
                    * TEMPERATURA_PADRAO
                    / (
                        resultado_z["Z"]
                        * PRESSAO_ATMOSFERICA_PADRAO
                        * temperatura_kelvin
                    )
                )


# ==============================================================
# MASSA
# ==============================================================

                massa = calcular_massa_por_densidade(
                    volume_cntp,
                    propriedade(
                        "densidade"
                    )
                )


# ==============================================================
# ENERGIA
# ==============================================================

                energia = calcular_energia_volume(
                    volume_cntp
                )


# ==============================================================
# ENERGIA EM kWh
# ==============================================================

                energia_kwh = (
                    energia
                    /
                    3.6
                )


# ==============================================================
# DENSIDADE APARENTE
# ==============================================================

                if volume_cntp > 0:

                    densidade_aparente = (
                        massa
                        /
                        volume_cntp
                    )

                else:

                    densidade_aparente = 0.0


# ==============================================================
# DENSIDADE
# ==============================================================

                densidade = calcular_massa_especifica(
                    massa,
                    volume_cntp
                )


# ==============================================================
# CONSTANTE ESPECÍFICA
# ==============================================================

                R_especifico = calcular_R_especifico(
                    calcular_massa_molar_mistura()
                )


# ==============================================================
# LINHA DA TABELA
# ==============================================================

                linha = {

                    "Temperatura_C": temperatura,

                    "Temperatura_K": temperatura_kelvin,

                    "Pressao_bar": pressao,

                    "Pressao_Absoluta": pressao_absoluta,

                    "Altitude": self.altitude,

                    "Pressao_Atmosferica": pressao_atm,

                    "Pr": resultado_z["Pr"],

                    "Tr": resultado_z["Tr"],

                    "Fator_Z": resultado_z["Z"],

                    "Volume_CNTP": volume_cntp,

                    "Massa": massa,

                    "Energia_MJ": energia,

                    "Energia_kWh": energia_kwh,

                    "Densidade": densidade,

                    "Densidade_Aparente": densidade_aparente,

                    "R_especifico": R_especifico

                }


                tabela.append(

                    linha

                )


                pressao += passo_pressao


            temperatura += passo_temperatura


        return tabela



# =============================================================================
# PARTE 82
# MOSTRAR TABELA
# =============================================================================

    def mostrar_tabela(

        self

    ):

        tabela = self.gerar_tabela_pressao_temperatura()

        print()

        print("=" * 75)

        print("TABELA PRESSÃO x TEMPERATURA")

        print("=" * 75)

        print()

        for linha in tabela:

            print(
		f"{linha['Temperatura_C']:6.1f} °C"
		f"    "
		f"{linha['Pressao_bar']:6.1f} bar"
	)


        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 92
# CLASSE ABASTECIMENTO
# =============================================================================

class Abastecimento:

    """
    Armazena todas as informações de um abastecimento.
    """

    def __init__(

        self,

        data,

        posto,

        cidade,

        odometro,

        volume_m3,

        preco_m3,

        temperatura,

        pressao,

        altitude,

        observacoes="",

        capacidade_cilindro_l=0.0,

        pressao_inicial=0.0,

        pressao_final=0.0,

        densidade_informada_kg_m3=0.0,

        metragem_teorica_m3=0.0,

        metragem_anp_m3=0.0,

        metragem_cientifica_m3=0.0

    ):

        self.data = data

        self.posto = posto

        self.cidade = cidade

        self.odometro = odometro

        self.volume_m3 = volume_m3

        self.preco_m3 = preco_m3

        self.temperatura = temperatura

        self.pressao = pressao

        self.altitude = altitude

        self.observacoes = observacoes

        self.capacidade_cilindro_l = capacidade_cilindro_l

        self.pressao_inicial = pressao_inicial

        self.pressao_final = pressao_final

        self.densidade_informada_kg_m3 = densidade_informada_kg_m3
        self.metragem_teorica_m3 = metragem_teorica_m3
        self.metragem_anp_m3 = metragem_anp_m3
        self.metragem_cientifica_m3 = metragem_cientifica_m3

        self.valor_total = (

            volume_m3 *

            preco_m3

        )


# =============================================================================
# PARTE 93
# RESUMO DO ABASTECIMENTO
# =============================================================================

    def resumo(self):

        print()

        print("=" * 75)

        print("DADOS DO ABASTECIMENTO")

        print("=" * 75)

        print()

        print(f"Data...............: {self.data}")

        print(f"Posto..............: {self.posto}")

        print(f"Cidade.............: {self.cidade}")

        print(f"Odômetro...........: {self.odometro:.1f} km")

        print(f"Volume.............: {self.volume_m3:.3f} m³")

        print(f"Preço por m³.......: R$ {self.preco_m3:.3f}")

        print(f"Valor Total........: R$ {self.valor_total:.2f}")

        print(f"Temperatura........: {self.temperatura:.1f} °C")

        print(f"Pressão............: {self.pressao:.1f} bar")

        print(f"Altitude...........: {self.altitude:.1f} m")

        print(f"Observações........: {self.observacoes}")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 94
# EXPORTAÇÃO PARA DICIONÁRIO
# =============================================================================

    def to_dict(self):

        return {

            "data": self.data,

            "posto": self.posto,

            "cidade": self.cidade,

            "odometro": self.odometro,

            "volume_m3": self.volume_m3,

            "preco_m3": self.preco_m3,

            "valor_total": self.valor_total,

            "temperatura": self.temperatura,

            "pressao": self.pressao,

            "altitude": self.altitude,

            "observacoes": self.observacoes,

            "capacidade_cilindro_l": self.capacidade_cilindro_l,

            "pressao_inicial": self.pressao_inicial,

            "pressao_final": self.pressao_final,

            "densidade_informada_kg_m3": self.densidade_informada_kg_m3,

            "metragem_teorica_m3": self.metragem_teorica_m3,

            "metragem_anp_m3": self.metragem_anp_m3,

            "metragem_cientifica_m3": self.metragem_cientifica_m3

        }


# =============================================================================
# PARTE 95
# HISTÓRICO DE ABASTECIMENTOS
# =============================================================================

class HistoricoAbastecimentos:

    """
    Armazena vários abastecimentos.
    """

    def __init__(self):

        self.abastecimentos = []


# =============================================================================
# PARTE 96
# ADICIONAR ABASTECIMENTO
# =============================================================================

    def adicionar(

        self,

        abastecimento

    ):

        if not isinstance(

            abastecimento,

            Abastecimento

        ):

            raise TypeError(

                "Objeto inválido."

            )

        self.abastecimentos.append(

            abastecimento

        )

        self.abastecimentos.sort(

            key=lambda x: (

                x.data,

                x.odometro

            )

        )



# =============================================================================
# PARTE 97
# TOTAL DE ABASTECIMENTOS
# =============================================================================

    def quantidade(self):

        return len(

            self.abastecimentos

        )


# =============================================================================
# PARTE 98
# LISTAR ABASTECIMENTOS
# =============================================================================

    def listar(self):

        print()

        print("=" * 75)

        print("HISTÓRICO DE ABASTECIMENTOS")

        print("=" * 75)

        print()

        for indice, abastecimento in enumerate(

            self.abastecimentos,

            start=1

        ):

            print(

                f"{indice:03d} - "

                f"{abastecimento.data} - "

                f"{abastecimento.posto} - "

                f"{abastecimento.volume_m3:.3f} m³ - "

                f"R$ {abastecimento.valor_total:.2f}"

            )

        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 99
# PROCURAR ABASTECIMENTO POR DATA
# =============================================================================

    def procurar_data(
        self,
        data
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.data == data:

                resultados.append(

                    abastecimento

                )

        return resultados


# =============================================================================
# PARTE 100
# PROCURAR ABASTECIMENTO POR POSTO
# =============================================================================

    def procurar_posto(
        self,
        posto
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.posto.lower() == posto.lower():

                resultados.append(

                    abastecimento

                )

        return resultados



# =============================================================================
# PARTE 101
# PROCURAR ABASTECIMENTO POR CIDADE
# =============================================================================

    def procurar_cidade(
        self,
        cidade
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.cidade.lower() == cidade.lower():

                resultados.append(

                    abastecimento

                )

        return resultados



# =============================================================================
# PARTE 102
# VALOR TOTAL GASTO
# =============================================================================

    def valor_total(self):

        total = 0.0

        for abastecimento in self.abastecimentos:

            total += abastecimento.valor_total

        return total


# =============================================================================
# PARTE 103
# VOLUME TOTAL ABASTECIDO
# =============================================================================

    def volume_total(self):

        total = 0.0

        for abastecimento in self.abastecimentos:

            total += abastecimento.volume_m3

        return total


# =============================================================================
# PARTE 104
# MÉDIA DO PREÇO DO m³
# =============================================================================

    def media_preco_m3(self):

        if len(self.abastecimentos) == 0:

            return 0.0

        soma = 0.0

        for abastecimento in self.abastecimentos:

            soma += abastecimento.preco_m3

        return soma / len(self.abastecimentos)


# =============================================================================
# PARTE 105
# MÉDIA DO VOLUME ABASTECIDO
# =============================================================================

    def media_volume(self):

        if len(self.abastecimentos) == 0:

            return 0.0

        return self.volume_total() / len(self.abastecimentos)


# =============================================================================
# PARTE 106
# MAIOR ABASTECIMENTO
# =============================================================================

    def maior_abastecimento(self):

        if len(self.abastecimentos) == 0:

            return None

        maior = self.abastecimentos[0]

        for abastecimento in self.abastecimentos:

            if abastecimento.volume_m3 > maior.volume_m3:

                maior = abastecimento

        return maior


# =============================================================================
# PARTE 107
# MENOR ABASTECIMENTO
# =============================================================================

    def menor_abastecimento(self):

        if len(self.abastecimentos) == 0:

            return None

        menor = self.abastecimentos[0]

        for abastecimento in self.abastecimentos:

            if abastecimento.volume_m3 < menor.volume_m3:

                menor = abastecimento

        return menor


# =============================================================================
# PARTE 108
# RELATÓRIO GERAL
# =============================================================================

    def relatorio(self):

        print()

        print("=" * 75)

        print("RELATÓRIO DO HISTÓRICO")

        print("=" * 75)

        print()

        print(f"Quantidade...............: {self.quantidade()}")

        print(f"Volume Total.............: {self.volume_total():.3f} m³")

        print(f"Valor Total..............: R$ {self.valor_total():.2f}")

        print(f"Média do Volume..........: {self.media_volume():.3f} m³")

        print(f"Média do Preço...........: R$ {self.media_preco_m3():.3f}")

        maior = self.maior_abastecimento()

        if maior is not None:

            print()

            print("Maior abastecimento")

            print("-------------------")

            print(f"Data.....................: {maior.data}")

            print(f"Posto....................: {maior.posto}")

            print(f"Volume...................: {maior.volume_m3:.3f} m³")

        menor = self.menor_abastecimento()

        if menor is not None:

            print()

            print("Menor abastecimento")

            print("-------------------")

            print(f"Data.....................: {menor.data}")

            print(f"Posto....................: {menor.posto}")

            print(f"Volume...................: {menor.volume_m3:.3f} m³")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 109
# EXPORTAR HISTÓRICO
# =============================================================================

    def exportar_lista(self):

        """
        Retorna todos os abastecimentos
        em formato de lista de dicionários.
        """

        lista = []

        for abastecimento in self.abastecimentos:

            lista.append(

                abastecimento.to_dict()

            )

        return lista


# =============================================================================
# PARTE 110
# LIMPAR HISTÓRICO
# =============================================================================

    def limpar(self):

        """
        Remove todos os abastecimentos
        da memória.
        """

        self.abastecimentos.clear()


# =============================================================================
# PARTE 111
# ÚLTIMO ABASTECIMENTO
# =============================================================================

    def ultimo(self):

        """
        Retorna o último abastecimento.
        """

        if len(self.abastecimentos) == 0:

            return None

        return self.abastecimentos[-1]


# =============================================================================
# PARTE 112
# PRIMEIRO ABASTECIMENTO
# =============================================================================

    def primeiro(self):

        """
        Retorna o primeiro abastecimento.
        """

        if len(self.abastecimentos) == 0:

            return None

        return self.abastecimentos[0]



# =============================================================================
# PARTE 113
# TOTAL DE QUILÔMETROS
# =============================================================================

    def quilometragem_total(self):

        """
        Calcula a quilometragem entre
        o primeiro e o último abastecimento.
        """

        if len(self.abastecimentos) < 2:

            return 0.0

        primeiro = self.primeiro()

        ultimo = self.ultimo()

        return (

            ultimo.odometro

            -

            primeiro.odometro

        )



# =============================================================================
# PARTE 114
# BANCO DE DADOS SQLITE
# =============================================================================

class BancoGNV:

    """
    Classe responsável pelo banco SQLite.
    """

    def __init__(

        self,

        nome_banco="gnv.db"

    ):

        self.nome_banco = nome_banco

        self.conexao = None

        self.cursor = None



# =============================================================================
# PARTE 115
# CONECTAR AO SQLITE
# =============================================================================

    def conectar(self):

        self.conexao = sqlite3.connect(

            self.nome_banco

        )

        self.cursor = self.conexao.cursor()



# =============================================================================
# PARTE 116
# FECHAR SQLITE
# =============================================================================

    def fechar(self):

        if self.conexao:

            self.conexao.close()



# =============================================================================
# PARTE 117
# CRIAR TABELA
# =============================================================================

    def criar_tabela(self):

        self.cursor.execute("""

        CREATE TABLE IF NOT EXISTS abastecimentos(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            data TEXT,

            posto TEXT,

            cidade TEXT,

            odometro REAL,

            volume REAL,

            preco REAL,

            valor REAL,

            temperatura REAL,

            pressao REAL,

            altitude REAL,

            observacoes TEXT,

            capacidade_cilindro_l REAL DEFAULT 0,

            pressao_inicial REAL DEFAULT 0,

            pressao_final REAL DEFAULT 0,

            densidade_informada_kg_m3 REAL DEFAULT 0,

            metragem_teorica_m3 REAL DEFAULT 0,

            metragem_anp_m3 REAL DEFAULT 0,

            metragem_cientifica_m3 REAL DEFAULT 0

        )

        """)

        self.conexao.commit()

        # Migração segura para bancos criados em versões anteriores.
        colunas_existentes = {
            linha[1] for linha in self.cursor.execute(
                "PRAGMA table_info(abastecimentos)"
            ).fetchall()
        }

        novas_colunas = {
            "capacidade_cilindro_l": "REAL DEFAULT 0",
            "pressao_inicial": "REAL DEFAULT 0",
            "pressao_final": "REAL DEFAULT 0",
            "densidade_informada_kg_m3": "REAL DEFAULT 0",
            "metragem_teorica_m3": "REAL DEFAULT 0",
            "metragem_anp_m3": "REAL DEFAULT 0",
            "metragem_cientifica_m3": "REAL DEFAULT 0",
        }

        for nome, definicao in novas_colunas.items():
            if nome not in colunas_existentes:
                self.cursor.execute(
                    f"ALTER TABLE abastecimentos ADD COLUMN {nome} {definicao}"
                )

        self.conexao.commit()



# =============================================================================
# PARTE 118
# SALVAR ABASTECIMENTO
# =============================================================================

    def salvar_abastecimento(
        self,
        abastecimento
    ):

        self.cursor.execute(

            """

            INSERT INTO abastecimentos(

                data,

                posto,

                cidade,

                odometro,

                volume,

                preco,

                valor,

                temperatura,

                pressao,

                altitude,

                observacoes,

                capacidade_cilindro_l,

                pressao_inicial,

                pressao_final,

                densidade_informada_kg_m3,

                metragem_teorica_m3,

                metragem_anp_m3,

                metragem_cientifica_m3

            )

            VALUES(

                ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?

            )

            """,

            (

                abastecimento.data,

                abastecimento.posto,

                abastecimento.cidade,

                abastecimento.odometro,

                abastecimento.volume_m3,

                abastecimento.preco_m3,

                abastecimento.valor_total,

                abastecimento.temperatura,

                abastecimento.pressao,

                abastecimento.altitude,

                abastecimento.observacoes,

                abastecimento.capacidade_cilindro_l,

                abastecimento.pressao_inicial,

                abastecimento.pressao_final,

                abastecimento.densidade_informada_kg_m3,

                abastecimento.metragem_teorica_m3,

                abastecimento.metragem_anp_m3,

                abastecimento.metragem_cientifica_m3

            )

        )

        self.conexao.commit()



# =============================================================================
# PARTE 119
# LISTAR TODOS OS ABASTECIMENTOS
# =============================================================================

    def listar_abastecimentos(self):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            ORDER BY id

            """

        )

        return self.cursor.fetchall()


# =============================================================================
# PARTE 120
# EXCLUIR ABASTECIMENTO
# =============================================================================

    def excluir_abastecimento(
        self,
        id_abastecimento
    ):

        self.cursor.execute(

            """

            DELETE FROM abastecimentos

            WHERE id = ?

            """,

            (

                id_abastecimento,

            )

        )

        self.conexao.commit()



# =============================================================================
# PARTE 121
# PESQUISAR ABASTECIMENTO
# =============================================================================

    def buscar_por_id(
        self,
        id_abastecimento
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE id = ?

            """,

            (

                id_abastecimento,

            )

        )

        return self.cursor.fetchone()


# =============================================================================
# PARTE 122
# ATUALIZAR ABASTECIMENTO
# =============================================================================

    def atualizar_abastecimento(
        self,
        id_abastecimento,
        abastecimento
    ):

        self.cursor.execute(

            """

            UPDATE abastecimentos

            SET

                data = ?,

                posto = ?,

                cidade = ?,

                odometro = ?,

                volume = ?,

                preco = ?,

                valor = ?,

                temperatura = ?,

                pressao = ?,

                altitude = ?,

                observacoes = ?,

                capacidade_cilindro_l = ?,

                pressao_inicial = ?,

                pressao_final = ?,

                densidade_informada_kg_m3 = ?

            WHERE id = ?

            """,

            (

                abastecimento.data,

                abastecimento.posto,

                abastecimento.cidade,

                abastecimento.odometro,

                abastecimento.volume_m3,

                abastecimento.preco_m3,

                abastecimento.valor_total,

                abastecimento.temperatura,

                abastecimento.pressao,

                abastecimento.altitude,

                abastecimento.observacoes,

                abastecimento.capacidade_cilindro_l,

                abastecimento.pressao_inicial,

                abastecimento.pressao_final,

                abastecimento.densidade_informada_kg_m3,

                id_abastecimento

            )

        )

        self.conexao.commit()


# =============================================================================
# PARTE 123
# PESQUISAR POR POSTO
# =============================================================================

    def buscar_por_posto(
        self,
        posto
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE posto LIKE ?

            ORDER BY data

            """,

            (

                "%" + posto + "%",

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 124
# PESQUISAR POR CIDADE
# =============================================================================

    def buscar_por_cidade(
        self,
        cidade
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE cidade LIKE ?

            ORDER BY data

            """,

            (

                "%" + cidade + "%",

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 125
# CONTAR ABASTECIMENTOS
# =============================================================================

    def contar_registros(self):

        self.cursor.execute(

            """

            SELECT COUNT(*)

            FROM abastecimentos

            """

        )

        return self.cursor.fetchone()[0]



# =============================================================================
# PARTE 126
# BUSCAR ENTRE DUAS DATAS
# =============================================================================

    def buscar_periodo(
        self,
        data_inicial,
        data_final
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE data BETWEEN ? AND ?

            ORDER BY data

            """,

            (

                data_inicial,

                data_final

            )

        )

        return self.cursor.fetchall()


# =============================================================================
# PARTE 127
# BUSCAR POR ODÔMETRO
# =============================================================================

    def buscar_odometro(
        self,
        km_inicial,
        km_final
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE odometro BETWEEN ? AND ?

            ORDER BY odometro

            """,

            (

                km_inicial,

                km_final

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 128
# CRIAR ÍNDICES
# =============================================================================

    def criar_indices(self):

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_data

            ON abastecimentos(data)

            """

        )

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_posto

            ON abastecimentos(posto)

            """

        )

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_odometro

            ON abastecimentos(odometro)

            """

        )

        self.conexao.commit()



# =============================================================================
# PARTE 129
# APAGAR TODOS OS REGISTROS
# =============================================================================

    def apagar_todos(self):

        self.cursor.execute(

            """

            DELETE FROM abastecimentos

            """

        )

        self.conexao.commit()


# =============================================================================
# PARTE 130
# TOTAL GASTO
# =============================================================================

    def total_gasto(self):

        self.cursor.execute(

            """

            SELECT SUM(valor)

            FROM abastecimentos

            """

        )

        resultado = self.cursor.fetchone()

        if resultado[0] is None:

            return 0.0

        return resultado[0]




# =============================================================================
# PARTE 401
# CLASSE PDF
# =============================================================================

class RelatorioPDF(FPDF):

    def header(self):

        self.set_font(

            "Arial",

            "B",

            14

        )

        self.cell(

            0,

            10,

            "RELATÓRIO DO SISTEMA GNV",

            0,

            1,

            "C"

        )

        self.ln(

            5

        )


# =============================================================================
# PARTE 402
# RODAPÉ
# =============================================================================

    def footer(self):

        self.set_y(

            -15

        )

        self.set_font(

            "Arial",

            "I",

            8

        )

        self.cell(

            0,

            10,

            f"Página {self.page_no()}",

            0,

            0,

            "C"

        )


# =============================================================================
# PARTE 403
# TÍTULO
# =============================================================================

    def titulo(

        self,

        texto

    ):

        self.set_font(

            "Arial",

            "B",

            12

        )

        self.cell(

            0,

            8,

            texto,

            0,

            1

        )



# =============================================================================
# PARTE 404
# LINHA
# =============================================================================

    def linha(

        self,

        titulo,

        valor

    ):

        self.set_font(

            "Arial",

            "",

            10

        )

        self.cell(

            70,

            7,

            titulo

        )

        self.cell(

            0,

            7,

            str(

                valor

            ),

            ln=True

        )

class InterfaceGNV:

    """
    Interface gráfica principal do sistema.
    """

    def __init__(

        self,

        janela

    ):

        self.janela = janela

        self.janela.title(

            "Sistema de Cálculo de GNV"

        )

        self.janela.geometry(

            "1400x850"

        )

        self.janela.minsize(
            900,
            600
        )




# =============================================================================
# PARTE 132
# NOTEBOOK
# =============================================================================

        self.notebook = ttk.Notebook(

            self.janela

        )

        self.notebook.pack(

            fill="both",

            expand=True
        )



# =============================================================================
# PARTE 133
# ABAS
# =============================================================================

        self.aba_calculos = ttk.Frame(

            self.notebook

        )

        self.aba_abastecimentos = ttk.Frame(

            self.notebook

        )

        self.aba_anp = ttk.Frame(

            self.notebook

        )

        self.aba_compressao = ttk.Frame(

            self.notebook

        )

        self.aba_historico = ttk.Frame(

            self.notebook

        )

        self.aba_sqlite = ttk.Frame(

            self.notebook

        )

        self.aba_excel = ttk.Frame(

            self.notebook

        )

        self.aba_graficos = ttk.Frame(

            self.notebook

        )

        self.aba_configuracoes = ttk.Frame(

            self.notebook

        )

        self.aba_formulas = ttk.Frame(

            self.notebook

        )

# =============================================================================
# PARTE 134
# ADICIONAR ABAS
# =============================================================================

        self.notebook.add(

            self.aba_calculos,

            text="Cálculos"

        )

        self.notebook.add(

            self.aba_abastecimentos,

            text="Abastecimentos"

        )

        self.notebook.add(

            self.aba_anp,

            text="ANP"

        )

        self.notebook.add(

            self.aba_compressao,

            text="Aquecimento / Compressão"

        )

        self.notebook.add(

            self.aba_historico,

            text="Histórico de Abastecimentos"

        )

        self.notebook.add(

            self.aba_sqlite,

            text="Banco SQLite"

        )

        self.notebook.add(

            self.aba_excel,

            text="Exportação / Excel"

        )

        self.notebook.add(

            self.aba_graficos,

            text="Gráficos de Abastecimento"

        )

        self.notebook.add(

            self.aba_configuracoes,

            text="Configurações do Sistema"

        )

        self.notebook.add(

            self.aba_formulas,

            text="Fórmulas e Física"

        )


# =============================================================================
# PARTE 177
# INICIALIZA O BANCO DE DADOS
# =============================================================================

        self.banco = BancoGNV()

        self.banco.conectar()

        self.banco.criar_tabela()

        self.banco.criar_indices()

# =============================================================================
# PARTE 178
# CARREGA O HISTÓRICO
# =============================================================================

        self.registros = self.banco.listar_abastecimentos()


# =============================================================================
# PARTE 179
# TOTAL DE REGISTROS
# =============================================================================

        self.total_registros = self.banco.contar_registros()

# =============================================================================
# PARTE 180A
# CRIA VARIÁVEL DA BARRA DE STATUS
# =============================================================================

        self.status = tk.StringVar(

            value=f"Sistema iniciado - {self.total_registros} abastecimentos cadastrados."

        )

# =============================================================================
# INTERFACE DAS ABAS PRINCIPAIS
# =============================================================================

# PARTE 135
# FRAME PRINCIPAL DA ABA CÁLCULOS
# =============================================================================

        self.frame_calculos = ttk.Frame(

            self.aba_calculos,

            padding=10

        )

        self.frame_calculos.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 136
# VOLUME DO CILINDRO
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Volume do cilindro (L):"

        ).grid(

            row=0,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_volume = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_volume.grid(

            row=0,

            column=1,

            padx=5,

            pady=5

        )

# =============================================================================
# PARTE 137
# QUANTIDADE DE CILINDROS
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Quantidade de cilindros:"

        ).grid(

            row=1,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_quantidade = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_quantidade.grid(

            row=1,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_quantidade.insert(

            0,

            "1"

        )


# =============================================================================
# PARTE 138
# PRESSÃO
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Pressão (bar):"

        ).grid(

            row=2,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_pressao = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_pressao.grid(

            row=2,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_pressao.insert(

            0,

            "200"

        )

# =============================================================================
# PARTE 139
# TEMPERATURA
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Temperatura (°C):"

        ).grid(

            row=3,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_temperatura = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_temperatura.grid(

            row=3,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_temperatura.insert(

            0,

            "20"
        )


# =============================================================================
# PARTE 140
# ALTITUDE
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Altitude (m):"

        ).grid(

            row=4,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_altitude = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_altitude.grid(

            row=4,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_altitude.insert(

            0,

            "0"

        )



# =============================================================================
# PARTE 141
# FATOR Z
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Fator Z:"

        ).grid(

            row=5,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_fator_z = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_fator_z.grid(

            row=5,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_fator_z.insert(

            0,

            "0.92"

        )


# =============================================================================
# PARTE 142
# MASSA MOLAR
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Massa molar (kg/mol):"

        ).grid(

            row=6,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_massa_molar = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_massa_molar.grid(

            row=6,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_massa_molar.insert(

            0,

            "0.01604"

        )

# =============================================================================
# DENSIDADE INFORMADA PELO POSTO
# =============================================================================

        ttk.Label(

            self.frame_calculos,

            text="Massa específica de referência (kg/m³):"

        ).grid(

            row=7,

            column=0,

            padx=5,

            pady=5,

            sticky="w"

        )

        self.entry_densidade_informada = ttk.Entry(

            self.frame_calculos,

            width=15

        )

        self.entry_densidade_informada.grid(

            row=7,

            column=1,

            padx=5,

            pady=5

        )

        self.entry_densidade_informada.insert(

            0,

            "0,76"

        )

# =============================================================================
# PARTE 143
# PAINEL DE BOTÕES E ÁREA DE RESULTADOS
# =============================================================================

        self.frame_botoes_calculos = ttk.Frame(
            self.frame_calculos
        )
        self.frame_botoes_calculos.grid(
            row=8,
            column=0,
            columnspan=4,
            padx=10,
            pady=(8, 8),
            sticky="ew"
        )

        for coluna in range(3):
            self.frame_botoes_calculos.grid_columnconfigure(
                coluna,
                weight=1
            )

        self.botao_calcular = ttk.Button(
            self.frame_botoes_calculos,
            text="Calcular",
            command=self.executar_calculo
        )
        self.botao_calcular.grid(
            row=0,
            column=0,
            padx=5,
            sticky="ew"
        )

        self.botao_limpar = ttk.Button(
            self.frame_botoes_calculos,
            text="Limpar Resultados",
            command=self.limpar_resultados
        )
        self.botao_limpar.grid(
            row=0,
            column=1,
            padx=5,
            sticky="ew"
        )

        self.botao_limpar_campos = ttk.Button(
            self.frame_botoes_calculos,
            text="Limpar Campos",
            command=self.limpar_campos
        )
        self.botao_limpar_campos.grid(
            row=0,
            column=2,
            padx=5,
            sticky="ew"
        )

        self.frame_resultados = ttk.Frame(
            self.frame_calculos
        )
        self.frame_resultados.grid(
            row=9,
            column=0,
            columnspan=4,
            padx=10,
            pady=(0, 10),
            sticky="nsew"
        )

        self.scroll_resultados = ttk.Scrollbar(
            self.frame_resultados,
            orient="vertical"
        )
        self.scroll_resultados.pack(
            side="right",
            fill="y"
        )

        self.texto_resultados = tk.Text(
            self.frame_resultados,
            font=("Consolas", 10),
            wrap="none",
            yscrollcommand=self.scroll_resultados.set
        )
        self.texto_resultados.pack(
            side="left",
            fill="both",
            expand=True
        )
        self.scroll_resultados.config(
            command=self.texto_resultados.yview
        )

        self.frame_calculos.grid_rowconfigure(
            9,
            weight=1
        )
        for coluna in range(4):
            self.frame_calculos.grid_columnconfigure(
                coluna,
                weight=1
            )


# =============================================================================
# PARTE 157
# FRAME ABA ABASTECIMENTOS
# =============================================================================

        self.frame_abastecimentos = ttk.Frame(

            self.aba_abastecimentos,

            padding=10

        )

        self.frame_abastecimentos.pack(

            fill="both",

            expand=True

        )

        ttk.Label(

            self.frame_abastecimentos,

            text="Cadastro e conferência física do abastecimento",

            font=("Arial", 14, "bold")

        ).grid(

            row=0,

            column=0,

            columnspan=3,

            pady=12

        )

        campos = [

            ("Data", "entry_data", 1),
            ("Posto", "entry_posto", 2),
            ("Cidade", "entry_cidade", 3),
            ("Odômetro (km)", "entry_odometro", 4),
            ("Capacidade do cilindro (L)", "entry_capacidade_cilindro", 5),
            ("Volume marcado pela bomba (m³)", "entry_volume_abastecido", 6),
            ("Preço por m³ (R$)", "entry_preco_m3", 7),
            ("Temperatura ambiente no abastecimento (°C)", "entry_temp_abastecimento", 8),
            ("Pressão inicial (bar)", "entry_pressao_inicial", 9),
            ("Pressão final (bar)", "entry_pressao_final", 10),
            ("Altitude (m)", "entry_altitude_abastecimento", 11),
            ("Massa específica do GNV (kg/m³)", "entry_densidade_informada_abastecimento", 12),

        ]

        for texto, atributo, linha in campos:

            ttk.Label(

                self.frame_abastecimentos,

                text=texto

            ).grid(

                row=linha,

                column=0,

                sticky="w",

                padx=5,

                pady=4

            )

            entrada = ttk.Entry(

                self.frame_abastecimentos,

                width=22

            )

            entrada.grid(

                row=linha,

                column=1,

                padx=5,

                pady=4

            )

            setattr(self, atributo, entrada)

        self.entry_pressao_abastecimento = self.entry_pressao_final
        self.entry_densidade_informada_abastecimento.insert(0, "0,76")

        ttk.Label(

            self.frame_abastecimentos,

            text="Observações"

        ).grid(

            row=12,

            column=0,

            sticky="nw",

            padx=5,

            pady=4

        )

        self.texto_observacoes = tk.Text(

            self.frame_abastecimentos,

            width=45,

            height=4

        )

        self.texto_observacoes.grid(

            row=12,

            column=1,

            padx=5,

            pady=4

        )

        self.botao_salvar = ttk.Button(

            self.frame_abastecimentos,

            text="Salvar Abastecimento",

            command=self.salvar_abastecimento

        )

        self.botao_salvar.grid(

            row=13,

            column=0,

            padx=10,

            pady=10,

            sticky="ew"

        )

        self.botao_novo = ttk.Button(

            self.frame_abastecimentos,

            text="Novo",

            command=self.novo_abastecimento

        )

        self.botao_novo.grid(

            row=13,

            column=1,

            padx=10,

            pady=10,

            sticky="ew"

        )

        self.texto_comparacao_abastecimento = tk.Text(

            self.frame_abastecimentos,

            width=78,

            height=12,

            font=("Consolas", 9)

        )

        self.texto_comparacao_abastecimento.grid(

            row=1,

            column=2,

            rowspan=14,

            padx=15,

            pady=5,

            sticky="nsew"

        )

        self.texto_comparacao_abastecimento.delete("1.0", "end")
        self.texto_comparacao_abastecimento.insert(
            tk.END,
            "RESULTADO DA ANÁLISE DO ABASTECIMENTO\n"
            "========================================\n\n"
            "Preencha os dados à esquerda e clique em 'Salvar Abastecimento'.\n"
            "A temperatura solicitada é a temperatura AMBIENTE, pois normalmente\n"
            "não temos a temperatura real do GNV dentro do cilindro durante o enchimento.\n"
        )

        self.frame_abastecimentos.grid_columnconfigure(2, weight=1)
        self.frame_abastecimentos.grid_rowconfigure(14, weight=1)


# =============================================================================
# ABA BANCO SQLITE
# =============================================================================

        self.frame_sqlite = ttk.Frame(
            self.aba_sqlite,
            padding=10
        )
        self.frame_sqlite.pack(fill="both", expand=True)
        ttk.Label(
            self.frame_sqlite,
            text="Registros armazenados no Banco SQLite",
            font=("Arial", 14, "bold")
        ).pack(pady=10)
        ttk.Button(
            self.frame_sqlite,
            text="Atualizar Banco SQLite",
            command=self.atualizar_sqlite
        ).pack(anchor="w", padx=5, pady=5)
        colunas_sqlite = (
            "ID", "Data", "Posto", "Cidade", "Odômetro",
            "Volume bomba (m³)", "Preço/m³", "Valor total",
            "Temp. ambiente (°C)", "Pressão registrada (bar)",
            "Altitude (m)", "Capacidade cilindro (L)",
            "Pressão inicial (bar)", "Pressão final (bar)",
            "Massa específica (kg/m³)",
            "Metragem teórica (m³)",
            "Metragem ANP (m³)",
            "Metragem científica (m³)"
        )
        self.tree_sqlite = ttk.Treeview(
            self.frame_sqlite,
            columns=colunas_sqlite,
            show="headings"
        )
        for coluna in colunas_sqlite:
            self.tree_sqlite.heading(coluna, text=coluna)
            self.tree_sqlite.column(coluna, width=110, anchor="center")
        self.tree_sqlite.pack(side="left", fill="both", expand=True)
        scrollbar_sqlite = ttk.Scrollbar(
            self.frame_sqlite,
            orient="vertical",
            command=self.tree_sqlite.yview
        )
        scrollbar_sqlite.pack(side="right", fill="y")
        self.tree_sqlite.configure(yscrollcommand=scrollbar_sqlite.set)

# =============================================================================
# ABA EXCEL
# =============================================================================

        self.frame_excel = ttk.Frame(
            self.aba_excel,
            padding=10
        )
        self.frame_excel.pack(fill="both", expand=True)
        ttk.Label(
            self.frame_excel,
            text="Exportação e integração com Excel",
            font=("Arial", 14, "bold")
        ).pack(pady=10)
        self.arquivo_excel = ""
        self.entry_arquivo_excel = ttk.Entry(
            self.frame_excel,
            width=80
        )
        self.entry_arquivo_excel.pack(fill="x", padx=10, pady=5)
        frame_excel_botoes = ttk.Frame(self.frame_excel)
        frame_excel_botoes.pack(pady=10)
        ttk.Button(
            frame_excel_botoes,
            text="Selecionar Excel",
            command=self.selecionar_arquivo_excel
        ).pack(side="left", padx=5)
        ttk.Button(
            frame_excel_botoes,
            text="Exportar Excel",
            command=self.exportar_excel
        ).pack(side="left", padx=5)
        ttk.Button(
            frame_excel_botoes,
            text="Gerar Relatório PDF",
            command=self.exportar_pdf
        ).pack(side="left", padx=5)

# =============================================================================
# ABA GRÁFICOS
# =============================================================================

        self.frame_graficos = ttk.Frame(
            self.aba_graficos,
            padding=10
        )
        self.frame_graficos.pack(fill="both", expand=True)
        ttk.Label(
            self.frame_graficos,
            text="Gráficos de abastecimentos",
            font=("Arial", 14, "bold")
        ).pack(pady=10)
        frame_controles_graficos = ttk.Frame(self.frame_graficos)
        frame_controles_graficos.pack(fill="x", pady=(0, 8))
        ttk.Label(frame_controles_graficos, text="Tipo de gráfico:").pack(side="left", padx=(0, 6))
        self.tipo_grafico = tk.StringVar(value="Gasto por posto")
        self.combo_tipo_grafico = ttk.Combobox(
            frame_controles_graficos,
            textvariable=self.tipo_grafico,
            values=("Gasto por posto", "Volume por posto", "Abastecimentos por posto", "Evolução do volume", "Km por m³", "Bomba × teórico"),
            state="readonly", width=28
        )
        self.combo_tipo_grafico.pack(side="left", padx=5)
        self.combo_tipo_grafico.bind("<<ComboboxSelected>>", lambda _event: self.atualizar_grafico())
        ttk.Button(frame_controles_graficos, text="Atualizar Gráfico", command=self.atualizar_grafico).pack(side="left", padx=5)
        ttk.Button(frame_controles_graficos, text="Limpar", command=lambda: self.canvas_grafico.delete("all")).pack(side="left", padx=5)
        self.canvas_grafico = tk.Canvas(self.frame_graficos, background="white", height=460, highlightthickness=1)
        self.canvas_grafico.pack(fill="both", expand=True, padx=10, pady=10)
        self.canvas_grafico.bind("<Configure>", lambda _event: self.atualizar_grafico())

# =============================================================================
# ABA CONFIGURAÇÕES
# =============================================================================

        self.frame_configuracoes = ttk.Frame(
            self.aba_configuracoes,
            padding=10
        )
        self.frame_configuracoes.pack(fill="both", expand=True)
        ttk.Label(
            self.frame_configuracoes,
            text="Configurações do Sistema",
            font=("Arial", 14, "bold")
        ).pack(pady=10)
        self.config_tema = tk.StringVar(value="claro")
        self.config_idioma = tk.StringVar(value="pt-BR")
        self.config_backup = tk.BooleanVar(value=True)
        ttk.Label(self.frame_configuracoes, text="Tema:").pack(anchor="w", padx=10, pady=5)
        ttk.Combobox(
            self.frame_configuracoes, textvariable=self.config_tema,
            values=("claro", "escuro"), state="readonly", width=20
        ).pack(anchor="w", padx=10)
        ttk.Label(self.frame_configuracoes, text="Idioma:").pack(anchor="w", padx=10, pady=5)
        ttk.Combobox(
            self.frame_configuracoes, textvariable=self.config_idioma,
            values=IDIOMAS_DISPONIVEIS, state="readonly", width=20
        ).pack(anchor="w", padx=10)
        ttk.Checkbutton(
            self.frame_configuracoes, text="Backup automático",
            variable=self.config_backup
        ).pack(anchor="w", padx=10, pady=10)
        frame_config_botoes = ttk.Frame(self.frame_configuracoes)
        frame_config_botoes.pack(pady=10)
        ttk.Button(
            frame_config_botoes, text="Salvar Configurações",
            command=self.salvar_configuracoes_tela
        ).pack(side="left", padx=5)
        ttk.Button(
            frame_config_botoes, text="Restaurar Padrão",
            command=self.reiniciar_configuracoes
        ).pack(side="left", padx=5)
        ttk.Button(
            frame_config_botoes, text="Exportar JSON",
            command=self.exportar_configuracoes
        ).pack(side="left", padx=5)
        ttk.Button(
            frame_config_botoes, text="Importar JSON",
            command=self.importar_configuracoes
        ).pack(side="left", padx=5)

# =============================================================================
# ABA FÓRMULAS E FÍSICA
# =============================================================================

        self.frame_formulas = ttk.Frame(
            self.aba_formulas,
            padding=12
        )
        self.frame_formulas.pack(
            fill="both",
            expand=True
        )

        ttk.Label(
            self.frame_formulas,
            text="Fórmulas, variáveis e fundamentos físicos",
            font=("Arial", 15, "bold")
        ).pack(
            anchor="w",
            pady=(0, 8)
        )

        self.scroll_formulas = ttk.Scrollbar(
            self.frame_formulas,
            orient="vertical"
        )
        self.scroll_formulas.pack(
            side="right",
            fill="y"
        )

        self.texto_formulas = tk.Text(
            self.frame_formulas,
            wrap="word",
            font=("Segoe UI", 10),
            yscrollcommand=self.scroll_formulas.set
        )
        self.texto_formulas.pack(
            side="left",
            fill="both",
            expand=True
        )
        self.scroll_formulas.config(
            command=self.texto_formulas.yview
        )

        texto_formulas = """FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV
============================================================

PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP
========================================

A ANP informa, para volumes médios comercializados de gás natural,
condições de referência de 20 °C e 1,033 kgf/cm². A condição padrão de
medição é definida como pressão absoluta de 0,101325 MPa e temperatura
de 20 °C.

IMPORTANTE: a ANP define a condição de referência. A fórmula implementada
neste programa é uma ESTIMATIVA FÍSICA DE CONVERSÃO e não afirma reproduzir
o algoritmo interno de um dispenser de GNV.

1. VOLUME FÍSICO DO CILINDRO
----------------------------
Vcil = capacidade(L) / 1000

26 L / 1000 = 0,026 m³.

Esse é o espaço físico interno do cilindro. Não é o volume normalizado
indicado pela bomba.

2. PRESSÃO ABSOLUTA
-------------------
Pabs = Pmanométrica + Patm

As equações de estado usam pressão absoluta.

3. TEMPERATURA ABSOLUTA
-----------------------
T(K) = T(°C) + 273,15

4. CONVERSÃO PARA 20 °C
-----------------------
Para uma quantidade de matéria fixa, em modelo ideal:

Vref = V × (P/Pref) × (Tref/T)

Tref = 293,15 K e Pref ≈ 1,01325 bar.

PARTE B — MODELO CIENTÍFICO DE GÁS REAL
========================================

1. EQUAÇÃO DE ESTADO
--------------------
P V = Z n R T

n = P V / (Z R T)

P = pressão absoluta (Pa)
V = volume físico (m³)
Z = fator de compressibilidade
n = quantidade de matéria (mol)
R = 8,314462618 J/(mol·K)
T = temperatura absoluta (K)

2. FATOR DE COMPRESSIBILIDADE Z
-------------------------------
Z = P V / (n R T)

Z = 1 representa o gás ideal. Para gás natural real, Z depende de
pressão, temperatura e composição. Um Z fixo informado pelo usuário é
uma aproximação, não uma determinação metrológica de Z.

3. QUANTIDADE DE GÁS ADICIONADA
-------------------------------
n_inicial = P_inicial_abs × V / (Z R T)
n_final   = P_final_abs × V / (Z R T)
Δn = n_final − n_inicial

4. MASSA ADICIONADA
-------------------
m = Δn × M

M é a massa molar do GNV em kg/mol.

5. VOLUME EQUIVALENTE NA TEMPERATURA INFORMADA
------------------------------------------------
O programa calcula primeiro os mols a partir de PV = Z n R T.
Depois pode expressar esses mesmos mols a uma pressão de referência:

Vref(T) = n R T / Pref

IMPORTANTE: neste programa, n foi calculado usando a própria temperatura
informada. Por isso, ao substituir n = P V / (Z R T), a temperatura cancela:

Vref(T) = P V / (Z Pref)

Assim, para um cilindro de 26 L a 220 bar e Z=0,92, o volume equivalente
a 1,01325 bar na temperatura informada pode permanecer praticamente
6,164 m³ tanto a 5 °C quanto a 20 °C ou 100 °C. Isso NÃO significa que
a mesma quantidade de gás teria o mesmo volume em duas temperaturas quando
n é mantido fixo; significa que o programa está recalculando a quantidade
de matéria para cada estado de pressão/temperatura informado.

6. CONVERSÃO CIENTÍFICA PARA 20 °C
----------------------------------
Os mesmos mols calculados pelo modelo Z informado são convertidos para 20 °C:

V20 = n R T20 / Pref

Se a temperatura informada for menor que 20 °C, V20 tende a ser maior.
Se for maior que 20 °C, V20 tende a ser menor, porque aqui n é mantido
fixo durante a conversão.

7. CONVERSÃO ANP/IDEALIZADA (Z=1)
---------------------------------
A aba ANP calcula separadamente uma estimativa com Z=1 e a condição de
referência de 20 °C e 1,033 kgf/cm² (aproximadamente 1,01325 bar).
Esse valor não deve ser confundido com o resultado científico que usa
Z=0,92, por exemplo. São dois modelos diferentes.

6. DENSIDADE DO GÁS REAL
------------------------
ρ = P M / (Z R T)

PARTE C — O QUE SERIA UM MODELO MAIS PRECISO
============================================

Em GNV a alta pressão, não é adequado considerar Z como uma constante
universal. Para elevar a precisão é necessário conhecer a composição do
gás e calcular suas propriedades termodinâmicas em função de P e T.

A ISO 12213 descreve métodos para o cálculo do fator de compressibilidade
de gás natural. A ISO 12213-2 usa composição molar; a ISO 12213-3 usa
propriedades físicas como poder calorífico, densidade relativa e CO₂,
além de pressão e temperatura.

AGA8 e GERG são modelos utilizados para propriedades de gás natural. O
NIST descreve AGA8 e GERG entre as equações de estado usadas em aplicações
de medição e propriedades termodinâmicas de gás natural.

Portanto, a evolução científica do programa deve ser:
1) obter composição do GNV;
2) calcular Z(P,T,composição), em vez de usar Z fixo;
3) considerar a temperatura real do gás durante o abastecimento;
4) conhecer as condições efetivamente usadas pelo medidor;
5) trabalhar com incerteza de medição.

PARTE D — TEMPERATURA DURANTE O ABASTECIMENTO
==============================================

Durante o enchimento existe entrada de massa e transferência de calor.
A aba Compressão / Temperatura usa uma compressão adiabática reversível
somente como cenário didático:

T₂/T₁ = (P₂/P₁)^((k−1)/k)
P·V^k = constante
V₂/V₁ = (P₁/P₂)^(1/k)

O abastecimento real é um sistema aberto, com troca de calor entre gás,
parede do cilindro, mangueira e ambiente. A temperatura calculada nessa
aba NÃO é uma medição da temperatura real do GNV.

PARTE E — COMPARAÇÃO COM A BOMBA
================================

Diferença = volume indicado pela bomba − volume calculado
Diferença percentual = diferença / volume calculado × 100

Uma diferença grande é um indício para investigação. Ela não constitui,
sozinha, prova metrológica de fraude. Uma conclusão técnica exige dados
do medidor, condições de referência, temperatura real do gás,
composição/Z, calibração e incerteza de medição.

FONTES E FUNDAMENTAÇÃO
=======================

ANP — Publicidade dos preços de gás natural:
https://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural

ANP — Glossário C / Condição Padrão de Medição:
https://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c

ISO 12213-2:2006 — cálculo do fator de compressibilidade por composição:
https://www.iso.org/standard/44411.html

ISO 12213-3:2006 — cálculo do fator de compressibilidade por propriedades:
https://www.iso.org/standard/44412.html

MIT OpenCourseWare — Thermodynamics: equação de estado do gás ideal PV = nRT:
https://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/

MIT OpenCourseWare — Materials at Equilibrium: propriedades de gases ideais e PV = nRT:
https://ocw.mit.edu/courses/3-20-materials-at-equilibrium-sma-5111-fall-2003/

Purdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics: gás ideal e fator de compressibilidade Z:
https://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf

Stanford University — Thermodynamics / Ideal Gas Law:
https://web.stanford.edu/~peastman/statmech/thermodynamics.html

Stanford University — Fundamentals of Compressible Flow, gases ideais e propriedades termodinâmicas:
https://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf

ITA — Departamento de Ciência e Tecnologia Aeroespacial: catálogo de graduação e disciplinas de Termodinâmica/Termodinâmica Aplicada:
https://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf

IME-USP — pesquisas acadêmicas envolvendo termodinâmica e sistemas de muitos corpos:
https://lattes.ime.usp.br/posmap/membro-1498618533380124.html

NIST — REFPROP / propriedades de misturas e AGA8:
https://www.nist.gov/srd/refprop

NIST — comparação de equações de estado para medição de gás natural:
https://www.nist.gov/publications/comparison-five-natural-gas-equations-state-used-flow-and-energy-measurement
"""
        self.texto_formulas.insert(
            tk.END,
            texto_formulas.strip()
        )
        self.texto_formulas.configure(
            state="disabled"
        )

# PARTE 406
# BOTÃO PDF
# =============================================================================

        self.botao_pdf = ttk.Button(

            self.aba_excel,

            text="Gerar Relatório PDF",

            command=self.exportar_pdf

        )

        self.botao_pdf.pack(

            pady=10

        )









# =============================================================================


# =============================================================================
# PARTE 180
# STATUS INICIAL
# =============================================================================

        if hasattr(self, "status"):

            self.status.set(

                f"Sistema iniciado - {self.total_registros} abastecimentos cadastrados."

            )


# =============================================================================
# PARTE 182
# FRAME DO HISTÓRICO
# =============================================================================

        self.frame_historico = ttk.Frame(

            self.aba_historico,

            padding=10

        )

        self.frame_historico.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 183
# TÍTULO
# =============================================================================

        ttk.Label(

            self.frame_historico,

            text="Histórico de Abastecimentos",

            font=("Arial",14,"bold")

        ).pack(

            pady=10

        )

# =============================================================================
# PARTE 192
# CAMPO DE PESQUISA
# =============================================================================

        self.frame_pesquisa = ttk.Frame(

            self.frame_historico

        )

        self.frame_pesquisa.pack(

            fill="x",

            pady=5

        )

        ttk.Label(

            self.frame_pesquisa,

            text="Pesquisar:"

        ).pack(

            side="left"

        )

        self.entry_pesquisa = ttk.Entry(

            self.frame_pesquisa,

            width=40

        )

        self.entry_pesquisa.pack(

            side="left",

            padx=5

        )

# =============================================================================
# PARTE 193
# BOTÃO PESQUISAR
# =============================================================================

        self.botao_pesquisar = ttk.Button(

            self.frame_pesquisa,

            text="Pesquisar",

            command=self.pesquisar_historico

        )

        self.botao_pesquisar.pack(

            side="left",

            padx=5

        )


# =============================================================================
# PARTE 194
# BOTÃO MOSTRAR TODOS
# =============================================================================

        self.botao_todos = ttk.Button(

            self.frame_pesquisa,

            text="Mostrar Todos",

            command=self.atualizar_historico

        )

        self.botao_todos.pack(

            side="left",

            padx=5

        )


# =============================================================================
# PARTE 196
# ENTER PESQUISA
# =============================================================================

        self.entry_pesquisa.bind(

            "<Return>",

            lambda event: self.pesquisar_historico()

        )



# =============================================================================
# ABA ANP - CONDIÇÃO DE REFERÊNCIA
# =============================================================================

        frame_anp = ttk.Frame(self.aba_anp, padding=12)
        frame_anp.pack(fill="both", expand=True)

        ttk.Label(
            frame_anp,
            text="Cálculo pela condição de referência da ANP",
            font=("Arial", 15, "bold")
        ).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 4))

        ttk.Label(
            frame_anp,
            text=(
                "A ANP define a condição de referência do gás natural. "
                "Os campos abaixo permitem aplicar uma estimativa física "
                "de conversão para essa condição. Este cálculo não reproduz "
                "o algoritmo interno do dispenser."
            ),
            wraplength=1100,
            justify="left"
        ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(0, 8))

        anp_campos = [
            ("Capacidade do cilindro (L)", "anp_capacidade", "26"),
            ("Pressão inicial manométrica (bar)", "anp_pi", "0"),
            ("Pressão final manométrica (bar)", "anp_pf", "220"),
            ("Temperatura ambiente (°C)", "anp_temp", "20"),
            ("Altitude (m)", "anp_alt", "0"),
            ("Volume indicado pela bomba (m³) — opcional", "anp_bomba", ""),
        ]
        for i, (rotulo, attr, valor) in enumerate(anp_campos):
            r = 2 + i // 2
            c = (i % 2) * 2
            ttk.Label(frame_anp, text=rotulo).grid(row=r, column=c, sticky="w", padx=(0, 5), pady=4)
            e = ttk.Entry(frame_anp, width=18)
            e.grid(row=r, column=c + 1, sticky="w", padx=(0, 18), pady=4)
            if valor:
                e.insert(0, valor)
            setattr(self, attr, e)

        frame_anp_botoes = ttk.Frame(frame_anp)
        frame_anp_botoes.grid(row=5, column=0, columnspan=4, sticky="ew", pady=8)

        texto_anp = tk.Text(frame_anp, wrap="word", font=("Consolas", 10), height=24)
        scroll_anp = ttk.Scrollbar(frame_anp, orient="vertical", command=texto_anp.yview)
        texto_anp.configure(yscrollcommand=scroll_anp.set)
        texto_anp.grid(row=6, column=0, columnspan=3, sticky="nsew", padx=(0, 5))
        scroll_anp.grid(row=6, column=3, sticky="ns")
        frame_anp.grid_columnconfigure(2, weight=1)
        frame_anp.grid_rowconfigure(6, weight=1)

        def copiar_dados_abastecimento_para_anp():
            pares = [
                (self.anp_capacidade, self.entry_capacidade_cilindro),
                (self.anp_pi, self.entry_pressao_inicial),
                (self.anp_pf, self.entry_pressao_final),
                (self.anp_temp, self.entry_temp_abastecimento),
                (self.anp_alt, self.entry_altitude_abastecimento),
                (self.anp_bomba, self.entry_volume_abastecido),
            ]
            for destino, origem in pares:
                destino.delete(0, tk.END)
                destino.insert(0, origem.get())

        def atualizar_calculo_anp_tela():
            try:
                c = converter_numero(self.anp_capacidade.get())
                pi = converter_numero(self.anp_pi.get())
                pf = converter_numero(self.anp_pf.get())
                temp = converter_numero(self.anp_temp.get())
                alt = converter_numero(self.anp_alt.get())
                bomba_texto = self.anp_bomba.get().strip()
                bomba = converter_numero(bomba_texto) if bomba_texto else None

                if c <= 0:
                    raise ValueError("A capacidade do cilindro deve ser maior que zero.")
                if pi < 0 or pf < 0:
                    raise ValueError("As pressões devem ser maiores ou iguais a zero.")
                if pf < pi:
                    raise ValueError("A pressão final deve ser maior ou igual à pressão inicial.")
                if temp <= -273.15:
                    raise ValueError("A temperatura deve ser maior que -273,15 °C.")

                v = calcular_volume_anp_referencia(c, pi, pf, temp, alt)
                patm = calcular_pressao_atmosferica(alt)
                pia = pi + patm
                pfa = pf + patm

                linhas = [
                    "CÁLCULO PELA CONDIÇÃO DE REFERÊNCIA DA ANP",
                    "=" * 62,
                    "",
                    "DADOS INFORMADOS",
                    f"Capacidade física do cilindro : {formatar_numero_br(c, 2)} L",
                    f"Pressão inicial manométrica   : {formatar_numero_br(pi, 4)} bar",
                    f"Pressão final manométrica     : {formatar_numero_br(pf, 4)} bar",
                    f"Pressão atmosférica estimada  : {formatar_numero_br(patm, 5)} bar",
                    f"Pressão inicial absoluta      : {formatar_numero_br(pia, 5)} bar",
                    f"Pressão final absoluta        : {formatar_numero_br(pfa, 5)} bar",
                    f"Temperatura ambiente          : {formatar_numero_br(temp, 2)} °C",
                    f"Altitude                      : {formatar_numero_br(alt, 1)} m",
                    "",
                    "CONDIÇÃO DE REFERÊNCIA ANP",
                    "Tref = 20 °C = 293,15 K",
                    "Pref = 1,033 kgf/cm² (aprox. 1,01325 bar)",
                    "",
                    "ESTIMATIVA FÍSICA ANP/IDEALIZADA (Z=1)",
                    "Vadd_ref = Vcil × (Pfinal_abs − Pinicial_abs) / Pref × Tref / Tgas",
                    "Tgas é a temperatura informada pelo usuário (normalmente ambiente).",
                    "",
                    f"VOLUME ADICIONADO EQUIVALENTE ANP/IDEALIZADO : {formatar_numero_br(v, 5)} m³",
                    f"VOLUME ADICIONADO EQUIVALENTE ANP/IDEALIZADO : {formatar_numero_br(v * 1000.0, 2)} L equivalentes",
                    "",
                    "INTERPRETAÇÃO",
                    "O valor acima NÃO é a capacidade física do cilindro.",
                    "É o volume equivalente da quantidade calculada na condição",
                    "de referência de 20 °C e aproximadamente 1,01325 bar.",
                    "",
                    "IMPORTANTE",
                    "A ANP estabelece a condição de referência, mas esta fórmula",
                    "não deve ser apresentada como o algoritmo do dispenser.",
                    "Para uma análise metrológica são necessários os dados e",
                    "condições efetivamente utilizados pelo sistema de medição.",
                ]

                if bomba is not None:
                    diferenca = bomba - v
                    pct = (diferenca / v * 100.0) if v > 0 else 0.0
                    linhas += [
                        "",
                        "COMPARAÇÃO OPCIONAL COM A BOMBA",
                        f"Volume indicado pela bomba       : {formatar_numero_br(bomba, 5)} m³",
                        f"Diferença bomba − referência    : {formatar_numero_br(diferenca, 5)} m³",
                        f"Diferença percentual             : {formatar_numero_br(pct, 2)} %",
                    ]

                linhas += [
                    "",
                    "FUNDAMENTAÇÃO",
                    "A ANP informa 20 °C e 1,033 kgf/cm² como condição de referência",
                    "para volumes médios comercializados de gás natural.",
                    "Consulte a aba Fórmulas e Física para as fontes oficiais.",
                ]

                texto_anp.delete("1.0", "end")
                texto_anp.insert(tk.END, "\n".join(linhas))
                texto_anp.see("1.0")

            except (ValueError, ZeroDivisionError) as e:
                messagebox.showerror("Cálculo ANP", str(e))

        ttk.Button(
            frame_anp_botoes,
            text="Calcular exclusivamente pela condição de referência ANP",
            command=atualizar_calculo_anp_tela
        ).pack(side="left", padx=(0, 8))
        ttk.Button(
            frame_anp_botoes,
            text="Copiar dados da aba Abastecimentos",
            command=copiar_dados_abastecimento_para_anp
        ).pack(side="left")

        texto_anp.insert(
            tk.END,
            "Preencha os campos acima ou use 'Copiar dados da aba Abastecimentos'.\n"
            "O volume da bomba é opcional e só é usado para comparação."
        )

# ABA COMPRESSÃO / TEMPERATURA
# =============================================================================

        fc=ttk.Frame(self.aba_compressao,padding=10); fc.pack(fill="both",expand=True)
        ttk.Label(fc,text="Aquecimento durante compressão — modelo termodinâmico idealizado",font=("Arial",15,"bold")).pack(anchor="w",pady=(0,4))
        ttk.Label(fc,text="O enchimento real não é uma compressão adiabática simples. Esta aba mostra o cenário idealizado e não mede a temperatura real do GNV.",wraplength=1250).pack(anchor="w",pady=(0,5))
        fi=ttk.Frame(fc); fi.pack(fill="x")
        for col,(lab,attr,val) in enumerate([("P inicial manométrica (bar)","comp_pi","1"),("P final manométrica (bar)","comp_pf","220"),("T inicial (°C)","comp_ti","24"),("Altitude (m)","comp_alt","0"),("k = Cp/Cv","comp_k","1,294")]):
            ttk.Label(fi,text=lab).grid(row=0,column=col,padx=3,sticky="w"); e=ttk.Entry(fi,width=13); e.grid(row=1,column=col,padx=3); e.insert(0,val); setattr(self,attr,e)
        self.texto_comp_info=tk.Text(fc,height=7,wrap="word",font=("Consolas",9)); self.texto_comp_info.pack(fill="x",pady=4)
        fg=ttk.Frame(fc); fg.pack(fill="both",expand=True); fg.grid_columnconfigure(0,weight=1); fg.grid_columnconfigure(1,weight=1); fg.grid_rowconfigure(0,weight=1); fg.grid_rowconfigure(1,weight=1)
        self.canvas_pt_comp=tk.Canvas(fg,background="white"); self.canvas_tv_comp=tk.Canvas(fg,background="white"); self.canvas_vp_comp=tk.Canvas(fg,background="white")
        self.canvas_pt_comp.grid(row=0,column=0,sticky="nsew",padx=3,pady=3); self.canvas_tv_comp.grid(row=0,column=1,sticky="nsew",padx=3,pady=3); self.canvas_vp_comp.grid(row=1,column=0,columnspan=2,sticky="nsew",padx=3,pady=3)
        def plot(cv,ds,xk,yk,title,xlab,ylab):
            cv.delete("all"); cv.update_idletasks(); w=max(cv.winfo_width(),450); h=max(cv.winfo_height(),170); xs=[d[xk] for d in ds]; ys=[d[yk] for d in ds]; xmin,xmax=min(xs),max(xs); ymin,ymax=min(ys),max(ys); xmax=xmax if xmax!=xmin else xmin+1; ymax=ymax if ymax!=ymin else ymin+1; ml,mr,mt,mb=55,15,25,30; X=lambda x:ml+(x-xmin)/(xmax-xmin)*(w-ml-mr); Y=lambda y:h-mb-(y-ymin)/(ymax-ymin)*(h-mt-mb); cv.create_text(w/2,12,text=title,font=("Arial",9,"bold")); cv.create_line(ml,mt,ml,h-mb); cv.create_line(ml,h-mb,w-mr,h-mb); cv.create_text(w/2,h-10,text=xlab); cv.create_text(12,h/2,text=ylab,angle=90); pts=[]
            for x,y in zip(xs,ys): pts += [X(x),Y(y)]
            cv.create_line(*pts,fill="#1f4e79",width=2)
        def atual_comp():
            try:
                pi=converter_numero(self.comp_pi.get()); pf=converter_numero(self.comp_pf.get()); ti=converter_numero(self.comp_ti.get()); alt=converter_numero(self.comp_alt.get()); k=converter_numero(self.comp_k.get()); m=calcular_compressao_ideal_adiabatica(pi,pf,ti,alt,k); ds=gerar_pontos_compressao_adiabatica(pi,pf,ti,alt,k,40)
                self.texto_comp_info.delete("1.0","end"); self.texto_comp_info.insert(tk.END,"MODELO ADIABÁTICO IDEALIZADO\n"+f"T inicial: {formatar_numero_br(ti,2)} °C | T final idealizada: {formatar_numero_br(m['temperatura_final_c'],2)} °C | ΔT: {formatar_numero_br(m['aumento_temperatura_c'],2)} °C\n"+f"V final/V inicial: {formatar_numero_br(m['volume_relativo_final']*100,2)} % | redução de V: {formatar_numero_br(m['reducao_volume_percentual'],2)} %\n\nT₂/T₁=(P₂/P₁)^((k−1)/k)   V₂/V₁=(P₁/P₂)^(1/k)   P·V^k=constante.\nForma logarítmica para análise de gráficos: ln(T₂/T₁)=((k−1)/k)·ln(P₂/P₁).\nATENÇÃO: cenário idealizado; o abastecimento real é um sistema aberto e troca calor com cilindro/ambiente.")
                plot(self.canvas_pt_comp,ds,"pressao_man_bar","temperatura_c","Pressão × temperatura","Pressão (bar)","T (°C)"); plot(self.canvas_tv_comp,ds,"temperatura_c","pressao_man_bar","Temperatura × pressão","T (°C)","P (bar)"); plot(self.canvas_vp_comp,ds,"pressao_man_bar","volume_relativo","Pressão × volume relativo","Pressão (bar)","V/V₀")
            except (ValueError,ZeroDivisionError) as e: messagebox.showerror("Compressão / Temperatura",str(e))
        ttk.Button(fc,text="Calcular e atualizar gráficos",command=atual_comp).pack(anchor="w",pady=4); self.janela.after(400,atual_comp)


# =============================================================================
# PARTE 184
# TREEVIEW
# =============================================================================

        self.tree = ttk.Treeview(

            self.frame_historico,

            columns=(

                "data",

                "posto",

                "cidade",

                "volume",

                "teorico",

                "valor"

            ),

            show="headings"

        )


# =============================================================================
# PARTE 185
# CABEÇALHOS
# =============================================================================

        self.tree.heading(

            "data",

            text="Data"

        )

        self.tree.heading(

            "posto",

            text="Posto"

        )

        self.tree.heading(

            "cidade",

            text="Cidade"

        )

        self.tree.heading(

            "volume",

            text="Volume"

        )

        self.tree.heading(

            "teorico",

            text="Teórico m³"

        )

        self.tree.heading(

            "valor",

            text="Valor"

        )


# =============================================================================
# PARTE 191
# LARGURA DAS COLUNAS
# =============================================================================

        self.tree.column(

            "data",

            width=120

        )

        self.tree.column(

            "posto",

            width=260

        )

        self.tree.column(

            "cidade",

            width=180

        )

        self.tree.column(

            "volume",

            width=100,

            anchor="center"

        )

        self.tree.column(

            "valor",

            width=120,

            anchor="e"

        )


# =============================================================================
# PARTE 186
# EXIBIR TREEVIEW
# =============================================================================

        self.tree.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 187
# SCROLLBAR DO HISTÓRICO
# =============================================================================

        self.scroll_historico = ttk.Scrollbar(

            self.frame_historico,

            orient="vertical",

            command=self.tree.yview

        )

        self.tree.configure(

            yscrollcommand=self.scroll_historico.set

        )

        self.scroll_historico.pack(

            side="right",

            fill="y"

        )

# =============================================================================
# =============================================================================
# PARTE 197
# TOTAL DE REGISTROS
# =============================================================================


# PARTE 190
# CARREGA O HISTÓRICO
# =============================================================================

# =============================================================================
# PARTE 197
# TOTAL DE REGISTROS
# =============================================================================

        self.label_total = ttk.Label(

            self.frame_historico,

            text="Total: 0 registros"

        )

        self.label_total.pack(

            anchor="e",

            pady=5

        )

        self.atualizar_historico()


# =============================================================================
# PARTE 199
# DUPLO CLIQUE
# =============================================================================

        self.tree.bind(

            "<Double-1>",

            self.abrir_registro

        )

# =============================================================================
# PARTE 201
# MENU CONTEXTUAL
# =============================================================================

        self.menu_tree = tk.Menu(

            self.tree,

            tearoff=False

        )



# =============================================================================
# PARTE 209
# MENU EDITAR
# =============================================================================

        self.menu_tree.add_command(

            label="Editar",

            command=self.editar_registro

        )



# =============================================================================
# PARTE 210
# MENU EXCLUIR
# =============================================================================

        self.menu_tree.add_command(

            label="Excluir",

            command=self.excluir_registro

        )



# =============================================================================
# PARTE 204
# BOTÃO DIREITO
# =============================================================================

        self.tree.bind(

            "<Button-3>",

            self.menu_contexto

        )


# =============================================================================
# PARTE 222
# BOTÃO EXPORTAR EXCEL
# =============================================================================

        self.botao_excel = ttk.Button(

            self.frame_historico,

            text="Exportar Excel",

            command=self.exportar_excel

        )

        self.botao_excel.pack(

            side="left",

            padx=5,

            pady=5

        )





# =============================================================================
# PARTE 254
# ABA ESTATÍSTICAS
# =============================================================================

        self.aba_estatisticas = ttk.Frame(

            self.notebook

        )

        self.notebook.add(

            self.aba_estatisticas,

            text="Total de Abastecimentos"

        )


# =============================================================================
# PARTE 255
# FRAME ESTATÍSTICAS
# =============================================================================

        self.frame_estatisticas = ttk.Frame(

            self.aba_estatisticas,

            padding=10

        )

        self.frame_estatisticas.pack(

            fill="both",

            expand=True

        )


# =============================================================================
# PARTE 256
# LABEL ESTATÍSTICAS
# =============================================================================

        self.label_estatisticas = ttk.Label(

            self.frame_estatisticas,

            text="Resumo Geral",

            font=(

                "Segoe UI",

                12,

                "bold"

            )

        )

        self.label_estatisticas.pack(

            pady=10

        )



# =============================================================================
# PARTE 257
# TEXTO ESTATÍSTICAS
# =============================================================================

        self.frame_texto_estatisticas = ttk.Frame(
            self.frame_estatisticas
        )
        self.frame_texto_estatisticas.pack(
            fill="both",
            expand=True,
            padx=5,
            pady=5
        )

        self.texto_estatisticas = tk.Text(
            self.frame_texto_estatisticas,
            width=90,
            height=10,
            wrap="none"
        )

        self.scroll_estatisticas = ttk.Scrollbar(
            self.frame_texto_estatisticas,
            orient="vertical",
            command=self.texto_estatisticas.yview
        )

        self.texto_estatisticas.configure(
            yscrollcommand=self.scroll_estatisticas.set
        )

        self.texto_estatisticas.pack(
            side="left",
            fill="both",
            expand=True
        )

        self.scroll_estatisticas.pack(
            side="right",
            fill="y"
        )



# =============================================================================
# PARTE 258
# BOTÃO ATUALIZAR
# =============================================================================

        self.botao_atualizar_estatisticas = ttk.Button(

            self.frame_estatisticas,

            text="Atualizar",

            command=self.atualizar_estatisticas

        )

        self.botao_atualizar_estatisticas.pack(

            pady=10

        )


# =============================================================================
# PARTE 260
# SEPARADOR
# =============================================================================

        self.separador_estatisticas = ttk.Separator(
            self.frame_estatisticas,
            orient="horizontal"
        )
        self.separador_estatisticas.pack_forget()


# =============================================================================
# PARTE 261
# TOTAL GASTO
# =============================================================================

        self.lbl_total_gasto = ttk.Label(

            self.frame_estatisticas,

            text="Total Gasto: R$ 0,00",

            font=(

                "Segoe UI",

                10,

                "bold"

            )

        )

        self.lbl_total_gasto.pack_forget()


# =============================================================================
# PARTE 262
# TOTAL M3
# =============================================================================

        self.lbl_total_m3 = ttk.Label(

            self.frame_estatisticas,

            text="Total Abastecido: 0,00 m³"

        )

        self.lbl_total_m3.pack_forget()


# =============================================================================
# PARTE 263
# MÉDIA PREÇO
# =============================================================================

        self.lbl_media_preco = ttk.Label(

            self.frame_estatisticas,

            text="Preço Médio: R$ 0,00"

        )

        self.lbl_media_preco.pack_forget()


# =============================================================================
# PARTE 264
# MÉDIA VOLUME
# =============================================================================

        self.lbl_media_volume = ttk.Label(

            self.frame_estatisticas,

            text="Volume Médio: 0,00 m³"

        )

        self.lbl_media_volume.pack_forget()


# =============================================================================
# PARTE 265
# MELHOR ABASTECIMENTO
# =============================================================================

        self.lbl_melhor = ttk.Label(

            self.frame_estatisticas,

            text="Melhor abastecimento:"

        )

        self.lbl_melhor.pack_forget()


# =============================================================================
# PARTE 266
# PIOR ABASTECIMENTO
# =============================================================================

        self.lbl_pior = ttk.Label(

            self.frame_estatisticas,

            text="Pior abastecimento:"

        )

        self.lbl_pior.pack_forget()



# =============================================================================
# PARTE 267
# MAIOR ABASTECIMENTO
# =============================================================================

        self.lbl_maior = ttk.Label(

            self.frame_estatisticas,

            text="Maior abastecimento:"

        )

        self.lbl_maior.pack(

            anchor="w",

            padx=10

        )


# =============================================================================
# PARTE 268
# MENOR ABASTECIMENTO
# =============================================================================

        self.lbl_menor = ttk.Label(

            self.frame_estatisticas,

            text="Menor abastecimento:"

        )

        self.lbl_menor.pack_forget()


# =============================================================================
# PARTE 269
# POSTOS
# =============================================================================

        self.lbl_postos = ttk.Label(

            self.frame_estatisticas,

            text="Postos cadastrados: 0"

        )

        self.lbl_postos.pack_forget()


# =============================================================================
# PARTE 270
# CIDADES
# =============================================================================

        self.lbl_cidades = ttk.Label(

            self.frame_estatisticas,

            text="Cidades cadastradas: 0"

        )

        self.lbl_cidades.pack_forget()






# =============================================================================
# PARTE 491
# CARREGAR CONFIGURAÇÕES
# =============================================================================

        self.carregar_configuracoes()


# =============================================================================
# PARTE 492
# INICIALIZAR CONFIGURAÇÕES
# =============================================================================

        self.inicializar_configuracoes()
        self.atualizar_configuracoes_tela()


# =============================================================================
# PARTE 493
# VERIFICAR CONFIGURAÇÕES
# =============================================================================

        if not self.configuracoes:

            self.criar_configuracao_padrao()


# =============================================================================
# PARTE 494
# ATUALIZAR STATUS
# =============================================================================

        self.status.set(

            "Sistema inicializado."

        )

# =============================================================================
# PARTE 495
# FIM DA INICIALIZAÇÃO
# =============================================================================

        self.janela.update_idletasks()

# =============================================================================
# PARTE 496
# VERIFICA BACKUP AUTOMÁTICO
# =============================================================================

        self.backup_automatico = self.ler_configuracao(

            "backup_automatico",

            True

        )


# =============================================================================
# PARTE 497
# CAMINHO BACKUP
# =============================================================================

        self.pasta_backup = self.ler_configuracao(

            "pasta_backup",

            "Backups"

        )


# =============================================================================
# PARTE 498
# CRIA PASTA BACKUP
# =============================================================================

        if not os.path.exists(

            self.pasta_backup

        ):

            os.makedirs(

                self.pasta_backup

            )


# =============================================================================
# PARTE 499
# TEMA ATUAL
# =============================================================================

        self.tema_atual = self.ler_configuracao(

            "tema",

            "claro"

        )



# =============================================================================
# PARTE 500
# IDIOMA
# =============================================================================

        self.idioma = self.ler_configuracao(

            "idioma",

            "pt-BR"

        )



# =============================================================================
# PARTE 501
# RECARREGAR CONFIGURAÇÕES
# =============================================================================

        self.carregar_configuracoes()



# =============================================================================
# PARTE 502
# ATUALIZAR INTERFACE
# =============================================================================

        self.atualizar_estatisticas()

        self.atualizar_historico()


# =============================================================================
# PARTE 503
# STATUS FINAL
# =============================================================================

        self.status.set(

            "Sistema atualizado."

        )


# =============================================================================
# PARTE 504
# FIM DO MÉTODO
# =============================================================================

        return

# =============================================================================
# PARTE 505
# ATUALIZAR INTERFACE
# =============================================================================

    def selecionar_arquivo_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar planilha Excel",
            filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if not arquivo:
            return
        self.arquivo_excel = arquivo
        self.entry_arquivo_excel.delete(0, tk.END)
        self.entry_arquivo_excel.insert(0, arquivo)
        self.atualizar_excel()

    def atualizar_grafico(self):
        """Atualiza gráficos de abastecimento usando os índices reais do SQLite."""
        self.canvas_grafico.delete("all")
        registros = self.banco.listar_abastecimentos()
        largura = max(self.canvas_grafico.winfo_width(), 820)
        altura = max(self.canvas_grafico.winfo_height(), 500)

        if not registros:
            self.canvas_grafico.create_text(
                largura / 2, altura / 2,
                text="Nenhum abastecimento cadastrado.",
                font=("Arial", 14, "bold")
            )
            return

        tipo = self.tipo_grafico.get() if hasattr(self, "tipo_grafico") else "Gasto por posto"
        margem_esq, margem_dir, margem_top, margem_base = 90, 35, 55, 75
        x0, y0 = margem_esq, altura - margem_base
        x1, y1 = largura - margem_dir, margem_top

        def numero(v, casas=2):
            return formatar_numero_br(float(v), casas)

        def eixos(titulo, unidade):
            self.canvas_grafico.create_text(
                largura / 2, 22, text=titulo,
                font=("Arial", 15, "bold")
            )
            self.canvas_grafico.create_line(x0, y0, x1, y0, width=2)
            self.canvas_grafico.create_line(x0, y0, x0, y1, width=2)
            self.canvas_grafico.create_text(
                x0, y1 - 18, text=unidade, anchor="w",
                font=("Arial", 9)
            )

        def barras(dados, titulo, unidade, fmt):
            eixos(titulo, unidade)
            if not dados:
                return
            maximo = max(v for _, v in dados) or 1.0
            aw, ah = x1 - x0, y0 - y1
            passo = aw / max(len(dados), 1)

            for i in range(6):
                valor = maximo * i / 5
                y = y0 - ah * i / 5
                self.canvas_grafico.create_line(x0, y, x1, y)
                self.canvas_grafico.create_text(
                    x0 - 8, y, text=fmt(valor), anchor="e",
                    font=("Arial", 8)
                )

            for i, (nome, valor) in enumerate(dados):
                cx = x0 + passo * (i + 0.5)
                bw = min(75, passo * 0.62)
                h = ah * valor / maximo
                self.canvas_grafico.create_rectangle(
                    cx - bw / 2, y0 - h, cx + bw / 2, y0,
                    fill="#5b8ff9", outline="#2f5fb3"
                )
                self.canvas_grafico.create_text(
                    cx, y0 - h - 10, text=fmt(valor),
                    font=("Arial", 9, "bold")
                )
                self.canvas_grafico.create_text(
                    cx, y0 + 12, text=str(nome)[:18],
                    font=("Arial", 8), anchor="n"
                )

        if tipo == "Gasto por posto":
            totais = {}
            for r in registros:
                posto = str(r[2] or "Sem posto")
                # SQLite: índice 7 = valor total em R$. Não multiplicar pela temperatura.
                totais[posto] = totais.get(posto, 0.0) + float(r[7] or 0.0)
            barras(
                sorted(totais.items(), key=lambda x: x[1], reverse=True)[:12],
                "Gasto total por posto", "R$",
                lambda v: "R$ " + numero(v, 0)
            )

        elif tipo == "Volume por posto":
            totais = {}
            for r in registros:
                posto = str(r[2] or "Sem posto")
                # SQLite: índice 5 = volume marcado pela bomba em m³.
                totais[posto] = totais.get(posto, 0.0) + float(r[5] or 0.0)
            barras(
                sorted(totais.items(), key=lambda x: x[1], reverse=True)[:12],
                "Volume abastecido por posto", "m³",
                lambda v: numero(v, 2) + " m³"
            )

        elif tipo == "Abastecimentos por posto":
            contagem = {}
            for r in registros:
                posto = str(r[2] or "Sem posto")
                contagem[posto] = contagem.get(posto, 0) + 1
            barras(
                sorted(contagem.items(), key=lambda x: x[1], reverse=True)[:12],
                "Quantidade de abastecimentos por posto", "registros",
                lambda v: numero(v, 0)
            )

        elif tipo == "Km por m³":
            pontos = []
            anteriores = None
            for r in registros:
                try:
                    odo = float(r[4] or 0.0)
                    volume = float(r[5] or 0.0)
                    if anteriores is not None and odo > anteriores and volume > 0:
                        pontos.append((str(r[1] or ""), (odo - anteriores) / volume))
                    if odo > 0:
                        anteriores = odo
                except (TypeError, ValueError):
                    continue
            pontos = pontos[-20:]
            eixos("Rendimento do veículo por abastecimento", "km por m³")
            if not pontos:
                self.canvas_grafico.create_text(
                    (x0 + x1) / 2, (y0 + y1) / 2,
                    text="São necessários pelo menos dois odômetros válidos.",
                    font=("Arial", 11)
                )
                return
            maximo = max(v for _, v in pontos) or 1.0
            aw, ah = x1 - x0, y0 - y1
            for i in range(6):
                valor = maximo * i / 5
                y = y0 - ah * i / 5
                self.canvas_grafico.create_line(x0, y, x1, y)
                self.canvas_grafico.create_text(x0 - 8, y, text=numero(valor, 1), anchor="e", font=("Arial", 8))
            coords = []
            for i, (data, valor) in enumerate(pontos):
                x = x0 if len(pontos) == 1 else x0 + aw * i / (len(pontos) - 1)
                y = y0 - ah * valor / maximo
                coords.append((x, y))
            for a, b in zip(coords, coords[1:]):
                self.canvas_grafico.create_line(a[0], a[1], b[0], b[1], width=3)
            for i, ((data, valor), (x, y)) in enumerate(zip(pontos, coords)):
                self.canvas_grafico.create_oval(x-5, y-5, x+5, y+5, fill="#d95f02", outline="#8c3d00")
                if i == 0 or i == len(coords)-1 or len(coords) <= 8:
                    self.canvas_grafico.create_text(x, y + 12, text=data[:10], anchor="n", font=("Arial", 8))
                    self.canvas_grafico.create_text(x, y - 12, text=numero(valor, 2), anchor="s", font=("Arial", 8, "bold"))

        elif tipo == "Bomba × teórico":
            eixos("Volume marcado pela bomba × volume científico", "m³")
            aw, ah = x1 - x0, y0 - y1
            pontos = []
            for i, r in enumerate(registros[-25:]):
                try:
                    bomba = float(r[5] or 0.0)
                    teorico = float(r[18] or 0.0) if len(r) > 18 else 0.0
                    if teorico > 0:
                        pontos.append((i + 1, bomba, teorico))
                except (TypeError, ValueError):
                    continue
            if not pontos:
                self.canvas_grafico.create_text((x0+x1)/2, (y0+y1)/2, text="Nenhum volume teórico disponível.", font=("Arial", 11))
                return
            maximo = max(max(p[1], p[2]) for p in pontos) or 1.0
            for i in range(6):
                valor = maximo * i / 5
                y = y0 - ah * i / 5
                self.canvas_grafico.create_line(x0, y, x1, y)
                self.canvas_grafico.create_text(x0-8, y, text=numero(valor, 1), anchor="e", font=("Arial", 8))
            passo = aw / max(len(pontos), 1)
            for idx, bomba, teorico in pontos:
                cx = x0 + passo * (idx - 0.5)
                hb = ah * bomba / maximo
                ht = ah * teorico / maximo
                bw = min(30, passo * .30)
                self.canvas_grafico.create_rectangle(cx-bw-2, y0-hb, cx-2, y0, fill="#5b8ff9")
                self.canvas_grafico.create_rectangle(cx+2, y0-ht, cx+bw+2, y0, fill="#59a14f")
                self.canvas_grafico.create_text(cx, y0+12, text=str(idx), anchor="n", font=("Arial", 8))
            self.canvas_grafico.create_text(x1-150, y1+5, text="Azul = bomba | Verde = científico", anchor="w", font=("Arial", 9))

        else:  # Evolução do volume
            pontos = []
            for r in registros:
                try:
                    pontos.append((str(r[1] or ""), float(r[5] or 0.0)))
                except (TypeError, ValueError):
                    continue
            pontos = pontos[-30:]
            eixos("Evolução dos últimos abastecimentos", "m³")
            if not pontos:
                return
            maximo = max(v for _, v in pontos) or 1.0
            aw, ah = x1 - x0, y0 - y1
            for i in range(6):
                valor = maximo * i / 5
                y = y0 - ah * i / 5
                self.canvas_grafico.create_line(x0, y, x1, y)
                self.canvas_grafico.create_text(x0 - 8, y, text=numero(valor, 1), anchor="e", font=("Arial", 8))
            coords = []
            for i, (data, valor) in enumerate(pontos):
                x = x0 if len(pontos) == 1 else x0 + aw * i / (len(pontos) - 1)
                y = y0 - ah * valor / maximo
                coords.append((x, y))
            for a, b in zip(coords, coords[1:]):
                self.canvas_grafico.create_line(a[0], a[1], b[0], b[1], width=3)
            for i, ((data, valor), (x, y)) in enumerate(zip(pontos, coords)):
                self.canvas_grafico.create_oval(x-4, y-4, x+4, y+4, fill="#d95f02", outline="#8c3d00")
                if i == 0 or i == len(coords)-1 or len(coords) <= 8:
                    self.canvas_grafico.create_text(x, y+12, text=data[:10], anchor="n", font=("Arial", 8))
                    self.canvas_grafico.create_text(x, y-12, text=numero(valor, 2), anchor="s", font=("Arial", 8, "bold"))

    def aplicar_idioma(self):
        """Atualiza nomes das abas e título principal conforme o idioma selecionado."""
        idioma=self.config_idioma.get() if hasattr(self,"config_idioma") else "pt-BR"
        nomes=IDIOMA_TABS.get(idioma,IDIOMA_TABS["pt-BR"])
        try:
            for idx,nome in enumerate(nomes):self.notebook.tab(idx,text=nome)
            titulos={"pt-BR":"Sistema de Cálculo de GNV","English":"CNG Calculation System","Español":"Sistema de Cálculo de GNV","Français":"Système de calcul GNV","Italiano":"Sistema di calcolo GNV","Deutsch":"GNV-Berechnungssystem"}
            self.janela.title(titulos.get(idioma,titulos["pt-BR"]))
        except Exception:pass

    def salvar_configuracoes_tela(self):
        self.definir_configuracao("tema", self.config_tema.get())
        self.definir_configuracao("idioma", self.config_idioma.get())
        self.definir_configuracao("backup_automatico", bool(self.config_backup.get()))
        self.aplicar_idioma()
        messagebox.showinfo("Configurações", "Configurações salvas com sucesso.")

    def atualizar_configuracoes_tela(self):
        self.config_tema.set(self.ler_configuracao("tema", "claro"))
        self.config_idioma.set(self.ler_configuracao("idioma", "pt-BR"))
        self.config_backup.set(bool(self.ler_configuracao("backup_automatico", True)))
        self.aplicar_idioma()

    def atualizar_interface(self):


# =============================================================================
# PARTE 506
# ATUALIZAR HISTÓRICO
# =============================================================================

        self.atualizar_historico()

# =============================================================================
# PARTE 507
# ATUALIZAR ESTATÍSTICAS
# =============================================================================

        self.atualizar_estatisticas()


# =============================================================================
# PARTE 508
# ATUALIZAR SQLITE
# =============================================================================

        if hasattr(

            self,

            "tree_sqlite"

        ):

            self.atualizar_sqlite()


# =============================================================================
# PARTE 509
# ATUALIZAR EXCEL
# =============================================================================

        if hasattr(

            self,

            "arquivo_excel"

        ):

            self.atualizar_excel()


# =============================================================================
# PARTE 510
# STATUS
# =============================================================================

        self.status.set(

            "Interface atualizada."

        )


# =============================================================================
# PARTE 511
# ATUALIZAR ABA CÁLCULOS
# =============================================================================

        self.janela.update_idletasks()


# =============================================================================
# PARTE 512
# ATUALIZAR JANELA
# =============================================================================

        self.update()


# =============================================================================
# PARTE 513
# LIMPAR STATUS
# =============================================================================

        self.after(

            3000,

            lambda: self.status.set("")

        )


# =============================================================================
# PARTE 514
# FIM DA ATUALIZAÇÃO
# =============================================================================

        return




# =============================================================================
# PARTE 515
# ATUALIZAR SQLITE
# =============================================================================

    def atualizar_sqlite(self):


# =============================================================================
# PARTE 516
# LIMPAR TREEVIEW SQLITE
# =============================================================================

        for item in self.tree_sqlite.get_children():

            self.tree_sqlite.delete(

                item

            )

# =============================================================================
# PARTE 517
# LER REGISTROS SQLITE
# =============================================================================

        registros = self.banco.listar_abastecimentos()



# =============================================================================
# PARTE 518
# INSERIR REGISTROS
# =============================================================================

        for registro in registros:

            # Índices 17 e 18 são os novos campos persistidos.
            # Para bancos antigos, calcula-se o modelo científico como fallback.
            try:
                metragem_anp = float(registro[17] or 0) if len(registro) > 17 else 0.0
            except (ValueError, TypeError):
                metragem_anp = 0.0

            try:
                metragem_cientifica = float(registro[18] or 0) if len(registro) > 18 else 0.0
            except (ValueError, TypeError):
                metragem_cientifica = 0.0

            if metragem_cientifica <= 0 and float(registro[12] or 0) > 0:
                try:
                    fator_z = converter_numero(self.entry_fator_z.get()) if hasattr(self, "entry_fator_z") else 0.92
                    massa_molar = converter_numero(self.entry_massa_molar.get()) if hasattr(self, "entry_massa_molar") else 0.01604
                    metragem_cientifica = calcular_comparacao_abastecimento(
                        float(registro[12] or 0), float(registro[13] or 0),
                        float(registro[14] or 0), float(registro[8] or 20),
                        float(registro[10] or 0), fator_z, massa_molar
                    )["volume_teorico_m3"]
                except (ValueError, TypeError, IndexError, ZeroDivisionError):
                    metragem_cientifica = 0.0

            if metragem_anp <= 0 and float(registro[12] or 0) > 0:
                try:
                    metragem_anp = calcular_volume_anp_referencia(
                        float(registro[12] or 0), float(registro[13] or 0),
                        float(registro[14] or 0), float(registro[8] or 20),
                        float(registro[10] or 0)
                    )
                except (ValueError, TypeError, IndexError, ZeroDivisionError):
                    metragem_anp = 0.0

            valores_sqlite = (
                registro[0], registro[1], registro[2], registro[3], registro[4],
                registro[5], registro[6], registro[7], registro[8], registro[9],
                registro[10], registro[12], registro[13], registro[14], registro[15],
                (float(registro[16] or 0) if len(registro) > 16 else 0.0),
                metragem_anp, metragem_cientifica
            )

            self.tree_sqlite.insert(
                "",
                tk.END,
                values=valores_sqlite
            )

# =============================================================================
# PARTE 519
# STATUS SQLITE
# =============================================================================

        self.status.set(

            "Tabela SQLite atualizada."

        )


# =============================================================================
# PARTE 520
# FIM SQLITE
# =============================================================================

        return


# =============================================================================
# PARTE 521
# ATUALIZAR EXCEL
# =============================================================================

    def atualizar_excel(self):


# =============================================================================
# PARTE 522
# VERIFICAR ARQUIVO
# =============================================================================

        if not hasattr(

            self,

            "arquivo_excel"

        ):

            return


# =============================================================================
# PARTE 523
# EXISTE O ARQUIVO
# =============================================================================

        if not os.path.exists(

            self.arquivo_excel

        ):

            return


# =============================================================================
# PARTE 524
# LER EXCEL
# =============================================================================

        self.df_excel = pd.read_excel(

            self.arquivo_excel

        )


# =============================================================================
# PARTE 525
# TOTAL DE REGISTROS
# =============================================================================

        self.total_excel = len(

            self.df_excel

        )

# =============================================================================
# PARTE 526
# STATUS EXCEL
# =============================================================================

        self.status.set(

            f"Excel atualizado ({self.total_excel} registros)."

        )


# =============================================================================
# PARTE 527
# ATUALIZAR TUDO
# =============================================================================

    def atualizar_tudo(self):


# =============================================================================
# PARTE 528
# ATUALIZAR HISTÓRICO
# =============================================================================

        self.atualizar_historico()


# =============================================================================
# PARTE 529
# ATUALIZAR ESTATÍSTICAS
# =============================================================================

        self.atualizar_estatisticas()



# =============================================================================
# PARTE 530
# ATUALIZAR SQLITE
# =============================================================================

        if hasattr(

            self,

            "tree_sqlite"

        ):

            self.atualizar_sqlite()


# =============================================================================
# PARTE 531
# ATUALIZAR EXCEL
# =============================================================================

        if hasattr(

            self,

            "arquivo_excel"

        ):

            self.atualizar_excel()



# =============================================================================
# PARTE 532
# FINALIZAR ATUALIZAÇÃO
# =============================================================================

        self.janela.update_idletasks()

        self.status.set(

            "Atualização completa."

        )

# =============================================================================
# PARTE 533
# LIMPAR CAMPOS
# =============================================================================

    def limpar_campos(self):


# =============================================================================
# PARTE 534
# LIMPAR ENTRADAS
# =============================================================================

        for widget in self.frame_calculos.winfo_children():

            if isinstance(

                widget,

                ttk.Entry

            ):

                widget.delete(

                    0,

                    tk.END

                )


# =============================================================================
# PARTE 535
# LIMPAR COMBOBOX
# =============================================================================

        for widget in self.frame_calculos.winfo_children():

            if isinstance(

                widget,

                ttk.Combobox

            ):

                widget.set("")


# =============================================================================
# PARTE 536
# LIMPAR TEXT
# =============================================================================

        for widget in self.frame_calculos.winfo_children():

            if isinstance(

                widget,

                tk.Text

            ):

                widget.delete(

                    "1.0",

                    tk.END

                )


# =============================================================================
# PARTE 537
# STATUS
# =============================================================================

        self.status.set(

            "Campos limpos."

        )


# =============================================================================
# PARTE 538
# FIM LIMPAR CAMPOS
# =============================================================================

        return


# =============================================================================
# PARTE 539
# LIMPAR RESULTADOS
# =============================================================================

    def limpar_resultados(self):

# =============================================================================
# PARTE 540
# LIMPAR ÁREA DE RESULTADOS
# =============================================================================

        if hasattr(

            self,

            "texto_resultados"

        ):

            self.texto_resultados.delete(

                "1.0",

                tk.END

            )


# =============================================================================
# PARTE 541
# LIMPAR ESTATÍSTICAS
# =============================================================================

        if hasattr(

            self,

            "label_estatisticas"

        ):

            self.label_estatisticas.configure(

                text=""

            )


# =============================================================================
# PARTE 542
# LIMPAR STATUS
# =============================================================================

        self.status.set("")


# =============================================================================
# PARTE 543
# ATUALIZAR TELA
# =============================================================================

        self.janela.update_idletasks()


# =============================================================================
# PARTE 544
# FIM LIMPAR RESULTADOS
# =============================================================================

        return





# =============================================================================
# PARTE 145
# EXECUTAR CÁLCULO
# =============================================================================

    def executar_calculo(self):

        try:

            volume = converter_numero(self.entry_volume.get())
            quantidade_float = converter_numero(self.entry_quantidade.get())
            quantidade = int(quantidade_float)
            if quantidade != quantidade_float:
                raise ValueError("A quantidade de cilindros deve ser inteira.")
            pressao = converter_numero(self.entry_pressao.get())
            temperatura = converter_numero(self.entry_temperatura.get())
            altitude = converter_numero(self.entry_altitude.get())
            fator_z = converter_numero(self.entry_fator_z.get())
            massa_molar = converter_numero(self.entry_massa_molar.get())
            densidade_informada = converter_numero(self.entry_densidade_informada.get())
            if densidade_informada <= 0:
                raise ValueError("Informe a massa específica do GNV em kg/m³.")
            if temperatura <= -273.15:
                raise ValueError("A temperatura deve ser maior que -273,15 °C.")

        except ValueError:

            self.texto_resultados.delete(
                "1.0",
                tk.END
            )

            self.texto_resultados.insert(
                tk.END,
                "Erro: verifique os valores informados."
            )

            return

        volume_total = volume * quantidade

        resultado = calcular_quantidade_gnv(
            volume_total,
            pressao,
            temperatura,
            altitude,
            fator_z,
            massa_molar,
            densidade_informada
        )

        self.texto_resultados.delete(
            "1.0",
            tk.END
        )

        self.texto_resultados.insert(
            tk.END,
            "=" * 70 + "\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "RELATÓRIO DOS CÁLCULOS DE GNV\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "=" * 70 + "\n\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Capacidade física total....: {formatar_numero_br(volume_total, 2)} L\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Quantidade de Cilindros...: {quantidade}\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Pressão...................: {pressao:.2f} bar\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Temperatura...............: {temperatura:.2f} °C\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Altitude..................: {altitude:.2f} m\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Fator Z...................: {fator_z:.4f}\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Massa Molar...............: {formatar_numero_br(massa_molar, 5)} kg/mol\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Massa específica informada: {formatar_numero_br(densidade_informada, 3)} kg/m³\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Volume TOTAL equivalente a 1,01325 bar na MESMA T informada: {formatar_numero_br(resultado['volume_equivalente_m3_temperatura_informada'], 3)} m³\n"
            "(quantidade TOTAL recalculada para o estado informado)\n"
            "ATENÇÃO: este número pode permanecer constante quando somente T é\n"
            "alterada, porque n também é recalculado pela mesma equação PV=ZnRT.\n"
        )

        self.texto_resultados.insert(
            tk.END,
            f"Volume TOTAL equivalente científico a 20 °C: {formatar_numero_br(resultado['volume_equivalente_m3_20c'], 3)} m³\n"
            "(os mesmos mols calculados no estado informado, convertidos para 20 °C)\n"
            f"Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C: {formatar_numero_br(resultado['volume_anp_ideal_m3_20c'], 3)} m³\n"
            "(somente a variação entre pressão inicial e final; não é o total armazenado)\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "Observação: o volume de referência não é o volume físico do cilindro.\n"
            "É o volume que a mesma quantidade de matéria ocuparia a 20 °C e\n"
            "1,01325 bar, usando Z de referência igual a 1.\n"
            "IMPORTANTE: esta aba calcula a quantidade TOTAL presente no cilindro.\n"
            "Na aba Abastecimentos calcula-se somente o GÁS ADICIONADO, isto é,\n"
            "a diferença entre a pressão final e a pressão inicial.\n"
        )

        self.texto_resultados.insert(
            tk.END,
            "\n"
        )

        if massa_molar > 0.030:
            self.texto_resultados.insert(
                tk.END,
                "ATENÇÃO: massa molar informada está muito alta para GNV\n"
                "e pode produzir uma massa calculada fisicamente incompatível.\n"
                "Verifique se o valor deveria ser, por exemplo, 0,01604 kg/mol.\n\n"
            )

        self.texto_resultados.insert(
            tk.END,
            "-" * 70 + "\n"
        )

        resultados_formatados = [

            ("Capacidade física do cilindro", resultado["volume_cilindro_l"], "L", 2),

            ("Capacidade física do cilindro", resultado["volume_cilindro_m3"], "m³", 6),

            ("Temperatura", resultado["temperatura_c"], "°C", 2),

            ("Temperatura absoluta", resultado["temperatura_k"], "K", 2),

            ("Pressão informada", resultado["pressao_bar"], "bar", 2),

            ("Pressão atmosférica", resultado["pressao_atm"], "bar", 5),

            ("Pressão absoluta", resultado["pressao_absoluta"], "bar", 5),

            ("Massa de GNV", resultado["massa"], "kg", 6),

            ("Densidade calculada no cilindro", resultado["densidade"], "kg/m³", 3),

            ("Massa específica de referência", densidade_informada, "kg/m³", 3),

            ("Massa estimada pela densidade de referência", resultado["massa_referencia_informada_kg"], "kg", 6),

            ("Diferença de massa entre modelos", resultado["diferenca_massa_referencia_kg"], "kg", 6),

            ("Quantidade de matéria", resultado["mols"], "mol", 6),

            ("Volume específico", resultado["volume_especifico"], "m³/kg", 9),

            ("Volume físico ocupado pelo modelo", resultado["volume_real"], "m³", 6),

            ("Volume total equivalente @ 1,01325 bar na mesma T", resultado["volume_equivalente_m3_temperatura_informada"], "m³", 3),

            ("Volume total equivalente a 20 °C", resultado["volume_equivalente_m3_20c"], "m³", 3),

        ]

        for nome, valor, unidade, casas in resultados_formatados:

            self.texto_resultados.insert(
                tk.END,
                f"{nome:<32}: {formatar_numero_br(valor, casas)} {unidade}\n"
            )

        self.texto_resultados.insert(
            tk.END,
            "\n" + "=" * 70 + "\n"
        )





# =============================================================================
# PARTE 258
# ATUALIZAR ESTATÍSTICAS
# =============================================================================

    def atualizar_estatisticas(self):
        """Atualiza a aba Total de Abastecimentos sem duplicar conteúdo."""
        registros = self.banco.listar_abastecimentos()
        self.texto_estatisticas.configure(state="normal")
        self.texto_estatisticas.delete("1.0", "end")

        total = len(registros)
        self.label_estatisticas.configure(text=f"Total de abastecimentos: {total}")

        volumes = [float(r[5] or 0.0) for r in registros]
        valores = [float(r[7] or 0.0) for r in registros]
        precos = [float(r[6] or 0.0) for r in registros]

        total_volume = sum(volumes)
        total_gasto = sum(valores)
        media_volume = total_volume / total if total else 0.0
        media_preco = sum(precos) / total if total else 0.0
        maior_volume = max(volumes, default=0.0)
        menor_volume = min(volumes, default=0.0)
        maior_valor = max(valores, default=0.0)
        menor_valor = min(valores, default=0.0)

        postos = {str(r[2] or "-") for r in registros}
        cidades = {str(r[3] or "-") for r in registros}

        ranking_postos = {}
        ranking_cidades = {}
        gasto_postos = {}
        gasto_cidades = {}
        volume_postos = {}
        for r in registros:
            posto = str(r[2] or "-")
            cidade = str(r[3] or "-")
            valor = float(r[7] or 0.0)
            volume = float(r[5] or 0.0)
            ranking_postos[posto] = ranking_postos.get(posto, 0) + 1
            ranking_cidades[cidade] = ranking_cidades.get(cidade, 0) + 1
            gasto_postos[posto] = gasto_postos.get(posto, 0.0) + valor
            gasto_cidades[cidade] = gasto_cidades.get(cidade, 0.0) + valor
            volume_postos[posto] = volume_postos.get(posto, 0.0) + volume

        posto_favorito = max(ranking_postos, key=ranking_postos.get) if ranking_postos else "-"
        cidade_favorita = max(ranking_cidades, key=ranking_cidades.get) if ranking_cidades else "-"
        posto_maior_gasto = max(gasto_postos, key=gasto_postos.get) if gasto_postos else "-"
        cidade_maior_gasto = max(gasto_cidades, key=gasto_cidades.get) if gasto_cidades else "-"
        posto_maior_volume = max(volume_postos, key=volume_postos.get) if volume_postos else "-"

        # Rendimento: cada abastecimento é comparado ao odômetro anterior válido.
        km_m3 = []
        odometro_anterior = None
        for r in registros:
            try:
                odo = float(r[4] or 0.0)
                volume = float(r[5] or 0.0)
            except (TypeError, ValueError):
                continue
            if odometro_anterior is not None and odo > odometro_anterior and volume > 0:
                km_m3.append((odo - odometro_anterior) / volume)
            if odo > 0:
                odometro_anterior = odo

        media_km_m3 = sum(km_m3) / len(km_m3) if km_m3 else 0.0
        melhor_km_m3 = max(km_m3, default=0.0)
        pior_km_m3 = min(km_m3, default=0.0)
        volume_medio_posto = total_volume / len(volume_postos) if volume_postos else 0.0

        primeira_data = registros[0][1] if registros else "-"
        ultima_data = registros[-1][1] if registros else "-"

        linhas = [
            "TOTAL DE ABASTECIMENTOS",
            "=" * 62,
            "",
            f"Quantidade de abastecimentos : {total}",
            f"Total de GNV abastecido      : {formatar_numero_br(total_volume, 3)} m³",
            f"Gasto total                  : R$ {formatar_numero_br(total_gasto, 2)}",
            f"Preço médio por m³           : R$ {formatar_numero_br(media_preco, 3)}",
            f"Volume médio por abastecimento: {formatar_numero_br(media_volume, 3)} m³",
            f"Maior abastecimento           : {formatar_numero_br(maior_volume, 3)} m³",
            f"Menor abastecimento           : {formatar_numero_br(menor_volume, 3)} m³",
            f"Maior valor pago              : R$ {formatar_numero_br(maior_valor, 2)}",
            f"Menor valor pago              : R$ {formatar_numero_br(menor_valor, 2)}",
            "",
            "RENDIMENTO DO VEÍCULO",
            "=" * 62,
            "km por m³ é calculado entre dois abastecimentos com odômetros válidos.",
            f"Média de rendimento           : {formatar_numero_br(media_km_m3, 2)} km/m³" if km_m3 else "Média de rendimento           : não calculável ainda",
            f"Melhor rendimento             : {formatar_numero_br(melhor_km_m3, 2)} km/m³" if km_m3 else "Melhor rendimento             : -",
            f"Menor rendimento              : {formatar_numero_br(pior_km_m3, 2)} km/m³" if km_m3 else "Menor rendimento              : -",
            "",
            "POSTOS E CIDADES",
            "=" * 62,
            f"Postos cadastrados            : {len(postos)}",
            f"Cidades cadastradas           : {len(cidades)}",
            f"Posto mais utilizado          : {posto_favorito}",
            f"Cidade mais utilizada         : {cidade_favorita}",
            f"Posto com maior gasto         : {posto_maior_gasto}",
            f"Cidade com maior gasto        : {cidade_maior_gasto}",
            f"Posto com maior volume        : {posto_maior_volume}",
            f"Volume médio por posto       : {formatar_numero_br(volume_medio_posto, 3)} m³",
            "",
            "PERÍODO",
            "=" * 62,
            f"Primeiro abastecimento        : {primeira_data}",
            f"Último abastecimento          : {ultima_data}",
            "",
            "OBSERVAÇÃO",
            "O rendimento em km/m³ depende do volume realmente medido pela bomba.",
            "Para comparar com a metragem teórica, use a aba Gráficos de Abastecimento.",
        ]
        self.texto_estatisticas.insert(tk.END, "\n".join(linhas))
        self.texto_estatisticas.see("1.0")

        # Os labels antigos são mantidos para compatibilidade com outras partes do programa.
        self.lbl_total_gasto.configure(text=f"Total Gasto: R$ {formatar_numero_br(total_gasto, 2)}")
        self.lbl_total_m3.configure(text=f"Total Abastecido: {formatar_numero_br(total_volume, 2)} m³")
        self.lbl_media_preco.configure(text=f"Preço Médio: R$ {formatar_numero_br(media_preco, 2)}")
        self.lbl_media_volume.configure(text=f"Volume Médio: {formatar_numero_br(media_volume, 2)} m³")
        self.lbl_maior.configure(text=f"Maior abastecimento: {formatar_numero_br(maior_volume, 2)} m³")
        self.lbl_menor.configure(text=f"Menor abastecimento: {formatar_numero_br(menor_volume, 2)} m³")
        self.lbl_postos.configure(text=f"Postos cadastrados: {len(postos)}")
        self.lbl_cidades.configure(text=f"Cidades cadastradas: {len(cidades)}")
        self.lbl_pior.configure(text=f"Menor valor pago: R$ {formatar_numero_br(menor_valor, 2)}")

        self.texto_estatisticas.configure(state="disabled")
        return

    def novo_abastecimento(self):

        self.entry_data.delete(0, tk.END)

        self.entry_posto.delete(0, tk.END)

        self.entry_cidade.delete(0, tk.END)

        self.entry_odometro.delete(0, tk.END)

        self.entry_capacidade_cilindro.delete(0, tk.END)

        self.entry_volume_abastecido.delete(0, tk.END)

        self.entry_preco_m3.delete(0, tk.END)

        self.entry_temp_abastecimento.delete(0, tk.END)

        self.entry_pressao_inicial.delete(0, tk.END)
        self.entry_pressao_final.delete(0, tk.END)

        self.entry_altitude_abastecimento.delete(0, tk.END)

        self.entry_densidade_informada_abastecimento.delete(0, tk.END)
        self.entry_densidade_informada_abastecimento.insert(0, "0,76")

        self.texto_observacoes.delete(

            "1.0",

            tk.END

        )



# =============================================================================
# PARTE 172
# SALVAR ABASTECIMENTO
# =============================================================================

    def salvar_abastecimento(self):

        try:

            observacoes = self.texto_observacoes.get(
                "1.0",
                tk.END
            ).strip()

            if not self.entry_data.get().strip():

                messagebox.showwarning(
                    "Atenção",
                    "Informe a data do abastecimento."
                )
                self.entry_data.focus()
                return

            if not self.entry_posto.get().strip():

                messagebox.showwarning(
                    "Atenção",
                    "Informe o posto."
                )
                self.entry_posto.focus()
                return

            if not self.entry_odometro.get().strip():

                messagebox.showwarning(
                    "Atenção",
                    "Informe o odômetro."
                )
                self.entry_odometro.focus()
                return

            abastecimento = Abastecimento(

                data=self.entry_data.get(),

                posto=self.entry_posto.get(),

                cidade=self.entry_cidade.get(),

                odometro=converter_numero(self.entry_odometro.get()),

                volume_m3=converter_numero(self.entry_volume_abastecido.get()),

                preco_m3=converter_numero(self.entry_preco_m3.get()),

                temperatura=converter_numero(self.entry_temp_abastecimento.get()),

                pressao=converter_numero(self.entry_pressao_final.get()),

                altitude=converter_numero(self.entry_altitude_abastecimento.get()),

                observacoes=observacoes,

                capacidade_cilindro_l=converter_numero(self.entry_capacidade_cilindro.get()),

                pressao_inicial=converter_numero(self.entry_pressao_inicial.get()),

                pressao_final=converter_numero(self.entry_pressao_final.get()),

                densidade_informada_kg_m3=converter_numero(self.entry_densidade_informada_abastecimento.get()),

                metragem_teorica_m3=0.0

            )

            if abastecimento.capacidade_cilindro_l <= 0:
                raise ValueError("Informe a capacidade do cilindro em litros.")

            comparacao = calcular_comparacao_abastecimento(
                abastecimento.capacidade_cilindro_l,
                abastecimento.pressao_inicial,
                abastecimento.pressao_final,
                abastecimento.temperatura,
                abastecimento.altitude,
                converter_numero(self.entry_fator_z.get()) if hasattr(self, "entry_fator_z") else 0.92,
                converter_numero(self.entry_massa_molar.get()) if hasattr(self, "entry_massa_molar") else 0.01604
            )

            volume_bomba = abastecimento.volume_m3
            volume_teorico = comparacao["volume_teorico_m3"]
            volume_anp = calcular_volume_anp_referencia(
                abastecimento.capacidade_cilindro_l,
                abastecimento.pressao_inicial,
                abastecimento.pressao_final,
                abastecimento.temperatura,
                abastecimento.altitude
            )
            abastecimento.metragem_teorica_m3 = volume_teorico
            abastecimento.metragem_anp_m3 = volume_anp
            abastecimento.metragem_cientifica_m3 = volume_teorico
            diferenca = volume_bomba - volume_teorico
            percentual = (diferenca / volume_teorico * 100.0) if volume_teorico > 0 else 0.0
            eficiencia = (volume_bomba / volume_teorico * 100.0) if volume_teorico > 0 else 0.0

            self.texto_comparacao_abastecimento.configure(state="normal")
            self.texto_comparacao_abastecimento.delete("1.0", "end")
            self.texto_comparacao_abastecimento.insert(
                tk.END,
                "ANÁLISE DO ABASTECIMENTO — COMPARAÇÃO DOS MODELOS\n"
                "==================================================\n\n"
                f"Capacidade do cilindro : {formatar_numero_br(abastecimento.capacidade_cilindro_l, 2)} L\n"
                f"Pressão inicial        : {formatar_numero_br(abastecimento.pressao_inicial, 2)} bar\n"
                f"Pressão final          : {formatar_numero_br(abastecimento.pressao_final, 2)} bar\n"
                f"Temperatura ambiente   : {formatar_numero_br(abastecimento.temperatura, 2)} °C\n"
                f"Massa específica posto : {formatar_numero_br(abastecimento.densidade_informada_kg_m3, 3)} kg/m³\n"
                f"Δ pressão              : {formatar_numero_br(comparacao['delta_pressao_bar'], 2)} bar\n\n"
                f"Volume indicado pela bomba              : {formatar_numero_br(volume_bomba, 3)} m³\n"
                f"Estimativa pela referência ANP (Z=1)    : {formatar_numero_br(calcular_volume_anp_referencia(abastecimento.capacidade_cilindro_l, abastecimento.pressao_inicial, abastecimento.pressao_final, abastecimento.temperatura, abastecimento.altitude), 3)} m³\n"
                f"Modelo científico com gás real (Z={formatar_numero_br(comparacao['fator_z'], 4)}) : {formatar_numero_br(volume_teorico, 3)} m³\n"                f"Observação: ANP/idealizado e modelo científico Z≠1 são modelos diferentes.\n"
                f"Diferença bomba - científico             : {formatar_numero_br(diferenca, 3)} m³\n"
                f"Diferença percentual                     : {formatar_numero_br(percentual, 2)} %\n"
                f"Relação bomba / científico               : {formatar_numero_br(eficiencia, 2)} %\n"
                f"Mols adicionados pelo modelo             : {formatar_numero_br(comparacao['mols_adicionados'], 3)} mol\n"
                f"Massa adicionada pelo modelo             : {formatar_numero_br(comparacao['massa_adicionada_kg'], 4)} kg\n\n"
                "REFERÊNCIAS E HIPÓTESES\n"
                "ANP: condição de referência de 20 °C / 1,033 kgf/cm².\n"
                "O cálculo ANP é uma conversão para essa condição; não reproduz o algoritmo do dispenser.\n"
                "Modelo científico: PV = Z n R T, usando Z informado.\n"
                "Temperatura disponível: AMBIENTE; não é a temperatura real medida do gás.\n"
                "O modelo científico usa a temperatura ambiente como aproximação para o gás.\n\n"
                "ATENÇÃO: esta comparação é um modelo físico. Para uma\n"
                "conclusão metrológica sobre fraude são necessários também\n"
                "dados de temperatura real do gás, composição/Z validado,\n"
                "condições de referência do medidor e evidências do posto.\n"
            )

        except ValueError as erro:

            messagebox.showerror(

                "Erro",

                str(erro)

            )

            return

# =============================================================================
# PARTE 173
# SALVAR NO SQLITE
# =============================================================================

        self.banco.salvar_abastecimento(

            abastecimento

        )

# =============================================================================
# PARTE 173A
# ATUALIZA TODA A INTERFACE
# =============================================================================

        self.atualizar_tudo()

# =============================================================================
# PARTE 174
# MENSAGEM DE SUCESSO
# =============================================================================

        messagebox.showinfo(

            "Sucesso",

            "Abastecimento gravado com sucesso."

        )


# =============================================================================
# PARTE 175
# LIMPAR FORMULÁRIO
# =============================================================================

        self.novo_abastecimento()



# =============================================================================
# PARTE 176
# ATUALIZAR BARRA DE STATUS
# =============================================================================

        self.status.set(

            "Abastecimento salvo e análise concluída."

        )





# =============================================================================
# PARTE 181
# FECHAR BANCO
# =============================================================================

    def fechar(self):

        if hasattr(

            self,

            "banco"

        ):

            self.banco.fechar()


# =============================================================================
# PARTE 188
# ATUALIZAR HISTÓRICO
# =============================================================================

    def atualizar_historico(self):

        for item in self.tree.get_children():

            self.tree.delete(

                item

            )

        registros = self.banco.listar_abastecimentos()

        for registro in registros:

            self.tree.insert(

                "",

                "end",

                values=(

                    registro[1],

                    registro[2],

                    registro[3],

                    registro[5],

                    (
                        float(registro[18] or 0)
                        if len(registro) > 18 and float(registro[18] or 0) > 0
                        else (
                            calcular_comparacao_abastecimento(
                                float(registro[12] or 0),
                                float(registro[13] or 0),
                                float(registro[14] or 0),
                                float(registro[8] or 20),
                                float(registro[10] or 0),
                                converter_numero(self.entry_fator_z.get()) if hasattr(self, "entry_fator_z") else 0.92,
                                converter_numero(self.entry_massa_molar.get()) if hasattr(self, "entry_massa_molar") else 0.01604
                            )["volume_teorico_m3"]
                            if len(registro) > 14 and float(registro[12] or 0) > 0
                            else "-"
                        )
                    ),

                    registro[7]

                )

            )

# =============================================================================
# PARTE 198
# ATUALIZA TOTAL
# =============================================================================

        self.label_total.configure(

            text=f"Total: {len(registros)} registros"

        )



# =============================================================================
# PARTE 221
# STATUS
# =============================================================================

        self.status.set(

            f"{len(registros)} registros carregados."

        )


# =============================================================================
# PARTE 223
# EXPORTAR EXCEL
# =============================================================================

    def exportar_excel(self):

        try:

            print(

                "Exportando para Excel..."

            )






# =============================================================================
# PARTE 224
# BUSCA DADOS
# =============================================================================

            registros = self.banco.listar_abastecimentos()

            print(

                f"{len(registros)} registros encontrados."

            )


# =============================================================================
# PARTE 225
# DATAFRAME
# =============================================================================

            df = pd.DataFrame(

                registros

            )


# =============================================================================
# PARTE 226
# COLUNAS
# =============================================================================

            df.columns = [

                "ID",

                "Data",

                "Posto",

                "Cidade",

                "Odometro",

                "Volume",

                "Preco",

                "Valor",

                "Temperatura",

                "Pressao",

                "Altitude"

            ]


# =============================================================================
# PARTE 227
# ESCOLHER ARQUIVO
# =============================================================================

            arquivo = filedialog.asksaveasfilename(

                title="Salvar planilha Excel",

                defaultextension=".xlsx",

                filetypes=[

                    ("Planilha Excel","*.xlsx")

                ],

                initialfile="Historico_GNV.xlsx"

            )

            if not arquivo:

                return

            self.arquivo_excel = arquivo

            if hasattr(self, "entry_arquivo_excel"):

                self.entry_arquivo_excel.delete(0, tk.END)

                self.entry_arquivo_excel.insert(0, arquivo)


# =============================================================================
# PARTE 228
# EXPORTAR DATAFRAME
# =============================================================================

            df.to_excel(

                arquivo,

                index=False

            )


# =============================================================================
# PARTE 229
# MENSAGEM
# =============================================================================

            messagebox.showinfo(

                "Excel",

                "Planilha exportada com sucesso."

            )


# =============================================================================
# PARTE 230
# STATUS
# =============================================================================

            self.status.set(

                "Planilha Excel exportada."

            )


# =============================================================================
# PARTE 232
# AJUSTA LARGURA DAS COLUNAS
# =============================================================================

            with pd.ExcelWriter(

                arquivo,

                engine="openpyxl"

            ) as writer:

                df.to_excel(

                    writer,

                    index=False,

                    sheet_name="Abastecimentos"

                )

                worksheet = writer.sheets["Abastecimentos"]

                for coluna in worksheet.columns:

                    tamanho = max(

                        len(str(c.value))

                        if c.value is not None

                        else 0

                        for c in coluna

                    )

                    worksheet.column_dimensions[

                        coluna[0].column_letter

                    ].width = tamanho + 3


# =============================================================================
# PARTE 233
# CONGELA CABEÇALHO
# =============================================================================

            worksheet.freeze_panes = "A2"

        #except Exception as e:

         #   messagebox.showerror(
          #      "Erro",
           #     f"Erro ao exportar Excel:\n{e}"
            #)


        
# =============================================================================
# PARTE 234
# FILTROS
# =============================================================================

            worksheet.auto_filter.ref = worksheet.dimensions

# =============================================================================
# PARTE 235
# CABEÇALHO
# =============================================================================

            from openpyxl.styles import Font

            for celula in worksheet[1]:

                celula.font = Font(

                    bold=True

                )


# =============================================================================
# PARTE 236
# FINALIZA EXPORTAÇÃO
# =============================================================================

            print()

            print("=" * 70)

            print("PLANILHA EXPORTADA COM SUCESSO")

            print("=" * 70)

            print(arquivo)

            print()


        except Exception as e:

            messagebox.showerror(

                "Erro",

                f"Erro ao exportar Excel:\n{e}"

            )
            
            
# =============================================================================
# PARTE 259
# ATUALIZAR ESTATÍSTICAS
# =============================================================================

                


# =============================================================================
# PARTE 407
# EXPORTAR PDF
# =============================================================================

    def exportar_pdf(self):

        pdf = RelatorioPDF()

        pdf.add_page()


# =============================================================================
# PARTE 408
# DATA
# =============================================================================

        pdf.titulo(

            "Informações Gerais"

        )

        pdf.linha(

            "Data:",

            datetime.now().strftime(

                "%d/%m/%Y"

            )

        )

        pdf.linha(

            "Hora:",

            datetime.now().strftime(

                "%H:%M:%S"

            )

        )


# =============================================================================
# PARTE 409
# TOTAL REGISTROS
# =============================================================================

        registros = self.banco.listar_abastecimentos()

        pdf.linha(

            "Abastecimentos:",

            len(

                registros

            )

        )


# =============================================================================
# PARTE 410
# ESPAÇO
# =============================================================================

        pdf.ln(

            5

        )


# =============================================================================
# PARTE 411
# TÍTULO HISTÓRICO
# =============================================================================

        pdf.titulo(

            "Histórico de Abastecimentos"

        )

        pdf.ln(

            3

        )


# =============================================================================
# PARTE 412
# LOOP
# =============================================================================

        for registro in registros:



# =============================================================================
# PARTE 413
# DATA
# =============================================================================

            pdf.linha(

                "Data",

                registro[1]

            )



# =============================================================================
# PARTE 414
# POSTO
# =============================================================================

            pdf.linha(

                "Posto",

                registro[2]

            )


# =============================================================================
# PARTE 415
# CIDADE
# =============================================================================

            pdf.linha(

                "Cidade",

                registro[3]

            )

# =============================================================================
# PARTE 416
# HODÔMETRO
# =============================================================================

            pdf.linha(

                "Hodômetro",

                f"{registro[4]} km"

            )


# =============================================================================
# PARTE 417
# VOLUME
# =============================================================================

            pdf.linha(

                "Volume",

                f"{registro[5]} m³"

            )


# =============================================================================
# PARTE 418
# PREÇO DO M³
# =============================================================================

            pdf.linha(

                "Preço do m³",

                f"R$ {registro[6]}"

            )

# =============================================================================
# PARTE 419
# VALOR TOTAL
# =============================================================================

            pdf.linha(

                "Valor Total",

                f"R$ {registro[7]}"

            )


# =============================================================================
# PARTE 420
# PRESSÃO
# =============================================================================

            pdf.linha(

                "Pressão",

                f"{registro[9]} bar"

            )

# =============================================================================
# PARTE 421
# TEMPERATURA
# =============================================================================

            pdf.linha(

                "Temperatura",

                f"{registro[8]} °C"

            )


# =============================================================================
# PARTE 422
# ALTITUDE
# =============================================================================

            pdf.linha(

                "Altitude",

                f"{registro[10]} m"

            )


# =============================================================================
# PARTE 423
# LINHA EM BRANCO
# =============================================================================

            pdf.ln(

                2

            )



# =============================================================================
# PARTE 424
# SEPARADOR
# =============================================================================

            pdf.cell(

                0,

                2,

                "_" * 80,

                ln=True

            )

# =============================================================================
# PARTE 425
# ESPAÇO
# =============================================================================

            pdf.ln(

                3

            )


# =============================================================================
# PARTE 426
# NOME DO ARQUIVO
# =============================================================================

        nome_arquivo = filedialog.asksaveasfilename(

            defaultextension=".pdf",

            filetypes=[

                (

                    "Arquivo PDF",

                    "*.pdf"

                )

            ],

            initialfile="Relatorio_GNV.pdf"

        )


# =============================================================================
# PARTE 427
# CANCELAMENTO
# =============================================================================

        if not nome_arquivo:

            return


# =============================================================================
# PARTE 428
# SALVAR PDF
# =============================================================================

        pdf.output(

            nome_arquivo

        )


# =============================================================================
# PARTE 429
# SUCESSO
# =============================================================================

        messagebox.showinfo(

            "PDF",

            "Relatório PDF gerado com sucesso!"

        )


# =============================================================================
# PARTE 430
# STATUS
# =============================================================================

        self.status.set(

            "Relatório PDF gerado com sucesso."

        )

        return




# =============================================================================
# PARTE 441
# EXPORTAR XML
# =============================================================================

        arquivo_xml = arquivo.replace(

            ".xlsx",

            ".xml"

        )

# =============================================================================
# PARTE 442
# CABEÇALHO XML
# =============================================================================

        with open(

            arquivo_xml,

            "w",

            encoding="utf-8"

        ) as xml:

            xml.write(

                "<?xml version='1.0' encoding='UTF-8'?>\n"

            )

            xml.write(

                "<abastecimentos>\n"

            )

# =============================================================================
# PARTE 443
# REGISTROS XML
# =============================================================================

            for registro in registros:

                xml.write(

                    "    <abastecimento>\n"

                )

# =============================================================================
# PARTE 444
# CAMPOS XML
# =============================================================================

                xml.write(

                    f"        <data>{registro[1]}</data>\n"

                )

                xml.write(

                    f"        <posto>{registro[2]}</posto>\n"

                )

                xml.write(

                    f"        <cidade>{registro[3]}</cidade>\n"

                )

                xml.write(

                    f"        <odometro>{registro[4]}</odometro>\n"

                )

                xml.write(

                    f"        <volume>{registro[5]}</volume>\n"

                )

                xml.write(

                    f"        <preco>{registro[6]}</preco>\n"

                )

                xml.write(

                    f"        <valor>{registro[7]}</valor>\n"

                )

                xml.write(

                    f"        <pressao>{registro[8]}</pressao>\n"

                )

                xml.write(

                    f"        <temperatura>{registro[9]}</temperatura>\n"

                )

                xml.write(

                    f"        <altitude>{registro[10]}</altitude>\n"

                )


# =============================================================================
# PARTE 445
# FECHAR REGISTRO XML
# =============================================================================

                xml.write(

                    "    </abastecimento>\n"

                )

            xml.write(

                "</abastecimentos>\n"

            )


# =============================================================================
# PARTE 446
# EXPORTAR TXT
# =============================================================================

        arquivo_txt = arquivo.replace(

            ".xlsx",

            ".txt"

        )



# =============================================================================
# PARTE 447
# ABRIR TXT
# =============================================================================

        with open(

            arquivo_txt,

            "w",

            encoding="utf-8"

        ) as txt:


# =============================================================================
# PARTE 448
# CABEÇALHO TXT
# =============================================================================

            txt.write(

                "RELATÓRIO DOS ABASTECIMENTOS\n"

            )

            txt.write(

                "=" * 80

            )

            txt.write(

                "\n\n"

            )


# =============================================================================
# PARTE 449
# REGISTROS TXT
# =============================================================================

            for registro in registros:


# =============================================================================
# PARTE 450
# DADOS TXT
# =============================================================================

                txt.write(

                    f"Data.............: {registro[1]}\n"

                )

                txt.write(

                    f"Posto............: {registro[2]}\n"

                )

                txt.write(

                    f"Cidade...........: {registro[3]}\n"

                )

                txt.write(

                    f"Hodômetro........: {registro[4]} km\n"

                )

                txt.write(

                    f"Volume...........: {registro[5]} m³\n"

                )

                txt.write(

                    f"Preço............: R$ {registro[6]}\n"

                )

                txt.write(

                    f"Valor............: R$ {registro[7]}\n"

                )

                txt.write(

                    f"Pressão..........: {registro[9]} bar\n"

                )

                txt.write(

                    f"Temperatura......: {registro[8]} °C\n"

                )

                txt.write(

                    f"Altitude.........: {registro[10]} m\n"

                )

                txt.write(

                    "-" * 80

                )

                txt.write(

                    "\n"

                )


# =============================================================================
# PARTE 451
# LINHA EM BRANCO TXT
# =============================================================================

                txt.write(

                    "\n"

                )


# =============================================================================
# PARTE 452
# TOTAL DE REGISTROS TXT
# =============================================================================

            txt.write(

                "\n"

            )

            txt.write(

                "=" * 80

            )

            txt.write(

                "\n"

            )

            txt.write(

                f"TOTAL DE ABASTECIMENTOS: {len(registros)}"

            )

            txt.write(

                "\n"

            )


# =============================================================================
# PARTE 453
# DATA EXPORTAÇÃO
# =============================================================================

            txt.write(

                f"Data da Exportação: "

                f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

            )

            txt.write(

                "\n"

            )


# =============================================================================
# PARTE 454
# FECHAR TXT
# =============================================================================

            txt.flush()


# =============================================================================
# PARTE 455
# STATUS TXT
# =============================================================================

        self.status.set(

            "Arquivos Excel, CSV, JSON, XML e TXT exportados."

        )


# =============================================================================
# PARTE 456
# BACKUP SQLITE
# =============================================================================

    def backup_sqlite(self):

        origem = self.banco.nome_banco



# =============================================================================
# PARTE 457
# ESCOLHER DESTINO
# =============================================================================

        destino = filedialog.asksaveasfilename(

            defaultextension=".db",

            filetypes=[

                (

                    "Banco SQLite",

                    "*.db"

                )

            ],

            initialfile="Backup_GNV.db"

        )

        if not destino:

            return


# =============================================================================
# PARTE 458
# COPIAR BANCO
# =============================================================================

        import shutil

        shutil.copy2(

            origem,

            destino

        )


# =============================================================================
# PARTE 459
# MENSAGEM
# =============================================================================

        messagebox.showinfo(

            "Backup",

            "Backup realizado com sucesso."

        )


# =============================================================================
# PARTE 460
# STATUS
# =============================================================================

        self.status.set(

            "Backup SQLite concluído."

        )


# =============================================================================
# PARTE 461
# RESTAURAR BACKUP
# =============================================================================

    def restaurar_backup(self):

        arquivo = filedialog.askopenfilename(

            filetypes=[

                (

                    "Banco SQLite",

                    "*.db"

                )

            ]

        )

        if not arquivo:

            return


# =============================================================================
# PARTE 462
# CONFIRMAR RESTAURAÇÃO
# =============================================================================

        resposta = messagebox.askyesno(

            "Restaurar",

            "Deseja substituir o banco de dados atual?"

        )

        if not resposta:

            return


# =============================================================================
# PARTE 463
# FECHAR CONEXÃO
# =============================================================================

        self.banco.conexao.close()


# =============================================================================
# PARTE 464
# COPIAR BACKUP
# =============================================================================

        import shutil

        shutil.copy2(

            arquivo,

            self.banco.nome_banco

        )


# =============================================================================
# PARTE 465
# RECONECTAR BANCO
# =============================================================================

        self.banco = BancoGNV()


# =============================================================================
# PARTE 466
# RECARREGAR HISTÓRICO
# =============================================================================

        self.atualizar_historico()


# =============================================================================
# PARTE 467
# RECARREGAR ESTATÍSTICAS
# =============================================================================

        self.atualizar_estatisticas()


# =============================================================================
# PARTE 468
# MENSAGEM RESTAURAÇÃO
# =============================================================================

        messagebox.showinfo(

            "Restauração",

            "Banco de dados restaurado com sucesso."

        )


# =============================================================================
# PARTE 469
# STATUS RESTAURAÇÃO
# =============================================================================

        self.status.set(

            "Backup restaurado com sucesso."

        )


# =============================================================================
# PARTE 470
# FIM RESTAURAÇÃO
# =============================================================================

        return

# =============================================================================
# PARTE 471
# CARREGAR CONFIGURAÇÕES
# =============================================================================

    def carregar_configuracoes(self):

        self.configuracoes = {}

        if os.path.exists(

            "configuracoes.json"

        ):

            with open(

                "configuracoes.json",

                "r",

                encoding="utf-8"

            ) as arquivo:

                self.configuracoes = json.load(

                    arquivo

                )


# =============================================================================
# PARTE 472
# SALVAR CONFIGURAÇÕES
# =============================================================================

    def salvar_configuracoes(self):

        with open(

            "configuracoes.json",

            "w",

            encoding="utf-8"

        ) as arquivo:

            json.dump(

                self.configuracoes,

                arquivo,

                indent=4,

                ensure_ascii=False

            )


# =============================================================================
# PARTE 473
# DEFINIR CONFIGURAÇÃO
# =============================================================================

    def definir_configuracao(

        self,

        chave,

        valor

    ):

        self.configuracoes[

            chave

        ] = valor


# =============================================================================
# PARTE 474
# GRAVAR CONFIGURAÇÃO
# =============================================================================

        self.salvar_configuracoes()



# =============================================================================
# PARTE 475
# LER CONFIGURAÇÃO
# =============================================================================

    def ler_configuracao(

        self,

        chave,

        padrao=None

    ):

        return self.configuracoes.get(

            chave,

            padrao

        )

# =============================================================================
# PARTE 477
# INICIALIZAR CONFIGURAÇÕES
# =============================================================================

    def inicializar_configuracoes(self):

        self.carregar_configuracoes()


# =============================================================================
# PARTE 478
# VERIFICAR TEMA
# =============================================================================

        tema = self.ler_configuracao(

            "tema",

            "claro"

        )


# =============================================================================
# PARTE 479
# APLICAR TEMA
# =============================================================================

        self.tema_atual = tema


# =============================================================================
# PARTE 480
# STATUS
# =============================================================================

        self.status.set(

            "Configurações carregadas."

        )


# =============================================================================
# PARTE 481
# CONFIGURAÇÃO PADRÃO
# =============================================================================

    def criar_configuracao_padrao(self):

        self.configuracoes = {

            "tema": "claro",

            "backup_automatico": True,

            "idioma": "pt-BR"

        }


# =============================================================================
# PARTE 482
# SALVAR PADRÃO
# =============================================================================

        self.salvar_configuracoes()


# =============================================================================
# PARTE 483
# REINICIAR CONFIGURAÇÕES
# =============================================================================

    def reiniciar_configuracoes(self):

        self.criar_configuracao_padrao()



# =============================================================================
# PARTE 484
# ATUALIZAR STATUS
# =============================================================================

        self.status.set(

            "Configurações restauradas."

        )


# =============================================================================
# PARTE 485
# MENSAGEM
# =============================================================================

        messagebox.showinfo(

            "Configurações",

            "Configurações restauradas com sucesso."

        )


# =============================================================================
# PARTE 486
# EXPORTAR CONFIGURAÇÕES
# =============================================================================

    def exportar_configuracoes(self):

        arquivo = filedialog.asksaveasfilename(

            defaultextension=".json",

            filetypes=[

                (

                    "Arquivo JSON",

                    "*.json"

                )

            ],

            initialfile="Configuracoes_GNV.json"

        )

        if not arquivo:

            return


# =============================================================================
# PARTE 487
# GRAVAR CONFIGURAÇÕES
# =============================================================================

        with open(

            arquivo,

            "w",

            encoding="utf-8"

        ) as arq:

            json.dump(

                self.configuracoes,

                arq,

                indent=4,

                ensure_ascii=False

            )


# =============================================================================
# PARTE 488
# IMPORTAR CONFIGURAÇÕES
# =============================================================================

    def importar_configuracoes(self):

        arquivo = filedialog.askopenfilename(

            filetypes=[

                (

                    "Arquivo JSON",

                    "*.json"

                )

            ]

        )

        if not arquivo:

            return

# =============================================================================
# PARTE 489
# LER CONFIGURAÇÕES
# =============================================================================

        with open(

            arquivo,

            "r",

            encoding="utf-8"

        ) as arq:

            self.configuracoes = json.load(

                arq

            )

        self.salvar_configuracoes()


# =============================================================================
# PARTE 490
# CONFIGURAÇÕES IMPORTADAS
# =============================================================================

        self.status.set(

            "Configurações importadas."

        )

        messagebox.showinfo(

            "Configurações",

            "Configurações importadas com sucesso."

        )








# =============================================================================
# PARTE 237
# EXPORTAR CSV
# =============================================================================

        arquivo_csv = arquivo.replace(

            ".xlsx",

            ".csv"

        )

        df.to_csv(

            arquivo_csv,

            index=False,

            sep=";",

            decimal=",",

            encoding="utf-8-sig"

        )



# =============================================================================
# PARTE 238
# STATUS CSV
# =============================================================================

        print(

            "CSV exportado:",

            arquivo_csv

        )

        self.status.set(

            "Excel e CSV exportados."

        )


# =============================================================================
# PARTE 239
# MENSAGEM COMPLETA
# =============================================================================

        messagebox.showinfo(

            "Exportação",

            "Arquivos Excel e CSV criados com sucesso."

        )






# =============================================================================
# PARTE 231
# TRATAMENTO DE ERROS
# =============================================================================

#        except Exception as erro:
#
#            messagebox.showerror(

#                "Erro",

#                str(erro)

#            )



# =============================================================================
# PARTE 240
# EXPORTAR PDF
# =============================================================================

        arquivo_pdf = arquivo.replace(

            ".xlsx",

            ".pdf"

        )

# =============================================================================
# PARTE 241
# CRIAR PDF
# =============================================================================

        pdf = FPDF(

            orientation="L",

            unit="mm",

            format="A4"

        )

        pdf.add_page()

        pdf.set_font(

            "Arial",

            "B",

            14

        )


# =============================================================================
# PARTE 242
# TÍTULO
# =============================================================================

        pdf.cell(

            0,

            10,

            "RELATORIO DE ABASTECIMENTOS GNV",

            ln=True,

            align="C"

        )

        pdf.ln(5)



# =============================================================================
# PARTE 243
# CABEÇALHO
# =============================================================================

        pdf.set_font(

            "Arial",

            "B",

            8

        )

        colunas = [

            ("Data",25),

            ("Posto",55),

            ("Cidade",40),

            ("Volume",25),

            ("Valor",25)

        ]

        for titulo, largura in colunas:

            pdf.cell(

                largura,

                8,

                titulo,

                border=1,

                align="C"

            )

        pdf.ln()



# =============================================================================
# PARTE 244
# DADOS
# =============================================================================

        pdf.set_font(

            "Arial",

            "",

            8

        )

        for registro in registros:

            pdf.cell(25,8,str(registro[1]),1)

            pdf.cell(55,8,str(registro[2]),1)

            pdf.cell(40,8,str(registro[3]),1)

            pdf.cell(25,8,str(registro[5]),1)

            pdf.cell(25,8,f"R$ {registro[7]:.2f}",1)

            pdf.ln()



# =============================================================================
# PARTE 245
# SALVAR PDF
# =============================================================================

        pdf.output(

            arquivo_pdf

        )

        print(

            "PDF exportado:",

            arquivo_pdf

        )

        self.status.set(

            "Excel, CSV e PDF exportados."

        )





# =============================================================================
# PARTE 247
# MENSAGEM FINAL
# =============================================================================

        messagebox.showinfo(

            "Exportação concluída",

            f"Arquivos gerados:\n\n"

            f"Excel:\n{arquivo}\n\n"

            f"CSV:\n{arquivo_csv}\n\n"

            f"PDF:\n{arquivo_pdf}"

        )



# =============================================================================
# PARTE 248
# LOG
# =============================================================================

        print()

        print("=" * 70)

        print("ARQUIVOS EXPORTADOS")

        print("=" * 70)

        print("Excel :", arquivo)

        print("CSV   :", arquivo_csv)

        print("PDF   :", arquivo_pdf)

        print("=" * 70)

        print()



# =============================================================================
# PARTE 249
# ABRIR PASTA
# =============================================================================

        pasta = os.path.dirname(

            arquivo

        )

        try:

            os.startfile(

                pasta

            )

        except Exception:

            pass



# =============================================================================
# PARTE 251
# TOTAL EXPORTADO
# =============================================================================

        print(

            f"Foram exportados {len(registros)} abastecimentos."

        )


# =============================================================================
# PARTE 252
# DATA E HORA
# =============================================================================

        print(

            datetime.now().strftime(

                "Exportado em %d/%m/%Y às %H:%M:%S"

            )

        )





# =============================================================================
# PARTE 195
# PESQUISAR HISTÓRICO
# =============================================================================

    def pesquisar_historico(self):

        texto = self.entry_pesquisa.get().strip()

        if texto == "":

            self.atualizar_historico()

            return

        for item in self.tree.get_children():

            self.tree.delete(

                item

            )

        registros = self.banco.buscar_por_posto(

            texto

        )

        for registro in registros:

            self.tree.insert(

                "",

                "end",

                values=(

                    registro[1],

                    registro[2],

                    registro[3],

                    registro[5],

                    "-",

                    registro[7]

                )

            )


# =============================================================================
# PARTE 200
# ABRIR REGISTRO
# =============================================================================

    def abrir_registro(

        self,

        event

    ):

        item = self.tree.focus()

        if not item:

            return

        valores = self.tree.item(

            item,

            "values"

        )

        print(

            valores

        )


    def editar_registro(self):

        item = self.tree.focus()

        if not item:
            return

        valores = self.tree.item(item, "values")

        print("Editar:", valores)

        messagebox.showinfo(
            "Editar registro",
            "O registro selecionado foi identificado para edição.\n\n"
            "A edição detalhada será implementada sem alterar os dados automaticamente."
        )


    def excluir_registro(self):

        item = self.tree.focus()

        if not item:
            return

        valores = self.tree.item(item, "values")

        resposta = messagebox.askyesno(
            "Excluir",
            "Deseja realmente excluir este abastecimento?"
        )

        if not resposta:
            return

        registros = self.banco.listar_abastecimentos()
        registro_id = None

        for registro in registros:
            if (
                str(registro[1]) == str(valores[0])
                and str(registro[2]) == str(valores[1])
                and str(registro[3]) == str(valores[2])
            ):
                registro_id = registro[0]
                break

        if registro_id is None:
            messagebox.showerror(
                "Excluir",
                "Não foi possível localizar o registro no banco de dados."
            )
            return

        self.banco.excluir_abastecimento(registro_id)
        self.recarregar_treeview()
        self.atualizar_estatisticas()


    def recarregar_treeview(self):

        self.atualizar_historico()


    def menu_contexto(
        self,
        event
    ):

        item = self.tree.identify_row(
            event.y
        )

        if item:

            self.tree.selection_set(
                item
            )

            self.tree.focus(
                item
            )

            self.menu_tree.post(
                event.x_root,
                event.y_root
            )

# =============================================================================
# INICIALIZAÇÃO DO PROGRAMA
# =============================================================================

if __name__ == "__main__":

    janela = tk.Tk()

    app = InterfaceGNV(

        janela

    )

    janela.mainloop()


# =============================================================================
# COMMIT GIT (PORTUGUÊS)
# =============================================================================
# feat: corrige modelos de referência e documentação física do GNV
#
# - distingue claramente modelo científico com Z informado de ANP/idealizado Z=1
# - corrige os rótulos de volume equivalente para evitar interpretação errada
# - documenta o cancelamento de T no volume equivalente a pressão fixa
# - adiciona referências acadêmicas do MIT OpenCourseWare e Purdue University
# - mantém referências ANP, ISO 12213 e NIST
# - simplifica o cabeçalho da análise de abastecimento
# - preserva a entrada de temperatura ambiente como aproximação no abastecimento
#
# feat: corrige referências físicas e responsividade das abas
#
# - aceita números com ponto ou vírgula decimal
# - remove duplicidade de volumes equivalentes e esclarece a condição de referência
# - reposiciona os botões acima da área de resultados
# - adiciona a aba "Fórmulas e Física"
# - documenta o fator Z e as equações utilizadas
# - usa temperatura ambiente no abastecimento como aproximação da temperatura do gás
# - compara densidade informada e densidade calculada
# - registra massa específica nos abastecimentos
# - adiciona aba ANP com campos próprios e condição de referência oficial documentada
# - separa cálculo de referência ANP de modelo científico com gás real
# - diferencia quantidade TOTAL armazenada de quantidade ADICIONADA no abastecimento
# - elimina repetição do relatório de análise física nas estatísticas
# - grava metragem cúbica teórica no SQLite
# - adiciona metragem cúbica teórica na visualização do SQLite




