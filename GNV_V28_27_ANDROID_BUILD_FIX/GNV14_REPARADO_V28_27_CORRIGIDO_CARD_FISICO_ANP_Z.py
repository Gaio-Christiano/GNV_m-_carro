# V28.18 — CONFIGURAÇÃO MULTIABA, CRÉDITOS E PALETA NEUTRA
# Correções: ANP/Fórmulas/Gráficos, aplicação sem reinício, créditos visíveis e escala branco-cinza-preto.
#
# V28.11 — CORREÇÃO DEFINITIVA DE CAMPOS, LINHAS 1/2, CORES E CONTRASTE
# Esta versão elimina as implementações duplicadas que faziam _save_config()
# procurar color_spinners, inexistente no novo seletor color_controls.
# Também força contraste escuro em campos claros e claro em campos escuros.
# Todas as alterações visuais estão comentadas para estudo futuro.

# =============================================================================
# ARQUIVO.....: GNV14_REPARADO_V28_ANDROID_NOTE_9_PRO_CORRIGIDO_V28_11_CORES_E_LETRAS_DEFINITIVAS.py
# AUTOR.......: Christiano T.Gaio
# OBJETIVO....: Calculadora de GNV
#
# DESCRIÇÃO
# ----------
# Este programa estima a quantidade de GNV em um cilindro usando um modelo
# de gás real simplificado, PV = Z n R T, com Z informado pelo usuário.
#
# OBSERVAÇÃO IMPORTANTE
# ---------------------
# O GNV é um gás REAL.
# Portanto, em pressões elevadas (200 a 220 bar) existe um erro quando usamos
# apenas a Lei dos Gases Ideais.
#
# Limitação importante: Z é fornecido como aproximação. Para maior precisão em
# alta pressão, é necessário calcular Z(P,T,composição) por um modelo validado,
# como AGA8/GERG, e conhecer a temperatura real do gás durante o abastecimento.
#
# Todo o código possui comentários para facilitar o aprendizado.
#
# V28.8 - REVISÃO VISUAL PROFUNDA: contraste automático, seletor arco-íris,
# intensidade/luminosidade, temas claro/escuro coerentes e resultados legíveis.
# =============================================================================






# =============================================================================
# PARTE 01 - IMPORTAÇÃO DAS BIBLIOTECAS
# =============================================================================

# Biblioteca matemática

import math
import csv
import os
import sqlite3
import json
import webbrowser
import urllib.parse
import colorsys

# =============================================================================
# PARTE 253
# IMPORT DATETIME
# =============================================================================

from datetime import datetime

# =============================================================================
# PARTE 246
# IMPORT FPDF
# =============================================================================
from fpdf import FPDF



# =============================================================================
# PARTE 02 - CONSTANTES FÍSICAS
# =============================================================================

# Pressão atmosférica padrão ao nível do mar (bar)
PRESSAO_ATMOSFERICA_PADRAO = 1.01325

# Temperatura padrão (Kelvin)
TEMPERATURA_PADRAO = 273.15

# Constante universal dos gases
#
# Unidade:
#
# J/(mol.K)
#
R = 8.314462618

# Massa molar aproximada do GNV
#
# O gás natural muda de composição conforme a distribuidora.
#
# Foi adotado um valor médio.
#

# -------------------------------------------------------------------------
# Estes valores mudam conforme a composição do gás fornecido pela
# distribuidora.
#
# O usuário poderá alterá-los na execução do programa.
# -------------------------------------------------------------------------

MASSA_MOLAR_GNV = None      # kg/mol

# Densidade média do GNV nas CNTP
#
# Este valor varia mensalmente.
#
DENSIDADE_GNV = None           # kg/m³

# Fator de compressibilidade
#
# Z = 1,000  -> gás ideal
#
# Para GNV normalmente varia aproximadamente entre
#
# 0,82 e 0,95
#
Z = None

# Idiomas disponíveis na interface.
IDIOMAS_DISPONIVEIS = ("pt-BR", "English", "Español", "Français", "Italiano", "Deutsch", "日本語", "中文")
IDIOMA_TABS = {
    # Ordem oficial das 12 abas do sistema.
    "pt-BR": [
        "Cálculos", "Abastecimentos", "ANP", "Aquecimento / Compressão",
        "Histórico de Abastecimentos", "Banco SQLite", "Exportação / Excel",
        "Gráficos de Abastecimento", "Configurações do Sistema",
        "Fórmulas e Física", "Total de Abastecimentos", "Créditos"
    ],
    "English": [
        "Calculations", "Refuelings", "ANP", "Heating / Compression",
        "Refueling History", "SQLite Database", "Export / Excel",
        "Refueling Charts", "System Settings", "Formulas & Physics",
        "Total Refuelings", "Credits"
    ],
    "Español": [
        "Cálculos", "Abastecimientos", "ANP", "Calentamiento / Compresión",
        "Historial de Abastecimientos", "Base SQLite", "Exportación / Excel",
        "Gráficos de Abastecimiento", "Configuración del Sistema",
        "Fórmulas y Física", "Total de Abastecimientos", "Créditos"
    ],
    "Français": [
        "Calculs", "Ravitaillements", "ANP", "Chauffage / Compression",
        "Historique des Ravitaillements", "Base SQLite", "Exportation / Excel",
        "Graphiques de Ravitaillement", "Paramètres du Système",
        "Formules et Physique", "Total des Ravitaillements", "Crédits"
    ],
    "Italiano": [
        "Calcoli", "Rifornimenti", "ANP", "Riscaldamento / Compressione",
        "Storico Rifornimenti", "Database SQLite", "Esportazione / Excel",
        "Grafici dei Rifornimenti", "Impostazioni di Sistema",
        "Formule e Fisica", "Totale Rifornimenti", "Crediti"
    ],
    "Deutsch": [
        "Berechnungen", "Tankvorgänge", "ANP", "Erwärmung / Kompression",
        "Tankvorgangsverlauf", "SQLite-Datenbank", "Export / Excel",
        "Tankdiagramme", "Systemeinstellungen", "Formeln & Physik",
        "Gesamte Tankvorgänge", "Credits"
    ],
    "日本語": [
        "計算", "給油", "ANP", "加熱 / 圧縮", "給油履歴", "SQLiteデータベース",
        "Excelエクスポート", "給油グラフ", "システム設定", "数式と物理", "給油総数", "クレジット"
    ],
    "中文": [
        "计算", "加气记录", "ANP", "加热 / 压缩", "加气历史", "SQLite数据库",
        "Excel导出", "加气图表", "系统设置", "公式与物理", "加气总数", "致谢"
    ],
}


# =============================================================================
# I18N V20 - tradução de interface, relatórios e fórmulas
# =============================================================================
I18N = {
"English": {
"Sistema de Cálculo de GNV":"CNG Calculation System","Cálculos":"Calculations","Abastecimentos":"Refuelings","Histórico de Abastecimentos":"Refueling History","Banco SQLite":"SQLite Database","Exportação / Excel":"Export / Excel","Gráficos de Abastecimento":"Refueling Charts","Configurações do Sistema":"System Settings","Fórmulas e Física":"Formulas & Physics","Total de Abastecimentos":"Total Refuelings","Aquecimento / Compressão":"Heating / Compression","Calcular":"Calculate","Limpar Resultados":"Clear Results","Limpar Campos":"Clear Fields","Salvar Abastecimento":"Save Refueling","Novo":"New","Observações":"Notes","Atualizar Banco SQLite":"Refresh SQLite Database","Selecionar Excel":"Select Excel","Exportar Excel":"Export Excel","Gerar Relatório PDF":"Generate PDF Report","Atualizar Gráfico":"Update Chart","Limpar":"Clear","Tema:":"Theme:","Idioma:":"Language:","Cores personalizadas:":"Custom colors:","Cor das abas":"Tab color","Cor das letras das abas":"Tab text color","Fundo da linha 1":"Row 1 background","Fundo da linha 2":"Row 2 background","Cor das letras das tabelas":"Table text color","Fundo geral das tabelas":"Table background","Backup automático":"Automatic backup","Salvar Configurações":"Save Settings","Restaurar Padrão":"Restore Defaults","Exportar JSON":"Export JSON","Importar JSON":"Import JSON","Escolher...":"Choose...","Aba atual:":"Current tab:","Pesquisar:":"Search:","Pesquisar":"Search","Mostrar Todos":"Show All","Resumo Geral":"General Summary","Atualizar":"Refresh","Calcular exclusivamente pela condição de referência ANP":"Calculate exclusively using the ANP reference condition","Copiar dados da aba Abastecimentos":"Copy data from Refuelings tab","Gerar Relatório PDF":"Generate PDF Report","Gráficos de abastecimentos":"Refueling charts","Tipo de gráfico:":"Chart type:","Atualizar Gráfico":"Update Chart","Consumo = distância entre odômetros válidos ÷ volume abastecido pela bomba.":"Consumption = distance between valid odometer readings ÷ volume supplied by the pump.","Registros armazenados no Banco SQLite":"Records stored in SQLite Database","Exportação e integração com Excel":"Export and Excel integration","Configurações":"Settings","Configurações salvas com sucesso.":"Settings saved successfully.","Configurações restauradas com sucesso.":"Settings restored successfully.","Volume do cilindro (L):":"Cylinder volume (L):","Quantidade de cilindros:":"Number of cylinders:","Pressão (bar):":"Pressure (bar):","Temperatura (°C):":"Temperature (°C):","Altitude (m):":"Altitude (m):","Fator Z:":"Compressibility factor Z:","Massa molar (kg/mol):":"Molar mass (kg/mol):","Massa específica de referência (kg/m³):":"Reference density (kg/m³):"},
"Español": {"Sistema de Cálculo de GNV":"Sistema de Cálculo de GNV","Cálculos":"Cálculos","Abastecimentos":"Abastecimientos","Histórico de Abastecimentos":"Historial de Abastecimientos","Banco SQLite":"Base de datos SQLite","Exportação / Excel":"Exportación / Excel","Gráficos de Abastecimento":"Gráficos de Abastecimiento","Configurações do Sistema":"Configuración del Sistema","Fórmulas e Física":"Fórmulas y Física","Total de Abastecimentos":"Total de Abastecimientos","Aquecimento / Compressão":"Calentamiento / Compresión","Calcular":"Calcular","Limpar Resultados":"Limpiar Resultados","Limpar Campos":"Limpiar Campos","Salvar Abastecimento":"Guardar Abastecimiento","Novo":"Nuevo","Observações":"Observaciones","Atualizar Banco SQLite":"Actualizar Base SQLite","Selecionar Excel":"Seleccionar Excel","Exportar Excel":"Exportar Excel","Gerar Relatório PDF":"Generar Informe PDF","Atualizar Gráfico":"Actualizar Gráfico","Limpar":"Limpiar","Tema:":"Tema:","Idioma:":"Idioma:","Cores personalizadas:":"Colores personalizados:","Cor das abas":"Color de pestañas","Cor das letras das abas":"Color del texto de pestañas","Fundo da linha 1":"Fondo de fila 1","Fundo da linha 2":"Fondo de fila 2","Cor das letras das tabelas":"Color del texto de tablas","Fundo geral das tabelas":"Fondo de tablas","Backup automático":"Copia de seguridad automática","Salvar Configurações":"Guardar Configuración","Restaurar Padrão":"Restaurar Predeterminados","Exportar JSON":"Exportar JSON","Importar JSON":"Importar JSON","Escolher...":"Elegir...","Aba atual:":"Pestaña actual:","Pesquisar:":"Buscar:","Pesquisar":"Buscar","Mostrar Todos":"Mostrar Todos","Resumo Geral":"Resumen General","Atualizar":"Actualizar","Configurações":"Configuración","Configurações salvas com sucesso.":"Configuración guardada correctamente.","Configurações restauradas com sucesso.":"Configuración restaurada correctamente.","Volume do cilindro (L):":"Volumen del cilindro (L):","Quantidade de cilindros:":"Cantidad de cilindros:","Pressão (bar):":"Presión (bar):","Temperatura (°C):":"Temperatura (°C):","Altitude (m):":"Altitud (m):","Fator Z:":"Factor de compresibilidad Z:","Massa molar (kg/mol):":"Masa molar (kg/mol):","Massa específica de referência (kg/m³):":"Densidad de referencia (kg/m³):"},
"Français": {"Sistema de Cálculo de GNV":"Système de calcul GNV","Cálculos":"Calculs","Abastecimentos":"Ravitaillements","Histórico de Abastecimentos":"Historique des ravitaillements","Banco SQLite":"Base SQLite","Exportação / Excel":"Exportation / Excel","Gráficos de Abastecimento":"Graphiques de ravitaillement","Configurações do Sistema":"Paramètres du système","Fórmulas e Física":"Formules et physique","Total de Abastecimentos":"Total des ravitaillements","Aquecimento / Compressão":"Chauffage / compression","Calcular":"Calculer","Limpar Resultados":"Effacer les résultats","Limpar Campos":"Effacer les champs","Salvar Abastecimento":"Enregistrer le ravitaillement","Novo":"Nouveau","Observações":"Observations","Atualizar Banco SQLite":"Actualiser la base SQLite","Selecionar Excel":"Sélectionner Excel","Exportar Excel":"Exporter Excel","Gerar Relatório PDF":"Générer le rapport PDF","Atualizar Gráfico":"Actualiser le graphique","Limpar":"Effacer","Tema:":"Thème :","Idioma:":"Langue :","Cores personalizadas:":"Couleurs personnalisées :","Cor das abas":"Couleur des onglets","Cor das letras das abas":"Couleur du texte des onglets","Fundo da linha 1":"Fond de ligne 1","Fundo da linha 2":"Fond de ligne 2","Cor das letras das tabelas":"Couleur du texte des tableaux","Fundo geral das tabelas":"Fond des tableaux","Backup automático":"Sauvegarde automatique","Salvar Configurações":"Enregistrer les paramètres","Restaurar Padrão":"Restaurer les valeurs par défaut","Exportar JSON":"Exporter JSON","Importar JSON":"Importer JSON","Escolher...":"Choisir...","Aba atual:":"Onglet actuel :","Pesquisar:":"Rechercher :","Pesquisar":"Rechercher","Mostrar Todos":"Tout afficher","Resumo Geral":"Résumé général","Atualizar":"Actualiser","Configurações":"Paramètres","Configurações salvas com sucesso.":"Paramètres enregistrés avec succès.","Configurações restauradas com sucesso.":"Paramètres restaurés avec succès.","Volume do cilindro (L):":"Volume du cylindre (L) :","Quantidade de cilindros:":"Nombre de cylindres :","Pressão (bar):":"Pression (bar) :","Temperatura (°C):":"Température (°C) :","Altitude (m):":"Altitude (m) :","Fator Z:":"Facteur de compressibilité Z :","Massa molar (kg/mol):":"Masse molaire (kg/mol) :","Massa específica de referência (kg/m³):":"Masse volumique de référence (kg/m³) :"},
"Italiano": {"Sistema de Cálculo de GNV":"Sistema di calcolo GNV","Cálculos":"Calcoli","Abastecimentos":"Rifornimenti","Histórico de Abastecimentos":"Storico rifornimenti","Banco SQLite":"Database SQLite","Exportação / Excel":"Esportazione / Excel","Gráficos de Abastecimento":"Grafici dei rifornimenti","Configurações do Sistema":"Impostazioni di sistema","Fórmulas e Física":"Formule e fisica","Total de Abastecimentos":"Totale rifornimenti","Aquecimento / Compressão":"Riscaldamento / compressione","Calcular":"Calcola","Limpar Resultados":"Cancella risultati","Limpar Campos":"Cancella campi","Salvar Abastecimento":"Salva rifornimento","Novo":"Nuovo","Observações":"Note","Atualizar Banco SQLite":"Aggiorna database SQLite","Selecionar Excel":"Seleziona Excel","Exportar Excel":"Esporta Excel","Gerar Relatório PDF":"Genera rapporto PDF","Atualizar Gráfico":"Aggiorna grafico","Limpar":"Cancella","Tema:":"Tema:","Idioma:":"Lingua:","Cores personalizadas:":"Colori personalizzati:","Cor das abas":"Colore schede","Cor das letras das abas":"Colore testo schede","Fundo da linha 1":"Sfondo riga 1","Fundo da linha 2":"Sfondo riga 2","Cor das letras das tabelas":"Colore testo tabelle","Fundo geral das tabelas":"Sfondo tabelle","Backup automático":"Backup automatico","Salvar Configurações":"Salva impostazioni","Restaurar Padrão":"Ripristina predefiniti","Exportar JSON":"Esporta JSON","Importar JSON":"Importa JSON","Escolher...":"Scegli...","Aba atual:":"Scheda attuale:","Pesquisar:":"Cerca:","Pesquisar":"Cerca","Mostrar Todos":"Mostra tutti","Resumo Geral":"Riepilogo generale","Atualizar":"Aggiorna","Configurações":"Impostazioni","Configurações salvas com sucesso.":"Impostazioni salvate con successo.","Configurações restauradas com sucesso.":"Impostazioni ripristinate con successo.","Volume do cilindro (L):":"Volume cilindro (L):","Quantidade de cilindros:":"Numero cilindri:","Pressão (bar):":"Pressione (bar):","Temperatura (°C):":"Temperatura (°C):","Altitude (m):":"Altitudine (m):","Fator Z:":"Fattore di comprimibilità Z:","Massa molar (kg/mol):":"Massa molare (kg/mol):","Massa específica de referência (kg/m³):":"Densità di riferimento (kg/m³):"},
"Deutsch": {"Sistema de Cálculo de GNV":"GNV-Berechnungssystem","Cálculos":"Berechnungen","Abastecimentos":"Tankvorgänge","Histórico de Abastecimentos":"Tankvorgangsverlauf","Banco SQLite":"SQLite-Datenbank","Exportação / Excel":"Export / Excel","Gráficos de Abastecimento":"Tankdiagramme","Configurações do Sistema":"Systemeinstellungen","Fórmulas e Física":"Formeln & Physik","Total de Abastecimentos":"Gesamte Tankvorgänge","Aquecimento / Compressão":"Erwärmung / Kompression","Calcular":"Berechnen","Limpar Resultados":"Ergebnisse löschen","Limpar Campos":"Felder löschen","Salvar Abastecimento":"Tankvorgang speichern","Novo":"Neu","Observações":"Hinweise","Atualizar Banco SQLite":"SQLite-Datenbank aktualisieren","Selecionar Excel":"Excel auswählen","Exportar Excel":"Excel exportieren","Gerar Relatório PDF":"PDF-Bericht erstellen","Atualizar Gráfico":"Diagramm aktualisieren","Limpar":"Löschen","Tema:":"Thema:","Idioma:":"Sprache:","Cores personalizadas:":"Benutzerdefinierte Farben:","Cor das abas":"Tab-Farbe","Cor das letras das abas":"Tab-Textfarbe","Fundo da linha 1":"Zeilenhintergrund 1","Fundo da linha 2":"Zeilenhintergrund 2","Cor das letras das tabelas":"Tabellentextfarbe","Fundo geral das tabelas":"Tabellenhintergrund","Backup automático":"Automatische Sicherung","Salvar Configurações":"Einstellungen speichern","Restaurar Padrão":"Standard wiederherstellen","Exportar JSON":"JSON exportieren","Importar JSON":"JSON importieren","Escolher...":"Auswählen...","Aba atual:":"Aktueller Tab:","Pesquisar:":"Suchen:","Pesquisar":"Suchen","Mostrar Todos":"Alle anzeigen","Resumo Geral":"Allgemeine Übersicht","Atualizar":"Aktualisieren","Configurações":"Einstellungen","Configurações salvas com sucesso.":"Einstellungen erfolgreich gespeichert.","Configurações restauradas com sucesso.":"Einstellungen erfolgreich wiederhergestellt.","Volume do cilindro (L):":"Zylindervolumen (L):","Quantidade de cilindros:":"Anzahl Zylinder:","Pressão (bar):":"Druck (bar):","Temperatura (°C):":"Temperatur (°C):","Altitude (m):":"Höhe (m):","Fator Z:":"Kompressibilitätsfaktor Z:","Massa molar (kg/mol):":"Molmasse (kg/mol):","Massa específica de referência (kg/m³):":"Referenzdichte (kg/m³):"},
"日本語": {"Sistema de Cálculo de GNV":"GNV計算システム","Cálculos":"計算","Abastecimentos":"給油","Histórico de Abastecimentos":"給油履歴","Banco SQLite":"SQLiteデータベース","Exportação / Excel":"Excelエクスポート","Gráficos de Abastecimento":"給油グラフ","Configurações do Sistema":"システム設定","Fórmulas e Física":"数式と物理","Total de Abastecimentos":"給油総数","Aquecimento / Compressão":"加熱 / 圧縮","Calcular":"計算","Limpar Resultados":"結果を消去","Limpar Campos":"入力を消去","Salvar Abastecimento":"給油を保存","Novo":"新規","Observações":"備考","Atualizar Banco SQLite":"SQLiteデータベースを更新","Selecionar Excel":"Excelを選択","Exportar Excel":"Excelをエクスポート","Gerar Relatório PDF":"PDFレポートを生成","Atualizar Gráfico":"グラフを更新","Limpar":"消去","Tema:":"テーマ:","Idioma:":"言語:","Cores personalizadas:":"カスタム色:","Cor das abas":"タブの色","Cor das letras das abas":"タブ文字色","Fundo da linha 1":"1行目の背景","Fundo da linha 2":"2行目の背景","Cor das letras das tabelas":"表文字色","Fundo geral das tabelas":"表の背景","Backup automático":"自動バックアップ","Salvar Configurações":"設定を保存","Restaurar Padrão":"既定値に戻す","Exportar JSON":"JSONをエクスポート","Importar JSON":"JSONをインポート","Escolher...":"選択...","Aba atual:":"現在のタブ:","Pesquisar:":"検索:","Pesquisar":"検索","Mostrar Todos":"すべて表示","Resumo Geral":"概要","Atualizar":"更新","Configurações":"設定","Configurações salvas com sucesso.":"設定を保存しました。","Configurações restauradas com sucesso.":"設定を既定値に戻しました。","Volume do cilindro (L):":"シリンダー容量 (L):","Quantidade de cilindros:":"シリンダー数:","Pressão (bar):":"圧力 (bar):","Temperatura (°C):":"温度 (°C):","Altitude (m):":"標高 (m):","Fator Z:":"圧縮係数 Z:","Massa molar (kg/mol):":"モル質量 (kg/mol):","Massa específica de referência (kg/m³):":"基準密度 (kg/m³):"},
"中文": {"Sistema de Cálculo de GNV":"GNV计算系统","Cálculos":"计算","Abastecimentos":"加气记录","Histórico de Abastecimentos":"加气历史","Banco SQLite":"SQLite数据库","Exportação / Excel":"Excel导出","Gráficos de Abastecimento":"加气图表","Configurações do Sistema":"系统设置","Fórmulas e Física":"公式与物理","Total de Abastecimentos":"加气总数","Aquecimento / Compressão":"加热 / 压缩","Calcular":"计算","Limpar Resultados":"清除结果","Limpar Campos":"清除字段","Salvar Abastecimento":"保存加气记录","Novo":"新建","Observações":"备注","Atualizar Banco SQLite":"更新SQLite数据库","Selecionar Excel":"选择Excel","Exportar Excel":"导出Excel","Gerar Relatório PDF":"生成PDF报告","Atualizar Gráfico":"更新图表","Limpar":"清除","Tema:":"主题:","Idioma:":"语言:","Cores personalizadas:":"自定义颜色:","Cor das abas":"标签页颜色","Cor das letras das abas":"标签文字颜色","Fundo da linha 1":"第1行背景","Fundo da linha 2":"第2行背景","Cor das letras das tabelas":"表格文字颜色","Fundo geral das tabelas":"表格背景","Backup automático":"自动备份","Salvar Configurações":"保存设置","Restaurar Padrão":"恢复默认值","Exportar JSON":"导出JSON","Importar JSON":"导入JSON","Escolher...":"选择...","Aba atual:":"当前标签:","Pesquisar:":"搜索:","Pesquisar":"搜索","Mostrar Todos":"显示全部","Resumo Geral":"总体摘要","Atualizar":"更新","Configurações":"设置","Configurações salvas com sucesso.":"设置已成功保存。","Configurações restauradas com sucesso.":"设置已恢复为默认值。","Volume do cilindro (L):":"气瓶容积 (L):","Quantidade de cilindros:":"气瓶数量:","Pressão (bar):":"压力 (bar):","Temperatura (°C):":"温度 (°C):","Altitude (m):":"海拔 (m):","Fator Z:":"压缩因子 Z:","Massa molar (kg/mol):":"摩尔质量 (kg/mol):","Massa específica de referência (kg/m³):":"参考密度 (kg/m³):"}
}
# Complemento de tradução para comparação detalhada do abastecimento.
for _lang, _map in {
    "English": {"Volume informado pela bomba":"Volume reported by pump","Volume real calculado — ANP/idealizado":"Calculated real volume — ANP/idealized","Diferença bomba − ANP":"Pump − ANP difference","Diferença percentual bomba × ANP":"Pump × ANP percentage difference","Volume real calculado — modelo físico":"Calculated real volume — physical model","Diferença bomba − físico":"Pump − physical difference","Diferença percentual bomba × físico":"Pump × physical percentage difference","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"INTERPRETATION: the system compares the pump reading with two independent models.","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"The classification is a comparative alert and is not, by itself, proof of fraud."},
    "Español": {"Volume informado pela bomba":"Volumen indicado por la bomba","Volume real calculado — ANP/idealizado":"Volumen real calculado — ANP/idealizado","Diferença bomba − ANP":"Diferencia bomba − ANP","Diferença percentual bomba × ANP":"Diferencia porcentual bomba × ANP","Volume real calculado — modelo físico":"Volumen real calculado — modelo físico","Diferença bomba − físico":"Diferencia bomba − físico","Diferença percentual bomba × físico":"Diferencia porcentual bomba × físico","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"INTERPRETACIÓN: el sistema compara la lectura de la bomba con dos modelos independientes.","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"La clasificación es una alerta comparativa y no constituye, por sí sola, prueba metrológica de fraude."},
    "Français": {"Volume informado pela bomba":"Volume indiqué par la pompe","Volume real calculado — ANP/idealizado":"Volume réel calculé — ANP/idéaliste","Diferença bomba − ANP":"Écart pompe − ANP","Diferença percentual bomba × ANP":"Écart en pourcentage pompe × ANP","Volume real calculado — modelo físico":"Volume réel calculé — modèle physique","Diferença bomba − físico":"Écart pompe − physique","Diferença percentual bomba × físico":"Écart en pourcentage pompe × physique","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"INTERPRÉTATION : le système compare la lecture de la pompe à deux modèles indépendants.","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"La classification est une alerte comparative et ne constitue pas, à elle seule, une preuve de fraude métrologique."},
    "Italiano": {"Volume informado pela bomba":"Volume indicato dalla pompa","Volume real calculado — ANP/idealizado":"Volume reale calcolato — ANP/idealizzato","Diferença bomba − ANP":"Differenza pompa − ANP","Diferença percentual bomba × ANP":"Differenza percentuale pompa × ANP","Volume real calculado — modelo físico":"Volume reale calcolato — modello fisico","Diferença bomba − físico":"Differenza pompa − fisico","Diferença percentual bomba × físico":"Differenza percentuale pompa × fisico","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"INTERPRETAZIONE: il sistema confronta la lettura della pompa con due modelli indipendenti.","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"La classificazione è un avviso comparativo e non costituisce, da sola, prova di frode metrologica."},
    "Deutsch": {"Volume informado pela bomba":"Von der Zapfsäule angezeigtes Volumen","Volume real calculado — ANP/idealizado":"Berechnetes reales Volumen — ANP/idealisiert","Diferença bomba − ANP":"Differenz Zapfsäule − ANP","Diferença percentual bomba × ANP":"Prozentuale Differenz Zapfsäule × ANP","Volume real calculado — modelo físico":"Berechnetes reales Volumen — physikalisches Modell","Diferença bomba − físico":"Differenz Zapfsäule − physikalisches Modell","Diferença percentual bomba × físico":"Prozentuale Differenz Zapfsäule × physikalisches Modell","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"INTERPRETATION: Das System vergleicht die Zapfsäulenanzeige mit zwei unabhängigen Modellen.","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"Die Klassifizierung ist ein Vergleichshinweis und allein kein metrologischer Nachweis eines Betrugs."},
    "日本語": {"Volume informado pela bomba":"ポンプ表示量","Volume real calculado — ANP/idealizado":"計算された実量 — ANP/理想モデル","Diferença bomba − ANP":"ポンプ − ANP 差分","Diferença percentual bomba × ANP":"ポンプ × ANP 差分率","Volume real calculado — modelo físico":"計算された実量 — 物理モデル","Diferença bomba − físico":"ポンプ − 物理モデル差分","Diferença percentual bomba × físico":"ポンプ × 物理モデル差分率","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"解釈：システムはポンプ表示値を2つの独立モデルと比較します。","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"分類は比較上の警告であり、それだけで計量上の不正の証明にはなりません。"},
    "中文": {"Volume informado pela bomba":"加气机显示体积","Volume real calculado — ANP/idealizado":"计算实际体积 — ANP/理想模型","Diferença bomba − ANP":"加气机 − ANP 差值","Diferença percentual bomba × ANP":"加气机 × ANP 百分比差异","Volume real calculado — modelo físico":"计算实际体积 — 物理模型","Diferença bomba − físico":"加气机 − 物理模型差值","Diferença percentual bomba × físico":"加气机 × 物理模型百分比差异","INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.":"说明：系统将加气机读数与两个独立模型进行比较。","A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude.":"分类只是比较警告，不能单独作为计量欺诈的证明。"}
}.items():
    I18N.setdefault(_lang, {}).update(_map)

# Complemento de tradução da interface móvel e dos relatórios.
for _lang, _pairs in {
    "English": {"Valores padrão":"Default values","RELATÓRIO DOS CÁLCULOS DE GNV":"CNG CALCULATION REPORT","Capacidade física total":"Total physical capacity","Quantidade de cilindros":"Number of cylinders","Pressão":"Pressure","Temperatura":"Temperature","Altitude":"Altitude","Fator Z":"Z factor","Massa molar":"Molar mass","Massa específica de referência":"Reference density","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"TOTAL equivalent volume at 1.01325 bar at the SAME entered temperature","Volume TOTAL equivalente científico a 20 °C":"Scientific TOTAL equivalent volume at 20 °C","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"ANP/idealized ADDED volume (Z=1) at 20 °C","Massa de GNV":"CNG mass","Densidade calculada no cilindro":"Calculated cylinder density","Quantidade de matéria":"Amount of substance","Tubulação alta pressão — mínimo":"High-pressure piping — minimum","Tubulação alta pressão — médio":"High-pressure piping — average","Tubulação alta pressão — máximo":"High-pressure piping — maximum","Válvulas de serviço":"Service valves","Redutor — câmaras de gás":"Reducer — gas chambers","Baixa pressão/filtro/flauta":"Low pressure/filter/rail","VOLUME TOTAL DO CIRCUITO":"TOTAL CIRCUIT VOLUME","Base geométrica da tubulação: V = pi × r² × L.":"Piping geometric basis: V = pi × r² × L.","Premissa: tubo externo 6 mm, parede 1,2 mm, interno 3,6 mm, comprimento 4,0–5,5 m.":"Assumption: 6 mm external tube, 1.2 mm wall, 3.6 mm internal diameter, length 4.0–5.5 m.","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"WARNING: engineering estimate; does not replace actual vehicle measurement.","Dados físicos inválidos.":"Invalid physical data.","Erro":"Error","Configurações salvas com sucesso.":"Settings saved successfully.","Configurações restauradas com sucesso.":"Settings restored successfully.","Bomba × Teórico":"Pump × Theoretical","Volume por abastecimento":"Volume per refueling","Consumo km/m³":"Consumption km/m³","Volume: bomba × teórico (m³)":"Volume: pump × theoretical (m³)","Volume por abastecimento (m³)":"Volume per refueling (m³)","Consumo: km/m³":"Consumption: km/m³"},
    "Español": {"Valores padrão":"Valores predeterminados","RELATÓRIO DOS CÁLCULOS DE GNV":"INFORME DE CÁLCULOS DE GNV","Capacidade física total":"Capacidad física total","Quantidade de cilindros":"Cantidad de cilindros","Pressão":"Presión","Temperatura":"Temperatura","Altitude":"Altitud","Fator Z":"Factor Z","Massa molar":"Masa molar","Massa específica de referência":"Densidad de referencia","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"Volumen TOTAL equivalente a 1,01325 bar a la MISMA T indicada","Volume TOTAL equivalente científico a 20 °C":"Volumen TOTAL equivalente científico a 20 °C","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"Volumen AÑADIDO ANP/idealizado (Z=1) a 20 °C","Massa de GNV":"Masa de GNV","Densidade calculada no cilindro":"Densidad calculada en el cilindro","Quantidade de matéria":"Cantidad de sustancia","Tubulação alta pressão — mínimo":"Tubería de alta presión — mínimo","Tubulação alta pressão — médio":"Tubería de alta presión — medio","Tubulação alta pressão — máximo":"Tubería de alta presión — máximo","Válvulas de serviço":"Válvulas de servicio","Redutor — câmaras de gás":"Reductor — cámaras de gas","Baixa pressão/filtro/flauta":"Baja presión/filtro/flauta","VOLUME TOTAL DO CIRCUITO":"VOLUMEN TOTAL DEL CIRCUITO","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"ATENCIÓN: estimación de ingeniería; no sustituye la medición real del vehículo.","Configurações salvas com sucesso.":"Configuración guardada correctamente.","Configurações restauradas com sucesso.":"Configuración restaurada correctamente.","Bomba × Teórico":"Bomba × Teórico","Volume por abastecimento":"Volumen por repostaje","Consumo km/m³":"Consumo km/m³","Erro":"Error"},
    "Français": {"Valores padrão":"Valeurs par défaut","RELATÓRIO DOS CÁLCULOS DE GNV":"RAPPORT DES CALCULS GNV","Capacidade física total":"Capacité physique totale","Quantidade de cilindros":"Nombre de cylindres","Pressão":"Pression","Temperatura":"Température","Altitude":"Altitude","Fator Z":"Facteur Z","Massa molar":"Masse molaire","Massa específica de referência":"Masse volumique de référence","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"Volume TOTAL équivalent à 1,01325 bar à la MÊME T indiquée","Volume TOTAL equivalente científico a 20 °C":"Volume TOTAL équivalent scientifique à 20 °C","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"Volume AJOUTÉ ANP/idéaliste (Z=1) à 20 °C","Massa de GNV":"Masse de GNV","Densidade calculada no cilindro":"Masse volumique calculée dans le cylindre","Quantidade de matéria":"Quantité de matière","Tubulação alta pressão — mínimo":"Tuyauterie haute pression — minimum","Tubulação alta pressão — médio":"Tuyauterie haute pression — moyenne","Tubulação alta pressão — máximo":"Tuyauterie haute pression — maximum","Válvulas de serviço":"Vannes de service","Redutor — câmaras de gás":"Réducteur — chambres de gaz","Baixa pressão/filtro/flauta":"Basse pression/filtre/rampe","VOLUME TOTAL DO CIRCUITO":"VOLUME TOTAL DU CIRCUIT","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"ATTENTION : estimation d’ingénierie ; ne remplace pas la mesure réelle du véhicule.","Configurações salvas com sucesso.":"Paramètres enregistrés avec succès.","Configurações restauradas com sucesso.":"Paramètres restaurés avec succès.","Bomba × Teórico":"Pompe × Théorique","Volume por abastecimento":"Volume par ravitaillement","Consumo km/m³":"Consommation km/m³","Erro":"Erreur"},
    "Italiano": {"Valores padrão":"Valori predefiniti","RELATÓRIO DOS CÁLCULOS DE GNV":"RAPPORTO DEI CALCOLI GNV","Capacidade física total":"Capacità fisica totale","Quantidade de cilindros":"Numero di cilindri","Pressão":"Pressione","Temperatura":"Temperatura","Altitude":"Altitudine","Fator Z":"Fattore Z","Massa molar":"Massa molare","Massa específica de referência":"Densità di riferimento","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"Volume TOTALE equivalente a 1,01325 bar alla STESSA T indicata","Volume TOTAL equivalente científico a 20 °C":"Volume TOTALE equivalente scientifico a 20 °C","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"Volume AGGIUNTO ANP/idealizzato (Z=1) a 20 °C","Massa de GNV":"Massa GNV","Densidade calculada no cilindro":"Densità calcolata nel cilindro","Quantidade de matéria":"Quantità di sostanza","Tubulação alta pressão — mínimo":"Tubazione alta pressione — minimo","Tubulação alta pressão — médio":"Tubazione alta pressione — medio","Tubulação alta pressão — máximo":"Tubazione alta pressione — massimo","Válvulas de serviço":"Valvole di servizio","Redutor — câmaras de gás":"Riduttore — camere del gas","Baixa pressão/filtro/flauta":"Bassa pressione/filtro/rail","VOLUME TOTAL DO CIRCUITO":"VOLUME TOTALE DEL CIRCUITO","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"ATTENZIONE: stima ingegneristica; non sostituisce la misura reale del veicolo.","Configurações salvas com sucesso.":"Impostazioni salvate con successo.","Configurações restauradas com sucesso.":"Impostazioni ripristinate con successo.","Bomba × Teórico":"Pompa × Teorico","Volume por abastecimento":"Volume per rifornimento","Consumo km/m³":"Consumo km/m³","Erro":"Errore"},
    "Deutsch": {"Valores padrão":"Standardwerte","RELATÓRIO DOS CÁLCULOS DE GNV":"GNV-BERECHNUNGSBERICHT","Capacidade física total":"Physisches Gesamtvolumen","Quantidade de cilindros":"Anzahl der Zylinder","Pressão":"Druck","Temperatura":"Temperatur","Altitude":"Höhe","Fator Z":"Z-Faktor","Massa molar":"Molmasse","Massa específica de referência":"Referenzdichte","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"Äquivalentes GESAMTVOLUMEN bei 1,01325 bar und derselben angegebenen T","Volume TOTAL equivalente científico a 20 °C":"Wissenschaftliches GESAMTÄQUIVALENTVOLUMEN bei 20 °C","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"ANP/idealisiertes HINZUGEFÜGTES Volumen (Z=1) bei 20 °C","Massa de GNV":"GNG-Masse","Densidade calculada no cilindro":"Berechnete Zylinderdichte","Quantidade de matéria":"Stoffmenge","Tubulação alta pressão — mínimo":"Hochdruckleitung — Minimum","Tubulação alta pressão — médio":"Hochdruckleitung — Mittelwert","Tubulação alta pressão — máximo":"Hochdruckleitung — Maximum","Válvulas de serviço":"Serviceventile","Redutor — câmaras de gás":"Druckminderer — Gaskammern","Baixa pressão/filtro/flauta":"Niederdruck/Filter/Rail","VOLUME TOTAL DO CIRCUITO":"GESAMTVOLUMEN DES KREISLAUFS","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"ACHTUNG: technische Schätzung; ersetzt nicht die tatsächliche Fahrzeugmessung.","Configurações salvas com sucesso.":"Einstellungen erfolgreich gespeichert.","Configurações restauradas com sucesso.":"Einstellungen auf Standard zurückgesetzt.","Bomba × Teórico":"Pumpe × Theoretisch","Volume por abastecimento":"Volumen pro Tankvorgang","Consumo km/m³":"Verbrauch km/m³","Erro":"Fehler"},
    "日本語": {"Valores padrão":"デフォルト値","RELATÓRIO DOS CÁLCULOS DE GNV":"GNV計算レポート","Capacidade física total":"物理的総容量","Quantidade de cilindros":"シリンダー数","Pressão":"圧力","Temperatura":"温度","Altitude":"標高","Fator Z":"Z係数","Massa molar":"モル質量","Massa específica de referência":"基準密度","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"入力した同じ温度における1.01325 bar換算の総等価容積","Volume TOTAL equivalente científico a 20 °C":"20 °Cにおける科学的総等価容積","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"20 °CにおけるANP/理想化追加容積 (Z=1)","Massa de GNV":"GNV質量","Densidade calculada no cilindro":"シリンダー内計算密度","Quantidade de matéria":"物質量","Tubulação alta pressão — mínimo":"高圧配管 — 最小","Tubulação alta pressão — médio":"高圧配管 — 平均","Tubulação alta pressão — máximo":"高圧配管 — 最大","Válvulas de serviço":"サービスバルブ","Redutor — câmaras de gás":"減圧器 — ガス室","Baixa pressão/filtro/flauta":"低圧/フィルター/レール","VOLUME TOTAL DO CIRCUITO":"回路総容積","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"注意：工学的推定値であり、実車の測定値に代わるものではありません。","Configurações salvas com sucesso.":"設定を保存しました。","Configurações restauradas com sucesso.":"設定を既定値に戻しました。","Bomba × Teórico":"ポンプ × 理論値","Volume por abastecimento":"給油ごとの容積","Consumo km/m³":"燃費 km/m³","Erro":"エラー"},
    "中文": {"Valores padrão":"默认值","RELATÓRIO DOS CÁLCULOS DE GNV":"GNV计算报告","Capacidade física total":"物理总容量","Quantidade de cilindros":"气瓶数量","Pressão":"压力","Temperatura":"温度","Altitude":"海拔","Fator Z":"Z因子","Massa molar":"摩尔质量","Massa específica de referência":"参考密度","Volume TOTAL equivalente a 1,01325 bar na MESMA T informada":"在相同输入温度下换算至1.01325 bar的总等效体积","Volume TOTAL equivalente científico a 20 °C":"20 °C下科学总等效体积","Volume ADICIONADO ANP/idealizado (Z=1) a 20 °C":"20 °C下ANP/理想化新增体积 (Z=1)","Massa de GNV":"GNV质量","Densidade calculada no cilindro":"气瓶内计算密度","Quantidade de matéria":"物质的量","Tubulação alta pressão — mínimo":"高压管路 — 最小","Tubulação alta pressão — médio":"高压管路 — 平均","Tubulação alta pressão — máximo":"高压管路 — 最大","Válvulas de serviço":"服务阀","Redutor — câmaras de gás":"减压器 — 气室","Baixa pressão/filtro/flauta":"低压/过滤器/喷轨","VOLUME TOTAL DO CIRCUITO":"回路总容积","ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.":"注意：工程估算值，不能替代车辆实际测量。","Configurações salvas com sucesso.":"设置保存成功。","Configurações restauradas com sucesso.":"设置已恢复默认值。","Bomba × Teórico":"加气机 × 理论值","Volume por abastecimento":"每次加气体积","Consumo km/m³":"油耗 km/m³","Erro":"错误"}
}.items():
    I18N.setdefault(_lang, {}).update(_pairs)

I18N_FORMULAS = {
"English":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"FORMULAS AND PHYSICS OF THE CNG CALCULATION SYSTEM","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"PART A — ANP REFERENCE CONDITION","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"PART B — SCIENTIFIC REAL-GAS MODEL","PRESSÃO ABSOLUTA":"ABSOLUTE PRESSURE","TEMPERATURA ABSOLUTA":"ABSOLUTE TEMPERATURE","FATOR DE COMPRESSIBILIDADE Z":"COMPRESSIBILITY FACTOR Z","QUANTIDADE DE GÁS ADICIONADA":"AMOUNT OF GAS ADDED","MASSA ADICIONADA":"MASS ADDED","VOLUME FÍSICO DO CILINDRO":"PHYSICAL CYLINDER VOLUME","CONVERSÃO CIENTÍFICA PARA 20 °C":"SCIENTIFIC CONVERSION TO 20 °C","CONVERSÃO ANP/IDEALIZADA (Z=1)":"ANP/IDEALIZED CONVERSION (Z=1)"},
"Español":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"FÓRMULAS Y FÍSICA DEL SISTEMA DE CÁLCULO DE GNV","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"PARTE A — CONDICIÓN DE REFERENCIA ANP","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"PARTE B — MODELO CIENTÍFICO DE GAS REAL","PRESSÃO ABSOLUTA":"PRESIÓN ABSOLUTA","TEMPERATURA ABSOLUTA":"TEMPERATURA ABSOLUTA","FATOR DE COMPRESSIBILIDADE Z":"FACTOR DE COMPRESIBILIDAD Z","QUANTIDADE DE GÁS ADICIONADA":"CANTIDAD DE GAS AÑADIDA","MASSA ADICIONADA":"MASA AÑADIDA","VOLUME FÍSICO DO CILINDRO":"VOLUMEN FÍSICO DEL CILINDRO","CONVERSÃO CIENTÍFICA PARA 20 °C":"CONVERSIÓN CIENTÍFICA A 20 °C","CONVERSÃO ANP/IDEALIZADA (Z=1)":"CONVERSIÓN ANP/IDEALIZADA (Z=1)"},
"Français":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"FORMULES ET PHYSIQUE DU SYSTÈME DE CALCUL GNV","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"PARTIE A — CONDITION DE RÉFÉRENCE ANP","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"PARTIE B — MODÈLE SCIENTIFIQUE DE GAZ RÉEL","PRESSÃO ABSOLUTA":"PRESSION ABSOLUE","TEMPERATURA ABSOLUTA":"TEMPÉRATURE ABSOLUE","FATOR DE COMPRESSIBILIDADE Z":"FACTEUR DE COMPRESSIBILITÉ Z","QUANTIDADE DE GÁS ADICIONADA":"QUANTITÉ DE GAZ AJOUTÉE","MASSA ADICIONADA":"MASSE AJOUTÉE","VOLUME FÍSICO DO CILINDRO":"VOLUME PHYSIQUE DU CYLINDRE","CONVERSÃO CIENTÍFICA PARA 20 °C":"CONVERSION SCIENTIFIQUE À 20 °C","CONVERSÃO ANP/IDEALIZADA (Z=1)":"CONVERSION ANP/IDÉALISÉE (Z=1)"},
"Italiano":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"FORMULE E FISICA DEL SISTEMA DI CALCOLO GNV","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"PARTE A — CONDIZIONE DI RIFERIMENTO ANP","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"PARTE B — MODELLO SCIENTIFICO DI GAS REALE","PRESSÃO ABSOLUTA":"PRESSIONE ASSOLUTA","TEMPERATURA ABSOLUTA":"TEMPERATURA ASSOLUTA","FATOR DE COMPRESSIBILIDADE Z":"FATTORE DI COMPRIMIBILITÀ Z","QUANTIDADE DE GÁS ADICIONADA":"QUANTITÀ DI GAS AGGIUNTA","MASSA ADICIONADA":"MASSA AGGIUNTA","VOLUME FÍSICO DO CILINDRO":"VOLUME FISICO DEL CILINDRO","CONVERSÃO CIENTÍFICA PARA 20 °C":"CONVERSIONE SCIENTIFICA A 20 °C","CONVERSÃO ANP/IDEALIZADA (Z=1)":"CONVERSIONE ANP/IDEALIZZATA (Z=1)"},
"Deutsch":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"FORMELN UND PHYSIK DES GNV-BERECHNUNGSSYSTEMS","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"TEIL A — ANP-REFERENZBEDINGUNG","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"TEIL B — WISSENSCHAFTLICHES ECHTGASMODELL","PRESSÃO ABSOLUTA":"ABSOLUTDRUCK","TEMPERATURA ABSOLUTA":"ABSOLUTTEMPERATUR","FATOR DE COMPRESSIBILIDADE Z":"KOMPRIMIERBARKEITSFAKTOR Z","QUANTIDADE DE GÁS ADICIONADA":"HINZUGEFÜGTE GAS-MENGE","MASSA ADICIONADA":"HINZUGEFÜGTE MASSE","VOLUME FÍSICO DO CILINDRO":"PHYSIKALISCHES ZYLINDERVOLUMEN","CONVERSÃO CIENTÍFICA PARA 20 °C":"WISSENSCHAFTLICHE UMRECHNUNG AUF 20 °C","CONVERSÃO ANP/IDEALIZADA (Z=1)":"ANP/IDEALISIERTE UMRECHNUNG (Z=1)"},
"日本語":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"GNV計算システムの数式と物理","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"パートA — ANP基準条件","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"パートB — 実在気体の科学モデル","PRESSÃO ABSOLUTA":"絶対圧力","TEMPERATURA ABSOLUTA":"絶対温度","FATOR DE COMPRESSIBILIDADE Z":"圧縮係数Z","QUANTIDADE DE GÁS ADICIONADA":"追加されたガス量","MASSA ADICIONADA":"追加質量","VOLUME FÍSICO DO CILINDRO":"シリンダーの物理容積","CONVERSÃO CIENTÍFICA PARA 20 °C":"20 °Cへの科学的換算","CONVERSÃO ANP/IDEALIZADA (Z=1)":"ANP/理想化換算 (Z=1)"},
"中文":{"FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV":"GNV计算系统公式与物理","PARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP":"A部分 — ANP参考条件","PARTE B — MODELO CIENTÍFICO DE GÁS REAL":"B部分 — 真实气体科学模型","PRESSÃO ABSOLUTA":"绝对压力","TEMPERATURA ABSOLUTA":"绝对温度","FATOR DE COMPRESSIBILIDADE Z":"压缩因子Z","QUANTIDADE DE GÁS ADICIONADA":"增加的气体量","MASSA ADICIONADA":"增加的质量","VOLUME FÍSICO DO CILINDRO":"气瓶物理容积","CONVERSÃO CIENTÍFICA PARA 20 °C":"换算到20 °C","CONVERSÃO ANP/IDEALIZADA (Z=1)":"ANP/理想气体换算 (Z=1)"}}

# =============================================================================
# PARTE 02A - CONDIÇÕES DE REFERÊNCIA DA ANP
# =============================================================================

# A ANP publica volumes de gás natural equivalentes a 20 °C e 1,033 kgf/cm².
# A condição padrão de medição também é expressa como 0,101325 MPa a 20 °C.
# Para os cálculos internos, usamos a forma SI equivalente: 1,01325 bar.
TEMPERATURA_REFERENCIA_ANP_C = 20.0
PRESSAO_REFERENCIA_ANP_KGF_CM2 = 1.033
PRESSAO_REFERENCIA_ANP_BAR = 1.01325

# =============================================================================
# PARTE 02A1 - MODELO DE REFERÊNCIA ANP
# =============================================================================

def calcular_volume_anp_mesma_temperatura(
    capacidade_cilindro_l, pressao_inicial_bar, pressao_final_bar,
    altitude_m=0.0, pressao_referencia_bar=1.01325
):
    """Volume ANP/idealizado expresso na propria temperatura informada.

    Nesta forma nao ha conversao para 20 °C; por isso T cancela.
    A conversao para 20 °C e apresentada separadamente.
    """
    V = capacidade_cilindro_l / 1000.0
    Patm = calcular_pressao_atmosferica(altitude_m)
    Pi = pressao_inicial_bar + Patm
    Pf = pressao_final_bar + Patm
    if V <= 0 or Pf < Pi or pressao_referencia_bar <= 0:
        raise ValueError("Dados invalidos para o calculo ANP na temperatura informada.")
    return V * ((Pf - Pi) / pressao_referencia_bar)


def calcular_volume_anp_referencia(
    capacidade_cilindro_l,
    pressao_inicial_bar,
    pressao_final_bar,
    temperatura_ambiente_c,
    altitude_m=0.0
):
    """Estima o volume equivalente usando a condição de referência da ANP.

    Este cálculo NÃO é uma reprodução do algoritmo interno do dispenser.
    É uma estimativa baseada na condição de referência publicada pela ANP,
    no volume físico do cilindro e na aproximação de gás ideal (Z=1).

    A temperatura disponível ao usuário é a temperatura ambiente; ela não é
    tratada como medição da temperatura real do gás durante a compressão.
    """
    V = capacidade_cilindro_l / 1000.0
    T = temperatura_ambiente_c + 273.15
    Tref = TEMPERATURA_REFERENCIA_ANP_C + 273.15
    Patm = calcular_pressao_atmosferica(altitude_m)
    Pi = pressao_inicial_bar + Patm
    Pf = pressao_final_bar + Patm

    if V <= 0 or T <= 0 or Pf < Pi:
        raise ValueError("Dados inválidos para o cálculo pela referência ANP.")

    return V * ((Pf - Pi) / PRESSAO_REFERENCIA_ANP_BAR) * (Tref / T)


def calcular_volume_cientifico_gas_real(
    capacidade_cilindro_l,
    pressao_inicial_bar,
    pressao_final_bar,
    temperatura_ambiente_c,
    altitude_m,
    fator_z,
    massa_molar,
    temperatura_referencia_c=20.0,
    pressao_referencia_bar=1.01325
):
    """Modelo físico de gás real usando Z informado pelo usuário.

    Usa PV = Z n R T, pressão absoluta e temperatura absoluta.
    Como a temperatura real do gás no interior do cilindro durante o
    abastecimento não é medida pelo usuário, a temperatura ambiente é
    utilizada como aproximação e isso é explicitado no resultado.

    O modelo é mais completo que a aproximação ideal porque permite Z != 1,
    mas NÃO é AGA8/GERG-2008: para alta precisão metrológica seriam
    necessários composição do gás e propriedades termodinâmicas validadas.
    """
    V = capacidade_cilindro_l / 1000.0
    T = temperatura_ambiente_c + 273.15
    Tref = temperatura_referencia_c + 273.15
    Patm = calcular_pressao_atmosferica(altitude_m)
    Pi = pressao_inicial_bar + Patm
    Pf = pressao_final_bar + Patm

    if V <= 0 or T <= 0 or fator_z <= 0 or massa_molar <= 0:
        raise ValueError("Parâmetros físicos inválidos no modelo de gás real.")
    if Pf < Pi:
        raise ValueError("A pressão final deve ser maior ou igual à pressão inicial.")

    # n = PV/(ZRT)
    n_i = (Pi * 100000.0 * V) / (fator_z * R * T)
    n_f = (Pf * 100000.0 * V) / (fator_z * R * T)
    delta_n = max(0.0, n_f - n_i)
    massa = delta_n * massa_molar

    # Conversão dos mols para a condição de referência, assumindo Zref=1.
    P_ref = pressao_referencia_bar * 100000.0
    volume_ref = delta_n * R * Tref / P_ref

    return {
        "capacidade_m3": V,
        "pressao_atmosferica_bar": Patm,
        "pressao_inicial_absoluta_bar": Pi,
        "pressao_final_absoluta_bar": Pf,
        "temperatura_kelvin": T,
        "mols_iniciais": n_i,
        "mols_finais": n_f,
        "mols_adicionados": delta_n,
        "massa_adicionada_kg": massa,
        "volume_referencia_m3": volume_ref,
        "volume_referencia_litros": volume_ref * 1000.0,
        "fator_z": fator_z,
        "temperatura_referencia_c": temperatura_referencia_c,
        "pressao_referencia_bar": pressao_referencia_bar,
        "temperatura_eh_ambiente": True
    }


# =============================================================================
# PARTE 02A - CONVERSÃO NUMÉRICA PT-BR / INTERNACIONAL
# =============================================================================

def converter_numero(valor):
    """
    Aceita números nos formatos mais comuns no Brasil e internacional.

    Exemplos aceitos:
        24,5
        24.5
        1.234,56
        1,234.56
        1234
    """
    texto = str(valor).strip().replace(" ", "")

    if not texto:
        raise ValueError("Valor numérico vazio.")

    if "," in texto and "." in texto:
        # O último separador é tratado como separador decimal.
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")

    elif "," in texto:
        texto = texto.replace(",", ".")

    return float(texto)


def formatar_numero_br(valor, casas=2):
    """Formata número para exibição brasileira, sem alterar o valor salvo."""
    return f"{float(valor):.{casas}f}".replace(".", ",")


# =============================================================================
# PARTE 02B - VOLUME EQUIVALENTE NA CONDIÇÃO DE REFERÊNCIA
# =============================================================================

def calcular_volume_referencia_m3(
    mols,
    temperatura_referencia_c=20.0,
    pressao_referencia_bar=1.01325,
    fator_z_referencia=1.0
):
    """Converte mols calculados para m³ na condição de referência adotada.

    Referência padrão do aplicativo: 20 °C e 1,01325 bar, com Z=1.
    O resultado é explicitamente chamado de volume equivalente, não de
    volume físico ocupado pelo cilindro.
    """
    temperatura_k = temperatura_referencia_c + 273.15
    pressao_pa = pressao_referencia_bar * 100000.0

    if mols <= 0 or temperatura_k <= 0 or pressao_pa <= 0:
        return 0.0

    return (
        mols * R * temperatura_k * fator_z_referencia
        / pressao_pa
    )


def calcular_volume_equivalente_na_temperatura_m3(mols, temperatura_c, pressao_referencia_bar=1.01325, fator_z_referencia=1.0):
    """Volume equivalente dos mesmos mols na temperatura informada."""
    temperatura_k = temperatura_c + 273.15
    pressao_pa = pressao_referencia_bar * 100000.0
    if mols <= 0 or temperatura_k <= 0 or pressao_pa <= 0:
        return 0.0
    return mols * R * temperatura_k * fator_z_referencia / pressao_pa


def calcular_compressao_ideal_adiabatica(pressao_inicial_manometrica_bar, pressao_final_manometrica_bar, temperatura_inicial_c, altitude_m=0.0, k=1.294):
    """Modelo didático de compressão reversível adiabática; não é o enchimento real do cilindro."""
    if k <= 1.0: raise ValueError("O expoente k deve ser maior que 1.")
    if temperatura_inicial_c <= -273.15: raise ValueError("A temperatura inicial deve ser maior que -273,15 °C.")
    patm=calcular_pressao_atmosferica(altitude_m); p1=pressao_inicial_manometrica_bar+patm; p2=pressao_final_manometrica_bar+patm; t1=temperatura_inicial_c+273.15
    if p1 <= 0 or p2 <= 0 or p2 < p1: raise ValueError("As pressões devem ser positivas e a final maior ou igual à inicial.")
    t2=t1*(p2/p1)**((k-1.0)/k); vr=(p1/p2)**(1.0/k)
    return {"pressao_atmosferica_bar":patm,"pressao_inicial_absoluta_bar":p1,"pressao_final_absoluta_bar":p2,"temperatura_inicial_c":temperatura_inicial_c,"temperatura_final_c":t2-273.15,"temperatura_final_k":t2,"aumento_temperatura_c":t2-t1,"volume_relativo_final":vr,"reducao_volume_percentual":(1-vr)*100,"k":k}


def gerar_pontos_compressao_adiabatica(pressao_inicial_manometrica_bar, pressao_final_manometrica_bar, temperatura_inicial_c, altitude_m=0.0, k=1.294, pontos=40):
    patm=calcular_pressao_atmosferica(altitude_m); p1=pressao_inicial_manometrica_bar+patm; p2=pressao_final_manometrica_bar+patm; t1=temperatura_inicial_c+273.15
    if p1 <= 0 or p2 < p1 or k <= 1: raise ValueError("Parâmetros inválidos para o gráfico de compressão.")
    n=max(2,int(pontos)); dados=[]
    for i in range(n):
        f=i/(n-1); pa=p1+(p2-p1)*f; tk=t1*(pa/p1)**((k-1)/k); vr=(p1/pa)**(1/k); dados.append({"pressao_man_bar":pa-patm,"temperatura_c":tk-273.15,"volume_relativo":vr})
    return dados


# =============================================================================
# PARTE 02C - ANÁLISE FÍSICA DO ABASTECIMENTO
# =============================================================================

def calcular_comparacao_abastecimento(
    capacidade_cilindro_l,
    pressao_inicial_bar,
    pressao_final_bar,
    temperatura_c,
    altitude_m,
    fator_z,
    massa_molar,
    temperatura_referencia_c=20.0,
    pressao_referencia_bar=1.01325
):
    """Compatibilidade: retorna o modelo científico de gás real."""
    resultado = calcular_volume_cientifico_gas_real(
        capacidade_cilindro_l,
        pressao_inicial_bar,
        pressao_final_bar,
        temperatura_c,
        altitude_m,
        fator_z,
        massa_molar,
        temperatura_referencia_c,
        pressao_referencia_bar
    )
    resultado["delta_pressao_bar"] = max(0.0, pressao_final_bar - pressao_inicial_bar)
    resultado["volume_teorico_m3"] = resultado["volume_referencia_m3"]
    resultado["massa_adicionada_kg"] = resultado["massa_adicionada_kg"]
    resultado["temperatura_referencia_c"] = temperatura_referencia_c
    resultado["pressao_referencia_bar"] = pressao_referencia_bar
    return resultado


# =============================================================================
# PARTE 03 - FUNÇÃO
# Calcula a pressão atmosférica em função da altitude.
# =============================================================================

def calcular_pressao_atmosferica(altitude):

    """
    Calcula a pressão atmosférica.

    Entrada

        altitude em metros

    Saída

        pressão em bar

    Fórmula aproximada válida para pequenas altitudes.
    """

    # Equação barométrica simplificada

    pressao = PRESSAO_ATMOSFERICA_PADRAO * \
              (1 - 2.25577e-5 * altitude) ** 5.25588

    return pressao





# =============================================================================
# PARTE 04 - FUNÇÃO
# Converte Celsius para Kelvin
# =============================================================================

def celsius_para_kelvin(temperatura):

    return temperatura + 273.15

# =============================================================================
# FUNÇÃO
# Converte litros para metros cúbicos
# =============================================================================

def litros_para_m3(litros):

    return litros / 1000.0

# =============================================================================
# FUNÇÃO
# Calcula pressão absoluta
# =============================================================================

def calcular_pressao_absoluta(pressao_manometrica,
                              pressao_atmosferica):

    """
    A pressão mostrada no manômetro é relativa.

    Para utilizar a Lei dos Gases devemos usar pressão absoluta.

    Pabs = Pman + Patm
    """

    return pressao_manometrica + pressao_atmosferica

# =============================================================================
# FUNÇÃO
# Calcula quantidade de mols
# =============================================================================

def calcular_mols(volume_m3,
                  pressao_bar,
                  temperatura_kelvin,
                  fator_z):

    """
    Calcula a quantidade de matéria.

    Equação

    PV = ZnRT

    Quanto menor o fator Z,
    maior será a quantidade de gás para
    a mesma pressão.
    """

    pressao_pa = pressao_bar * 100000

    mols = (
        pressao_pa *
        volume_m3
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return mols

# =============================================================================
# FUNÇÃO
# Calcula massa do gás
# =============================================================================

def calcular_massa(mols):

    """
    Massa = mol × Massa molar
    """

    return mols * MASSA_MOLAR_GNV



# =============================================================================
# FUNÇÃO
# Calcula volume equivalente em condições normais
# =============================================================================

def calcular_volume_equivalente(massa):

    """
    m³ = massa / densidade
    """

    return massa / DENSIDADE_GNV




# =============================================================================
# FUNÇÃO PARA CALCULAR O ABASTECIMENTO
# =============================================================================

def calcular_abastecimento(
    volume_abastecido,
    preco_m3,
    odometro_anterior,
    odometro_atual,
    massa_gnv
):
    """
    Calcula todas as informações referentes ao abastecimento.

    Parâmetros
    ----------
    volume_abastecido : float
        Volume abastecido em m³.

    preco_m3 : float
        Preço do GNV por metro cúbico.

    odometro_anterior : float
        Quilometragem anterior.

    odometro_atual : float
        Quilometragem atual.

    massa_gnv : float
        Massa estimada do GNV.

    Retorno
    -------
    dict
        Dicionário contendo todos os resultados.
    """

    valor_total = volume_abastecido * preco_m3

    distancia = odometro_atual - odometro_anterior

    if distancia < 0:

        distancia = 0

    if volume_abastecido > 0:

        rendimento = distancia / volume_abastecido

    else:

        rendimento = 0

    if distancia > 0:

        custo_km = valor_total / distancia

    else:

        custo_km = 0

    if massa_gnv > 0:

        custo_kg = valor_total / massa_gnv

    else:

        custo_kg = 0

    custo_100km = custo_km * 100

    return {

        "valor_total": valor_total,

        "distancia": distancia,

        "rendimento": rendimento,

        "custo_km": custo_km,

        "custo_100km": custo_100km,

        "custo_kg": custo_kg

    }




# =============================================================================
# PARTE 21
# FUNÇÃO PARA CALCULAR O VALOR ENERGÉTICO DO GNV
# =============================================================================

def calcular_energia_gnv(
    volume_m3,
    pcs=39.5
):
    """
    Calcula a energia contida no GNV.

    volume_m3
        Volume do gás em metros cúbicos.

    pcs
        Poder Calorífico Superior
        MJ/m³

    Retorna

        energia_mj
        energia_kwh
    """

    energia_mj = volume_m3 * pcs

    energia_kwh = energia_mj / 3.6

    return {

        "MJ": energia_mj,

        "kWh": energia_kwh

    }


# =============================================================================
# PARTE 22
# FUNÇÃO PARA CALCULAR A MASSA DO GNV
# =============================================================================

def calcular_massa_gnv(
    volume_m3,
    densidade=0.717
):
    """
    Calcula a massa do GNV.

    densidade padrão em CNTP.

    Retorna a massa em quilogramas.
    """

    return volume_m3 * densidade


# =============================================================================
# PARTE 25
# CÁLCULO DA DENSIDADE REAL DO GNV
# =============================================================================

def calcular_densidade_real(

    massa,

    volume

):
    """
    Calcula a densidade real do gás.

    densidade = massa / volume

    Retorna kg/m³.
    """

    if volume <= 0:

        return 0.0

    return massa / volume

# =============================================================================
# PARTE 26
# PESO DO GNV
# =============================================================================

def calcular_peso(

    massa

):
    """
    Calcula o peso do gás.

    Retorna Newton.
    """

    G = 9.80665

    return massa * G


# =============================================================================
# PARTE 27
# CÁLCULO DA ENERGIA PELO PESO DO GNV
# =============================================================================

def calcular_energia_por_massa(
    massa_kg,
    pci=50.0
):
    """
    Calcula a energia disponível no GNV.

    Parâmetros
    ----------
    massa_kg : float
        Massa do GNV em quilogramas.

    pci : float
        Poder calorífico inferior em MJ/kg.
        Valor padrão aproximado para GNV.

    Retorno
    -------
    dict contendo:

        energia_mj
        energia_kwh
    """

    energia_mj = massa_kg * pci

    energia_kwh = energia_mj / 3.6

    return {

        "energia_mj": energia_mj,

        "energia_kwh": energia_kwh

    }

# =============================================================================
# PARTE 28
# CÁLCULO DO CONSUMO DO VEÍCULO
# =============================================================================

def calcular_consumo(

    distancia_km,

    volume_m3,

    massa_kg

):
    """
    Calcula diversos índices de consumo.

    Retorna um dicionário contendo:

        km_por_m3

        m3_por_100km

        km_por_kg

        kg_por_100km
    """

    resultado = {}

    if volume_m3 > 0:

        resultado["km_por_m3"] = (
            distancia_km /
            volume_m3
        )

        resultado["m3_por_100km"] = (
            volume_m3 /
            distancia_km
        ) * 100 if distancia_km > 0 else 0

    else:

        resultado["km_por_m3"] = 0

        resultado["m3_por_100km"] = 0

    if massa_kg > 0:

        resultado["km_por_kg"] = (
            distancia_km /
            massa_kg
        )

        resultado["kg_por_100km"] = (
            massa_kg /
            distancia_km
        ) * 100 if distancia_km > 0 else 0

    else:

        resultado["km_por_kg"] = 0

        resultado["kg_por_100km"] = 0

    return resultado


# =============================================================================
# PARTE 29
# CÁLCULO DA MASSA DO GNV PELA DENSIDADE
# =============================================================================

def calcular_massa_por_densidade(

    volume_m3,

    densidade

):
    """
    Calcula a massa do GNV.

    massa = volume × densidade

    volume_m3
        Volume em metros cúbicos

    densidade
        kg/m³

    Retorna
        Massa em quilogramas.
    """

    return volume_m3 * densidade


# =============================================================================
# PARTE 30
# DENSIDADE APARENTE
# =============================================================================

def calcular_densidade_aparente(

    massa,

    volume

):
    """
    Calcula a densidade aparente.

    densidade = massa / volume
    """

    if volume <= 0:

        return 0

    return massa / volume


# =============================================================================
# PARTE 31
# ENERGIA ESPECÍFICA
# =============================================================================

def calcular_energia_especifica(

    energia_kwh,

    massa

):
    """
    Calcula a energia específica.

    kWh/kg
    """

    if massa <= 0:

        return 0

    return energia_kwh / massa

# =============================================================================
# PARTE 32
# CÁLCULO DO VOLUME LIVRE DO CILINDRO
# =============================================================================

def calcular_volume_livre(

    volume_total_litros,

    percentual_ocupacao

):
    """
    Calcula o volume livre existente
    no cilindro.

    percentual_ocupacao

        0 a 100 %
    """

    if percentual_ocupacao < 0:

        percentual_ocupacao = 0

    if percentual_ocupacao > 100:

        percentual_ocupacao = 100

    volume_ocupado = (

        volume_total_litros

        *

        percentual_ocupacao

        /

        100

    )

    volume_livre = (

        volume_total_litros

        -

        volume_ocupado

    )

    return volume_livre


# =============================================================================
# PARTE 33
# PERCENTUAL DE ENCHIMENTO
# =============================================================================

def calcular_percentual_enchimento(

    pressao_bar,

    pressao_maxima=220

):
    """
    Retorna o percentual de enchimento
    do cilindro baseado na pressão.

    A pressão máxima pode ser alterada.
    """

    if pressao_maxima <= 0:

        return 0

    percentual = (

        pressao_bar

        /

        pressao_maxima

    ) * 100

    if percentual < 0:

        percentual = 0

    if percentual > 100:

        percentual = 100

    return percentual


# =============================================================================
# PARTE 34
# PRESSÃO RESTANTE
# =============================================================================

def calcular_pressao_restante(

    pressao_atual,

    pressao_minima=5

):
    """
    Calcula a pressão realmente
    utilizável do cilindro.
    """

    restante = (

        pressao_atual

        -

        pressao_minima

    )

    if restante < 0:

        restante = 0

    return restante


# =============================================================================
# PARTE 35
# CÁLCULO DA MASSA CONSUMIDA
# =============================================================================

def calcular_massa_consumida(

    massa_inicial,

    massa_final

):
    """
    Calcula a massa consumida.

    Retorna sempre um valor positivo.
    """

    massa = (

        massa_inicial

        -

        massa_final

    )

    if massa < 0:

        massa = 0

    return massa



# =============================================================================
# PARTE 36
# CÁLCULO DO VOLUME CONSUMIDO
# =============================================================================

def calcular_volume_consumido(

    volume_inicial,

    volume_final

):
    """
    Calcula o volume efetivamente consumido.
    """

    volume = (

        volume_inicial

        -

        volume_final

    )

    if volume < 0:

        volume = 0

    return volume


# =============================================================================
# PARTE 37
# PERCENTUAL CONSUMIDO
# =============================================================================

def calcular_percentual_consumido(

    valor_inicial,

    valor_final

):
    """
    Calcula o percentual consumido.

    Exemplo

    Inicial = 16 m³

    Final = 4 m³

    Consumo = 75%
    """

    if valor_inicial <= 0:

        return 0

    percentual = (

        (

            valor_inicial

            -

            valor_final

        )

        /

        valor_inicial

    ) * 100

    if percentual < 0:

        percentual = 0

    if percentual > 100:

        percentual = 100

    return percentual


# =============================================================================
# PARTE 38
# EFICIÊNCIA DO ABASTECIMENTO
# =============================================================================

def calcular_eficiencia(

    volume_teorico,

    volume_real

):
    """
    Calcula a eficiência do abastecimento.

    Quanto mais próximo de 100%,
    mais próximo do cálculo teórico.
    """

    if volume_teorico <= 0:

        return 0

    eficiencia = (

        volume_real

        /

        volume_teorico

    ) * 100

    return eficiencia


# =============================================================================
# PARTE 39
# PRESSÃO DE ENCHIMENTO
# =============================================================================

def calcular_pressao_enchimento(

    pressao_inicial,

    pressao_final

):
    """
    Calcula quanto a pressão aumentou
    durante o abastecimento.

    Retorna o aumento em bar.
    """

    aumento = (

        pressao_final

        -

        pressao_inicial

    )

    if aumento < 0:

        aumento = 0

    return aumento



# =============================================================================
# PARTE 40
# TAXA DE COMPRESSÃO
# =============================================================================

def calcular_taxa_compressao(

    pressao_absoluta,

    pressao_atmosferica

):
    """
    Calcula a taxa de compressão.

    Exemplo

    201 bar absolutos

    1,01325 bar atmosféricos

    Resultado ≈ 198
    """

    if pressao_atmosferica <= 0:

        return 0

    return (

        pressao_absoluta

        /

        pressao_atmosferica

    )



# =============================================================================
# PARTE 41
# ENERGIA POR METRO CÚBICO
# =============================================================================

def calcular_energia_m3(

    energia_total,

    volume_m3

):
    """
    Calcula a energia existente
    em cada metro cúbico.
    """

    if volume_m3 <= 0:

        return 0

    return (

        energia_total

        /

        volume_m3

    )


# =============================================================================
# PARTE 42
# FATOR DE UTILIZAÇÃO
# =============================================================================

def calcular_fator_utilizacao(

    volume_utilizado,

    volume_total

):
    """
    Calcula o percentual realmente
    utilizado do cilindro.
    """

    if volume_total <= 0:

        return 0

    fator = (

        volume_utilizado

        /

        volume_total

    ) * 100

    if fator < 0:

        fator = 0

    if fator > 100:

        fator = 100

    return fator


# =============================================================================
# PARTE 43
# BIBLIOTECA DE PROPRIEDADES DO GNV
# =============================================================================

def obter_propriedades_gnv():

    """
    Retorna um dicionário contendo as
    propriedades médias do GNV.

    Todos os valores poderão ser alterados
    futuramente conforme a composição do gás.
    """

    propriedades = {

        "massa_molar": 0.01812,      # kg/mol

        "densidade": 0.717,          # kg/m³ CNTP

        "pcs": 39.50,                # MJ/m³

        "pci": 35.80,                # MJ/m³

        "cp": 2.20,                  # kJ/(kg.K)

        "cv": 1.70,                  # kJ/(kg.K)

        "k": 1.294,                  # Cp/Cv

        "R_especifico": 518.3,       # J/(kg.K)

        "temperatura_critica": 190.6,# Kelvin

        "pressao_critica": 45.99     # bar

    }

    return propriedades


# =============================================================================
# PARTE 44
# MOSTRAR AS PROPRIEDADES
# =============================================================================

def mostrar_propriedades_gnv():

    propriedades = obter_propriedades_gnv()

    print()

    print("=" * 75)

    print("PROPRIEDADES DO GNV")

    print("=" * 75)

    print()

    for chave, valor in propriedades.items():

        print(f"{chave:<25} : {valor}")

    print()

    print("=" * 75)

    print()


# =============================================================================
# PARTE 45
# OBTER UMA PROPRIEDADE
# =============================================================================

def propriedade(nome):

    """
    Retorna uma propriedade específica.

    Exemplo

    pcs = propriedade("pcs")
    """

    propriedades = obter_propriedades_gnv()

    return propriedades.get(nome)


# =============================================================================
# PARTE 46
# ENERGIA UTILIZANDO A BIBLIOTECA
# =============================================================================

def calcular_energia_volume(

    volume_m3

):
    """
    Calcula automaticamente a energia
    utilizando o PCS armazenado na biblioteca.
    """

    pcs = propriedade(

        "pcs"

    )

    energia = (

        volume_m3

        *

        pcs

    )

    return energia



# =============================================================================
# PARTE 47
# COMPOSIÇÃO DO GNV
# =============================================================================

def obter_composicao_padrao():

    """
    Retorna uma composição típica do GNV.

    Os valores são em porcentagem (%).
    """

    composicao = {

        "CH4": 89.00,

        "C2H6": 6.00,

        "C3H8": 2.00,

        "CO2": 1.50,

        "N2": 1.00,

        "O2": 0.30,

        "Outros": 0.20

    }

    return composicao




# =============================================================================
# PARTE 48
# MOSTRAR COMPOSIÇÃO DO GNV
# =============================================================================

def mostrar_composicao():

    composicao = obter_composicao_padrao()

    print()

    print("=" * 75)

    print("COMPOSIÇÃO DO GNV")

    print("=" * 75)

    print()

    for gas, percentual in composicao.items():

        print(f"{gas:<10} : {percentual:6.2f} %")

    print()

    print("=" * 75)

    print()



# =============================================================================
# PARTE 49
# MASSA MOLAR DA MISTURA
# =============================================================================

def calcular_massa_molar_mistura():

    """
    Calcula a massa molar média
    da composição do GNV.
    """

    massas = {

        "CH4":16.043,

        "C2H6":30.070,

        "C3H8":44.097,

        "CO2":44.010,

        "N2":28.014,

        "O2":31.999,

        "Outros":20.000

    }

    composicao = obter_composicao_padrao()

    massa = 0

    for gas in composicao:

        massa += (

            composicao[gas]

            *

            massas[gas]

        )

    massa /= 100

    return massa / 1000


# =============================================================================
# PARTE 50
# RESUMO DAS PROPRIEDADES DO GNV
# =============================================================================

def resumo_gnv():

    print()

    print("=" * 75)

    print("RESUMO DO GNV")

    print("=" * 75)

    print()

    print(
        f"Massa molar média : "
        f"{calcular_massa_molar_mistura():.6f} kg/mol"
    )

    print(
        f"Densidade........ : "
        f"{propriedade('densidade'):.3f} kg/m³"
    )

    print(
        f"PCS............... : "
        f"{propriedade('pcs'):.2f} MJ/m³"
    )

    print(
        f"PCI............... : "
        f"{propriedade('pci'):.2f} MJ/m³"
    )

    print()

    print("=" * 75)

    print()



# =============================================================================
# PARTE 51
# TEMPERATURA REDUZIDA
# =============================================================================

def calcular_temperatura_reduzida(

    temperatura_kelvin,

    temperatura_critica

):
    """
    Calcula a temperatura reduzida.

    Tr = T / Tc
    """

    if temperatura_critica <= 0:

        return 0

    return (

        temperatura_kelvin

        /

        temperatura_critica

    )


# =============================================================================
# PARTE 52
# PRESSÃO REDUZIDA
# =============================================================================

def calcular_pressao_reduzida(

    pressao_bar,

    pressao_critica

):
    """
    Calcula a pressão reduzida.

    Pr = P / Pc
    """

    if pressao_critica <= 0:

        return 0

    return (

        pressao_bar

        /

        pressao_critica

    )



# =============================================================================
# PARTE 53
# CONSTANTE ESPECÍFICA DO GNV
# =============================================================================

def calcular_R_especifico(

    massa_molar

):
    """
    Calcula a constante específica.

    R = 8,314462618 / M
    """

    R_UNIVERSAL = 8.314462618

    if massa_molar <= 0:

        return 0

    return (

        R_UNIVERSAL

        /

        massa_molar

    )



# =============================================================================
# PARTE 54
# MASSA ESPECÍFICA
# =============================================================================

def calcular_massa_especifica(

    massa,

    volume

):
    """
    Calcula a massa específica.

    kg/m³
    """

    if volume <= 0:

        return 0

    return (

        massa

        /

        volume

    )


# =============================================================================
# PARTE 55
# FATOR Z APROXIMADO
# =============================================================================

def calcular_Z_aproximado(

    pressao_reduzida,

    temperatura_reduzida

):
    """
    Calcula uma aproximação do fator Z.

    Esta função será substituída futuramente
    pela Equação de Peng-Robinson ou
    Dranchuk-Abou-Kassem.

    Retorna um fator Z entre aproximadamente
    0,65 e 1,10.
    """

    Z = (

        1

        +

        0.08

        *

        pressao_reduzida

        /

        temperatura_reduzida

    )

    if Z < 0.65:

        Z = 0.65

    if Z > 1.10:

        Z = 1.10

    return Z


# =============================================================================
# PARTE 56
# CÁLCULO AUTOMÁTICO DE Tr E Pr
# =============================================================================

def calcular_propriedades_reduzidas(

    pressao_absoluta,

    temperatura_kelvin

):

    propriedades = obter_propriedades_gnv()

    Tr = calcular_temperatura_reduzida(

        temperatura_kelvin,

        propriedades["temperatura_critica"]

    )

    Pr = calcular_pressao_reduzida(

        pressao_absoluta,

        propriedades["pressao_critica"]

    )

    return {

        "Tr": Tr,

        "Pr": Pr

    }



# =============================================================================
# PARTE 57
# CÁLCULO COMPLETO DO FATOR Z
# =============================================================================

def calcular_Z(

    pressao_absoluta,

    temperatura_kelvin

):

    reduzidas = calcular_propriedades_reduzidas(

        pressao_absoluta,

        temperatura_kelvin

    )

    Z = calcular_Z_aproximado(

        reduzidas["Pr"],

        reduzidas["Tr"]

    )

    return {

        "Z": Z,

        "Pr": reduzidas["Pr"],

        "Tr": reduzidas["Tr"]

    }


# =============================================================================
# PARTE 83
# CÁLCULO DA MASSA DO GÁS REAL
# =============================================================================

def calcular_massa_gas_real(
    pressao_absoluta_bar,
    volume_m3,
    temperatura_kelvin,
    fator_z,
    massa_molar,
    densidade_informada_kg_m3=None
):
    """
    Calcula a massa do GNV utilizando
    a equação dos gases reais.

    Entradas
    --------
    pressao_absoluta_bar : Pressão absoluta (bar)
    volume_m3            : Volume (m³)
    temperatura_kelvin   : Temperatura (K)
    fator_z              : Fator de compressibilidade
    massa_molar          : Massa molar (kg/mol)

    Retorna
    -------
    Massa em kg
    """

    R = 8.314462618

    pressao_pa = pressao_absoluta_bar * 100000

    massa = (
        pressao_pa *
        volume_m3 *
        massa_molar
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return massa


# =============================================================================
# PARTE 84
# CÁLCULO DO NÚMERO DE MOLS
# =============================================================================

def calcular_numero_mols(
    massa_kg,
    massa_molar
):
    """
    Calcula a quantidade de matéria (mol).

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    massa_molar : Massa molar (kg/mol)

    Retorno
    -------
    Quantidade de mols.
    """

    if massa_molar <= 0:

        return 0.0

    numero_mols = (

        massa_kg

        /

        massa_molar

    )

    return numero_mols


# =============================================================================
# PARTE 85
# CÁLCULO DA DENSIDADE DO GÁS REAL
# =============================================================================

def calcular_densidade_gas_real(
    pressao_absoluta_bar,
    temperatura_kelvin,
    fator_z,
    massa_molar
):
    """
    Calcula a densidade do GNV utilizando
    a equação dos gases reais.

    Entradas
    --------
    pressao_absoluta_bar : Pressão absoluta (bar)

    temperatura_kelvin : Temperatura (K)

    fator_z : Fator de compressibilidade

    massa_molar : Massa molar (kg/mol)

    Retorna
    -------
    Densidade em kg/m³
    """

    R = 8.314462618

    pressao_pa = pressao_absoluta_bar * 100000

    if (
        temperatura_kelvin <= 0
        or fator_z <= 0
        or massa_molar <= 0
    ):

        return 0.0

    densidade = (
        pressao_pa *
        massa_molar
    ) / (
        fator_z *
        R *
        temperatura_kelvin
    )

    return densidade


# =============================================================================
# PARTE 86
# CÁLCULO DO VOLUME ESPECÍFICO DO GÁS REAL
# =============================================================================

def calcular_volume_especifico_gas_real(
    massa_kg,
    volume_m3
):
    """
    Calcula o volume específico do GNV.

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    volume_m3 : Volume ocupado pelo gás (m³)

    Retorna
    -------
    Volume específico em m³/kg
    """

    if massa_kg <= 0:

        return 0.0

    volume_especifico = (

        volume_m3

        /

        massa_kg

    )

    return volume_especifico



# =============================================================================
# PARTE 87
# CÁLCULO DO VOLUME REAL DO GNV
# =============================================================================

def calcular_volume_real_gnv(
    massa_kg,
    densidade_kg_m3
):
    """
    Calcula o volume realmente ocupado
    pelo GNV.

    Entradas
    --------
    massa_kg : Massa do gás (kg)

    densidade_kg_m3 : Densidade (kg/m³)

    Retorna
    -------
    Volume em m³
    """

    if densidade_kg_m3 <= 0:

        return 0.0

    volume = (

        massa_kg

        /

        densidade_kg_m3

    )

    return volume




# =============================================================================
# PARTE 88
# CALCULA A QUANTIDADE DE GNV NO CILINDRO
# =============================================================================

def calcular_quantidade_gnv(
    volume_cilindro_litros,
    pressao_bar,
    temperatura_c,
    altitude_m,
    fator_z,
    massa_molar,
    densidade_informada_kg_m3=None
):
    """
    Calcula automaticamente diversas propriedades
    do GNV armazenado no cilindro.

    Retorna um dicionário contendo todos os resultados.
    """

    # ---------------------------------------------------------
    # Conversões
    # ---------------------------------------------------------

    volume_m3 = volume_cilindro_litros / 1000.0

    temperatura_k = temperatura_c + 273.15

    pressao_atm = calcular_pressao_atmosferica(
        altitude_m
    )

    pressao_abs = pressao_bar + pressao_atm

    # ---------------------------------------------------------
    # Massa
    # ---------------------------------------------------------

    massa = calcular_massa_gas_real(
        pressao_abs,
        volume_m3,
        temperatura_k,
        fator_z,
        massa_molar
    )

    # ---------------------------------------------------------
    # Densidade
    # ---------------------------------------------------------

    densidade = calcular_densidade_gas_real(
        pressao_abs,
        temperatura_k,
        fator_z,
        massa_molar
    )

    # ---------------------------------------------------------
    # Número de mols
    # ---------------------------------------------------------

    mols = calcular_numero_mols(
        massa,
        massa_molar
    )

    # ---------------------------------------------------------
    # Volume específico
    # ---------------------------------------------------------

    volume_especifico = calcular_volume_especifico_gas_real(
        massa,
        volume_m3
    )

    # ---------------------------------------------------------
    # Volume real
    # ---------------------------------------------------------

    volume_real = calcular_volume_real_gnv(
        massa,
        densidade
    )

    # ANP / idealizado:
    # Ambos os resultados abaixo são apresentados na referência ANP de 20 °C.
    #
    # 1) "T informada" = quantidade estimada usando a temperatura informada
    #    no cilindro e convertida para a referência ANP de 20 °C.
    # 2) "20 °C" = cenário recalculado supondo o cilindro a 20 °C.
    #
    # Isso evita a interpretação errada de que "ANP mesma T" seja a
    # quantidade padronizada de GNV. A expressão V*ΔP/Pref, sem Tref/T,
    # é volume equivalente na própria temperatura e NÃO é o valor ANP
    # padronizado a 20 °C.
    volume_anp_t_informada = calcular_volume_anp_referencia(
        volume_cilindro_litros, 0.0, pressao_bar, temperatura_c, altitude_m
    )
    volume_anp_20c = calcular_volume_anp_referencia(
        volume_cilindro_litros, 0.0, pressao_bar, 20.0, altitude_m
    )

    # Modelo físico com Z informado:
    # "T informada" é calculado pelo PV=ZnRT usando a temperatura informada
    # e convertido para 20 °C.
    fisico_z_t_informada_20c = calcular_volume_referencia_m3(
        mols, 20.0, 1.01325, 1.0
    )
    # "20 °C" recalcula a quantidade de matéria como se o cilindro estivesse
    # a 20 °C, mantendo o mesmo Z informado.
    temperatura_20_k = 293.15
    # Como estamos calculando o GAS ADICIONADO ao abastecer de 0 bar
    # manométrico até a pressão final, a variação de pressão é ΔP = Pfinal - Pinicial.
    # A pressão atmosférica cancela na diferença de pressão absoluta.
    n_20c = (
        pressao_bar * 100000.0 * volume_m3
        / (fator_z * R * temperatura_20_k)
    )
    fisico_z_20c = calcular_volume_referencia_m3(
        n_20c, 20.0, 1.01325, 1.0
    )

    # ---------------------------------------------------------
    # Retorno
    # ---------------------------------------------------------

    return {

        "volume_cilindro_l": volume_cilindro_litros,

        "volume_cilindro_m3": volume_m3,

        "temperatura_c": temperatura_c,

        "temperatura_k": temperatura_k,

        "pressao_bar": pressao_bar,

        "pressao_atm": pressao_atm,

        "pressao_absoluta": pressao_abs,

        "massa": massa,

        "densidade": densidade,

        "mols": mols,

        "volume_especifico": volume_especifico,

        "volume_real": volume_real,

        # Modelo físico Z — T informada, expresso na referência de 20 °C.
        "volume_equivalente_m3_temperatura_informada": fisico_z_t_informada_20c,

        # Modelo físico Z — cenário recalculado a 20 °C.
        "volume_equivalente_m3_20c": fisico_z_20c,

        # ANP/idealizado — ambos já estão expressos na referência ANP de 20 °C.
        "volume_anp_ideal_m3_temperatura_informada": volume_anp_t_informada,
        "volume_anp_ideal_m3_20c": volume_anp_20c,

        "densidade_informada_kg_m3": densidade_informada_kg_m3,

        "massa_referencia_informada_kg": (
            calcular_volume_referencia_m3(
                mols,
                20.0,
                1.01325,
                1.0
            ) * densidade_informada_kg_m3
            if densidade_informada_kg_m3 is not None else None
        ),

        "diferenca_massa_referencia_kg": (
            (
                calcular_volume_referencia_m3(
                    mols,
                    20.0,
                    1.01325,
                    1.0
                ) * densidade_informada_kg_m3
            ) - massa
            if densidade_informada_kg_m3 is not None else None
        ),

        "volume_equivalente_litros_20c": (
            calcular_volume_referencia_m3(
                mols,
                20.0,
                1.01325,
                1.0
            ) * 1000.0
        ),

        # O circuito acessório é separado do cilindro.
        # Os valores abaixo são referências de engenharia do estudo técnico.
        "volume_circuito_gnv": estimar_faixas_volume_circuito_gnv()

    }


# =============================================================================
# PARTE 88A
# VOLUME GEOMÉTRICO DO CIRCUITO ACESSÓRIO DE GNV
# =============================================================================

def calcular_volume_circuito_gnv(
    diametro_interno_alta_mm=3.6,
    comprimento_alta_m=4.0,
    volume_valvulas_ml=3.0,
    volume_redutor_ml=32.0,
    volume_baixa_ml=0.0
):
    """
    Estima o volume geométrico interno do circuito acessório de GNV,
    EXCLUINDO RIGOROSAMENTE O CILINDRO/RESERVATÓRIO.

    O cálculo é uma estimativa geométrica de engenharia baseada no
    RELATÓRIO TÉCNICO fornecido para este projeto. Não substitui a
    medição física do veículo nem o desenho técnico do fabricante.

    Componentes considerados:
      1. Tubulação de alta pressão;
      2. Corpo/canais internos das válvulas de serviço;
      3. Câmaras de passagem de gás do redutor;
      4. Linha de baixa pressão, filtro e flauta/injetores.

    Para a tubulação cilíndrica:
        V = pi * r² * L

    Retorna volumes em mL e L.
    """
    d_m = float(diametro_interno_alta_mm) / 1000.0
    L_m = float(comprimento_alta_m)
    r_m = d_m / 2.0

    volume_tubo_m3 = math.pi * (r_m ** 2) * L_m
    volume_tubo_ml = volume_tubo_m3 * 1_000_000.0

    valvulas = float(volume_valvulas_ml)
    redutor = float(volume_redutor_ml)
    baixa = float(volume_baixa_ml)
    total_ml = volume_tubo_ml + valvulas + redutor + baixa

    return {
        "diametro_interno_mm": float(diametro_interno_alta_mm),
        "comprimento_alta_m": L_m,
        "raio_interno_mm": float(diametro_interno_alta_mm) / 2.0,
        "volume_tubulacao_ml": volume_tubo_ml,
        "volume_valvulas_ml": valvulas,
        "volume_redutor_ml": redutor,
        "volume_baixa_ml": baixa,
        "volume_total_ml": total_ml,
        "volume_total_l": total_ml / 1000.0,
    }


def estimar_faixas_volume_circuito_gnv():
    """
    Consolida exatamente a tabela de faixas do RELATÓRIO TÉCNICO
    fornecido para este projeto.

    Os valores são referências/estimativas de engenharia do estudo.
    A função calcular_volume_circuito_gnv() permite substituir essas
    referências por dimensões reais de um veículo específico.
    """
    def faixa(tubo, valvulas, redutor, baixa):
        total = float(tubo) + float(valvulas) + float(redutor) + float(baixa)
        return {
            "volume_tubulacao_ml": float(tubo),
            "volume_valvulas_ml": float(valvulas),
            "volume_redutor_ml": float(redutor),
            "volume_baixa_ml": float(baixa),
            "volume_total_ml": total,
            "volume_total_l": total / 1000.0,
        }

    return {
        "minimo": faixa(40, 3, 32, 0),
        "medio": faixa(55, 4, 46, 25),
        "maximo": faixa(70, 5, 65, 140),
    }


# =============================================================================
# PARTE 89
# COMPARAÇÃO DO ABASTECIMENTO
# =============================================================================

def comparar_abastecimento(
    volume_informado_m3,
    volume_calculado_m3,
    massa_informada_kg=0.0,
    massa_calculada_kg=0.0
):
    """
    Compara os valores informados pelo posto
    com os valores calculados pelo programa.

    Retorna um dicionário contendo as diferenças.
    """

    diferenca_volume = (
        volume_informado_m3 -
        volume_calculado_m3
    )

    if volume_calculado_m3 > 0:

        erro_volume = (
            diferenca_volume /
            volume_calculado_m3
        ) * 100

    else:

        erro_volume = 0.0

    diferenca_massa = (
        massa_informada_kg -
        massa_calculada_kg
    )

    if massa_calculada_kg > 0:

        erro_massa = (
            diferenca_massa /
            massa_calculada_kg
        ) * 100

    else:

        erro_massa = 0.0

    return {

        "volume_informado": volume_informado_m3,

        "volume_calculado": volume_calculado_m3,

        "diferenca_volume": diferenca_volume,

        "erro_volume_percentual": erro_volume,

        "massa_informada": massa_informada_kg,

        "massa_calculada": massa_calculada_kg,

        "diferenca_massa": diferenca_massa,

        "erro_massa_percentual": erro_massa

    }


# =============================================================================
# PARTE 90
# CÁLCULO DA ENERGIA TOTAL DO GNV
# =============================================================================

def calcular_energia_gnv_por_massa(
    massa_kg,
    pcs_mj_kg=50.02,
    pci_mj_kg=45.00
):
    """
    Calcula a energia contida no GNV.

    Entradas
    --------
    massa_kg : Massa do GNV (kg)

    pcs_mj_kg : Poder Calorífico Superior (MJ/kg)

    pci_mj_kg : Poder Calorífico Inferior (MJ/kg)

    Retorna
    -------
    Dicionário contendo a energia.
    """

    energia_pcs = massa_kg * pcs_mj_kg

    energia_pci = massa_kg * pci_mj_kg

    energia_pcs_kwh = energia_pcs / 3.6

    energia_pci_kwh = energia_pci / 3.6

    return {

        "PCS_MJ": energia_pcs,

        "PCI_MJ": energia_pci,

        "PCS_kWh": energia_pcs_kwh,

        "PCI_kWh": energia_pci_kwh

    }



# =============================================================================
# PARTE 91
# CÁLCULO COMPLETO DO CONSUMO DO VEÍCULO
# =============================================================================

def calcular_consumo_veiculo(
    distancia_km,
    volume_m3,
    massa_kg,
    valor_abastecimento,
    energia_mj
):
    """
    Calcula diversos indicadores de consumo.

    Entradas
    --------
    distancia_km : Distância percorrida (km)

    volume_m3 : Volume abastecido (m³)

    massa_kg : Massa abastecida (kg)

    valor_abastecimento : Valor pago (R$)

    energia_mj : Energia do abastecimento (MJ)

    Retorna
    -------
    Dicionário contendo os indicadores.
    """

    if distancia_km <= 0:

        return None

    if volume_m3 > 0:

        km_m3 = distancia_km / volume_m3

    else:

        km_m3 = 0.0

    if massa_kg > 0:

        km_kg = distancia_km / massa_kg

    else:

        km_kg = 0.0

    if energia_mj > 0:

        km_mj = distancia_km / energia_mj

        mj_km = energia_mj / distancia_km

    else:

        km_mj = 0.0

        mj_km = 0.0

    custo_km = valor_abastecimento / distancia_km

    custo_m3 = 0.0

    if volume_m3 > 0:

        custo_m3 = valor_abastecimento / volume_m3

    custo_kg = 0.0

    if massa_kg > 0:

        custo_kg = valor_abastecimento / massa_kg

    custo_mj = 0.0

    if energia_mj > 0:

        custo_mj = valor_abastecimento / energia_mj

    return {

        "km_m3": km_m3,

        "km_kg": km_kg,

        "km_MJ": km_mj,

        "MJ_km": mj_km,

        "custo_km": custo_km,

        "custo_m3": custo_m3,

        "custo_kg": custo_kg,

        "custo_MJ": custo_mj

    }


# =============================================================================
# PARTE 58
# CLASSE GNV
# =============================================================================

class GNV:

    """
    Classe principal para armazenar
    todos os dados referentes ao GNV.
    """

    def __init__(

        self,

        volume_m3,

        pressao_bar,

        temperatura_c,

        altitude

    ):

        self.volume_m3 = volume_m3

        self.pressao_bar = pressao_bar

        self.temperatura_c = temperatura_c

        self.altitude = altitude

        self.temperatura_k = (

            temperatura_c

            +

            273.15

        )


# =============================================================================
# PARTE 59
# PRESSÃO ATMOSFÉRICA
# =============================================================================

    def pressao_atmosferica(self):

        return calcular_pressao_atmosferica(

            self.altitude

        )


# =============================================================================
# PARTE 60
# PRESSÃO ABSOLUTA
# =============================================================================

    def pressao_absoluta(self):

        return (

            self.pressao_bar

            +

            self.pressao_atmosferica()

        )


# =============================================================================
# PARTE 61
# FATOR Z
# =============================================================================

    def fator_Z(self):

        resultado = calcular_Z(

            self.pressao_absoluta(),

            self.temperatura_k

        )

        return resultado


# =============================================================================
# PARTE 62
# MASSA DO GNV
# =============================================================================

    def massa(self):

        densidade = propriedade(

            "densidade"

        )

        return calcular_massa_por_densidade(

            self.volume_m3,

            densidade

        )


# =============================================================================
# PARTE 63
# ENERGIA DO GNV
# =============================================================================

    def energia(self):

        pcs = propriedade(

            "pcs"

        )

        energia_mj = (

            self.volume_m3

            *

            pcs

        )

        energia_kwh = (

            energia_mj

            /

            3.6

        )

        return {

            "MJ": energia_mj,

            "kWh": energia_kwh

        }


# =============================================================================
# PARTE 64
# AUTONOMIA
# =============================================================================

    def autonomia(

        self,

        consumo_km_m3

    ):

        return (

            self.volume_m3

            *

            consumo_km_m3

        )


# =============================================================================
# PARTE 65
# DENSIDADE
# =============================================================================

    def densidade(self):

        return propriedade(

            "densidade"

        )


# =============================================================================
# PARTE 66
# MASSA MOLAR
# =============================================================================

    def massa_molar(self):

        return calcular_massa_molar_mistura()


# =============================================================================
# PARTE 67
# RELATÓRIO COMPLETO
# =============================================================================

    def relatorio(self, consumo_km_m3=15.0):
        """
        Exibe um relatório completo do estado do GNV.

        consumo_km_m3:
            Rendimento médio do veículo em km por m³.
        """

        pressao_atm = self.pressao_atmosferica()

        pressao_abs = self.pressao_absoluta()

        fator = self.fator_Z()

        massa = self.massa()

        energia = self.energia()

        autonomia = self.autonomia(consumo_km_m3)

        print()

        print("=" * 75)

        print("RELATÓRIO COMPLETO DO GNV")

        print("=" * 75)

        print(f"Volume................: {self.volume_m3:.3f} m³")

        print(f"Pressão...............: {self.pressao_bar:.2f} bar")

        print(f"Pressão Atmosférica...: {pressao_atm:.4f} bar")

        print(f"Pressão Absoluta......: {pressao_abs:.4f} bar")

        print(f"Temperatura...........: {self.temperatura_c:.2f} °C")

        print(f"Temperatura...........: {self.temperatura_k:.2f} K")

        print(f"Pressão Reduzida......: {fator['Pr']:.4f}")

        print(f"Temperatura Reduzida..: {fator['Tr']:.4f}")

        print(f"Fator Z...............: {fator['Z']:.5f}")

        print(f"Massa................: {massa:.3f} kg")

        print(f"Energia..............: {energia['MJ']:.2f} MJ")

        print(f"Energia..............: {energia['kWh']:.2f} kWh")

        print(f"Autonomia Estimada...: {autonomia:.2f} km")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 68
# DICIONÁRIO COMPLETO
# =============================================================================

    def dados(self, consumo_km_m3=15.0):
        """
        Retorna todos os cálculos em formato
        de dicionário.

        Essa função será utilizada futuramente
        para:

            Excel

            SQLite

            PDF

            API

            Interface Gráfica
        """

        pressao_atm = self.pressao_atmosferica()

        pressao_abs = self.pressao_absoluta()

        fator = self.fator_Z()

        energia = self.energia()

        return {

            "Volume_m3": self.volume_m3,

            "Pressao_bar": self.pressao_bar,

            "Pressao_Atmosferica": pressao_atm,

            "Pressao_Absoluta": pressao_abs,

            "Temperatura_C": self.temperatura_c,

            "Temperatura_K": self.temperatura_k,

            "Pressao_Reduzida": fator["Pr"],

            "Temperatura_Reduzida": fator["Tr"],

            "Fator_Z": fator["Z"],

            "Massa": self.massa(),

            "Energia_MJ": energia["MJ"],

            "Energia_kWh": energia["kWh"],

            "Autonomia": self.autonomia(
                consumo_km_m3
            )

        }


# =============================================================================
# PARTE 69
# MOSTRAR DADOS
# =============================================================================

    def mostrar_dados(self):

        dados = self.dados()

        print()

        print("=" * 75)

        print("DADOS DA CLASSE GNV")

        print("=" * 75)

        print()

        for chave, valor in dados.items():

            print(

                f"{chave:<30}: {valor}"

            )

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 70
# EXPORTAR DADOS
# =============================================================================

    def exportar(self):

        return self.dados()

# =============================================================================
# PARTE 71
# EXPORTAÇÃO JSON
# =============================================================================

    def json(self):

        import json

        return json.dumps(

            self.dados(),

            indent=4,

            ensure_ascii=False

        )


# =============================================================================
# PARTE 72
# SIMULAÇÃO DA PRESSÃO EM OUTRA TEMPERATURA
# =============================================================================

    def simular_pressao(

        self,

        nova_temperatura_c

    ):
        """
        Simula a nova pressão considerando
        que a quantidade de gás permanece
        constante.

        Utiliza uma aproximação baseada
        na Lei dos Gases Ideais.

        Retorna a pressão em bar.
        """

        temperatura1 = self.temperatura_k

        temperatura2 = (

            nova_temperatura_c

            +

            273.15

        )

        pressao_absoluta = (

            self.pressao_absoluta()

            *

            temperatura2

            /

            temperatura1

        )

        pressao_manometrica = (

            pressao_absoluta

            -

            self.pressao_atmosferica()

        )

        return pressao_manometrica



# =============================================================================
# PARTE 73
# RESFRIAMENTO
# =============================================================================

    def simular_resfriamento(

        self,

        temperatura_final

    ):

        return self.simular_pressao(

            temperatura_final

        )


# =============================================================================
# PARTE 74
# AQUECIMENTO
# =============================================================================

    def simular_aquecimento(

        self,

        temperatura_final

    ):

        return self.simular_pressao(

            temperatura_final

        )


# =============================================================================
# PARTE 75
# TABELA DE PRESSÃO x TEMPERATURA
# =============================================================================

    def tabela_temperatura(

        self,

        temperatura_inicial=-20,

        temperatura_final=60,

        passo=2

    ):

        print()

        print("=" * 75)

        print("SIMULAÇÃO DE PRESSÃO")

        print("=" * 75)

        print()

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            pressao = self.simular_pressao(

                temperatura

            )

            print(

                f"{temperatura:6.1f} °C   "

                f"{pressao:8.2f} bar"

            )

            temperatura += passo

        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 76
# SIMULAÇÃO DE ALTITUDE
# =============================================================================

    def simular_altitude(

        self,

        nova_altitude

    ):
        """
        Simula a pressão atmosférica para
        outra altitude.
        """

        return calcular_pressao_atmosferica(

            nova_altitude

        )


# =============================================================================
# PARTE 77
# PRESSÃO ABSOLUTA EM OUTRA ALTITUDE
# =============================================================================

    def pressao_absoluta_altitude(

        self,

        nova_altitude

    ):

        pressao_atm = calcular_pressao_atmosferica(

            nova_altitude

        )

        return (

            self.pressao_bar

            +

            pressao_atm

        )


# =============================================================================
# PARTE 78
# SIMULAÇÃO COMPLETA
# =============================================================================

    def simular(

        self,

        temperatura,

        altitude

    ):

        pressao = self.simular_pressao(

            temperatura

        )

        pressao_atm = calcular_pressao_atmosferica(

            altitude

        )

        pressao_abs = (

            pressao

            +

            pressao_atm

        )

        return {

            "temperatura": temperatura,

            "altitude": altitude,

            "pressao_atmosferica": pressao_atm,

            "pressao_manometrica": pressao,

            "pressao_absoluta": pressao_abs

        }


# =============================================================================
# PARTE 79
# RELATÓRIO DA SIMULAÇÃO
# =============================================================================

    def relatorio_simulacao(

        self,

        temperatura,

        altitude

    ):

        dados = self.simular(

            temperatura,

            altitude

        )

        print()

        print("=" * 75)

        print("SIMULAÇÃO")

        print("=" * 75)

        print()

        print(f"Temperatura...........: {dados['temperatura']:.2f} °C")

        print(f"Altitude..............: {dados['altitude']:.2f} m")

        print(f"Pressão Atmosférica...: {dados['pressao_atmosferica']:.4f} bar")

        print(f"Pressão Manométrica...: {dados['pressao_manometrica']:.4f} bar")

        print(f"Pressão Absoluta......: {dados['pressao_absoluta']:.4f} bar")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 80
# SIMULAÇÃO EM FAIXA
# =============================================================================

    def simular_faixa_temperatura(

        self,

        temperatura_inicial,

        temperatura_final,

        incremento=2

    ):

        resultados = []

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            resultados.append(

                self.simular(

                    temperatura,

                    self.altitude

                )

            )

            temperatura += incremento

        return resultados



# =============================================================================
# PARTE 81
# GERAÇÃO DA TABELA PRESSÃO x TEMPERATURA
# =============================================================================

    def gerar_tabela_pressao_temperatura(

        self,

        temperatura_inicial=-20,

        temperatura_final=60,

        passo_temperatura=2,

        pressao_inicial=0,

        pressao_final=300,

        passo_pressao=5

    ):
        """
        Gera uma tabela contendo todas as
        combinações entre temperatura e pressão.

        Retorna uma lista de dicionários.
        """

        tabela = []

        temperatura = temperatura_inicial

        while temperatura <= temperatura_final:

            pressao = pressao_inicial

            while pressao <= pressao_final:

                """linha = {

                    "Temperatura": temperatura,

                    "Pressao": pressao

                }

                tabela.append(

                    linha

                )"""


# ==============================================================
# CONVERSÕES
# ==============================================================

                temperatura_kelvin = (
                    temperatura
                    +
                    273.15
                )

                pressao_atm = calcular_pressao_atmosferica(
                    self.altitude
                )

                pressao_absoluta = (
                    pressao
                    +
                    pressao_atm
                )


# ==============================================================
# FATOR Z
# ==============================================================

                resultado_z = calcular_Z(
                    pressao_absoluta,
                    temperatura_kelvin
                )


# ==============================================================
# VOLUME NAS CNTP
# ==============================================================

                # Volume equivalente nas condições de referência (CNTP)
                # usando a equação dos gases reais: P V = Z n R T.
                volume_cntp = (
                    self.volume_m3
                    * pressao_absoluta
                    * TEMPERATURA_PADRAO
                    / (
                        resultado_z["Z"]
                        * PRESSAO_ATMOSFERICA_PADRAO
                        * temperatura_kelvin
                    )
                )


# ==============================================================
# MASSA
# ==============================================================

                massa = calcular_massa_por_densidade(
                    volume_cntp,
                    propriedade(
                        "densidade"
                    )
                )


# ==============================================================
# ENERGIA
# ==============================================================

                energia = calcular_energia_volume(
                    volume_cntp
                )


# ==============================================================
# ENERGIA EM kWh
# ==============================================================

                energia_kwh = (
                    energia
                    /
                    3.6
                )


# ==============================================================
# DENSIDADE APARENTE
# ==============================================================

                if volume_cntp > 0:

                    densidade_aparente = (
                        massa
                        /
                        volume_cntp
                    )

                else:

                    densidade_aparente = 0.0


# ==============================================================
# DENSIDADE
# ==============================================================

                densidade = calcular_massa_especifica(
                    massa,
                    volume_cntp
                )


# ==============================================================
# CONSTANTE ESPECÍFICA
# ==============================================================

                R_especifico = calcular_R_especifico(
                    calcular_massa_molar_mistura()
                )


# ==============================================================
# LINHA DA TABELA
# ==============================================================

                linha = {

                    "Temperatura_C": temperatura,

                    "Temperatura_K": temperatura_kelvin,

                    "Pressao_bar": pressao,

                    "Pressao_Absoluta": pressao_absoluta,

                    "Altitude": self.altitude,

                    "Pressao_Atmosferica": pressao_atm,

                    "Pr": resultado_z["Pr"],

                    "Tr": resultado_z["Tr"],

                    "Fator_Z": resultado_z["Z"],

                    "Volume_CNTP": volume_cntp,

                    "Massa": massa,

                    "Energia_MJ": energia,

                    "Energia_kWh": energia_kwh,

                    "Densidade": densidade,

                    "Densidade_Aparente": densidade_aparente,

                    "R_especifico": R_especifico

                }


                tabela.append(

                    linha

                )


                pressao += passo_pressao


            temperatura += passo_temperatura


        return tabela



# =============================================================================
# PARTE 82
# MOSTRAR TABELA
# =============================================================================

    def mostrar_tabela(

        self

    ):

        tabela = self.gerar_tabela_pressao_temperatura()

        print()

        print("=" * 75)

        print("TABELA PRESSÃO x TEMPERATURA")

        print("=" * 75)

        print()

        for linha in tabela:

            print(
		f"{linha['Temperatura_C']:6.1f} °C"
		f"    "
		f"{linha['Pressao_bar']:6.1f} bar"
	)


        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 92
# CLASSE ABASTECIMENTO
# =============================================================================

class Abastecimento:

    """
    Armazena todas as informações de um abastecimento.
    """

    def __init__(

        self,

        data,

        posto,

        cidade,

        odometro,

        volume_m3,

        preco_m3,

        temperatura,

        pressao,

        altitude,

        observacoes="",

        capacidade_cilindro_l=0.0,

        pressao_inicial=0.0,

        pressao_final=0.0,

        densidade_informada_kg_m3=0.0,

        metragem_teorica_m3=0.0,

        metragem_anp_m3=0.0,

        metragem_cientifica_m3=0.0,

        status_bomba="NAO_CLASSIFICADO",
        divergencia_anp_m3=0.0,
        divergencia_fisica_m3=0.0,
        divergencia_anp_pct=0.0,
        divergencia_fisica_pct=0.0

    ):

        self.data = data

        self.posto = posto

        self.cidade = cidade

        self.odometro = odometro

        self.volume_m3 = volume_m3

        self.preco_m3 = preco_m3

        self.temperatura = temperatura

        self.pressao = pressao

        self.altitude = altitude

        self.observacoes = observacoes

        self.capacidade_cilindro_l = capacidade_cilindro_l

        self.pressao_inicial = pressao_inicial

        self.pressao_final = pressao_final

        self.densidade_informada_kg_m3 = densidade_informada_kg_m3
        self.metragem_teorica_m3 = metragem_teorica_m3
        self.metragem_anp_m3 = metragem_anp_m3
        self.metragem_cientifica_m3 = metragem_cientifica_m3
        self.status_bomba = status_bomba
        self.divergencia_anp_m3 = divergencia_anp_m3
        self.divergencia_fisica_m3 = divergencia_fisica_m3
        self.divergencia_anp_pct = divergencia_anp_pct
        self.divergencia_fisica_pct = divergencia_fisica_pct

        self.valor_total = (

            volume_m3 *

            preco_m3

        )


# =============================================================================
# PARTE 93
# RESUMO DO ABASTECIMENTO
# =============================================================================

    def resumo(self):

        print()

        print("=" * 75)

        print("DADOS DO ABASTECIMENTO")

        print("=" * 75)

        print()

        print(f"Data...............: {self.data}")

        print(f"Posto..............: {self.posto}")

        print(f"Cidade.............: {self.cidade}")

        print(f"Odômetro...........: {self.odometro:.1f} km")

        print(f"Volume.............: {self.volume_m3:.3f} m³")

        print(f"Preço por m³.......: R$ {self.preco_m3:.3f}")

        print(f"Valor Total........: R$ {self.valor_total:.2f}")

        print(f"Temperatura........: {self.temperatura:.1f} °C")

        print(f"Pressão............: {self.pressao:.1f} bar")

        print(f"Altitude...........: {self.altitude:.1f} m")

        print(f"Observações........: {self.observacoes}")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 94
# EXPORTAÇÃO PARA DICIONÁRIO
# =============================================================================

    def to_dict(self):

        return {

            "data": self.data,

            "posto": self.posto,

            "cidade": self.cidade,

            "odometro": self.odometro,

            "volume_m3": self.volume_m3,

            "preco_m3": self.preco_m3,

            "valor_total": self.valor_total,

            "temperatura": self.temperatura,

            "pressao": self.pressao,

            "altitude": self.altitude,

            "observacoes": self.observacoes,

            "capacidade_cilindro_l": self.capacidade_cilindro_l,

            "pressao_inicial": self.pressao_inicial,

            "pressao_final": self.pressao_final,

            "densidade_informada_kg_m3": self.densidade_informada_kg_m3,

            "metragem_teorica_m3": self.metragem_teorica_m3,

            "metragem_anp_m3": self.metragem_anp_m3,

            "metragem_cientifica_m3": self.metragem_cientifica_m3,
            "status_bomba": self.status_bomba,
            "divergencia_anp_m3": self.divergencia_anp_m3,
            "divergencia_fisica_m3": self.divergencia_fisica_m3,
            "divergencia_anp_pct": self.divergencia_anp_pct,
            "divergencia_fisica_pct": self.divergencia_fisica_pct

        }


# =============================================================================
# PARTE 95
# HISTÓRICO DE ABASTECIMENTOS
# =============================================================================

class HistoricoAbastecimentos:

    """
    Armazena vários abastecimentos.
    """

    def __init__(self):

        self.abastecimentos = []


# =============================================================================
# PARTE 96
# ADICIONAR ABASTECIMENTO
# =============================================================================

    def adicionar(

        self,

        abastecimento

    ):

        if not isinstance(

            abastecimento,

            Abastecimento

        ):

            raise TypeError(

                "Objeto inválido."

            )

        self.abastecimentos.append(

            abastecimento

        )

        self.abastecimentos.sort(

            key=lambda x: (

                x.data,

                x.odometro

            )

        )



# =============================================================================
# PARTE 97
# TOTAL DE ABASTECIMENTOS
# =============================================================================

    def quantidade(self):

        return len(

            self.abastecimentos

        )


# =============================================================================
# PARTE 98
# LISTAR ABASTECIMENTOS
# =============================================================================

    def listar(self):

        print()

        print("=" * 75)

        print("HISTÓRICO DE ABASTECIMENTOS")

        print("=" * 75)

        print()

        for indice, abastecimento in enumerate(

            self.abastecimentos,

            start=1

        ):

            print(

                f"{indice:03d} - "

                f"{abastecimento.data} - "

                f"{abastecimento.posto} - "

                f"{abastecimento.volume_m3:.3f} m³ - "

                f"R$ {abastecimento.valor_total:.2f}"

            )

        print()

        print("=" * 75)

        print()



# =============================================================================
# PARTE 99
# PROCURAR ABASTECIMENTO POR DATA
# =============================================================================

    def procurar_data(
        self,
        data
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.data == data:

                resultados.append(

                    abastecimento

                )

        return resultados


# =============================================================================
# PARTE 100
# PROCURAR ABASTECIMENTO POR POSTO
# =============================================================================

    def procurar_posto(
        self,
        posto
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.posto.lower() == posto.lower():

                resultados.append(

                    abastecimento

                )

        return resultados



# =============================================================================
# PARTE 101
# PROCURAR ABASTECIMENTO POR CIDADE
# =============================================================================

    def procurar_cidade(
        self,
        cidade
    ):

        resultados = []

        for abastecimento in self.abastecimentos:

            if abastecimento.cidade.lower() == cidade.lower():

                resultados.append(

                    abastecimento

                )

        return resultados



# =============================================================================
# PARTE 102
# VALOR TOTAL GASTO
# =============================================================================

    def valor_total(self):

        total = 0.0

        for abastecimento in self.abastecimentos:

            total += abastecimento.valor_total

        return total


# =============================================================================
# PARTE 103
# VOLUME TOTAL ABASTECIDO
# =============================================================================

    def volume_total(self):

        total = 0.0

        for abastecimento in self.abastecimentos:

            total += abastecimento.volume_m3

        return total


# =============================================================================
# PARTE 104
# MÉDIA DO PREÇO DO m³
# =============================================================================

    def media_preco_m3(self):

        if len(self.abastecimentos) == 0:

            return 0.0

        soma = 0.0

        for abastecimento in self.abastecimentos:

            soma += abastecimento.preco_m3

        return soma / len(self.abastecimentos)


# =============================================================================
# PARTE 105
# MÉDIA DO VOLUME ABASTECIDO
# =============================================================================

    def media_volume(self):

        if len(self.abastecimentos) == 0:

            return 0.0

        return self.volume_total() / len(self.abastecimentos)


# =============================================================================
# PARTE 106
# MAIOR ABASTECIMENTO
# =============================================================================

    def maior_abastecimento(self):

        if len(self.abastecimentos) == 0:

            return None

        maior = self.abastecimentos[0]

        for abastecimento in self.abastecimentos:

            if abastecimento.volume_m3 > maior.volume_m3:

                maior = abastecimento

        return maior


# =============================================================================
# PARTE 107
# MENOR ABASTECIMENTO
# =============================================================================

    def menor_abastecimento(self):

        if len(self.abastecimentos) == 0:

            return None

        menor = self.abastecimentos[0]

        for abastecimento in self.abastecimentos:

            if abastecimento.volume_m3 < menor.volume_m3:

                menor = abastecimento

        return menor


# =============================================================================
# PARTE 108
# RELATÓRIO GERAL
# =============================================================================

    def relatorio(self):

        print()

        print("=" * 75)

        print("RELATÓRIO DO HISTÓRICO")

        print("=" * 75)

        print()

        print(f"Quantidade...............: {self.quantidade()}")

        print(f"Volume Total.............: {self.volume_total():.3f} m³")

        print(f"Valor Total..............: R$ {self.valor_total():.2f}")

        print(f"Média do Volume..........: {self.media_volume():.3f} m³")

        print(f"Média do Preço...........: R$ {self.media_preco_m3():.3f}")

        maior = self.maior_abastecimento()

        if maior is not None:

            print()

            print("Maior abastecimento")

            print("-------------------")

            print(f"Data.....................: {maior.data}")

            print(f"Posto....................: {maior.posto}")

            print(f"Volume...................: {maior.volume_m3:.3f} m³")

        menor = self.menor_abastecimento()

        if menor is not None:

            print()

            print("Menor abastecimento")

            print("-------------------")

            print(f"Data.....................: {menor.data}")

            print(f"Posto....................: {menor.posto}")

            print(f"Volume...................: {menor.volume_m3:.3f} m³")

        print()

        print("=" * 75)

        print()


# =============================================================================
# PARTE 109
# EXPORTAR HISTÓRICO
# =============================================================================

    def exportar_lista(self):

        """
        Retorna todos os abastecimentos
        em formato de lista de dicionários.
        """

        lista = []

        for abastecimento in self.abastecimentos:

            lista.append(

                abastecimento.to_dict()

            )

        return lista


# =============================================================================
# PARTE 110
# LIMPAR HISTÓRICO
# =============================================================================

    def limpar(self):

        """
        Remove todos os abastecimentos
        da memória.
        """

        self.abastecimentos.clear()


# =============================================================================
# PARTE 111
# ÚLTIMO ABASTECIMENTO
# =============================================================================

    def ultimo(self):

        """
        Retorna o último abastecimento.
        """

        if len(self.abastecimentos) == 0:

            return None

        return self.abastecimentos[-1]


# =============================================================================
# PARTE 112
# PRIMEIRO ABASTECIMENTO
# =============================================================================

    def primeiro(self):

        """
        Retorna o primeiro abastecimento.
        """

        if len(self.abastecimentos) == 0:

            return None

        return self.abastecimentos[0]



# =============================================================================
# PARTE 113
# TOTAL DE QUILÔMETROS
# =============================================================================

    def quilometragem_total(self):

        """
        Calcula a quilometragem entre
        o primeiro e o último abastecimento.
        """

        if len(self.abastecimentos) < 2:

            return 0.0

        primeiro = self.primeiro()

        ultimo = self.ultimo()

        return (

            ultimo.odometro

            -

            primeiro.odometro

        )



# =============================================================================
# PARTE 114
# BANCO DE DADOS SQLITE
# =============================================================================

def normalizar_cidade(nome):
    """Normaliza nomes de cidades para evitar duplicidade por maiúsculas/minúsculas.

    Exemplos: Rio de Janeiro, rio de janeiro e RIO DE JANEIRO
    passam a ser armazenados como "Rio de Janeiro".
    """
    texto = " ".join(str(nome or "").strip().split())
    if not texto:
        return ""
    palavras_pequenas = {"de", "da", "do", "das", "dos", "e", "em"}
    partes = []
    for i, palavra in enumerate(texto.lower().split(" ")):
        if i > 0 and palavra in palavras_pequenas:
            partes.append(palavra)
        else:
            partes.append(palavra[:1].upper() + palavra[1:])
    return " ".join(partes)


def chave_cidade(nome):
    """Chave case-insensitive para agrupamento estatístico de cidades."""
    return " ".join(str(nome or "").strip().casefold().split())


class BancoGNV:

    """
    Classe responsável pelo banco SQLite.
    """

    def __init__(

        self,

        nome_banco="gnv.db"

    ):

        self.nome_banco = nome_banco

        self.conexao = None

        self.cursor = None



# =============================================================================
# PARTE 115
# CONECTAR AO SQLITE
# =============================================================================

    def conectar(self):

        self.conexao = sqlite3.connect(

            self.nome_banco

        )

        self.cursor = self.conexao.cursor()



# =============================================================================
# PARTE 116
# FECHAR SQLITE
# =============================================================================

    def fechar(self):

        if self.conexao:

            self.conexao.close()



# =============================================================================
# PARTE 117
# CRIAR TABELA
# =============================================================================

    def criar_tabela(self):

        self.cursor.execute("""

        CREATE TABLE IF NOT EXISTS abastecimentos(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            data TEXT,

            posto TEXT,

            cidade TEXT,

            odometro REAL,

            volume REAL,

            preco REAL,

            valor REAL,

            temperatura REAL,

            pressao REAL,

            altitude REAL,

            observacoes TEXT,

            capacidade_cilindro_l REAL DEFAULT 0,

            pressao_inicial REAL DEFAULT 0,

            pressao_final REAL DEFAULT 0,

            densidade_informada_kg_m3 REAL DEFAULT 0,

            metragem_teorica_m3 REAL DEFAULT 0,

            metragem_anp_m3 REAL DEFAULT 0,

            metragem_cientifica_m3 REAL DEFAULT 0,

            status_bomba TEXT DEFAULT 'NAO_CLASSIFICADO',
            divergencia_anp_m3 REAL DEFAULT 0,
            divergencia_fisica_m3 REAL DEFAULT 0,
            divergencia_anp_pct REAL DEFAULT 0,
            divergencia_fisica_pct REAL DEFAULT 0

        )

        """)

        self.conexao.commit()

        # Migração segura para bancos criados em versões anteriores.
        colunas_existentes = {
            linha[1] for linha in self.cursor.execute(
                "PRAGMA table_info(abastecimentos)"
            ).fetchall()
        }

        novas_colunas = {
            "capacidade_cilindro_l": "REAL DEFAULT 0",
            "pressao_inicial": "REAL DEFAULT 0",
            "pressao_final": "REAL DEFAULT 0",
            "densidade_informada_kg_m3": "REAL DEFAULT 0",
            "metragem_teorica_m3": "REAL DEFAULT 0",
            "metragem_anp_m3": "REAL DEFAULT 0",
            "metragem_cientifica_m3": "REAL DEFAULT 0",
            "status_bomba": "TEXT DEFAULT 'NAO_CLASSIFICADO'",
            "divergencia_anp_m3": "REAL DEFAULT 0",
            "divergencia_fisica_m3": "REAL DEFAULT 0",
            "divergencia_anp_pct": "REAL DEFAULT 0",
            "divergencia_fisica_pct": "REAL DEFAULT 0",
        }

        for nome, definicao in novas_colunas.items():
            if nome not in colunas_existentes:
                self.cursor.execute(
                    f"ALTER TABLE abastecimentos ADD COLUMN {nome} {definicao}"
                )

        self.conexao.commit()

        # Normaliza cidades já existentes no banco, sem alterar os demais dados.
        try:
            cidades = self.cursor.execute("SELECT id, cidade FROM abastecimentos").fetchall()
            for registro_id, cidade in cidades:
                cidade_normalizada = normalizar_cidade(cidade)
                if cidade_normalizada and cidade_normalizada != cidade:
                    self.cursor.execute(
                        "UPDATE abastecimentos SET cidade=? WHERE id=?",
                        (cidade_normalizada, registro_id)
                    )
            self.conexao.commit()
        except sqlite3.Error:
            pass



# =============================================================================
# PARTE 118
# SALVAR ABASTECIMENTO
# =============================================================================

    def salvar_abastecimento(
        self,
        abastecimento
    ):

        self.cursor.execute(

            """

            INSERT INTO abastecimentos(

                data,

                posto,

                cidade,

                odometro,

                volume,

                preco,

                valor,

                temperatura,

                pressao,

                altitude,

                observacoes,

                capacidade_cilindro_l,

                pressao_inicial,

                pressao_final,

                densidade_informada_kg_m3,

                metragem_teorica_m3,

                metragem_anp_m3,

                metragem_cientifica_m3,
                status_bomba,
                divergencia_anp_m3,
                divergencia_fisica_m3,
                divergencia_anp_pct,
                divergencia_fisica_pct

            )

            VALUES(

                ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?

            )

            """,

            (

                abastecimento.data,

                abastecimento.posto,

                abastecimento.cidade,

                abastecimento.odometro,

                abastecimento.volume_m3,

                abastecimento.preco_m3,

                abastecimento.valor_total,

                abastecimento.temperatura,

                abastecimento.pressao,

                abastecimento.altitude,

                abastecimento.observacoes,

                abastecimento.capacidade_cilindro_l,

                abastecimento.pressao_inicial,

                abastecimento.pressao_final,

                abastecimento.densidade_informada_kg_m3,

                abastecimento.metragem_teorica_m3,

                abastecimento.metragem_anp_m3,

                abastecimento.metragem_cientifica_m3,
                abastecimento.status_bomba,
                abastecimento.divergencia_anp_m3,
                abastecimento.divergencia_fisica_m3,
                abastecimento.divergencia_anp_pct,
                abastecimento.divergencia_fisica_pct

            )

        )

        self.conexao.commit()



# =============================================================================
# PARTE 119
# LISTAR TODOS OS ABASTECIMENTOS
# =============================================================================

    def listar_abastecimentos(self):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            ORDER BY id

            """

        )

        return self.cursor.fetchall()


# =============================================================================
# PARTE 120
# EXCLUIR ABASTECIMENTO
# =============================================================================

    def excluir_abastecimento(
        self,
        id_abastecimento
    ):

        self.cursor.execute(

            """

            DELETE FROM abastecimentos

            WHERE id = ?

            """,

            (

                id_abastecimento,

            )

        )

        self.conexao.commit()



# =============================================================================
# PARTE 121
# PESQUISAR ABASTECIMENTO
# =============================================================================

    def buscar_por_id(
        self,
        id_abastecimento
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE id = ?

            """,

            (

                id_abastecimento,

            )

        )

        return self.cursor.fetchone()


# =============================================================================
# PARTE 122
# ATUALIZAR ABASTECIMENTO
# =============================================================================

    def atualizar_abastecimento(
        self,
        id_abastecimento,
        abastecimento
    ):

        self.cursor.execute(

            """

            UPDATE abastecimentos

            SET

                data = ?,

                posto = ?,

                cidade = ?,

                odometro = ?,

                volume = ?,

                preco = ?,

                valor = ?,

                temperatura = ?,

                pressao = ?,

                altitude = ?,

                observacoes = ?,

                capacidade_cilindro_l = ?,

                pressao_inicial = ?,

                pressao_final = ?,

                densidade_informada_kg_m3 = ?,
                metragem_teorica_m3 = ?,
                metragem_anp_m3 = ?,
                metragem_cientifica_m3 = ?

            WHERE id = ?

            """,

            (

                abastecimento.data,

                abastecimento.posto,

                abastecimento.cidade,

                abastecimento.odometro,

                abastecimento.volume_m3,

                abastecimento.preco_m3,

                abastecimento.valor_total,

                abastecimento.temperatura,

                abastecimento.pressao,

                abastecimento.altitude,

                abastecimento.observacoes,

                abastecimento.capacidade_cilindro_l,

                abastecimento.pressao_inicial,

                abastecimento.pressao_final,

                abastecimento.densidade_informada_kg_m3,

                id_abastecimento

            )

        )

        self.conexao.commit()


# =============================================================================
# PARTE 123
# PESQUISAR POR POSTO
# =============================================================================

    def buscar_por_posto(
        self,
        posto
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE posto LIKE ?

            ORDER BY data

            """,

            (

                "%" + posto + "%",

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 124
# PESQUISAR POR CIDADE
# =============================================================================

    def buscar_por_cidade(
        self,
        cidade
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE cidade LIKE ?

            ORDER BY data

            """,

            (

                "%" + cidade + "%",

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 125
# CONTAR ABASTECIMENTOS
# =============================================================================

    def contar_registros(self):

        self.cursor.execute(

            """

            SELECT COUNT(*)

            FROM abastecimentos

            """

        )

        return self.cursor.fetchone()[0]



# =============================================================================
# PARTE 126
# BUSCAR ENTRE DUAS DATAS
# =============================================================================

    def buscar_periodo(
        self,
        data_inicial,
        data_final
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE data BETWEEN ? AND ?

            ORDER BY data

            """,

            (

                data_inicial,

                data_final

            )

        )

        return self.cursor.fetchall()


# =============================================================================
# PARTE 127
# BUSCAR POR ODÔMETRO
# =============================================================================

    def buscar_odometro(
        self,
        km_inicial,
        km_final
    ):

        self.cursor.execute(

            """

            SELECT *

            FROM abastecimentos

            WHERE odometro BETWEEN ? AND ?

            ORDER BY odometro

            """,

            (

                km_inicial,

                km_final

            )

        )

        return self.cursor.fetchall()



# =============================================================================
# PARTE 128
# CRIAR ÍNDICES
# =============================================================================

    def criar_indices(self):

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_data

            ON abastecimentos(data)

            """

        )

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_posto

            ON abastecimentos(posto)

            """

        )

        self.cursor.execute(

            """

            CREATE INDEX IF NOT EXISTS idx_odometro

            ON abastecimentos(odometro)

            """

        )

        self.conexao.commit()



# =============================================================================
# PARTE 129
# APAGAR TODOS OS REGISTROS
# =============================================================================

    def apagar_todos(self):

        self.cursor.execute(

            """

            DELETE FROM abastecimentos

            """

        )

        self.conexao.commit()


# =============================================================================
# PARTE 130
# TOTAL GASTO
# =============================================================================

    def total_gasto(self):

        self.cursor.execute(

            """

            SELECT SUM(valor)

            FROM abastecimentos

            """

        )

        resultado = self.cursor.fetchone()

        if resultado[0] is None:

            return 0.0

        return resultado[0]




# =============================================================================
# PARTE 401
# CLASSE PDF
# =============================================================================

class RelatorioPDF(FPDF):

    def header(self):

        self.set_font(

            "Arial",

            "B",

            14

        )

        self.cell(

            0,

            10,

            "RELATÓRIO DO SISTEMA GNV",

            0,

            1,

            "C"

        )

        self.ln(

            5

        )


# =============================================================================
# PARTE 402
# RODAPÉ
# =============================================================================

    def footer(self):

        self.set_y(

            -15

        )

        self.set_font(

            "Arial",

            "I",

            8

        )

        self.cell(

            0,

            10,

            f"Página {self.page_no()}",

            0,

            0,

            "C"

        )


# =============================================================================
# PARTE 403
# TÍTULO
# =============================================================================

    def titulo(

        self,

        texto

    ):

        self.set_font(

            "Arial",

            "B",

            12

        )

        self.cell(

            0,

            8,

            texto,

            0,

            1

        )



# =============================================================================
# PARTE 404
# LINHA
# =============================================================================

    def linha(

        self,

        titulo,

        valor

    ):

        self.set_font(

            "Arial",

            "",

            10

        )

        self.cell(

            70,

            7,

            titulo

        )

        self.cell(

            0,

            7,

            str(

                valor

            ),

            ln=True

        )

def numero_sqlite(valor, casas=2):
    try:
        return formatar_numero_br(float(valor), casas)
    except (TypeError, ValueError):
        return "-"


FORMULAS_I18N_COMPLETAS = {'Français': 'FORMULES ET PHYSIQUE DU SYSTÈME DE CALCUL DU GNV\n============================================================\n\nPARTIE A — CONDITION DE RÉFÉRENCE DE L’ANP\n============================================\n\nL’ANP utilise des conditions de référence de 20 °C et 1,033 kgf/cm² pour les volumes moyens de gaz naturel commercialisés. La condition standard de mesure est exprimée comme une pression absolue de 0,101325 MPa à 20 °C.\n\nIMPORTANT : l’ANP définit la condition de référence. La formule implémentée par ce programme est une ESTIMATION PHYSIQUE DE CONVERSION et ne prétend pas reproduire l’algorithme interne d’un distributeur de GNV.\n\n1. VOLUME PHYSIQUE DU CYLINDRE\n------------------------------\nVcil = capacité(L) / 1000\n\n26 L / 1000 = 0,026 m³.\n\nIl s’agit de l’espace physique interne du cylindre. Ce n’est pas le volume normalisé indiqué par la pompe.\n\n2. PRESSION ABSOLUE\n-------------------\nPabs = pression manométrique + Patm\n\nLes équations d’état utilisent la pression absolue.\n\n3. TEMPÉRATURE ABSOLUE\n----------------------\nT(K) = T(°C) + 273,15\n\n4. CONVERSION À 20 °C\n---------------------\nPour une quantité de matière fixe, dans le modèle du gaz parfait :\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293,15 K et Pref ≈ 1,01325 bar.\n\nPARTIE B — MODÈLE SCIENTIFIQUE DE GAZ RÉEL\n===========================================\n\n1. ÉQUATION D’ÉTAT\n------------------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = pression absolue (Pa)\nV = volume physique (m³)\nZ = facteur de compressibilité\nn = quantité de matière (mol)\nR = 8,314462618 J/(mol·K)\nT = température absolue (K)\n\n2. FACTEUR DE COMPRESSIBILITÉ Z\n-------------------------------\nZ = P V / (n R T)\n\nZ = 1 représente un gaz parfait. Pour le gaz naturel réel, Z dépend de la pression, de la température et de la composition. Une valeur Z fixe saisie par l’utilisateur est une approximation et non une détermination métrologique de Z.\n\n3. QUANTITÉ DE GAZ AJOUTÉE\n---------------------------\nn_initial = P_initial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_initial\n\n4. MASSE AJOUTÉE\n----------------\nm = Δn × M\n\nM est la masse molaire du GNV en kg/mol.\n\n5. VOLUME ÉQUIVALENT À LA TEMPÉRATURE SAISIE\n---------------------------------------------\nLe programme calcule d’abord la quantité de matière à partir de PV = Z n R T. Il peut ensuite exprimer cette même quantité de matière à une pression de référence :\n\nVref(T) = n R T / Pref\n\nIMPORTANT : dans ce programme, n est calculé avec la température saisie. En remplaçant n = P V / (Z R T), la température se simplifie :\n\nVref(T) = P V / (Z Pref)\n\nAinsi, pour un cylindre de 26 L à 220 bar et Z=0,92, le volume équivalent à 1,01325 bar à la température saisie peut rester approximativement égal à 6,164 m³ à 5 °C, 20 °C ou 100 °C. Cela NE signifie PAS qu’une même quantité fixe de gaz possède le même volume à différentes températures ; cela signifie que le programme recalcule la quantité de matière pour chaque état de pression/température saisi.\n\n6. CONVERSION SCIENTIFIQUE À 20 °C\n----------------------------------\nLes mêmes moles calculées avec le modèle Z saisi sont converties à 20 °C :\n\nV20 = n R T20 / Pref\n\nSi la température saisie est inférieure à 20 °C, V20 tend à être plus grand. Si elle est supérieure à 20 °C, V20 tend à être plus petit, car n est maintenu fixe pendant cette conversion.\n\n7. CONVERSION ANP/IDÉALISÉE (Z=1)\n----------------------------------\nL’onglet ANP calcule séparément une estimation avec Z=1 et la condition de référence 20 °C / 1,033 kgf/cm² (environ 1,01325 bar). Cette valeur ne doit pas être confondue avec le résultat scientifique utilisant Z=0,92. Il s’agit de deux modèles différents.\n\n8. DENSITÉ DU GAZ RÉEL\n----------------------\nρ = P M / (Z R T)\n\nPARTIE C — CE QU’EXIGERAIT UN MODÈLE PLUS PRÉCIS\n=================================================\n\nPour le GNV à haute pression, considérer Z comme une constante universelle n’est pas adéquat. Une précision supérieure exige la composition du gaz et les propriétés thermodynamiques en fonction de P et T.\n\nL’ISO 12213 décrit des méthodes pour calculer le facteur de compressibilité du gaz naturel. L’ISO 12213-2 utilise la composition molaire ; l’ISO 12213-3 utilise des propriétés physiques telles que le pouvoir calorifique, la densité relative et le CO₂, avec la pression et la température.\n\nAGA8 et GERG sont des modèles utilisés pour les propriétés du gaz naturel. Le NIST décrit AGA8 et GERG parmi les équations d’état utilisées pour la mesure du gaz naturel et les applications de propriétés thermodynamiques.\n\nPar conséquent, l’évolution scientifique du programme devrait être :\n1) obtenir la composition du GNV ;\n2) calculer Z(P,T,composition) au lieu d’utiliser un Z fixe ;\n3) considérer la température réelle du gaz pendant le ravitaillement ;\n4) connaître les conditions réellement utilisées par le compteur ;\n5) prendre en compte l’incertitude de mesure.\n\nPARTIE D — TEMPÉRATURE PENDANT LE RAVITAILLEMENT\n=================================================\n\nPendant le remplissage, de la masse entre et des échanges thermiques se produisent. L’onglet Chauffage / Compression utilise une compression adiabatique réversible uniquement comme scénario pédagogique :\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = constante\nV₂/V₁ = (P₁/P₂)^(1/k)\n\nLe ravitaillement réel est un système ouvert avec transfert de chaleur entre le gaz, la paroi du cylindre, le flexible et l’environnement. La température calculée dans cet onglet n’est PAS une mesure de la température réelle du GNV.\n\nPARTIE E — COMPARAISON AVEC LA POMPE\n=====================================\n\nÉcart = volume indiqué par la pompe − volume calculé\nÉcart en pourcentage = écart / volume calculé × 100\n\nUn écart important constitue un indice nécessitant une investigation. À lui seul, il ne constitue pas une preuve métrologique de fraude. Une conclusion technique exige les données du compteur, les conditions de référence, la température réelle du gaz, la composition/Z, l’étalonnage et l’incertitude de mesure.\n\nPARTIE F — VOLUME GÉOMÉTRIQUE DU CIRCUIT ACCESSOIRE DE GNV\n============================================================\n\nOBJECTIF\n--------\nEstimer les volumes internes minimum, moyen et maximum du circuit GNV haute et basse pression, en EXCLUANT STRICTEMENT LE CYLINDRE/RÉSERVOIR. L’étude considère l’espace interne des conduites, vannes, régulateur/réducteur et composants basse pression.\n\nIMPORTANT SUR LA NATURE DE CE CALCUL\n-------------------------------------\nCes valeurs sont des estimations de référence d’ingénierie basées sur le RAPPORT TECHNIQUE fourni pour ce projet. Elles ne constituent pas des spécifications universelles de tous les kits GNV. Une mesure spécifique au véhicule doit utiliser le diamètre intérieur réel, la longueur, le modèle de réducteur, les vannes, le filtre, la rampe d’injecteurs et les autres composants.\n\n1. CONDUITE HAUTE PRESSION\n--------------------------\nL’hypothèse de l’étude est une conduite de diamètre extérieur 6 mm et d’épaisseur de paroi 1,2 mm, donnant un diamètre intérieur de 3,6 mm.\nD_interne = D_externe - 2 × épaisseur de paroi\nD_interne = 6,0 - 2 × 1,2 = 3,6 mm\nr = 1,8 mm = 0,0018 m\n\nPour une conduite cylindrique droite :\nV_t = pi × r² × L\n\nAvec L = 4,0 m : V_t ≈ 40,7 mL\nAvec L = 5,5 m : V_t ≈ 70,1 mL\n\nLa plage géométrique approximative est donc de 40 à 70 mL.\n\n2. VANNES DE SERVICE\n--------------------\nLe rapport technique estime le volume mort combiné des vannes du cylindre et de service de remplissage à 3 à 5 mL.\n\n3. RÉDUCTEUR / RÉGULATEUR DE PRESSION\n--------------------------------------\nL’étude considère les chambres utilisées exclusivement pour le passage du gaz dans le régulateur/réducteur, à l’exclusion du circuit d’eau de chauffage. La plage estimée est de 30 à 65 mL.\n\n4. BASSE PRESSION / FILTRE / RAMPE D’INJECTEURS\n------------------------------------------------\nPour les systèmes modernes, le rapport considère les flexibles basse pression, le filtre de phase gazeuse et la rampe/collecteur d’injecteurs. La plage estimée est de 50 à 140 mL. Si cette section n’est pas présente, sa contribution peut être considérée comme 0 mL.\n\n5. CONSOLIDATION DE L’ÉTUDE\n---------------------------\nComposant                         Minimum    Moyenne    Maximum\nConduite haute pression          40 mL       55 mL       70 mL\nVannes de service                 3 mL        4 mL        5 mL\nRéducteur — chambres de gaz      32 mL       46 mL       65 mL\nBasse pression/filtre/rampe       0 mL       25 mL      140 mL\nVOLUME TOTAL DU CIRCUIT          75 mL      130 mL      280 mL\n\nL’étude adopte 130 mL comme référence géométrique moyenne représentative, sans affirmer que cette valeur soit une mesure universelle de la flotte.\n\n6. POURQUOI LE CIRCUIT DOIT-IL ÊTRE TRAITÉ SÉPARÉMENT DU CYLINDRE ?\n------------------------------------------------------------------\nLe cylindre est un grand réservoir physique de gaz. Les conduites, vannes, régulateur/réducteur et composants basse pression constituent des volumes internes supplémentaires du circuit. Ainsi, dans une analyse géométrique, ces espaces peuvent être comptabilisés séparément.\n\nC’est particulièrement important lorsqu’on compare le volume indiqué par le distributeur, le volume calculé, le volume stocké et le volume interne du circuit. Ces grandeurs ne doivent pas être additionnées sans définir clairement la grandeur physique mesurée.\n\n7. COMMENT TRANSFORMER L’ESTIMATION EN MESURE SPÉCIFIQUE DU VÉHICULE\n---------------------------------------------------------------------\nPour obtenir un résultat spécifique au véhicule, fournir :\n- diamètre intérieur réel de la conduite ;\n- longueur réelle de la conduite haute pression ;\n- nombre et volume interne des vannes ;\n- fabricant/modèle du réducteur ;\n- nombre d’étages du réducteur ;\n- volume interne des chambres de gaz ;\n- volume du filtre de phase gazeuse ;\n- volume de la rampe/du collecteur d’injecteurs ;\n- volume interne des injecteurs et raccords ;\n- documentation technique du fabricant, lorsqu’elle est disponible.\n\n8. RÉFÉRENCES DE L’ÉTUDE TECHNIQUE\n----------------------------------\nINMETRO — Portaria nº 111/2022 et exigences relatives aux composants GNV/GNC :\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1, systèmes de véhicules au gaz naturel :\nhttps://www.abntcatalogo.com.br/\n\nUSP — École Polytechnique / Bibliothèque numérique des thèses et dissertations :\nhttps://teses.usp.br/\n\nUFRGS — Centre de référence pour l’enseignement de la physique (CREF) :\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / Dépôt institutionnel :\nhttps://pantheon.ufrj.br/\n\nITA — Institut technologique de l’aéronautique :\nhttps://www.ita.br/\n\nIME — Institut militaire d’ingénierie :\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science :\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering :\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics :\nhttps://ocw.mit.edu/\n\nNIST — REFPROP et propriétés thermophysiques :\nhttps://www.nist.gov/srd/refprop\n\nSOURCES ET FONDEMENTS TECHNIQUES\n================================\nANP — Publication des prix du gaz naturel :\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — Glossaire / Condition standard de mesure :\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — calcul du facteur de compressibilité par composition :\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — calcul à partir de propriétés physiques :\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics : équation du gaz parfait PV = nRT :\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics :\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law :\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow :\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — Programme académique relatif à la thermodynamique :\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / équations d’état du gaz naturel :\nhttps://www.nist.gov/srd/refprop\n', 'Italiano': 'FORMULE E FISICA DEL SISTEMA DI CALCOLO DEL GNV\n============================================================\n\nPARTE A — CONDIZIONE DI RIFERIMENTO ANP\n========================================\n\nL’ANP utilizza condizioni di riferimento di 20 °C e 1,033 kgf/cm² per i volumi medi commercializzati di gas naturale. La condizione standard di misura è espressa come pressione assoluta di 0,101325 MPa a 20 °C.\n\nIMPORTANTE: l’ANP definisce la condizione di riferimento. La formula implementata dal programma è una STIMA FISICA DI CONVERSIONE e non dichiara di riprodurre l’algoritmo interno di un distributore GNV.\n\n1. VOLUME FISICO DEL CILINDRO\n-----------------------------\nVcil = capacità(L) / 1000\n\n26 L / 1000 = 0,026 m³.\n\nQuesto è lo spazio fisico interno del cilindro. Non è il volume normalizzato indicato dalla pompa.\n\n2. PRESSIONE ASSOLUTA\n---------------------\nPabs = pressione manometrica + Patm\n\nLe equazioni di stato utilizzano la pressione assoluta.\n\n3. TEMPERATURA ASSOLUTA\n-----------------------\nT(K) = T(°C) + 273,15\n\n4. CONVERSIONE A 20 °C\n----------------------\nPer una quantità di sostanza fissa, nel modello di gas ideale:\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293,15 K e Pref ≈ 1,01325 bar.\n\nPARTE B — MODELLO SCIENTIFICO DEL GAS REALE\n============================================\n\n1. EQUAZIONE DI STATO\n---------------------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = pressione assoluta (Pa)\nV = volume fisico (m³)\nZ = fattore di comprimibilità\nn = quantità di sostanza (mol)\nR = 8,314462618 J/(mol·K)\nT = temperatura assoluta (K)\n\n2. FATTORE DI COMPRIMIBILITÀ Z\n------------------------------\nZ = P V / (n R T)\n\nZ = 1 rappresenta un gas ideale. Per il gas naturale reale, Z dipende da pressione, temperatura e composizione. Un valore Z fisso inserito dall’utente è un’approssimazione e non una determinazione metrologica di Z.\n\n3. QUANTITÀ DI GAS AGGIUNTA\n---------------------------\nn_iniziale = P_iniziale_abs × V / (Z R T)\nn_finale   = P_finale_abs × V / (Z R T)\nΔn = n_finale − n_iniziale\n\n4. MASSA AGGIUNTA\n-----------------\nm = Δn × M\n\nM è la massa molare del GNV in kg/mol.\n\n5. VOLUME EQUIVALENTE ALLA TEMPERATURA INSERITA\n------------------------------------------------\nIl programma calcola prima la quantità di sostanza da PV = Z n R T. Può quindi esprimere la stessa quantità di sostanza a una pressione di riferimento:\n\nVref(T) = n R T / Pref\n\nIMPORTANTE: in questo programma n viene calcolato usando la temperatura inserita. Sostituendo n = P V / (Z R T), la temperatura si semplifica:\n\nVref(T) = P V / (Z Pref)\n\nPertanto, per un cilindro da 26 L a 220 bar e Z=0,92, il volume equivalente a 1,01325 bar alla temperatura inserita può rimanere circa 6,164 m³ a 5 °C, 20 °C o 100 °C. Ciò NON significa che la stessa quantità fissa di gas abbia lo stesso volume a temperature diverse; significa che il programma ricalcola la quantità di sostanza per ogni stato di pressione/temperatura inserito.\n\n6. CONVERSIONE SCIENTIFICA A 20 °C\n----------------------------------\nLe stesse moli calcolate dal modello con Z inserito vengono convertite a 20 °C:\n\nV20 = n R T20 / Pref\n\nSe la temperatura inserita è inferiore a 20 °C, V20 tende a essere maggiore. Se è superiore a 20 °C, V20 tende a essere minore, perché n viene mantenuto fisso durante la conversione.\n\n7. CONVERSIONE ANP/IDEALIZZATA (Z=1)\n-------------------------------------\nLa scheda ANP calcola separatamente una stima con Z=1 e la condizione di riferimento 20 °C / 1,033 kgf/cm² (circa 1,01325 bar). Questo valore non deve essere confuso con il risultato scientifico che utilizza Z=0,92. Sono modelli differenti.\n\n8. DENSITÀ DEL GAS REALE\n------------------------\nρ = P M / (Z R T)\n\nPARTE C — COSA RICHIEDEREBBE UN MODELLO PIÙ PRECISO\n====================================================\n\nPer il GNV ad alta pressione, considerare Z come una costante universale non è adeguato. Una maggiore precisione richiede la composizione del gas e le proprietà termodinamiche in funzione di P e T.\n\nLa ISO 12213 descrive metodi per calcolare il fattore di comprimibilità del gas naturale. ISO 12213-2 utilizza la composizione molare; ISO 12213-3 utilizza proprietà fisiche come potere calorifico, densità relativa e CO₂, insieme a pressione e temperatura.\n\nAGA8 e GERG sono modelli utilizzati per le proprietà del gas naturale. Il NIST descrive AGA8 e GERG tra le equazioni di stato utilizzate nelle applicazioni di misura e proprietà termodinamiche del gas naturale.\n\nPertanto, l’evoluzione scientifica del programma dovrebbe essere:\n1) ottenere la composizione del GNV;\n2) calcolare Z(P,T,composizione) invece di usare un Z fisso;\n3) considerare la temperatura reale del gas durante il rifornimento;\n4) conoscere le condizioni effettivamente utilizzate dal misuratore;\n5) considerare l’incertezza di misura.\n\nPARTE D — TEMPERATURA DURANTE IL RIFORNIMENTO\n==============================================\n\nDurante il riempimento entra massa e avviene trasferimento di calore. La scheda Riscaldamento / Compressione usa la compressione adiabatica reversibile solo come scenario didattico:\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = costante\nV₂/V₁ = (P₁/P₂)^(1/k)\n\nIl rifornimento reale è un sistema aperto con trasferimento di calore tra gas, parete del cilindro, tubo flessibile e ambiente. La temperatura calcolata in questa scheda NON è una misura della temperatura reale del GNV.\n\nPARTE E — CONFRONTO CON LA POMPA\n=================================\n\nDifferenza = volume indicato dalla pompa − volume calcolato\nDifferenza percentuale = differenza / volume calcolato × 100\n\nUna differenza elevata è un’indicazione per un’indagine. Da sola non costituisce una prova metrologica di frode. Una conclusione tecnica richiede dati del misuratore, condizioni di riferimento, temperatura reale del gas, composizione/Z, taratura e incertezza di misura.\n\nPARTE F — VOLUME GEOMETRICO DEL CIRCUITO ACCESSORIO GNV\n=========================================================\n\nOBIETTIVO\n---------\nStimare i volumi interni minimo, medio e massimo del circuito GNV ad alta e bassa pressione, ESCLUDENDO RIGOROSAMENTE IL CILINDRO/SERBATOIO. Lo studio considera lo spazio interno di tubazioni, valvole, regolatore/riduttore e componenti a bassa pressione.\n\nIMPORTANTE SULLA NATURA DEL CALCOLO\n-----------------------------------\nQuesti valori sono stime di riferimento ingegneristiche basate sul RAPPORTO TECNICO fornito per questo progetto. Non sono specifiche universali per ogni kit GNV. Una misurazione specifica del veicolo deve utilizzare diametro interno reale, lunghezza, modello del riduttore, valvole, filtro, flauto iniettori e altri componenti.\n\n1. TUBAZIONE AD ALTA PRESSIONE\n------------------------------\nLa premessa dello studio è un tubo con diametro esterno di 6 mm e spessore della parete di 1,2 mm, con diametro interno di 3,6 mm.\nD_interno = D_esterno - 2 × spessore parete\nD_interno = 6,0 - 2 × 1,2 = 3,6 mm\nr = 1,8 mm = 0,0018 m\n\nPer un tubo cilindrico rettilineo:\nV_t = pi × r² × L\n\nCon L = 4,0 m: V_t ≈ 40,7 mL\nCon L = 5,5 m: V_t ≈ 70,1 mL\n\nPertanto, l’intervallo geometrico approssimativo è da 40 a 70 mL.\n\n2. VALVOLE DI SERVIZIO\n----------------------\nIl rapporto tecnico stima il volume morto combinato delle valvole del cilindro e del servizio di rifornimento in 3–5 mL.\n\n3. RIDUTTORE / REGOLATORE DI PRESSIONE\n--------------------------------------\nLo studio considera le camere utilizzate esclusivamente per il passaggio del gas nel regolatore/riduttore, escludendo il circuito dell’acqua di riscaldamento. L’intervallo stimato è 30–65 mL.\n\n4. BASSA PRESSIONE / FILTRO / FLAUTO INIETTORI\n-----------------------------------------------\nPer i sistemi moderni, il rapporto considera tubi flessibili a bassa pressione, filtro della fase gassosa e flauto/collettore degli iniettori. L’intervallo stimato è 50–140 mL. Se questa sezione non è presente, il contributo può essere considerato 0 mL.\n\n5. CONSOLIDAMENTO DELLO STUDIO\n------------------------------\nComponente                         Minimo    Medio     Massimo\nTubazione alta pressione           40 mL      55 mL      70 mL\nValvole di servizio                 3 mL       4 mL       5 mL\nRiduttore — camere gas             32 mL      46 mL      65 mL\nBassa pressione/filtro/flauto       0 mL      25 mL     140 mL\nVOLUME TOTALE CIRCUITO             75 mL     130 mL     280 mL\n\nLo studio adotta 130 mL come riferimento geometrico medio rappresentativo, senza affermare che tale valore sia una misura universale del parco veicoli.\n\n6. PERCHÉ IL CIRCUITO DEVE ESSERE TRATTATO SEPARATAMENTE DAL CILINDRO?\n---------------------------------------------------------------------\nIl cilindro è un grande serbatoio fisico di gas. Tubazioni, valvole, regolatore/riduttore e componenti a bassa pressione costituiscono volumi interni aggiuntivi del circuito. Pertanto, in un’analisi del volume geometrico, questi spazi possono essere contabilizzati separatamente.\n\nQuesto è particolarmente importante quando si confrontano volume indicato dal distributore, volume calcolato, volume immagazzinato e volume interno del circuito. Queste grandezze non devono essere sommate senza definire chiaramente quale grandezza fisica viene misurata.\n\n7. COME TRASFORMARE LA STIMA IN UNA MISURA SPECIFICA DEL VEICOLO\n-----------------------------------------------------------------\nPer un risultato specifico del veicolo, fornire:\n- diametro interno reale della tubazione;\n- lunghezza reale della linea ad alta pressione;\n- numero e volume interno delle valvole;\n- produttore/modello del riduttore;\n- numero di stadi del riduttore;\n- volume interno delle camere del gas;\n- volume del filtro della fase gassosa;\n- volume del flauto/collettore degli iniettori;\n- volume interno degli iniettori e raccordi;\n- documentazione tecnica del produttore, quando disponibile.\n\n8. RIFERIMENTI DELLO STUDIO TECNICO\n-----------------------------------\nINMETRO — Portaria nº 111/2022 e requisiti relativi ai componenti GNV/GNC:\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1, sistemi veicolari a gas naturale:\nhttps://www.abntcatalogo.com.br/\n\nUSP — Escola Politécnica / Biblioteca Digitale di Tesi e Dissertazioni:\nhttps://teses.usp.br/\n\nUFRGS — Centro di riferimento per l’insegnamento della Fisica (CREF):\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / Repository istituzionale:\nhttps://pantheon.ufrj.br/\n\nITA — Instituto Tecnológico de Aeronáutica:\nhttps://www.ita.br/\n\nIME — Instituto Militar de Engenharia:\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science:\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering:\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics:\nhttps://ocw.mit.edu/\n\nNIST — REFPROP e proprietà termofisiche:\nhttps://www.nist.gov/srd/refprop\n\nFONTI E FONDAMENTI TECNICI\n===========================\nANP — Pubblicazione dei prezzi del gas naturale:\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — Glossario / Condizione standard di misura:\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — calcolo del fattore di comprimibilità tramite composizione:\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — calcolo tramite proprietà fisiche:\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics: equazione del gas ideale PV = nRT:\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics:\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law:\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow:\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — percorso accademico relativo alla termodinamica:\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / equazioni di stato del gas naturale:\nhttps://www.nist.gov/srd/refprop\n', 'Español': 'FÓRMULAS Y FÍSICA DEL SISTEMA DE CÁLCULO DE GNV\n============================================================\n\nPARTE A — CONDICIÓN DE REFERENCIA DE LA ANP\n============================================\n\nLa ANP utiliza condiciones de referencia de 20 °C y 1,033 kgf/cm² para los volúmenes medios comercializados de gas natural. La condición estándar de medición se expresa como una presión absoluta de 0,101325 MPa a 20 °C.\n\nIMPORTANTE: la ANP define la condición de referencia. La fórmula implementada por este programa es una ESTIMACIÓN FÍSICA DE CONVERSIÓN y no afirma reproducir el algoritmo interno de un surtidor de GNV.\n\n1. VOLUMEN FÍSICO DEL CILINDRO\n------------------------------\nVcil = capacidad(L) / 1000\n\n26 L / 1000 = 0,026 m³.\n\nEste es el espacio físico interno del cilindro. No es el volumen normalizado indicado por la bomba.\n\n2. PRESIÓN ABSOLUTA\n-------------------\nPabs = presión manométrica + Patm\n\nLas ecuaciones de estado utilizan presión absoluta.\n\n3. TEMPERATURA ABSOLUTA\n-----------------------\nT(K) = T(°C) + 273,15\n\n4. CONVERSIÓN A 20 °C\n---------------------\nPara una cantidad fija de sustancia, en el modelo de gas ideal:\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293,15 K y Pref ≈ 1,01325 bar.\n\nPARTE B — MODELO CIENTÍFICO DE GAS REAL\n========================================\n\n1. ECUACIÓN DE ESTADO\n---------------------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = presión absoluta (Pa)\nV = volumen físico (m³)\nZ = factor de compresibilidad\nn = cantidad de sustancia (mol)\nR = 8,314462618 J/(mol·K)\nT = temperatura absoluta (K)\n\n2. FACTOR DE COMPRESIBILIDAD Z\n------------------------------\nZ = P V / (n R T)\n\nZ = 1 representa un gas ideal. Para gas natural real, Z depende de presión, temperatura y composición. Un Z fijo introducido por el usuario es una aproximación, no una determinación metrológica de Z.\n\n3. CANTIDAD DE GAS AÑADIDA\n--------------------------\nn_inicial = P_inicial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_inicial\n\n4. MASA AÑADIDA\n---------------\nm = Δn × M\n\nM es la masa molar del GNV en kg/mol.\n\n5. VOLUMEN EQUIVALENTE A LA TEMPERATURA INTRODUCIDA\n----------------------------------------------------\nEl programa calcula primero la cantidad de sustancia mediante PV = Z n R T. Después puede expresar esa misma cantidad a una presión de referencia:\n\nVref(T) = n R T / Pref\n\nIMPORTANTE: en este programa, n se calcula utilizando la temperatura introducida. Por tanto, al sustituir n = P V / (Z R T), la temperatura se cancela:\n\nVref(T) = P V / (Z Pref)\n\nAsí, para un cilindro de 26 L a 220 bar y Z=0,92, el volumen equivalente a 1,01325 bar a la temperatura introducida puede permanecer aproximadamente en 6,164 m³ a 5 °C, 20 °C o 100 °C. Esto NO significa que la misma cantidad fija de gas tenga el mismo volumen a diferentes temperaturas; significa que el programa recalcula la cantidad de sustancia para cada estado de presión/temperatura introducido.\n\n6. CONVERSIÓN CIENTÍFICA A 20 °C\n--------------------------------\nLos mismos moles calculados por el modelo con Z introducido se convierten a 20 °C:\n\nV20 = n R T20 / Pref\n\nSi la temperatura introducida es inferior a 20 °C, V20 tiende a ser mayor. Si es superior a 20 °C, V20 tiende a ser menor porque n se mantiene fijo durante esta conversión.\n\n7. CONVERSIÓN ANP/IDEALIZADA (Z=1)\n-----------------------------------\nLa pestaña ANP calcula por separado una estimación usando Z=1 y la condición de referencia de 20 °C / 1,033 kgf/cm² (aproximadamente 1,01325 bar). Este valor no debe confundirse con el resultado científico que utiliza Z=0,92. Son modelos diferentes.\n\n8. DENSIDAD DEL GAS REAL\n------------------------\nρ = P M / (Z R T)\n\nPARTE C — QUÉ REQUERIRÍA UN MODELO MÁS PRECISO\n===============================================\n\nPara GNV a alta presión, no es adecuado considerar Z como una constante universal. Una mayor precisión requiere conocer la composición del gas y las propiedades termodinámicas en función de P y T.\n\nISO 12213 describe métodos para calcular el factor de compresibilidad del gas natural. ISO 12213-2 utiliza composición molar; ISO 12213-3 utiliza propiedades físicas como poder calorífico, densidad relativa y CO₂, junto con presión y temperatura.\n\nAGA8 y GERG son modelos utilizados para las propiedades del gas natural. NIST describe AGA8 y GERG entre las ecuaciones de estado utilizadas en medición y propiedades termodinámicas del gas natural.\n\nPor tanto, la evolución científica del programa debería ser:\n1) obtener la composición del GNV;\n2) calcular Z(P,T,composición) en lugar de usar Z fijo;\n3) considerar la temperatura real del gas durante el repostaje;\n4) conocer las condiciones realmente utilizadas por el medidor;\n5) considerar la incertidumbre de medición.\n\nPARTE D — TEMPERATURA DURANTE EL REPOSTAJE\n===========================================\n\nDurante el llenado entra masa y se produce transferencia de calor. La pestaña Calentamiento / Compresión utiliza compresión adiabática reversible únicamente como escenario didáctico:\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = constante\nV₂/V₁ = (P₁/P₂)^(1/k)\n\nEl repostaje real es un sistema abierto, con transferencia de calor entre gas, pared del cilindro, manguera y entorno. La temperatura calculada en esta pestaña NO es una medición de la temperatura real del GNV.\n\nPARTE E — COMPARACIÓN CON LA BOMBA\n===================================\n\nDiferencia = volumen indicado por la bomba − volumen calculado\nDiferencia porcentual = diferencia / volumen calculado × 100\n\nUna diferencia grande es un indicio para investigar. Por sí sola no constituye una prueba metrológica de fraude. Una conclusión técnica requiere datos del medidor, condiciones de referencia, temperatura real del gas, composición/Z, calibración e incertidumbre de medición.\n\nPARTE F — VOLUMEN GEOMÉTRICO DEL CIRCUITO ACCESORIO DE GNV\n============================================================\n\nOBJETIVO\n--------\nEstimar los volúmenes internos mínimo, medio y máximo del circuito GNV de alta y baja presión, EXCLUYENDO ESTRICTAMENTE EL CILINDRO/DEPÓSITO. El estudio considera el espacio interno de tuberías, válvulas, regulador/reductor y componentes de baja presión.\n\nIMPORTANTE SOBRE LA NATURALEZA DE ESTE CÁLCULO\n-----------------------------------------------\nEstos valores son estimaciones de referencia de ingeniería basadas en el INFORME TÉCNICO suministrado para este proyecto. No son especificaciones universales de todos los kits GNV. Una medición específica del vehículo debe utilizar el diámetro interno, longitud, modelo del reductor, válvulas, filtro, flauta de inyectores y demás componentes reales.\n\n1. TUBERÍA DE ALTA PRESIÓN\n--------------------------\nLa premisa del estudio es una tubería de 6 mm de diámetro exterior y 1,2 mm de espesor de pared, resultando en un diámetro interior de 3,6 mm.\nD_interno = D_externo - 2 × espesor\nD_interno = 6,0 - 2 × 1,2 = 3,6 mm\nr = 1,8 mm = 0,0018 m\n\nPara una tubería cilíndrica recta:\nV_t = pi × r² × L\n\nCon L = 4,0 m: V_t ≈ 40,7 mL\nCon L = 5,5 m: V_t ≈ 70,1 mL\n\nPor tanto, el intervalo geométrico aproximado es de 40 a 70 mL.\n\n2. VÁLVULAS DE SERVICIO\n-----------------------\nEl informe técnico estima el volumen muerto combinado de las válvulas del cilindro y de servicio de repostaje entre 3 y 5 mL.\n\n3. REDUCTOR / REGULADOR DE PRESIÓN\n----------------------------------\nEl estudio considera las cámaras utilizadas exclusivamente para el paso del gas en el regulador/reductor, excluyendo el circuito de agua de calentamiento. El intervalo estimado es de 30 a 65 mL.\n\n4. BAJA PRESIÓN / FILTRO / FLAUTA DE INYECTORES\n------------------------------------------------\nPara sistemas modernos, el informe considera mangueras de baja presión, filtro de fase gaseosa y flauta/colector de inyectores. El intervalo estimado es de 50 a 140 mL. Si esta sección no está presente, su contribución puede considerarse 0 mL.\n\n5. CONSOLIDACIÓN DEL ESTUDIO\n----------------------------\nComponente                         Mínimo    Medio     Máximo\nTubería alta presión               40 mL      55 mL      70 mL\nVálvulas de servicio                3 mL       4 mL       5 mL\nReductor — cámaras de gas          32 mL      46 mL      65 mL\nBaja presión/filtro/flauta          0 mL      25 mL     140 mL\nVOLUMEN TOTAL DEL CIRCUITO         75 mL     130 mL     280 mL\n\nEl estudio adopta 130 mL como referencia geométrica media representativa, sin afirmar que este valor sea una medición universal de la flota.\n\n6. ¿POR QUÉ EL CIRCUITO DEBE TRATARSE SEPARADAMENTE DEL CILINDRO?\n------------------------------------------------------------------\nEl cilindro es un gran depósito físico de gas. Las tuberías, válvulas, regulador/reductor y componentes de baja presión forman volúmenes internos adicionales del circuito. Por ello, en un análisis geométrico, estos espacios pueden contabilizarse por separado.\n\nEsto es especialmente importante al comparar volumen indicado por el surtidor, volumen calculado, volumen almacenado y volumen interno del circuito. Estas magnitudes no deben sumarse sin definir claramente qué magnitud física se está midiendo.\n\n7. CÓMO TRANSFORMAR LA ESTIMACIÓN EN UNA MEDICIÓN ESPECÍFICA DEL VEHÍCULO\n------------------------------------------------------------------------\nPara un resultado específico del vehículo, informar:\n- diámetro interior real de la tubería;\n- longitud real de la línea de alta presión;\n- cantidad y volumen interno de las válvulas;\n- fabricante/modelo del reductor;\n- número de etapas del reductor;\n- volumen interno de las cámaras de gas;\n- volumen del filtro de fase gaseosa;\n- volumen de la flauta/colector de inyectores;\n- volumen interno de inyectores y conexiones;\n- documentación técnica del fabricante, cuando esté disponible.\n\n8. REFERENCIAS DEL ESTUDIO TÉCNICO\n----------------------------------\nINMETRO — Portaria nº 111/2022 y requisitos relacionados con componentes GNV/GNC:\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1, sistemas de vehículos a gas natural:\nhttps://www.abntcatalogo.com.br/\n\nUSP — Escola Politécnica / Biblioteca Digital de Tesis y Disertaciones:\nhttps://teses.usp.br/\n\nUFRGS — Centro de Referencia para la Enseñanza de Física (CREF):\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / Repositorio Institucional:\nhttps://pantheon.ufrj.br/\n\nITA — Instituto Tecnológico de Aeronáutica:\nhttps://www.ita.br/\n\nIME — Instituto Militar de Engenharia:\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science:\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering:\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics:\nhttps://ocw.mit.edu/\n\nNIST — REFPROP y propiedades termofísicas:\nhttps://www.nist.gov/srd/refprop\n\nFUENTES Y FUNDAMENTO TÉCNICO\n============================\nANP — Publicación de precios del gas natural:\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — Glosario / Condición estándar de medición:\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — cálculo del factor de compresibilidad mediante composición:\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — cálculo mediante propiedades físicas:\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics: ecuación del gas ideal PV = nRT:\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics:\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law:\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow:\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — programa académico relacionado con termodinámica:\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / ecuaciones de estado del gas natural:\nhttps://www.nist.gov/srd/refprop\n', 'Deutsch': 'FORMELN UND PHYSIK DES GNV-BERECHNUNGSSYSTEMS\n============================================================\n\nTEIL A — ANP-REFERENZBEDINGUNG\n==============================\n\nDie ANP verwendet für durchschnittlich vermarktete Erdgasvolumina Referenzbedingungen von 20 °C und 1,033 kgf/cm². Die Standardmessbedingung wird als Absolutdruck von 0,101325 MPa bei 20 °C angegeben.\n\nWICHTIG: Die ANP legt die Referenzbedingung fest. Die in diesem Programm implementierte Formel ist eine PHYSIKALISCHE UMWERTUNGSSCHÄTZUNG und beansprucht nicht, den internen Algorithmus einer GNV-Tankstelle nachzubilden.\n\n1. PHYSIKALISCHES ZYLINDERVOLUMEN\n---------------------------------\nVcil = Kapazität(L) / 1000\n\n26 L / 1000 = 0,026 m³.\n\nDies ist der physikalische Innenraum des Zylinders. Es ist nicht das normierte Volumen, das von der Zapfsäule angezeigt wird.\n\n2. ABSOLUTDRUCK\n---------------\nPabs = Manometerdruck + Patm\n\nZustandsgleichungen verwenden den Absolutdruck.\n\n3. ABSOLUTTEMPERATUR\n--------------------\nT(K) = T(°C) + 273,15\n\n4. UMRECHNUNG AUF 20 °C\n-----------------------\nFür eine feste Stoffmenge im idealen Gasmodell:\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293,15 K und Pref ≈ 1,01325 bar.\n\nTEIL B — WISSENSCHAFTLICHES ECHTGASMODELL\n==========================================\n\n1. ZUSTANDSGLEICHUNG\n--------------------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = Absolutdruck (Pa)\nV = physikalisches Volumen (m³)\nZ = Kompressibilitätsfaktor\nn = Stoffmenge (mol)\nR = 8,314462618 J/(mol·K)\nT = absolute Temperatur (K)\n\n2. KOMPRESSIBILITÄTSFAKTOR Z\n----------------------------\nZ = P V / (n R T)\n\nZ = 1 entspricht einem idealen Gas. Bei realem Erdgas hängt Z von Druck, Temperatur und Zusammensetzung ab. Ein vom Benutzer eingegebener fester Z-Wert ist eine Näherung und keine metrologische Bestimmung von Z.\n\n3. HINZUGEFÜGTE GAS-MENGE\n-------------------------\nn_initial = P_initial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_initial\n\n4. HINZUGEFÜGTE MASSE\n---------------------\nm = Δn × M\n\nM ist die molare Masse des GNV in kg/mol.\n\n5. ÄQUIVALENTES VOLUMEN BEI DER EINGEGEBENEN TEMPERATUR\n--------------------------------------------------------\nDas Programm berechnet zunächst die Stoffmenge aus PV = Z n R T. Danach kann dieselbe Stoffmenge bei einem Referenzdruck ausgedrückt werden:\n\nVref(T) = n R T / Pref\n\nWICHTIG: In diesem Programm wird n mit der eingegebenen Temperatur berechnet. Durch Einsetzen von n = P V / (Z R T) kürzt sich die Temperatur heraus:\n\nVref(T) = P V / (Z Pref)\n\nFür einen 26-L-Zylinder bei 220 bar und Z=0,92 kann das äquivalente Volumen bei 1,01325 bar und der eingegebenen Temperatur daher bei etwa 6,164 m³ bei 5 °C, 20 °C oder 100 °C bleiben. Dies bedeutet NICHT, dass dieselbe feste Gasmenge bei unterschiedlichen Temperaturen dasselbe Volumen besitzt; es bedeutet, dass das Programm die Stoffmenge für jeden eingegebenen Druck-/Temperaturzustand neu berechnet.\n\n6. WISSENSCHAFTLICHE UMRECHNUNG AUF 20 °C\n------------------------------------------\nDieselben mit dem eingegebenen Z-Modell berechneten Mol werden auf 20 °C umgerechnet:\n\nV20 = n R T20 / Pref\n\nLiegt die eingegebene Temperatur unter 20 °C, ist V20 tendenziell größer. Liegt sie über 20 °C, ist V20 tendenziell kleiner, weil n während dieser Umrechnung konstant gehalten wird.\n\n7. ANP/IDEALISIERTE UMRECHNUNG (Z=1)\n------------------------------------\nDie ANP-Registerkarte berechnet separat eine Schätzung mit Z=1 und der Referenzbedingung 20 °C / 1,033 kgf/cm² (etwa 1,01325 bar). Dieser Wert darf nicht mit dem wissenschaftlichen Ergebnis mit Z=0,92 verwechselt werden. Es handelt sich um unterschiedliche Modelle.\n\n8. DICHTE DES REALGASES\n-----------------------\nρ = P M / (Z R T)\n\nTEIL C — WAS EIN GENAUERES MODELL ERFORDERN WÜRDE\n=================================================\n\nBei Hochdruck-GNV ist es nicht angemessen, Z als universelle Konstante zu betrachten. Für höhere Genauigkeit sind Gaszusammensetzung und thermodynamische Eigenschaften als Funktionen von P und T erforderlich.\n\nISO 12213 beschreibt Verfahren zur Berechnung des Kompressibilitätsfaktors von Erdgas. ISO 12213-2 verwendet die molare Zusammensetzung; ISO 12213-3 verwendet physikalische Eigenschaften wie Heizwert, relative Dichte und CO₂ zusammen mit Druck und Temperatur.\n\nAGA8 und GERG sind Modelle für Erdgas-Eigenschaften. Das NIST beschreibt AGA8 und GERG als Zustandsgleichungen für Anwendungen der Erdgasmesstechnik und thermodynamischer Eigenschaften.\n\nDaher sollte die wissenschaftliche Weiterentwicklung des Programms folgende Schritte umfassen:\n1) GNV-Zusammensetzung bestimmen;\n2) Z(P,T,Zusammensetzung) statt eines festen Z berechnen;\n3) die tatsächliche Gastemperatur während der Betankung berücksichtigen;\n4) die vom Messgerät tatsächlich verwendeten Bedingungen kennen;\n5) Messunsicherheit berücksichtigen.\n\nTEIL D — TEMPERATUR WÄHREND DER BETANKUNG\n==========================================\n\nWährend der Befüllung tritt Masse ein und Wärme wird übertragen. Die Registerkarte Erwärmung / Kompression verwendet reversible adiabatische Kompression ausschließlich als Lehrszenario:\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = konstant\nV₂/V₁ = (P₁/P₂)^(1/k)\n\nDie reale Betankung ist ein offenes System mit Wärmeübertragung zwischen Gas, Zylinderwand, Schlauch und Umgebung. Die in dieser Registerkarte berechnete Temperatur ist KEINE Messung der tatsächlichen GNV-Temperatur.\n\nTEIL E — VERGLEICH MIT DER ZAPFSÄULE\n=====================================\n\nDifferenz = von der Zapfsäule angezeigtes Volumen − berechnetes Volumen\nProzentuale Differenz = Differenz / berechnetes Volumen × 100\n\nEine große Differenz ist ein Hinweis für eine Untersuchung. Allein stellt sie keinen metrologischen Nachweis eines Betrugs dar. Eine technische Schlussfolgerung erfordert Messgerätedaten, Referenzbedingungen, tatsächliche Gastemperatur, Zusammensetzung/Z, Kalibrierung und Messunsicherheit.\n\nTEIL F — GEOMETRISCHES VOLUMEN DES GNV-ZUBEHÖRKREISLAUFS\n=========================================================\n\nZIEL\n----\nSchätzung der minimalen, mittleren und maximalen Innenvolumina des Hoch- und Niederdruckkreislaufs des GNV-Systems, wobei ZYLINDER/RESERVOIR STRENG AUSGESCHLOSSEN WIRD. Die Studie berücksichtigt den Innenraum von Leitungen, Ventilen, Regler/Druckminderer und Niederdruckkomponenten.\n\nWICHTIG ZUR NATUR DIESER BERECHNUNG\n-----------------------------------\nDiese Werte sind technische Referenzschätzungen auf Grundlage des für dieses Projekt bereitgestellten TECHNISCHEN BERICHTS. Sie sind keine universellen Spezifikationen für alle GNV-Kits. Für eine fahrzeugspezifische Messung müssen tatsächlicher Innendurchmesser, Länge, Druckminderermodell, Ventile, Filter, Injektorleiste und weitere reale Komponenten verwendet werden.\n\n1. HOCHDRUCKLEITUNG\n-------------------\nDie Studienannahme ist ein Rohr mit 6 mm Außendurchmesser und 1,2 mm Wandstärke, wodurch ein Innendurchmesser von 3,6 mm entsteht.\nD_innen = D_außen - 2 × Wandstärke\nD_innen = 6,0 - 2 × 1,2 = 3,6 mm\nr = 1,8 mm = 0,0018 m\n\nFür ein gerades zylindrisches Rohr:\nV_t = pi × r² × L\n\nBei L = 4,0 m: V_t ≈ 40,7 mL\nBei L = 5,5 m: V_t ≈ 70,1 mL\n\nDer ungefähre geometrische Bereich beträgt daher 40 bis 70 mL.\n\n2. SERVICEVENTILE\n-----------------\nDer technische Bericht schätzt das kombinierte Totvolumen der Zylinder- und Betankungsventile auf 3 bis 5 mL.\n\n3. DRUCKMINDERER / DRUCKREGLER\n------------------------------\nDie Studie berücksichtigt die ausschließlich für den Gasdurchgang verwendeten Kammern des Reglers/Druckminderers, ohne den Wasser-Heizkreislauf. Der geschätzte Bereich beträgt 30 bis 65 mL.\n\n4. NIEDERDRUCK / FILTER / INJEKTORLEISTE\n-----------------------------------------\nFür moderne Systeme berücksichtigt der Bericht Niederdruckschläuche, Gasphasenfilter und Injektorleiste/Verteiler. Der geschätzte Bereich beträgt 50 bis 140 mL. Falls dieser Abschnitt nicht vorhanden ist, kann sein Beitrag mit 0 mL angesetzt werden.\n\n5. ZUSAMMENFASSUNG DER STUDIE\n-----------------------------\nKomponente                         Minimum    Mittel    Maximum\nHochdruckleitung                    40 mL      55 mL      70 mL\nServiceventile                       3 mL       4 mL       5 mL\nDruckminderer — Gaskammern          32 mL      46 mL      65 mL\nNiederdruck/Filter/Leiste            0 mL      25 mL     140 mL\nGESAMTVOLUMEN KREISLAUF             75 mL     130 mL     280 mL\n\nDie Studie verwendet 130 mL als repräsentativen mittleren geometrischen Referenzwert, ohne zu behaupten, dass dieser Wert eine universelle Messung der Fahrzeugflotte ist.\n\n6. WARUM MUSS DER KREISLAUF GETRENNT VOM ZYLINDER BETRACHTET WERDEN?\n--------------------------------------------------------------------\nDer Zylinder ist ein großer physischer Gasspeicher. Leitungen, Ventile, Regler/Druckminderer und Niederdruckkomponenten bilden zusätzliche Innenvolumina des Kreislaufs. Deshalb können diese Räume bei einer geometrischen Volumenanalyse getrennt erfasst werden.\n\nDies ist besonders wichtig beim Vergleich von Zapfsäulenanzeige, berechnetem Volumen, gespeichertem Volumen und internem Kreislaufvolumen. Diese Größen dürfen nicht addiert werden, ohne eindeutig festzulegen, welche physikalische Größe gemessen wird.\n\n7. WIE AUS DER SCHÄTZUNG EINE FAHRZEUGSPEZIFISCHE MESSUNG WIRD\n---------------------------------------------------------------\nFür ein fahrzeugspezifisches Ergebnis angeben:\n- tatsächlicher Innendurchmesser der Leitung;\n- tatsächliche Länge der Hochdruckleitung;\n- Anzahl und Innenvolumen der Ventile;\n- Hersteller/Modell des Druckminderers;\n- Anzahl der Druckmindererstufen;\n- Innenvolumen der Gaskammern;\n- Volumen des Gasphasenfilters;\n- Volumen der Injektorleiste/des Verteilers;\n- Innenvolumen der Injektoren und Anschlüsse;\n- technische Herstellerdokumentation, sofern vorhanden.\n\n8. REFERENZEN DER TECHNISCHEN STUDIE\n------------------------------------\nINMETRO — Portaria nº 111/2022 und Anforderungen an GNV/GNC-Komponenten:\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1, Erdgas-Fahrzeugsysteme:\nhttps://www.abntcatalogo.com.br/\n\nUSP — Escola Politécnica / Digitale Bibliothek für Dissertationen und Abschlussarbeiten:\nhttps://teses.usp.br/\n\nUFRGS — Zentrum für Physikunterricht (CREF):\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / Institutionelles Repository:\nhttps://pantheon.ufrj.br/\n\nITA — Instituto Tecnológico de Aeronáutica:\nhttps://www.ita.br/\n\nIME — Instituto Militar de Engenharia:\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science:\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering:\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics:\nhttps://ocw.mit.edu/\n\nNIST — REFPROP und thermophysikalische Eigenschaften:\nhttps://www.nist.gov/srd/refprop\n\nQUELLEN UND TECHNISCHE GRUNDLAGEN\n=================================\nANP — Veröffentlichung von Erdgaspreisen:\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — Glossar / Standardmessbedingung:\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — Berechnung des Kompressibilitätsfaktors anhand der Zusammensetzung:\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — Berechnung anhand physikalischer Eigenschaften:\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics: ideale Gasgleichung PV = nRT:\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics:\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law:\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow:\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — akademischer Lehrplan mit Bezug zur Thermodynamik:\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / Zustandsgleichungen für Erdgas:\nhttps://www.nist.gov/srd/refprop\n', '日本語': 'GNV計算システムの数式と物理\n============================================================\n\n第A部 — ANP基準条件\n====================\n\nANPは、平均的に販売される天然ガスの体積について、20 °Cおよび1.033 kgf/cm²を基準条件として使用しています。標準測定条件は、20 °Cにおける絶対圧力0.101325 MPaとして表されます。\n\n重要：ANPは基準条件を定めています。本プログラムで実装している式は「物理的な換算の推定」であり、GNVディスペンサー内部のアルゴリズムを再現するものではありません。\n\n1. シリンダーの物理容積\n------------------------\nVcil = 容量(L) / 1000\n\n26 L / 1000 = 0.026 m³。\n\nこれはシリンダー内部の物理的な空間です。ポンプが表示する基準化体積ではありません。\n\n2. 絶対圧力\n-----------\nPabs = ゲージ圧力 + Patm\n\n状態方程式では絶対圧力を使用します。\n\n3. 絶対温度\n-----------\nT(K) = T(°C) + 273.15\n\n4. 20 °Cへの換算\n-----------------\n物質量を一定とした理想気体モデルでは：\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293.15 K、Pref ≈ 1.01325 bar。\n\n第B部 — 実在気体の科学モデル\n============================\n\n1. 状態方程式\n-------------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = 絶対圧力 (Pa)\nV = 物理容積 (m³)\nZ = 圧縮係数\nn = 物質量 (mol)\nR = 8.314462618 J/(mol·K)\nT = 絶対温度 (K)\n\n2. 圧縮係数Z\n------------\nZ = P V / (n R T)\n\nZ = 1は理想気体を表します。実在する天然ガスではZは圧力、温度、組成によって変化します。ユーザーが入力した固定Zは近似値であり、計量学的に決定されたZではありません。\n\n3. 追加されたガス量\n--------------------\nn_initial = P_initial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_initial\n\n4. 追加質量\n-----------\nm = Δn × M\n\nMはGNVのモル質量 (kg/mol) です。\n\n5. 入力温度における等価体積\n----------------------------\nプログラムはまずPV = Z n R Tから物質量を計算します。その後、同じ物質量を基準圧力で表すことができます：\n\nVref(T) = n R T / Pref\n\n重要：本プログラムではnを入力温度を使用して計算します。そのためn = P V / (Z R T)を代入すると温度が相殺されます：\n\nVref(T) = P V / (Z Pref)\n\nしたがって、26 L、220 bar、Z=0.92のシリンダーでは、入力温度における1.01325 bar換算の体積は5 °C、20 °C、100 °Cで約6.164 m³のままになる場合があります。これは、同じ固定されたガス量が異なる温度で同じ体積になるという意味ではありません。入力された圧力・温度状態ごとにプログラムが物質量を再計算しているという意味です。\n\n6. 20 °Cへの科学的換算\n-----------------------\n入力されたZモデルで計算した同じモル数を20 °Cへ換算します：\n\nV20 = n R T20 / Pref\n\n入力温度が20 °C未満ならV20は大きくなる傾向があります。20 °Cを超える場合は、換算中にnを一定とするためV20は小さくなる傾向があります。\n\n7. ANP/理想化換算 (Z=1)\n-------------------------\nANPタブでは、Z=1および20 °C / 1.033 kgf/cm²（約1.01325 bar）の基準条件を使用した推定値を別に計算します。この値をZ=0.92を使用する科学モデルの結果と混同してはいけません。異なるモデルです。\n\n8. 実在気体の密度\n------------------\nρ = P M / (Z R T)\n\n第C部 — より高精度なモデルに必要なもの\n========================================\n\n高圧GNVではZを普遍的な定数として扱うことは適切ではありません。高精度化には、PとTの関数としてのガス組成および熱力学的性質が必要です。\n\nISO 12213は天然ガスの圧縮係数を計算する方法を規定しています。ISO 12213-2はモル組成を使用し、ISO 12213-3は発熱量、相対密度、CO₂などの物性値と圧力・温度を使用します。\n\nAGA8およびGERGは天然ガスの性質に使用されるモデルです。NISTは天然ガスの計量および熱力学的性質の用途で使用される状態方程式としてAGA8とGERGを説明しています。\n\nしたがって、本プログラムの科学的発展には次が必要です：\n1) GNVの組成を取得する;\n2) 固定ZではなくZ(P,T,組成)を計算する;\n3) 給油中の実際のガス温度を考慮する;\n4) メーターが実際に使用する条件を把握する;\n5) 測定不確かさを考慮する。\n\n第D部 — 給油中の温度\n=====================\n\n充填中は質量が流入し、熱移動が発生します。加熱/圧縮タブでは可逆断熱圧縮を教育用シナリオとしてのみ使用します：\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = constant\nV₂/V₁ = (P₁/P₂)^(1/k)\n\n実際の給油は、ガス、シリンダー壁、ホース、周囲環境の間で熱が移動する開放系です。このタブで計算される温度は実際のGNV温度の測定値ではありません。\n\n第E部 — ポンプとの比較\n=======================\n\n差 = ポンプ表示体積 − 計算体積\n差の百分率 = 差 / 計算体積 × 100\n\n大きな差は調査のきっかけになります。しかし、それだけで不正の計量学的証明にはなりません。技術的な結論には、メーターのデータ、基準条件、実ガス温度、組成/Z、校正、測定不確かさが必要です。\n\n第F部 — GNV付属回路の幾何学的体積\n==================================\n\n目的\n----\nシリンダー/リザーバーを厳密に除外し、GNVの高圧・低圧回路の内部体積の最小・平均・最大値を推定します。配管、バルブ、レギュレーター/減圧器、低圧部品の内部空間を対象とします。\n\nこの計算の性質についての重要事項\n----------------------------------\nこれらの値は、本プロジェクトで提供された技術報告書に基づく工学的な参考推定値です。すべてのGNVキットに共通する仕様ではありません。車両ごとの測定には実際の内径、長さ、減圧器モデル、バルブ、フィルター、インジェクターレールなどの実部品データを使用する必要があります。\n\n1. 高圧配管\n-----------\n研究上の前提は、外径6 mm、肉厚1.2 mmの配管で、内径3.6 mmとなります。\nD_internal = D_external - 2 × wall thickness\nD_internal = 6.0 - 2 × 1.2 = 3.6 mm\nr = 1.8 mm = 0.0018 m\n\n直線円筒管では：\nV_t = pi × r² × L\n\nL = 4.0 m：V_t ≈ 40.7 mL\nL = 5.5 m：V_t ≈ 70.1 mL\n\nしたがって、概算の幾何学的範囲は40～70 mLです。\n\n2. サービスバルブ\n-----------------\n技術報告書では、シリンダーバルブと充填サービスバルブの合計デッドボリュームを3～5 mLと推定しています。\n\n3. 圧力減圧器 / レギュレーター\n------------------------------\nガスの通過専用のレギュレーター/減圧器内部チャンバーを対象とし、水による加熱回路は除外します。推定範囲は30～65 mLです。\n\n4. 低圧 / フィルター / インジェクターレール\n---------------------------------------------\n現代的なシステムでは、低圧ホース、気相フィルター、インジェクターレール/マニホールドを対象とします。推定範囲は50～140 mLです。この部分が存在しない構成では0 mLとできます。\n\n5. 研究結果の統合\n------------------\n構成部品                         最小       平均       最大\n高圧配管                         40 mL      55 mL      70 mL\nサービスバルブ                   3 mL       4 mL       5 mL\n減圧器 — ガスチャンバー         32 mL      46 mL      65 mL\n低圧/フィルター/レール            0 mL      25 mL     140 mL\n回路総体積                       75 mL     130 mL     280 mL\n\n本研究では130 mLを代表的な平均幾何学基準値として採用します。ただし、車両全体に普遍的な測定値であることを意味しません。\n\n6. なぜ回路をシリンダーと分けて扱うのか\n------------------------------------------\nシリンダーは大きな物理的ガス貯蔵容器です。一方、配管、バルブ、レギュレーター/減圧器、低圧部品は回路内に追加の内部体積を形成します。したがって、幾何学的体積解析ではこれらを分離して計上できます。\n\nこれは、給油機表示体積、計算体積、貯蔵体積、回路内部体積を比較する場合に特に重要です。測定対象の物理量を明確に定義せず、これらの体積を単純に合計してはいけません。\n\n7. 推定値を車両固有の測定値にする方法\n--------------------------------------\n車両固有の結果には次を入力します：\n- 配管の実内径;\n- 高圧配管の実長;\n- バルブの数量と内部体積;\n- 減圧器のメーカー/モデル;\n- 減圧器の段数;\n- ガスチャンバーの内部体積;\n- 気相フィルターの体積;\n- インジェクターレール/マニホールドの体積;\n- インジェクターおよび継手の内部体積;\n- 入手可能な場合はメーカーの技術資料。\n\n8. 技術研究の参考資料\n----------------------\nINMETRO — Portaria nº 111/2022およびGNV/GNC部品関連要件：\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1、天然ガス自動車システム：\nhttps://www.abntcatalogo.com.br/\n\nUSP — Escola Politécnica / 学位論文・学術論文デジタルライブラリ：\nhttps://teses.usp.br/\n\nUFRGS — 物理教育リファレンスセンター (CREF)：\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / 機関リポジトリ：\nhttps://pantheon.ufrj.br/\n\nITA — 航空技術研究所：\nhttps://www.ita.br/\n\nIME — 軍事工学研究所：\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science：\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering：\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics：\nhttps://ocw.mit.edu/\n\nNIST — REFPROPおよび熱物性：\nhttps://www.nist.gov/srd/refprop\n\n参考資料と技術的根拠\n====================\nANP — 天然ガス価格の公表：\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — 用語集 / 標準測定条件：\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — 組成による圧縮係数計算：\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — 物性値による計算：\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics：理想気体方程式 PV = nRT：\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics：\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law：\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow：\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — 熱力学関連の学術カリキュラム：\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / 天然ガス状態方程式：\nhttps://www.nist.gov/srd/refprop\n', '中文': 'GNV计算系统公式与物理\n============================================================\n\nA部分 — ANP参考条件\n====================\n\nANP针对平均商业化天然气体积采用20 °C和1.033 kgf/cm²的参考条件。标准计量条件表示为20 °C时的绝对压力0.101325 MPa。\n\n重要：ANP规定参考条件。本程序中的公式属于“物理换算估算”，并不声称复制GNV加气机内部算法。\n\n1. 气瓶物理容积\n----------------\nVcil = 容量(L) / 1000\n\n26 L / 1000 = 0.026 m³。\n\n这是气瓶内部的物理空间，不是加气机显示的标准化体积。\n\n2. 绝对压力\n------------\nPabs = 表压 + Patm\n\n状态方程必须使用绝对压力。\n\n3. 绝对温度\n------------\nT(K) = T(°C) + 273.15\n\n4. 换算到20 °C\n---------------\n对于固定物质的量，在理想气体模型中：\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293.15 K，Pref ≈ 1.01325 bar。\n\nB部分 — 真实气体科学模型\n==========================\n\n1. 状态方程\n-----------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = 绝对压力 (Pa)\nV = 物理体积 (m³)\nZ = 压缩因子\nn = 物质的量 (mol)\nR = 8.314462618 J/(mol·K)\nT = 绝对温度 (K)\n\n2. 压缩因子Z\n------------\nZ = P V / (n R T)\n\nZ = 1表示理想气体。真实天然气的Z取决于压力、温度和组成。用户输入的固定Z只是近似值，不是计量学意义上的Z测定值。\n\n3. 增加的气体量\n----------------\nn_initial = P_initial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_initial\n\n4. 增加的质量\n--------------\nm = Δn × M\n\nM为GNV摩尔质量，单位kg/mol。\n\n5. 输入温度下的等效体积\n------------------------\n程序首先通过PV = Z n R T计算物质的量，然后可以在参考压力下表示相同物质的量：\n\nVref(T) = n R T / Pref\n\n重要：在本程序中，n使用输入温度计算。因此代入n = P V / (Z R T)后，温度会相消：\n\nVref(T) = P V / (Z Pref)\n\n因此，对于26 L、220 bar、Z=0.92的气瓶，在输入温度下换算到1.01325 bar的等效体积可能在5 °C、20 °C或100 °C时都约为6.164 m³。这并不意味着同一固定气体量在不同温度下体积相同，而是因为程序针对每个输入的压力/温度状态重新计算了物质的量。\n\n6. 科学换算到20 °C\n-------------------\n使用输入Z模型计算出的相同摩尔数换算到20 °C：\n\nV20 = n R T20 / Pref\n\n如果输入温度低于20 °C，V20趋于更大；如果高于20 °C，V20趋于更小，因为在此换算过程中n保持不变。\n\n7. ANP/理想化换算 (Z=1)\n------------------------\nANP选项卡另外使用Z=1和20 °C / 1.033 kgf/cm²（约1.01325 bar）的参考条件进行估算。该值不能与使用Z=0.92的科学结果混淆，它们属于不同模型。\n\n8. 真实气体密度\n----------------\nρ = P M / (Z R T)\n\nC部分 — 更精确模型需要什么\n============================\n\n对于高压GNV，把Z视为通用常数并不合适。更高精度需要获得气体组成，并根据P和T计算热力学性质。\n\nISO 12213描述天然气压缩因子的计算方法。ISO 12213-2使用摩尔组成；ISO 12213-3使用热值、相对密度和CO₂等物理性质以及压力和温度。\n\nAGA8和GERG是天然气性质计算中使用的模型。NIST将AGA8和GERG列为天然气计量和热力学性质应用中使用的状态方程模型。\n\n因此，本程序的科学升级应包括：\n1) 获取GNV组成；\n2) 计算Z(P,T,组成)，而不是使用固定Z；\n3) 考虑加气过程中的实际气体温度；\n4) 获得计量器实际采用的条件；\n5) 考虑测量不确定度。\n\nD部分 — 加气过程中的温度\n=========================\n\n充装过程中质量进入系统，同时发生热量传递。加热/压缩选项卡仅将可逆绝热压缩作为教学场景：\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = 常数\nV₂/V₁ = (P₁/P₂)^(1/k)\n\n实际加气是开放系统，气体、气瓶壁、软管和环境之间存在热传递。本选项卡计算的温度不是实际GNV温度的测量值。\n\nE部分 — 与加气机比较\n=====================\n\n差值 = 加气机显示体积 − 计算体积\n百分比差值 = 差值 / 计算体积 × 100\n\n较大的差异是需要调查的迹象，但单凭此结果不能证明计量欺诈。技术结论需要计量器数据、参考条件、实际气体温度、组成/Z、校准以及测量不确定度。\n\nF部分 — GNV附属回路的几何体积\n==============================\n\n目标\n----\n估算GNV高压和低压回路内部体积的最小、平均和最大值，并严格排除气瓶/储气罐。研究范围包括管路、阀门、调压器/减压器和低压部件的内部空间。\n\n关于计算性质的重要说明\n------------------------\n这些数值是基于本项目提供的技术报告得到的工程参考估算，并不是所有GNV套件的通用规格。针对具体车辆的测量必须采用实际内径、长度、减压器型号、阀门、过滤器、喷油器轨等真实部件数据。\n\n1. 高压管路\n-----------\n研究假设管道外径6 mm、壁厚1.2 mm，因此内径为3.6 mm。\nD_internal = D_external - 2 × 壁厚\nD_internal = 6.0 - 2 × 1.2 = 3.6 mm\nr = 1.8 mm = 0.0018 m\n\n直圆柱管道：\nV_t = pi × r² × L\n\nL = 4.0 m：V_t ≈ 40.7 mL\nL = 5.5 m：V_t ≈ 70.1 mL\n\n因此，近似几何范围为40～70 mL。\n\n2. 服务阀门\n-----------\n技术报告估计气瓶阀和加气服务阀的合计死体积为3～5 mL。\n\n3. 压力减压器/调压器\n--------------------\n研究考虑调压器/减压器中专门用于气体通过的腔室，排除水加热回路。估计范围为30～65 mL。\n\n4. 低压/过滤器/喷油器轨\n------------------------\n对于现代系统，报告考虑低压软管、气相过滤器和喷油器轨/歧管。估计范围为50～140 mL。如果没有这一部分，可按0 mL计算。\n\n5. 研究汇总\n-----------\n部件                              最小       平均       最大\n高压管路                          40 mL      55 mL      70 mL\n服务阀门                           3 mL       4 mL       5 mL\n减压器 — 气体腔室                 32 mL      46 mL      65 mL\n低压/过滤器/轨                     0 mL      25 mL     140 mL\n回路总容积                        75 mL     130 mL     280 mL\n\n本研究采用130 mL作为具有代表性的平均几何参考值，但不声称它是整个车辆群体的通用测量值。\n\n6. 为什么回路必须与气瓶分开处理？\n----------------------------------\n气瓶是大型物理储气容器。管路、阀门、调压器/减压器和低压部件构成回路中的额外内部体积。因此，在几何体积分析中，这些空间可以单独计量。\n\n当比较加气机显示体积、计算体积、储存体积和回路内部体积时，这一点尤其重要。在没有明确测量物理量的情况下，不应简单地把这些体积相加。\n\n7. 如何把估算转化为车辆实际测量\n--------------------------------\n要得到具体车辆结果，需要提供：\n- 管道实际内径；\n- 高压管路实际长度；\n- 阀门数量及内部体积；\n- 减压器制造商/型号；\n- 减压器级数；\n- 气体腔室内部体积；\n- 气相过滤器体积；\n- 喷油器轨/歧管体积；\n- 喷油器和接头内部体积；\n- 有条件时提供制造商技术资料。\n\n8. 技术研究参考资料\n--------------------\nINMETRO — Portaria nº 111/2022以及GNV/GNC部件相关要求：\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1，天然气车辆系统：\nhttps://www.abntcatalogo.com.br/\n\nUSP — Escola Politécnica / 学位论文与学位论文数字图书馆：\nhttps://teses.usp.br/\n\nUFRGS — 物理教学参考中心(CREF)：\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / 机构知识库：\nhttps://pantheon.ufrj.br/\n\nITA — 航空技术学院：\nhttps://www.ita.br/\n\nIME — 军事工程学院：\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science：\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering：\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics：\nhttps://ocw.mit.edu/\n\nNIST — REFPROP和热物性：\nhttps://www.nist.gov/srd/refprop\n\n资料来源与技术依据\n==================\nANP — 天然气价格公布：\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — 术语表 / 标准测量条件：\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — 根据组成计算压缩因子：\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — 根据物理性质计算：\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics：理想气体状态方程PV = nRT：\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics：\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law：\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow：\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — 与热力学相关的学术课程：\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / 天然气状态方程：\nhttps://www.nist.gov/srd/refprop\n', 'English': 'FORMULAS AND PHYSICS OF THE CNG CALCULATION SYSTEM\n============================================================\n\nPART A — ANP REFERENCE CONDITION\n================================\n\nANP uses reference conditions of 20 °C and 1.033 kgf/cm² for average marketed natural-gas volumes. The standard measurement condition is expressed as an absolute pressure of 0.101325 MPa at 20 °C.\n\nIMPORTANT: ANP defines the reference condition. The formula implemented by this program is a PHYSICAL CONVERSION ESTIMATE and does not claim to reproduce the internal algorithm of a CNG dispenser.\n\n1. PHYSICAL CYLINDER VOLUME\n----------------------------\nVcil = capacity(L) / 1000\n\n26 L / 1000 = 0.026 m³.\n\nThis is the physical internal space of the cylinder. It is not the normalized volume indicated by the pump.\n\n2. ABSOLUTE PRESSURE\n--------------------\nPabs = gauge pressure + Patm\n\nEquations of state use absolute pressure.\n\n3. ABSOLUTE TEMPERATURE\n-----------------------\nT(K) = T(°C) + 273.15\n\n4. CONVERSION TO 20 °C\n-----------------------\nFor a fixed amount of substance, in the ideal-gas model:\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293.15 K and Pref ≈ 1.01325 bar.\n\nPART B — SCIENTIFIC REAL-GAS MODEL\n===================================\n\n1. EQUATION OF STATE\n--------------------\nP V = Z n R T\nn = P V / (Z R T)\n\nP = absolute pressure (Pa)\nV = physical volume (m³)\nZ = compressibility factor\nn = amount of substance (mol)\nR = 8.314462618 J/(mol·K)\nT = absolute temperature (K)\n\n2. COMPRESSIBILITY FACTOR Z\n----------------------------\nZ = P V / (n R T)\n\nZ = 1 represents an ideal gas. For real natural gas, Z depends on pressure, temperature and composition. A fixed Z entered by the user is an approximation, not a metrological determination of Z.\n\n3. AMOUNT OF GAS ADDED\n----------------------\nn_initial = P_initial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_initial\n\n4. MASS ADDED\n-------------\nm = Δn × M\n\nM is the CNG molar mass in kg/mol.\n\n5. EQUIVALENT VOLUME AT THE ENTERED TEMPERATURE\n-----------------------------------------------\nThe program first calculates the amount of substance from PV = Z n R T. It can then express the same amount of substance at a reference pressure:\n\nVref(T) = n R T / Pref\n\nIMPORTANT: in this program, n is calculated using the entered temperature. Therefore, substituting n = P V / (Z R T) causes temperature to cancel:\n\nVref(T) = P V / (Z Pref)\n\nThus, for a 26 L cylinder at 220 bar and Z=0.92, the equivalent volume at 1.01325 bar at the entered temperature may remain approximately 6.164 m³ at 5 °C, 20 °C or 100 °C. This does NOT mean that the same fixed amount of gas has the same volume at different temperatures; it means that the program recalculates the amount of substance for each entered pressure/temperature state.\n\n6. SCIENTIFIC CONVERSION TO 20 °C\n---------------------------------\nThe same moles calculated by the entered-Z model are converted to 20 °C:\n\nV20 = n R T20 / Pref\n\nIf the entered temperature is below 20 °C, V20 tends to be larger. If it is above 20 °C, V20 tends to be smaller because n is held fixed during this conversion.\n\n7. ANP/IDEALIZED CONVERSION (Z=1)\n---------------------------------\nThe ANP tab separately calculates an estimate using Z=1 and the 20 °C / 1.033 kgf/cm² reference condition (approximately 1.01325 bar). This value must not be confused with the scientific result using Z=0.92. They are different models.\n\n8. REAL-GAS DENSITY\n--------------------\nρ = P M / (Z R T)\n\nPART C — WHAT A MORE PRECISE MODEL WOULD REQUIRE\n=================================================\n\nFor high-pressure CNG, treating Z as a universal constant is not adequate. Higher precision requires the gas composition and thermodynamic properties as functions of P and T.\n\nISO 12213 describes methods for calculating the compressibility factor of natural gas. ISO 12213-2 uses molar composition; ISO 12213-3 uses physical properties such as calorific value, relative density and CO₂, together with pressure and temperature.\n\nAGA8 and GERG are models used for natural-gas properties. NIST describes AGA8 and GERG among equations of state used in natural-gas measurement and thermodynamic-property applications.\n\nTherefore, the scientific evolution of the program should be:\n1) obtain CNG composition;\n2) calculate Z(P,T,composition) instead of using a fixed Z;\n3) consider the actual gas temperature during refueling;\n4) know the conditions actually used by the meter;\n5) account for measurement uncertainty.\n\nPART D — TEMPERATURE DURING REFUELING\n=====================================\n\nDuring filling, mass enters and heat transfer occurs. The Heating / Compression tab uses reversible adiabatic compression only as an educational scenario:\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = constant\nV₂/V₁ = (P₁/P₂)^(1/k)\n\nActual refueling is an open system with heat transfer among the gas, cylinder wall, hose and surroundings. The temperature calculated in this tab is NOT a measurement of the actual CNG temperature.\n\nPART E — COMPARISON WITH THE PUMP\n==================================\n\nDifference = volume indicated by the pump − calculated volume\nPercentage difference = difference / calculated volume × 100\n\nA large difference is an indication for investigation. By itself, it is not metrological proof of fraud. A technical conclusion requires meter data, reference conditions, actual gas temperature, composition/Z, calibration and measurement uncertainty.\n\nPART F — GEOMETRIC VOLUME OF THE CNG ACCESSORY CIRCUIT\n=======================================================\n\nOBJECTIVE\n---------\nEstimate the minimum, average and maximum internal volumes of the high- and low-pressure CNG circuit, STRICTLY EXCLUDING THE CYLINDER/RESERVOIR. The study considers the internal space of lines, valves, regulator/reducer and low-pressure components.\n\nIMPORTANT ABOUT THE NATURE OF THIS CALCULATION\n-----------------------------------------------\nThese values are engineering reference estimates based on the TECHNICAL REPORT supplied for this project. They are not universal specifications for every CNG kit. A vehicle-specific measurement must use the actual internal diameter, length, reducer model, valves, filter, injector rail and other components.\n\n1. HIGH-PRESSURE PIPE\n---------------------\nThe study premise is a pipe with 6 mm outside diameter and 1.2 mm wall thickness, giving an internal diameter of 3.6 mm.\nD_internal = D_external - 2 × wall thickness\nD_internal = 6.0 - 2 × 1.2 = 3.6 mm\nr = 1.8 mm = 0.0018 m\n\nFor a straight cylindrical pipe:\nV_t = pi × r² × L\n\nWith L = 4.0 m: V_t ≈ 40.7 mL\nWith L = 5.5 m: V_t ≈ 70.1 mL\n\nTherefore, the approximate geometric range is 40 to 70 mL.\n\n2. SERVICE VALVES\n------------------\nThe technical report estimates the combined dead volume of the cylinder and filling-service valves at 3 to 5 mL.\n\n3. PRESSURE REDUCER / REGULATOR\n-------------------------------\nThe study considers the chambers used exclusively for gas passage in the regulator/reducer, excluding the water-heating circuit. The estimated range is 30 to 65 mL.\n\n4. LOW PRESSURE / FILTER / INJECTOR RAIL\n----------------------------------------\nFor modern systems, the report considers low-pressure hoses, gas-phase filter and injector rail/manifold. The estimated range is 50 to 140 mL. If this section is not present, its contribution may be considered 0 mL.\n\n5. STUDY CONSOLIDATION\n----------------------\nComponent                         Minimum    Average    Maximum\nHigh-pressure pipe                40 mL      55 mL       70 mL\nService valves                     3 mL       4 mL        5 mL\nReducer — gas chambers            32 mL      46 mL       65 mL\nLow pressure/filter/rail           0 mL      25 mL      140 mL\nTOTAL CIRCUIT VOLUME              75 mL     130 mL      280 mL\n\nThe study adopts 130 mL as a representative average geometric reference, without claiming that this value is a universal measurement for the fleet.\n\n6. WHY THE CIRCUIT MUST BE TREATED SEPARATELY FROM THE CYLINDER\n----------------------------------------------------------------\nThe cylinder is a large physical gas reservoir. The lines, valves, regulator/reducer and low-pressure components form additional internal volumes in the circuit. Therefore, in a geometric volume analysis, these spaces can be accounted for separately.\n\nThis is particularly important when comparing dispenser-indicated volume, calculated volume, stored volume and internal circuit volume. These quantities must not be added without clearly defining which physical quantity is being measured.\n\n7. HOW TO TURN THE ESTIMATE INTO A VEHICLE-SPECIFIC MEASUREMENT\n----------------------------------------------------------------\nFor a vehicle-specific result, provide:\n- actual internal pipe diameter;\n- actual high-pressure line length;\n- number and internal volume of valves;\n- reducer manufacturer/model;\n- reducer number of stages;\n- internal gas-chamber volume;\n- gas-phase filter volume;\n- injector rail/manifold volume;\n- internal volume of injectors and fittings;\n- manufacturer technical documentation, when available.\n\n8. TECHNICAL STUDY REFERENCES\n-----------------------------\nINMETRO — Portaria nº 111/2022 and requirements related to CNG/NGV components:\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1, natural-gas vehicle systems:\nhttps://www.abntcatalogo.com.br/\n\nUSP — Polytechnic School / Digital Library of Theses and Dissertations:\nhttps://teses.usp.br/\n\nUFRGS — Center for Physics Teaching Reference (CREF):\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / Institutional Repository:\nhttps://pantheon.ufrj.br/\n\nITA — Aeronautics Institute of Technology:\nhttps://www.ita.br/\n\nIME — Military Institute of Engineering:\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science:\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering:\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics:\nhttps://ocw.mit.edu/\n\nNIST — REFPROP and thermophysical properties:\nhttps://www.nist.gov/srd/refprop\n\nSOURCES AND TECHNICAL BASIS\n===========================\nANP — Natural-gas price publication:\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — Glossary / Standard Measurement Condition:\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — compressibility factor calculation by composition:\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — calculation using physical properties:\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics: ideal-gas equation PV = nRT:\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics:\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law:\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow:\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — Thermodynamics-related academic curriculum:\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nNIST — REFPROP / natural-gas equations of state:\nhttps://www.nist.gov/srd/refprop\n'}




# =============================================================================
# V28.18 — CONFIGURAÇÃO MULTI-ABA + CRÉDITOS + ESCALA DE CINZA
# =============================================================================
# Correções: aplicar configurações em todas as telas sem reconstrução;
# resultados ANP/Compressão/SQLite/Fórmulas com linhas alternadas e alinhamento;
# restauração de tema sem reintroduzir cores antigas; escala branco-cinza-preto;
# nova aba de créditos e propósito do projeto.
# =============================================================================
# V28.10 MOBILE ANDROID — REDMI NOTE 9 PRO
# Interface Kivy responsiva para Android.
# Mantém o núcleo científico, histórico, SQLite e estudo do circuito GNV.
# =============================================================================

import os
import json
import math
from datetime import datetime
from pathlib import Path

from kivy.app import App
from kivy.clock import Clock
from kivy.metrics import dp, sp
from kivy.core.window import Window
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner, SpinnerOption
from kivy.uix.slider import Slider
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.widget import Widget
from openpyxl import Workbook
from kivy.graphics import Color, Rectangle, Line

APP_VERSION = "V28.26"
APP_TITLE = f"Sistema de Cálculos e Análise da Capacidade do Cilindro de GNV - {APP_VERSION}"
ANDROID_TARGET = "Redmi Note 9 Pro — 6,67\" / 2400×1080 — interface responsiva"

MOBILE_TABS = IDIOMA_TABS


def mobile_text(idioma, texto):
    return I18N.get(idioma, {}).get(texto, texto)


def mobile_formula_text(idioma):
    if idioma == "pt-BR":
        return getattr(MobileGNVApp.instance, "formula_pt", "")
    return FORMULAS_I18N_COMPLETAS.get(idioma, getattr(MobileGNVApp.instance, "formula_pt", ""))


class ConfigSpinnerOption(SpinnerOption):
    """
    Opção do menu suspenso com contraste alto.

    O problema da versão anterior era que o menu herdava cores que podiam
    ficar semelhantes ao fundo. Por isso esta classe usa sempre um fundo
    escuro e texto claro.
    """
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ''
        self.background_down = ''
        self.background_color = (0.12, 0.08, 0.06, 1)
        self.color = (1, 1, 1, 1)
        self.font_size = sp(13)


class ColorPickerPopup(Popup):
    """
    Seletor visual de cores da aba Configurações.

    A versão anterior oferecia apenas uma lista com nomes de cores. Isso
    não atende ao que foi solicitado: escolher visualmente qualquer cor,
    incluindo intensidade e luminosidade.

    Este seletor possui três formas de escolha:

    1. Uma grade com cores do arco-íris;
    2. Controle de SATURAÇÃO (intensidade da cor);
    3. Controle de LUMINOSIDADE/VALOR (claro -> escuro).

    O usuário pode clicar diretamente em uma cor ou ajustar os controles
    com precisão. A cor escolhida é devolvida ao aplicativo em #RRGGBB.
    """

    HUES = [
        (0.00, 'Vermelho'), (0.08, 'Laranja'), (0.16, 'Amarelo'),
        (0.25, 'Verde'), (0.33, 'Verde-lima'), (0.42, 'Turquesa'),
        (0.50, 'Ciano'), (0.58, 'Azul'), (0.67, 'Índigo'),
        (0.75, 'Roxo'), (0.83, 'Magenta'), (0.92, 'Rosa')
    ]

    def __init__(self, title, initial_hex, on_apply, **kwargs):
        super().__init__(
            title=f'Escolher cor — {title}',
            size_hint=(0.94, 0.92),
            auto_dismiss=False,
            **kwargs
        )
        self.on_apply = on_apply
        self.initial_hex = initial_hex
        self._suspend = False

        h, sat, val = self._hex_to_hsv(initial_hex)
        self.hue = h
        self.saturation = sat
        self.value = val

        root = BoxLayout(orientation='vertical', spacing=dp(8), padding=dp(10))

        # ---------------------------------------------------------------
        # Pré-visualização grande da cor atualmente selecionada.
        # ---------------------------------------------------------------
        self.preview = Label(
            text=f'COR SELECIONADA\n{self._current_hex()}',
            size_hint_y=None, height=dp(72),
            bold=True, font_size=sp(14),
            halign='center', valign='middle'
        )
        root.add_widget(self.preview)

        # ---------------------------------------------------------------
        # Grade arco-íris: cada coluna é uma tonalidade e cada linha é
        # uma luminosidade diferente. A primeira linha é muito clara;
        # a última é muito escura.
        # ---------------------------------------------------------------
        root.add_widget(Label(
            text='Arco-íris — clique em qualquer quadrado para escolher a cor',
            size_hint_y=None, height=dp(30), font_size=sp(11)
        ))

        values = [1.00, 0.97, 0.94, 0.90, 0.85, 0.80, 0.74, 0.68,
          0.60, 0.52, 0.44, 0.36, 0.28, 0.20, 0.12, 0.06]
        rainbow = GridLayout(
            cols=len(self.HUES), rows=len(values), spacing=dp(2),
            size_hint_y=None, height=dp(len(values) * 24 + 12)
        )

        # Sete níveis para dar uma seleção mais fina do claro ao escuro.
        # Escala neutra obrigatória: branco → cinzas → preto.
        # Cada quadrado recebe ação própria e entra no mesmo fluxo do seletor.
        gray_values = [1.00, 0.88, 0.75, 0.60, 0.45, 0.30, 0.15, 0.00]
        gray_box = BoxLayout(size_hint_y=None, height=dp(34), spacing=dp(2))
        for gv in gray_values:
            ghx = self._hsv_to_hex(0.0, 0.0, gv)
            gbtn = Button(text='', size_hint_x=1, background_normal='', background_down='')
            gbtn.background_color = self._hex_rgba(ghx)
            gbtn.bind(on_release=lambda _b, vv=gv: self._choose_swatch(0.0, 0.0, vv))
            gray_box.add_widget(gbtn)
        root.add_widget(Label(text='ESCALA NEUTRA — Branco → Cinza → Preto', size_hint_y=None, height=dp(26), font_size=sp(10)))
        root.add_widget(gray_box)

        for value in values:
            for hue, name in self.HUES:
                swatch_sat = 0.86 if value < 0.92 else max(0.20, 0.86 * (1.0 - (value - 0.92) / 0.08))
                hx = self._hsv_to_hex(hue, swatch_sat, value)
                btn = Button(
                    text='', size_hint=(1, None), height=dp(30),
                    background_normal='', background_down=''
                )
                btn.background_color = self._hex_rgba(hx)
                btn.bind(on_release=lambda _btn, hh=hue, vv=value:
                          self._choose_swatch(hh, (0.86 if vv < 0.92 else max(0.20, 0.86 * (1.0 - (vv - 0.92) / 0.08))), vv))
                rainbow.add_widget(btn)
        root.add_widget(rainbow)

        # ---------------------------------------------------------------
        # Escala neutra completa: branco -> cinza -> preto.
        # São 11 níveis para facilitar a escolha de fundos e textos.
        # ---------------------------------------------------------------
        root.add_widget(Label(text='ESCALA NEUTRA — BRANCO → CINZAS → PRETO', size_hint_y=None, height=dp(28), font_size=sp(11), bold=True))
        neutral = GridLayout(cols=11, rows=1, spacing=dp(2), size_hint_y=None, height=dp(34))
        for i in range(11):
            level = round(255 * (1 - i / 10))
            hx = f'#{level:02X}{level:02X}{level:02X}'
            btn = Button(text='', size_hint=(1,None), height=dp(34), background_normal='', background_down='')
            btn.background_color = self._hex_rgba(hx)
            btn.bind(on_release=lambda _btn, value=hx: self._apply_neutral(value))
            neutral.add_widget(btn)
        root.add_widget(neutral)

        # ---------------------------------------------------------------
        # Controles finos. Saturação = intensidade da cor.
        # Valor = luminosidade (1 claro, valores menores mais escuros).
        # Matiz = posição no arco-íris.
        # ---------------------------------------------------------------
        self.hue_slider = self._slider_row(root, 'Matiz / Cor', self.hue)
        self.saturation_slider = self._slider_row(root, 'Intensidade / Saturação', self.saturation)
        self.value_slider = self._slider_row(root, 'Luminosidade / Claro → Escuro', self.value)

        # ---------------------------------------------------------------
        # Código hexadecimal para estudo e conferência.
        # ---------------------------------------------------------------
        self.hex_label = Label(
            text=f'Hexadecimal: {self._current_hex()}',
            size_hint_y=None, height=dp(30), font_size=sp(12), bold=True
        )
        root.add_widget(self.hex_label)

        # ---------------------------------------------------------------
        # Botões finais.
        # ---------------------------------------------------------------
        buttons = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        cancel = Button(text='Cancelar', font_size=sp(12))
        apply = Button(text='Aplicar cor', font_size=sp(12), bold=True)
        cancel.bind(on_release=lambda *_: self.dismiss())
        apply.bind(on_release=self._apply)
        buttons.add_widget(cancel)
        buttons.add_widget(apply)
        root.add_widget(buttons)

        self.content = root
        self._update_preview()

    def _apply_neutral(self, hx):
        """Aplica um tom neutro da escala branco-cinza-preto."""
        self._suspend = True
        h, sat, val = self._hex_to_hsv(hx)
        self.hue, self.saturation, self.value = h, sat, val
        self.hue_slider.value = h
        self.saturation_slider.value = sat
        self.value_slider.value = val
        self._suspend = False
        self._update_preview()

    def _slider_row(self, parent, caption, value):
        """Cria uma linha de controle deslizante e retorna o Slider."""
        row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(6))
        label = Label(text=caption, size_hint_x=0.38, font_size=sp(10),
                      halign='left', valign='middle')
        slider = Slider(min=0, max=1, value=value, step=0.001, size_hint_x=0.62)
        slider.bind(value=lambda *_: self._slider_changed())
        row.add_widget(label)
        row.add_widget(slider)
        parent.add_widget(row)
        return slider

    @staticmethod
    def _hex_to_hsv(value):
        """Converte #RRGGBB para HSV normalizado."""
        try:
            v = str(value).strip().lstrip('#')
            if len(v) == 3:
                v = ''.join(ch * 2 for ch in v)
            rgb = tuple(int(v[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
            return colorsys.rgb_to_hsv(*rgb)
        except Exception:
            return 0.08, 0.25, 0.90

    @staticmethod
    def _hsv_to_hex(h, s, v):
        """Converte HSV normalizado para #RRGGBB."""
        rgb = colorsys.hsv_to_rgb(float(h), float(s), float(v))
        return '#' + ''.join(f'{max(0, min(255, round(x * 255))):02X}' for x in rgb)

    @staticmethod
    def _hex_rgba(value):
        try:
            v = str(value).strip().lstrip('#')
            if len(v) == 3:
                v = ''.join(ch * 2 for ch in v)
            return [int(v[i:i + 2], 16) / 255.0 for i in (0, 2, 4)] + [1]
        except Exception:
            return [0, 0, 0, 1]

    def _current_hex(self):
        return self._hsv_to_hex(self.hue, self.saturation, self.value)

    def _choose_swatch(self, hue, saturation, value):
        """Recebe um clique da grade arco-íris."""
        self._suspend = True
        self.hue = hue
        self.saturation = saturation
        self.value = value
        self.hue_slider.value = hue
        self.saturation_slider.value = saturation
        self.value_slider.value = value
        self._suspend = False
        self._update_preview()

    def _slider_changed(self):
        if not self._suspend:
            self.hue = self.hue_slider.value
            self.saturation = self.saturation_slider.value
            self.value = self.value_slider.value
            self._update_preview()

    def _update_preview(self):
        """Atualiza a amostra e garante contraste do texto."""
        hx = self._current_hex()
        rgba = self._hex_rgba(hx)
        self.preview.background_color = rgba
        lum = 0.299 * rgba[0] + 0.587 * rgba[1] + 0.114 * rgba[2]
        self.preview.color = (0.05, 0.04, 0.03, 1) if lum > 0.56 else (1, 1, 1, 1)
        self.preview.text = f'COR SELECIONADA\n{hx}'
        self.hex_label.text = f'Hexadecimal: {hx}'

    def _apply(self, *_args):
        """Confirma a cor e devolve o valor para a tela de Configurações."""
        if callable(self.on_apply):
            self.on_apply(self._current_hex())
        self.dismiss()


class ScrollText(ScrollView):
    """
    Área de resultados rolável com linhas alternadas.

    Cada linha é um Label independente. Isso é importante porque o Kivy
    não permite aplicar uma cor de fundo diferente a cada linha de um
    único Label. A classe também mantém ``self.label`` por compatibilidade
    com o código de PDF da versão anterior.
    """
    def __init__(self, text="", horizontal=False, font_size=13, **kwargs):
        super().__init__(**kwargs)
        self.horizontal_mode = horizontal
        self.do_scroll_y = True
        self.do_scroll_x = horizontal
        self.bar_width = dp(10)
        self.font_size_value = font_size
        self._text_color = (1,1,1,1)
        self._row1 = (0.20,0.20,0.20,1)
        self._row2 = (0.14,0.14,0.14,1)
        self._alignment = "center"
        self._row_text1 = self._text_color
        self._row_text2 = self._text_color

        # ``label`` continua existindo para exportação/PDF.
        self.label = Label(text=str(text))
        self._text_value = str(text)

        self.container = BoxLayout(orientation='vertical', spacing=0, padding=0,
                                   size_hint=(None,None))
        self.container.bind(minimum_height=self.container.setter('height'))
        self.add_widget(self.container)
        self.bind(size=self._sync)
        self.set_text(text)

    def _make_row(self, text, index):
        """Cria uma linha visual com a cor alternada configurada."""
        lab = Label(
            text=str(text), font_size=sp(self.font_size_value),
            size_hint=(None,None), valign='middle',
            padding=(dp(10),dp(5))
        )
        lab.color = self._text_color
        lab._row_text_color = self._text_color
        lab._row_bg_hex = self._row1 if index % 2 == 0 else self._row2
        lab.halign = self._alignment
        lab.texture_update()

        viewport = max(self.width-dp(5), dp(300))
        natural_w = lab.texture_size[0] + dp(20)
        # Mesmo em modo horizontal, a linha mantém no mínimo a largura da tela
        # para que o alinhamento selecionado seja perceptível.
        lab.width = max(viewport, natural_w if self.horizontal_mode else viewport)
        lab.text_size = (max(lab.width-dp(20),dp(1)),None)
        lab.texture_update()
        lab.height = max(lab.texture_size[1] + dp(10), dp(28))

        with lab.canvas.before:
            bgc = Color(*(self._row1 if index % 2 == 0 else self._row2))
            rect = Rectangle(pos=lab.pos,size=lab.size)
        lab._row_color = bgc
        lab._row_rect = rect
        lab.bind(pos=lambda w,*a:setattr(w._row_rect,'pos',w.pos),
                 size=lambda w,*a:setattr(w._row_rect,'size',w.size))
        return lab

    def _sync(self,*_):
        """
        Atualiza largura e alinhamento das linhas.

        V28.13: mesmo com rolagem horizontal, cada linha tem pelo menos a
        largura da janela. Assim Esquerda/Centro/Direita realmente funcionam.
        Se o texto for maior que a tela, a linha cresce e continua rolável.
        """
        if not hasattr(self,'container'): return
        viewport = max(self.width-dp(5), dp(300))
        max_width = viewport
        for child in reversed(self.container.children):
            child.texture_update()
            natural = child.texture_size[0] + dp(20)
            child.width = max(viewport, natural if self.horizontal_mode else viewport)
            child.text_size = (max(child.width-dp(20),dp(1)), None)
            child.halign=self._alignment
            child.texture_update()
            child.height=max(child.texture_size[1]+dp(10),dp(28))
            max_width = max(max_width, child.width)
            child._row_rect.pos=child.pos
            child._row_rect.size=child.size
        self.container.width = max_width

    def set_colors(self, text_color=None, background_color=None, row1=None, row2=None, alignment=None, row_text1=None, row_text2=None):
        """Aplica FUNDO e LETRA de cada linha separadamente.

        CORREÇÃO V28.11:
        O Kivy guarda os filhos do BoxLayout em ordem inversa. Em versões
        anteriores isso facilitava a associação errada entre a linha visual e
        a cor escolhida. Agora cada Label recebe também a cor no atributo
        ``_row_bg_hex`` e a pintura é refeita explicitamente.
        """
        self._row1 = row1 if row1 is not None else (background_color or self._row1)
        self._row2 = row2 if row2 is not None else (background_color or self._row2)
        if row_text1 is not None:
            self._row_text1 = row_text1
        if row_text2 is not None:
            self._row_text2 = row_text2
        if alignment in ('left', 'center', 'right'):
            self._alignment = alignment

        # ``reversed(children)`` recupera a ordem original das linhas.
        for idx, child in enumerate(reversed(self.container.children)):
            bg = self._row1 if idx % 2 == 0 else self._row2
            txt = row_text1 if idx % 2 == 0 else row_text2
            if txt is None:
                txt = self._row_text1 if idx % 2 == 0 else self._row_text2
            if txt is None:
                txt = text_color or self._text_color

            child._row_bg_hex = bg
            child._row_text_color = txt
            child.color = txt
            child.halign = self._alignment
            child._row_color.rgba = bg
            child._row_rect.pos = child.pos
            child._row_rect.size = child.size

        self._sync()

    def set_alignment(self,alignment):
        if alignment in ('left','center','right'):
            self._alignment=alignment
            self._sync()

    def set_text(self,text):
        self._text_value=str(text)
        self.label.text=self._text_value
        self.container.clear_widgets()
        lines=self._text_value.split('\n') if self._text_value else ['']
        for idx,line in enumerate(lines):
            self.container.add_widget(self._make_row(line,idx))
        self._sync()
        # Reaplica imediatamente as cores/alinhamento já configurados.
        # Isso impede que qualquer novo resultado volte para o padrão da classe.
        self.set_colors(row1=self._row1, row2=self._row2,
                        alignment=self._alignment,
                        row_text1=getattr(self, "_row_text1", self._text_color),
                        row_text2=getattr(self, "_row_text2", self._text_color))
        self.scroll_y=1
        if self.horizontal_mode:
            self.scroll_x=0


class GradientTextInput(TextInput):
    """Campo de entrada com degradê horizontal configurável."""
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        self.background_normal=''
        self.background_active=''
        self.background_color=(0,0,0,0)
        self.gradient_start='#FFFFFF'
        self.gradient_end='#E8D2BF'
        self.gradient_enabled=True
        self._gradient_parts=[]
        for _ in range(20):
            with self.canvas.before:
                c=Color(1,1,1,1); r=Rectangle(pos=self.pos,size=self.size)
            self._gradient_parts.append((c,r))
        self.bind(pos=self._redraw_gradient,size=self._redraw_gradient)
        self._redraw_gradient()

    def _rgb(self,v):
        try:
            v=str(v).strip().lstrip('#')
            if len(v)==3: v=''.join(x*2 for x in v)
            return tuple(int(v[i:i+2],16)/255 for i in (0,2,4))
        except Exception:
            return (1,1,1)

    def set_gradient(self,start,end,enabled=True):
        self.gradient_start=start; self.gradient_end=end; self.gradient_enabled=bool(enabled)
        self._redraw_gradient()

    def _redraw_gradient(self,*_):
        a=self._rgb(self.gradient_start); b=self._rgb(self.gradient_end)
        if not self.gradient_enabled: b=a
        n=len(self._gradient_parts)
        for i,(c,r) in enumerate(self._gradient_parts):
            f=i/max(n-1,1)
            rgb=tuple(a[k]+(b[k]-a[k])*f for k in range(3))
            c.rgba=(*rgb,1)
            r.pos=(self.x+self.width*i/n,self.y)
            r.size=(self.width/n+1,self.height)


class GNVLineChart(Widget):
    """Gráfico Android com linhas, pontos, valores numéricos e legenda."""
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.series=[]; self.labels=[]; self.x_values=None; self.title=""; self.series_names=[]; self.value_labels=[]
        self.bind(pos=self._draw, size=self._draw)

    def set_data(self, series, labels, title, series_names=None, x_values=None):
        self.series=series or []; self.labels=labels or []; self.x_values=x_values; self.title=title; self.series_names=series_names or []
        for group in list(self.value_labels):
            for w in group: self.remove_widget(w)
        self.value_labels=[]
        self._draw()

    def _app_text_color(self):
        try:
            app=App.get_running_app()
            return app._hex_rgba(app.colors.get("text", "#FFFFFF"))
        except Exception:
            return (1,1,1,1)

    def _app_chart_bg(self):
        try:
            app=App.get_running_app()
            return app._hex_rgba(app.colors.get("result", "#181818"))
        except Exception:
            return (0.06,0.06,0.06,1)

    def _draw(self, *_):
        self.canvas.clear()
        for group in self.value_labels:
            for w in group: w.opacity=0
        if self.width<dp(100) or self.height<dp(100): return
        vals=[v for ser in self.series for v in ser if isinstance(v,(int,float)) and math.isfinite(v)]
        if not vals: return
        lo=min(0.0,min(vals)); hi=max(vals)
        if hi<=lo: hi=lo+1.0
        plot_x=self.x+dp(60); plot_y=self.y+dp(48); plot_w=max(dp(100),self.width-dp(80)); plot_h=max(dp(100),self.height-dp(100))
        colors=[(0.20,0.65,1,1),(1,0.55,0.15,1),(0.25,0.9,0.45,1),(0.85,0.35,0.9,1)]
        with self.canvas:
            Color(*self._app_chart_bg()); Rectangle(pos=self.pos,size=self.size)
            Color(0.35,0.35,0.35,1); Line(rectangle=(plot_x,plot_y,plot_w,plot_h),width=1)
            for idx,ser in enumerate(self.series):
                if not ser: continue
                pts=[]; col=colors[idx%len(colors)]
                for j,v in enumerate(ser):
                    if not isinstance(v,(int,float)) or not math.isfinite(v): continue
                    xv=(self.x_values[j] if self.x_values is not None and j < len(self.x_values) else j)
                    xmin=(min(self.x_values) if self.x_values else 0); xmax=(max(self.x_values) if self.x_values else max(1,len(ser)-1))
                    if xmax==xmin: xmax=xmin+1.0
                    xx=plot_x+((xv-xmin)/(xmax-xmin))*plot_w; yy=plot_y+((v-lo)/(hi-lo))*plot_h; pts.extend([xx,yy])
                if len(pts)>=2:
                    Color(*col); Line(points=pts,width=1.8)
                while len(self.value_labels)<=idx: self.value_labels.append([])
                while len(self.value_labels[idx])<len(ser):
                    lab=Label(font_size=sp(8),size_hint=(None,None),size=(dp(72),dp(18)),halign='center'); self.add_widget(lab); self.value_labels[idx].append(lab)
                for j,v in enumerate(ser):
                    if not isinstance(v,(int,float)) or not math.isfinite(v): continue
                    xv=(self.x_values[j] if self.x_values is not None and j < len(self.x_values) else j)
                    xmin=(min(self.x_values) if self.x_values else 0); xmax=(max(self.x_values) if self.x_values else max(1,len(ser)-1))
                    if xmax==xmin: xmax=xmin+1.0
                    xx=plot_x+((xv-xmin)/(xmax-xmin))*plot_w; yy=plot_y+((v-lo)/(hi-lo))*plot_h
                    Color(*col); Rectangle(pos=(xx-dp(3),yy-dp(3)),size=(dp(6),dp(6)))
                    lab=self.value_labels[idx][j]
                    lab.text=f"{v:.3f}"
                    lab.color=self._app_text_color()
                    lab.bold=True
                    lab.pos=(xx-dp(36),yy+dp(7))
                    lab.opacity=1
        if not hasattr(self,'title_label'):
            self.title_label=Label(font_size=sp(11),size_hint=(None,None),size=(dp(360),dp(22)),halign='left'); self.add_widget(self.title_label)
        self.title_label.text=self.title; self.title_label.color=self._app_text_color(); self.title_label.pos=(self.x+dp(4),self.top-dp(27))
        if not hasattr(self,'legend_label'):
            self.legend_label=Label(font_size=sp(8),size_hint=(None,None),size=(dp(500),dp(20)),halign='left'); self.add_widget(self.legend_label)
        names=self.series_names or [f"Série {i+1}" for i in range(len(self.series))]
        self.legend_label.text="   ".join(f"{i+1}: {n}" for i,n in enumerate(names)); self.legend_label.color=self._app_text_color(); self.legend_label.pos=(self.x+dp(60),self.y+dp(18))


class MobileScreen(Screen):
    def __init__(self, title_key, **kwargs):
        super().__init__(**kwargs)
        self.title_key = title_key


class MobileGNVApp(App):
    instance = None

    def build(self):
        MobileGNVApp.instance = self
        self.title = APP_TITLE
        self.idioma = "pt-BR"
        self.base_dir = Path(self.user_data_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = str(self.base_dir / "gnv_dados.db")
        self.config_path = self.base_dir / "configuracoes.json"
        self.banco = BancoGNV(self.db_path)
        self.banco.conectar()
        self.banco.criar_tabela()
        self.banco.criar_indices()
        self.formula_pt = self._load_formula_pt()
        self.colors = self._default_colors()
        self._colors_personalizadas = False
        self._visual_ready = False

        self.sm = ScreenManager()
        self.screen_names = []
        for i, key in enumerate([
            "Cálculos", "Abastecimentos", "ANP", "Aquecimento / Compressão",
            "Histórico de Abastecimentos", "Banco SQLite", "Exportação / Excel",
            "Gráficos de Abastecimento", "Configurações do Sistema",
            "Fórmulas e Física", "Total de Abastecimentos", "Créditos"
        ]):
            name = f"screen_{i}"
            self.screen_names.append(name)
            self.sm.add_widget(MobileScreen(key, name=name))

        root = BoxLayout(orientation="vertical", spacing=dp(4), padding=dp(5))
        self.header = Label(text=APP_TITLE, size_hint_y=None, height=dp(42), font_size=sp(15), bold=True)
        root.add_widget(self.header)

        nav = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(4))
        self.tab_spinner = Spinner(text="Cálculos", values=tuple(self.screen_names_for_language()), size_hint_x=0.78)
        self.tab_spinner.bind(text=self._go_from_spinner)
        nav.add_widget(self.tab_spinner)
        self.lang_spinner = Spinner(text="pt-BR", values=IDIOMAS_DISPONIVEIS, size_hint_x=0.22)
        self.lang_spinner.bind(text=self.change_language)
        nav.add_widget(self.lang_spinner)
        root.add_widget(nav)
        root.add_widget(self.sm)

        self.footer = Label(
            text="Analista de Sistemas e Pesquisador - Christiano T.Gaio - Desenvolvedor | Projeto iniciado o Desenvolvimento em 06/2026",
            size_hint_y=None, height=dp(34), font_size=sp(9), halign="center", valign="middle"
        )
        self.footer.bind(size=lambda *_: setattr(self.footer, "text_size", self.footer.size))
        root.add_widget(self.footer)

        self._load_config()
        self._build_all_screens()
        self._apply_language()
        # IMPORTANTE: self.root ainda não existe enquanto build() está montando a árvore.
        # A aplicação visual é agendada para o próximo ciclo do Kivy, quando
        # self.root já estará disponível. Isso corrige a causa estrutural das
        # versões anteriores em que a configuração parecia não fazer efeito.
        Clock.schedule_once(lambda _dt: self._mark_visual_ready(), 0)
        return root

    def _load_formula_pt(self):
        # Reconstitui o texto PT-BR do próprio código fonte.
        # No APK, este bloco é substituído abaixo pela constante embarcada.
        return 'FÓRMULAS E FÍSICA DO SISTEMA DE CÁLCULO DE GNV\n============================================================\n\nPARTE A — CONDIÇÃO DE REFERÊNCIA DA ANP\n========================================\n\nA ANP informa, para volumes médios comercializados de gás natural,\ncondições de referência de 20 °C e 1,033 kgf/cm². A condição padrão de\nmedição é definida como pressão absoluta de 0,101325 MPa e temperatura\nde 20 °C.\n\nIMPORTANTE: a ANP define a condição de referência. A fórmula implementada\nneste programa é uma ESTIMATIVA FÍSICA DE CONVERSÃO e não afirma reproduzir\no algoritmo interno de um dispenser de GNV.\n\n1. VOLUME FÍSICO DO CILINDRO\n----------------------------\nVcil = capacidade(L) / 1000\n\n26 L / 1000 = 0,026 m³.\n\nEsse é o espaço físico interno do cilindro. Não é o volume normalizado\nindicado pela bomba.\n\n2. PRESSÃO ABSOLUTA\n-------------------\nPabs = Pmanométrica + Patm\n\nAs equações de estado usam pressão absoluta.\n\n3. TEMPERATURA ABSOLUTA\n-----------------------\nT(K) = T(°C) + 273,15\n\n4. CONVERSÃO PARA 20 °C\n-----------------------\nPara uma quantidade de matéria fixa, em modelo ideal:\n\nVref = V × (P/Pref) × (Tref/T)\n\nTref = 293,15 K e Pref ≈ 1,01325 bar.\n\nPARTE B — MODELO CIENTÍFICO DE GÁS REAL\n========================================\n\n1. EQUAÇÃO DE ESTADO\n--------------------\nP V = Z n R T\n\nn = P V / (Z R T)\n\nP = pressão absoluta (Pa)\nV = volume físico (m³)\nZ = fator de compressibilidade\nn = quantidade de matéria (mol)\nR = 8,314462618 J/(mol·K)\nT = temperatura absoluta (K)\n\n2. FATOR DE COMPRESSIBILIDADE Z\n-------------------------------\nZ = P V / (n R T)\n\nZ = 1 representa o gás ideal. Para gás natural real, Z depende de\npressão, temperatura e composição. Um Z fixo informado pelo usuário é\numa aproximação, não uma determinação metrológica de Z.\n\n3. QUANTIDADE DE GÁS ADICIONADA\n-------------------------------\nn_inicial = P_inicial_abs × V / (Z R T)\nn_final   = P_final_abs × V / (Z R T)\nΔn = n_final − n_inicial\n\n4. MASSA ADICIONADA\n-------------------\nm = Δn × M\n\nM é a massa molar do GNV em kg/mol.\n\n5. VOLUME EQUIVALENTE NA TEMPERATURA INFORMADA\n------------------------------------------------\nO programa calcula primeiro os mols a partir de PV = Z n R T.\nDepois pode expressar esses mesmos mols a uma pressão de referência:\n\nVref(T) = n R T / Pref\n\nIMPORTANTE: neste programa, n foi calculado usando a própria temperatura\ninformada. Por isso, ao substituir n = P V / (Z R T), a temperatura cancela:\n\nVref(T) = P V / (Z Pref)\n\nAssim, para um cilindro de 26 L a 220 bar e Z=0,92, o volume equivalente\na 1,01325 bar na temperatura informada pode permanecer praticamente\n6,164 m³ tanto a 5 °C quanto a 20 °C ou 100 °C. Isso NÃO significa que\na mesma quantidade de gás teria o mesmo volume em duas temperaturas quando\nn é mantido fixo; significa que o programa está recalculando a quantidade\nde matéria para cada estado de pressão/temperatura informado.\n\n6. CONVERSÃO CIENTÍFICA PARA 20 °C\n----------------------------------\nOs mesmos mols calculados pelo modelo Z informado são convertidos para 20 °C:\n\nV20 = n R T20 / Pref\n\nSe a temperatura informada for menor que 20 °C, V20 tende a ser maior.\nSe for maior que 20 °C, V20 tende a ser menor, porque aqui n é mantido\nfixo durante a conversão.\n\n7. CONVERSÃO ANP/IDEALIZADA (Z=1)\n---------------------------------\nA aba ANP calcula separadamente uma estimativa com Z=1 e a condição de\nreferência de 20 °C e 1,033 kgf/cm² (aproximadamente 1,01325 bar).\nEsse valor não deve ser confundido com o resultado científico que usa\nZ=0,92, por exemplo. São dois modelos diferentes.\n\n6. DENSIDADE DO GÁS REAL\n------------------------\nρ = P M / (Z R T)\n\nPARTE C — O QUE SERIA UM MODELO MAIS PRECISO\n============================================\n\nEm GNV a alta pressão, não é adequado considerar Z como uma constante\nuniversal. Para elevar a precisão é necessário conhecer a composição do\ngás e calcular suas propriedades termodinâmicas em função de P e T.\n\nA ISO 12213 descreve métodos para o cálculo do fator de compressibilidade\nde gás natural. A ISO 12213-2 usa composição molar; a ISO 12213-3 usa\npropriedades físicas como poder calorífico, densidade relativa e CO₂,\nalém de pressão e temperatura.\n\nAGA8 e GERG são modelos utilizados para propriedades de gás natural. O\nNIST descreve AGA8 e GERG entre as equações de estado usadas em aplicações\nde medição e propriedades termodinâmicas de gás natural.\n\nPortanto, a evolução científica do programa deve ser:\n1) obter composição do GNV;\n2) calcular Z(P,T,composição), em vez de usar Z fixo;\n3) considerar a temperatura real do gás durante o abastecimento;\n4) conhecer as condições efetivamente usadas pelo medidor;\n5) trabalhar com incerteza de medição.\n\nPARTE D — TEMPERATURA DURANTE O ABASTECIMENTO\n==============================================\n\nDurante o enchimento existe entrada de massa e transferência de calor.\nA aba Compressão / Temperatura usa uma compressão adiabática reversível\nsomente como cenário didático:\n\nT₂/T₁ = (P₂/P₁)^((k−1)/k)\nP·V^k = constante\nV₂/V₁ = (P₁/P₂)^(1/k)\n\nO abastecimento real é um sistema aberto, com troca de calor entre gás,\nparede do cilindro, mangueira e ambiente. A temperatura calculada nessa\naba NÃO é uma medição da temperatura real do GNV.\n\nPARTE E — COMPARAÇÃO COM A BOMBA\n================================\n\nDiferença = volume indicado pela bomba − volume calculado\nDiferença percentual = diferença / volume calculado × 100\n\nUma diferença grande é um indício para investigação. Ela não constitui,\nsozinha, prova metrológica de fraude. Uma conclusão técnica exige dados\ndo medidor, condições de referência, temperatura real do gás,\ncomposição/Z, calibração e incerteza de medição.\n\nPARTE F — VOLUME GEOMÉTRICO DO CIRCUITO ACESSÓRIO DE GNV\n==============================================================\n\nOBJETIVO\n--------\nDeterminar uma estimativa dos volumes internos mínimos, médios e máximos\ndo circuito de alta e baixa pressão do sistema GNV, DESCONSIDERANDO\nRIGOROSAMENTE O CILINDRO/RESERVATÓRIO. O estudo considera o espaço interno\ndas linhas, válvulas, redutor e componentes de baixa pressão.\n\nIMPORTANTE SOBRE A NATUREZA DO CÁLCULO\n--------------------------------------\nOs valores desta seção são referências/estimativas geométricas de engenharia baseadas no RELATÓRIO TÉCNICO fornecido para o projeto. Eles não devem ser\ntratados como uma especificação universal de todos os kits GNV. Para uma\nmedição específica devem ser utilizados o diâmetro interno, comprimento,\nmodelo do redutor, válvulas, filtro, flauta e demais componentes reais.\n\n1. TUBULAÇÃO DE ALTA PRESSÃO\n----------------------------\nA premissa do estudo é uma tubulação com 6 mm de diâmetro externo e parede\nde 1,2 mm, resultando em diâmetro interno de 3,6 mm.\n\nD_interno = D_externo - 2 x espessura\nD_interno = 6,0 - 2 x 1,2 = 3,6 mm\nr = 1,8 mm = 0,0018 m\n\nPara um tubo cilíndrico reto:\n\nV_t = pi x r² x L\n\nCom L = 4,0 m:\nV_t ≈ 40,7 mL\n\nCom L = 5,5 m:\nV_t ≈ 70,1 mL\n\nAssim, o intervalo geométrico aproximado do estudo é de 40 a 70 mL.\n\n2. VÁLVULAS DE SERVIÇO\n----------------------\nO relatório técnico considera canais internos restritos e estima o espaço\nmorto combinado das válvulas de cilindro e abastecimento entre 3 e 5 mL.\n\n3. REDUTOR DE PRESSÃO\n---------------------\nO estudo considera as câmaras destinadas exclusivamente à passagem do gás\nno regulador/redutor, excluindo as passagens do circuito de água de\naquecimento. A referência estimada é de 30 a 65 mL.\n\n4. BAIXA PRESSÃO / FILTRO / FLAUTA\n----------------------------------\nPara sistemas modernos, o relatório considera mangueiras de baixa pressão,\nfiltro de fase gasosa e flauta/manifold dos injetores. A faixa estimada é\nde 50 a 140 mL. Em uma configuração sem esse trecho, a contribuição pode ser considerada 0 mL.\n\n5. CONSOLIDAÇÃO DO ESTUDO\n-------------------------\nComponente                         Mínimo    Médio    Máximo\nTubulação alta pressão             40 mL     55 mL    70 mL\nVálvulas de serviço                 3 mL      4 mL     5 mL\nRedutor — câmaras de gás           32 mL     46 mL    65 mL\nBaixa pressão/filtro/flauta          0 mL     25 mL   140 mL\nVOLUME TOTAL DO CIRCUITO            75 mL    130 mL   280 mL\n\nO estudo adota 130 mL como referência média geométrica representativa,\nsem afirmar que esse valor seja uma medição universal da frota.\n\n6. POR QUE O CIRCUITO DEVE SER TRATADO SEPARADAMENTE DO CILINDRO?\n-----------------------------------------------------------------\nO cilindro é um reservatório de grande volume físico. Já a tubulação,\nválvulas, redutor e baixa pressão constituem volumes internos adicionais\ndo circuito. Portanto, em uma análise de volume geométrico do sistema,\nesses espaços podem ser contabilizados separadamente.\n\nIsso é particularmente importante quando uma análise compara volume\nindicado, volume calculado, volume armazenado e volume interno do circuito.\nNão se deve somar esses volumes sem definir claramente qual grandeza está\nsendo medida.\n\n7. COMO TRANSFORMAR A ESTIMATIVA EM MEDIÇÃO DO VEÍCULO\n-------------------------------------------------------\nPara sair da estimativa e obter um resultado específico do veículo, informe:\n\n- diâmetro interno real da tubulação;\n- comprimento real da linha de alta pressão;\n- quantidade e volume interno das válvulas;\n- fabricante/modelo do redutor;\n- número de estágios do redutor;\n- volume interno das câmaras de gás;\n- volume do filtro de fase gasosa;\n- volume da flauta/manifold;\n- volume interno dos injetores e conexões;\n- documentação técnica do fabricante, quando disponível.\n\n8. REFERÊNCIAS DO ESTUDO TÉCNICO\n--------------------------------\nINMETRO — Portaria nº 111/2022 e requisitos relacionados a componentes\nde sistemas GNV/GNC:\nhttps://registro.inmetro.gov.br/objetos/\n\nABNT — NBR 11353-1, sistemas de gás natural veicular:\nhttps://www.abntcatalogo.com.br/\n\nUSP — Escola Politécnica / Biblioteca Digital de Teses e Dissertações:\nhttps://teses.usp.br/\n\nUFRGS — Centro de Referência para o Ensino de Física (CREF):\nhttps://cref.if.ufrgs.br/\n\nUFRJ — Pantheon / Repositório Institucional:\nhttps://pantheon.ufrj.br/\n\nITA — Instituto Tecnológico de Aeronáutica:\nhttps://www.ita.br/\n\nIME — Instituto Militar de Engenharia:\nhttps://www.ime.eb.mil.br/\n\nUniversity of Oxford — Engineering Science:\nhttps://eng.ox.ac.uk/\n\nUniversity of Maryland — A. James Clark School of Engineering:\nhttps://eng.umd.edu/\n\nMIT OpenCourseWare — Engineering / Thermodynamics / Fluid Mechanics:\nhttps://ocw.mit.edu/\n\nNIST — REFPROP e propriedades termofísicas:\nhttps://www.nist.gov/srd/refprop\n\nANP — Agência Nacional do Petróleo, Gás Natural e Biocombustíveis:\nhttps://www.gov.br/anp/\n\nFONTES E FUNDAMENTAÇÃO\n=======================\n\nANP — Publicidade dos preços de gás natural:\nhttps://www.gov.br/anp/pt-br/assuntos/movimentacao-estocagem-e-comercializacao-de-gas-natural/acompanhamento-do-mercado-de-gas-natural/publicidade-dos-precos-de-gas-natural\n\nANP — Glossário C / Condição Padrão de Medição:\nhttps://www.gov.br/anp/pt-br/acesso-a-informacao/glossario/c\n\nISO 12213-2:2006 — cálculo do fator de compressibilidade por composição:\nhttps://www.iso.org/standard/44411.html\n\nISO 12213-3:2006 — cálculo do fator de compressibilidade por propriedades:\nhttps://www.iso.org/standard/44412.html\n\nMIT OpenCourseWare — Thermodynamics: equação de estado do gás ideal PV = nRT:\nhttps://ocw.mit.edu/courses/5-60-thermodynamics-kinetics-spring-2008/\n\nMIT OpenCourseWare — Materials at Equilibrium: propriedades de gases ideais e PV = nRT:\nhttps://ocw.mit.edu/courses/3-20-materials-at-equilibrium-sma-5111-fall-2003/\n\nPurdue University — Thermodynamics, Fluid Mechanics and Gas Dynamics: gás ideal e fator de compressibilidade Z:\nhttps://engineering.purdue.edu/~wassgren/teaching/ME20000/NotesAndReading/Lec11_Reading_Wassgren.pdf\n\nStanford University — Thermodynamics / Ideal Gas Law:\nhttps://web.stanford.edu/~peastman/statmech/thermodynamics.html\n\nStanford University — Fundamentals of Compressible Flow, gases ideais e propriedades termodinâmicas:\nhttps://web.stanford.edu/~cantwell/AA210A_Course_Material/AA210A_Lectures/AA210A_Chapter_2_Thermo_of_gases_Brian_J_Cantwell.pdf\n\nITA — Departamento de Ciência e Tecnologia Aeroespacial: catálogo de graduação e disciplinas de Termodinâmica/Termodinâmica Aplicada:\nhttps://www.ita.br/sites/default/files/pages/collection/Cat%C3%A1logo%20dos%20Cursos%20de%20Gradua%C3%A7%C3%A3o%202026%20-%20digital%20Rev.26.02.24.pdf\n\nIME-USP — pesquisas acadêmicas envolvendo termodinâmica e sistemas de muitos corpos:\nhttps://lattes.ime.usp.br/posmap/membro-1498618533380124.html\n\nNIST — REFPROP / propriedades de misturas e AGA8:\nhttps://www.nist.gov/srd/refprop\n\nNIST — comparação de equações de estado para medição de gás natural:\nhttps://www.nist.gov/publications/comparison-five-natural-gas-equations-state-used-flow-and-energy-measurement'

    def screen_names_for_language(self):
        return MOBILE_TABS.get(self.idioma, MOBILE_TABS["pt-BR"])

    def _make_scroll(self):
        return ScrollView(do_scroll_x=False, do_scroll_y=True)

    def _content_box(self, padding=10):
        box=GridLayout(cols=1, spacing=dp(7), padding=dp(padding), size_hint_y=None)
        box.bind(minimum_height=box.setter("height"))
        return box

    def _label(self, text, h=32, bold=False):
        return Label(text=self._t(text), size_hint_y=None, height=dp(h), font_size=sp(12), bold=bold, halign="left", valign="middle")

    def _entry(self, value="", h=40):
        """Cria um campo de entrada já com a configuração visual atual."""
        e = GradientTextInput(
            text=str(value), multiline=False, size_hint_y=None,
            height=dp(h), font_size=sp(14), padding=[dp(8), dp(8)]
        )
        e._visual_role = "input"
        try:
            hx = self.colors.get('input_text', '#241713')
            e.foreground_color = self._hex_rgba(hx)
            e.disabled_foreground_color = e.foreground_color
            e.cursor_color = e.foreground_color
            e.hint_text_color = [*e.foreground_color[:3], 0.65]
            e.selection_color = [*e.foreground_color[:3], 0.30]
            e.set_gradient(
                self.colors.get('field_start', '#FFFFFF'),
                self.colors.get('field_end', '#E6CDB8'),
                self.colors.get('gradient_enabled', True)
            )
        except Exception:
            pass
        return e

    def _button(self, text, callback, h=42):
        """Cria botão sólido usando diretamente a cor configurada."""
        b = Button(
            text=self._t(text), size_hint_y=None, height=dp(h), font_size=sp(12),
            background_normal='', background_down=''
        )
        b._visual_role = "button"
        try:
            hx = self.colors.get('button', '#B88962')
            b.background_color = self._hex_rgba(hx)
            b.color = self._contrast_color(hx)
        except Exception:
            pass
        b.bind(on_release=callback)
        return b

    def _t(self, text):
        return mobile_text(self.idioma, text)

    def _hex_rgba(self, value, alpha=1.0):
        try:
            v=str(value).strip().lstrip("#")
            if len(v)==3: v=''.join(c*2 for c in v)
            if len(v)!=6: raise ValueError
            return [int(v[i:i+2],16)/255.0 for i in (0,2,4)] + [alpha]
        except Exception:
            return [0,0,0,1]

    def _relative_luminance(self, hex_color):
        """Calcula a luminância relativa WCAG da cor #RRGGBB.

        A luminância é usada somente para decidir se texto escuro ou claro
        terá contraste suficiente. Ela não altera a cor escolhida pelo usuário.
        """
        rgb = self._hex_rgba(hex_color)[:3]
        linear = []
        for channel in rgb:
            channel = float(channel)
            linear.append(channel / 12.92 if channel <= 0.03928 else ((channel + 0.055) / 1.055) ** 2.4)
        return 0.2126 * linear[0] + 0.7152 * linear[1] + 0.0722 * linear[2]

    def _contrast_ratio(self, color_a, color_b):
        """Retorna a razão de contraste WCAG entre duas cores."""
        la = self._relative_luminance(color_a)
        lb = self._relative_luminance(color_b)
        hi, lo = max(la, lb), min(la, lb)
        return (hi + 0.05) / (lo + 0.05)

    def _contrast_color(self, hex_color):
        """Escolhe preto ou branco com a maior razão de contraste."""
        black = '#111111'
        white = '#FFFDF8'
        return self._hex_rgba(black if self._contrast_ratio(hex_color, black) >= self._contrast_ratio(hex_color, white) else white)

    def _preferred_text_for_background(self, preferred_hex, background_hex):
        """
        Usa a cor de letra escolhida pelo usuário quando ela é legível.

        Regra definitiva: se a cor escolhida não atingir contraste WCAG de
        aproximadamente 4.5:1, ela é substituída automaticamente pela melhor
        alternativa escura/clara. Assim o usuário pode escolher a cor da letra,
        mas nunca consegue deixar o programa ilegível por acidente.
        """
        try:
            preferred = str(preferred_hex).upper()
            if self._contrast_ratio(preferred, background_hex) >= 4.5:
                return self._hex_rgba(preferred)
        except Exception:
            pass
        return self._contrast_color(background_hex)

    def _input_text_color_exact(self, preferred_hex):
        """
        V28.13 - Retorna EXATAMENTE a cor escolhida pelo usuário para os
        caracteres dos campos de entrada.

        Versões anteriores substituíam automaticamente cores claras quando
        o contraste era considerado insuficiente. Isso fazia parecer que a
        configuração não funcionava. Os padrões dos temas já são legíveis;
        quando o usuário escolhe outra cor, essa cor é aplicada literalmente.
        """
        return self._hex_rgba(preferred_hex)

    def _result_text_color_exact(self, preferred_hex):
        """V28.13 - Retorna EXATAMENTE a cor escolhida para os resultados."""
        return self._hex_rgba(preferred_hex)

    def _find_widgets_of_type(self, root, widget_type):
        """Retorna todos os widgets de um tipo sem depender da ordem dos filhos."""
        found = []
        def visit(w):
            if isinstance(w, widget_type):
                found.append(w)
            for child in getattr(w, 'children', []):
                visit(child)
        if root is not None:
            visit(root)
        return found

    def _mark_visual_ready(self):
        """Marca a árvore Kivy como pronta e aplica a configuração uma única vez."""
        self._visual_ready = True
        self._apply_colors()

    def _apply_colors(self):
        """
        Aplica a paleta atual de forma determinística.

        PRINCÍPIO DESTA VERSÃO:
        1. O usuário escolhe uma cor.
        2. O programa tenta respeitar a cor escolhida para o texto.
        3. Se ela ficar ilegível sobre o fundo, o programa troca somente a
           cor do texto por uma alternativa de alto contraste.

        Assim, tema claro e tema escuro não dependem de uma cor fixa de letra.
        """
        if getattr(self, 'root', None) is None:
            return
        c = self.colors
        bg_hex = c.get('bg', '#F5EFEA')
        text_hex = c.get('text', '#2E211B')
        input_text_hex = c.get('input_text', '#241713')
        result_text_hex = c.get('result_text', '#2E211B')
        button_hex = c.get('button', '#B88962')
        field_hex = c.get('field', '#FFF8F0')
        field_start = c.get('field_start', field_hex)
        field_end = c.get('field_end', field_hex)
        result_hex = c.get('result', '#F0E2D7')
        row1_hex = c.get('row1', result_hex)
        row2_hex = c.get('row2', result_hex)
        alignment = c.get('alignment', 'center')

        Window.clearcolor = self._hex_rgba(bg_hex)

        # Cada tipo de componente recebe sua própria cor de texto.
        bg_text = self._preferred_text_for_background(text_hex, bg_hex)
        button_text = self._preferred_text_for_background(text_hex, button_hex)
        # IMPORTANTE: os campos e os resultados não usam mais a cor geral.
        # Cada grupo possui sua própria caixa de cor na Configuração.
        # V28.13: as duas caixas de cor de texto são independentes e literais.
        # Não substituímos a escolha do usuário por marrom/preto/branco automático.
        field_text = self._input_text_color_exact(input_text_hex)
        row1_text = self._result_text_color_exact(result_text_hex)
        row2_text = self._result_text_color_exact(result_text_hex)

        def walk(w):
            try:
                if isinstance(w, GradientTextInput):
                    # TextInput precisa de foreground_color; color não é o
                    # atributo correto para o texto digitado.
                    w.foreground_color = field_text
                    w.hint_text_color = [field_text[0], field_text[1], field_text[2], 0.65]
                    w.cursor_color = field_text
                    w.selection_color = [field_text[0], field_text[1], field_text[2], 0.30]
                    w.disabled_foreground_color = field_text
                    w.set_gradient(field_start, field_end, c.get('gradient_enabled', True))

                elif isinstance(w, TextInput):
                    w.foreground_color = field_text
                    w.hint_text_color = [field_text[0], field_text[1], field_text[2], 0.65]
                    w.cursor_color = field_text
                    w.selection_color = [field_text[0], field_text[1], field_text[2], 0.30]
                    w.disabled_foreground_color = field_text

                elif isinstance(w, ScrollText):
                    # AQUI está a correção definitiva do efeito zebrado:
                    # fundo e texto são calculados separadamente para cada linha.
                    w.set_colors(
                        background_color=self._hex_rgba(result_hex),
                        row1=self._hex_rgba(row1_hex),
                        row2=self._hex_rgba(row2_hex),
                        alignment=alignment,
                        row_text1=row1_text,
                        row_text2=row2_text
                    )

                elif isinstance(w, Button):
                    # Os controles da paleta são tratados novamente abaixo,
                    # depois da caminhada, para preservar a cor individual.
                    w.background_normal = ''
                    w.background_down = ''
                    w.background_color = self._hex_rgba(button_hex)
                    w.color = button_text

                elif isinstance(w, Spinner):
                    w.background_normal = ''
                    w.background_color = self._hex_rgba(button_hex)
                    w.color = button_text

                elif isinstance(w, Label):
                    # Labels que pertencem a um ScrollText NÃO podem receber a
                    # cor geral. Isso era o motivo de a cor da Linha 1/Linha 2
                    # ser perdida depois da caminhada pela árvore de widgets.
                    if hasattr(w, '_row_text_color'):
                        w.color = w._row_text_color
                    else:
                        w.color = bg_text

            except Exception:
                # Falha visual isolada jamais deve derrubar o aplicativo.
                pass

            for child in getattr(w, 'children', []):
                walk(child)

        if getattr(self, 'root', None) is not None:
            walk(self.root)

        # REAPLICAÇÃO FINAL DAS LINHAS 1 E 2
        # ---------------------------------------------------------------
        # Esta segunda passagem é proposital. Ela garante que nenhum Label
        # genérico, Spinner ou outro componente tenha sobrescrito as cores
        # individuais escolhidas pelo usuário para os resultados.
        for _scroll in self._find_widgets_of_type(self.root, ScrollText):
            try:
                _scroll.set_colors(
                    row1=self._hex_rgba(row1_hex),
                    row2=self._hex_rgba(row2_hex),
                    alignment=alignment,
                    row_text1=row1_text,
                    row_text2=row2_text
                )
            except Exception:
                pass

        # Cabeçalho e rodapé usam o fundo principal como referência.
        for attr in ('header', 'footer'):
            w = getattr(self, attr, None)
            if w is not None:
                try:
                    w.color = bg_text
                except Exception:
                    pass

        # Navegação usa a cor do botão.
        for attr in ('tab_spinner', 'lang_spinner'):
            w = getattr(self, attr, None)
            if w is not None:
                try:
                    w.color = button_text
                    w.background_normal = ''
                    w.background_color = self._hex_rgba(button_hex)
                except Exception:
                    pass

        # V28.12: cartão m³ destacado e com contraste garantido.
        try:
            self._style_gnv_m3_card(row1_hex, result_text_hex)
        except Exception:
            pass

        self._refresh_color_controls()

    def _apply_result_styles_to_all_tabs(self):
        """
        V28.19 — Aplica as configurações de resultados explicitamente
        em todas as abas que possuem ScrollText.

        Isso evita depender somente da caminhada genérica da árvore Kivy.
        Cada aba mantém sua própria referência ao resultado.
        """
        if getattr(self, "root", None) is None:
            return

        c = self.colors
        row1 = self._hex_rgba(c.get("row1", c.get("result", "#F7ECE3")))
        row2 = self._hex_rgba(c.get("row2", c.get("result", "#E1C8B5")))
        result_bg = self._hex_rgba(c.get("result", "#F0E2D7"))
        result_text = self._hex_rgba(c.get("result_text", "#2E211B"))
        alignment = c.get("alignment", "center")

        # Referências de resultados das diversas abas.
        attrs = (
            "calc_result",
            "ab_compare",
            "anp_result",
            "comp_result",
            "comp_chart_result",
            "hist_result",
            "sqlite_result",
            "chart_result",
            "formula_view",
            "total_result",
        )

        for attr in attrs:
            widget = getattr(self, attr, None)
            if isinstance(widget, ScrollText):
                try:
                    widget.set_colors(
                        background_color=result_bg,
                        row1=row1,
                        row2=row2,
                        alignment=alignment,
                        row_text1=result_text,
                        row_text2=result_text,
                    )
                except Exception:
                    pass

        # Cobertura adicional: qualquer ScrollText criado futuramente.
        try:
            for widget in self._find_widgets_of_type(self.root, ScrollText):
                widget.set_colors(
                    background_color=result_bg,
                    row1=row1,
                    row2=row2,
                    alignment=alignment,
                    row_text1=result_text,
                    row_text2=result_text,
                )
        except Exception:
            pass

    def _goto_creditos(self, *_args):
        """Abre diretamente a aba Créditos."""
        try:
            self.sm.current = self.screen_names[11]
            names = self.screen_names_for_language()
            if len(names) > 11:
                self.tab_spinner.text = names[11]
        except Exception:
            pass

    def _refresh_color_controls(self):
        """
        Atualiza SOMENTE os controles da paleta arco-íris.

        Este método é chamado no final de _apply_colors, depois da caminhada
        pelos widgets. Isso impede que um Button genérico sobrescreva a cor
        escolhida para Linha 1, Linha 2 ou qualquer outro item da paleta.
        """
        for key, control in getattr(self, 'color_controls', {}).items():
            try:
                hx = str(self.colors.get(key, '#FFFFFF')).upper()
                control.background_normal = ''
                control.background_down = ''
                control.background_color = self._hex_rgba(hx)
                control.color = self._contrast_color(hx)
                control.text = f'{self._color_name(hx, self.palette_map)}\n{hx}'
            except Exception:
                pass

    def _default_colors(self):
        """
        Tema inicial claro e legível.

        O resultado não é preto e os campos têm um degradê suave. O usuário
        pode modificar cada uma dessas cores no seletor arco-íris.
        """
        return {
            'theme': 'claro',
            'bg': '#F5EFEA',
            'text': '#2E211B',
            # Cor exclusiva dos caracteres digitados nos campos de entrada.
            'input_text': '#241713',
            # Cor exclusiva dos textos produzidos pelo sistema/resultados.
            'result_text': '#2E211B',
            'button': '#B88962',
            'field': '#FFF8F0',
            'field_start': '#FFFFFF',
            'field_end': '#E6CDB8',
            'gradient_enabled': True,
            'result': '#F0E2D7',
            'row1': '#F7ECE3',
            'row2': '#E1C8B5',
            'alignment': 'center'
        }


    # =====================================================================
    # V28.18 — APLICAÇÃO VISUAL MULTIABA EXPLÍCITA
    # ---------------------------------------------------------------------
    # Em versões anteriores, a configuração dependia de percorrer a árvore
    # inteira do Kivy. Isso funcionava em algumas telas e falhava em outras.
    # A V28.18 mantém uma lista explícita de TODOS os ScrollText criados e
    # reaplica fundo/linha/alinhamento diretamente a cada instância.
    # O mesmo método é chamado imediatamente quando o usuário altera, salva
    # ou restaura uma configuração. Nenhum reinício do programa é necessário.
    # =====================================================================
    def _all_result_widgets(self):
        """Retorna todos os relatórios/textos que devem receber o estilo zebrado."""
        attrs = [
            'calc_result', 'anp_result', 'comp_result', 'comp_chart_result',
            'ab_compare', 'hist_result', 'sqlite_result', 'chart_result', 'formula_view',
            'total_result'
        ]
        out = []
        seen = set()
        for attr in attrs:
            w = getattr(self, attr, None)
            if isinstance(w, ScrollText) and id(w) not in seen:
                out.append(w); seen.add(id(w))
        return out

    def _apply_all_visual_now(self):
        """Aplica toda a configuração visual AGORA, sem reconstruir telas."""
        c = self.colors
        # Janela.
        Window.clearcolor = self._hex_rgba(c.get('bg', '#F5EFEA'))

        # Campos de entrada: aplica diretamente em cada instância viva.
        input_text = self._hex_rgba(c.get('input_text', '#241713'))
        start = c.get('field_start', c.get('field', '#FFFFFF'))
        end = c.get('field_end', c.get('field', '#FFFFFF'))
        enabled = bool(c.get('gradient_enabled', True))
        for w in self._find_widgets_of_type(self.root, GradientTextInput):
            try:
                w.foreground_color = input_text
                w.disabled_foreground_color = input_text
                w.cursor_color = input_text
                w.hint_text_color = [*input_text[:3], 0.65]
                w.selection_color = [*input_text[:3], 0.30]
                w.set_gradient(start, end, enabled)
            except Exception:
                pass

        # Botões gerais.
        button_hex = c.get('button', '#B88962')
        for w in self._find_widgets_of_type(self.root, Button):
            try:
                w.background_normal = ''
                w.background_down = ''
                w.background_color = self._hex_rgba(button_hex)
                w.color = self._contrast_color(button_hex)
            except Exception:
                pass

        # Relatórios: esta é a parte que precisava ser multiaba.
        row1 = self._hex_rgba(c.get('row1', '#F7ECE3'))
        row2 = self._hex_rgba(c.get('row2', '#E1C8B5'))
        result_text = self._hex_rgba(c.get('result_text', '#2E211B'))
        alignment = c.get('alignment', 'center')
        for w in self._all_result_widgets():
            try:
                w.set_colors(
                    background_color=self._hex_rgba(c.get('result', '#F0E2D7')),
                    row1=row1, row2=row2,
                    alignment=alignment,
                    row_text1=result_text, row_text2=result_text
                )
            except Exception:
                pass

        # Labels comuns.
        general = self._hex_rgba(c.get('text', '#2E211B'))
        for w in self._find_widgets_of_type(self.root, Label):
            try:
                if not hasattr(w, '_row_text_color') and w is not getattr(self, 'calc_cylinder_m3_card', None):
                    w.color = general
            except Exception:
                pass

        # Cartão de m³.
        try:
            self._style_gnv_m3_card(c.get('row1', '#F7ECE3'), c.get('result_text', '#2E211B'))
        except Exception:
            pass

        # Botões da própria paleta precisam continuar mostrando a cor que
        # representam, em vez da cor geral dos botões.
        self._refresh_color_controls()

    def _update_config_controls_from_colors(self):
        """Sincroniza tema, alinhamento, degradê e amostras sem reconstruir abas."""
        if hasattr(self, 'theme_spinner'):
            self.theme_spinner.text = self.colors.get('theme', 'claro')
        if hasattr(self, 'alignment_spinner'):
            self.alignment_spinner.text = {
                'left':'Esquerda','center':'Centro','right':'Direita'
            }.get(self.colors.get('alignment','center'),'Centro')
        if hasattr(self, 'gradient_spinner'):
            self.gradient_spinner.text = 'Ativado' if self.colors.get('gradient_enabled', True) else 'Desativado'
        self._refresh_color_controls()

    def _build_all_screens(self):
        self._build_calculos()
        self._build_abastecimentos()
        self._build_anp()
        self._build_compressao()
        self._build_historico()
        self._build_sqlite()
        self._build_excel()
        self._build_graficos()
        self._build_config()
        self._build_formulas()
        self._build_total()
        self._build_creditos()

    def _set_screen_content(self, idx, widget, horizontal=False):
        scr=self.sm.get_screen(self.screen_names[idx])
        scr.clear_widgets()
        if isinstance(widget, ScrollText):
            scr.add_widget(widget)
            return
        sv=ScrollView(do_scroll_y=True, do_scroll_x=horizontal, bar_width=dp(10), size_hint=(1,1))
        sv.add_widget(widget)
        scr.add_widget(sv)

    # ---------------- CALCULOS ----------------
    def _build_calculos(self):
        box=self._content_box(); self.calc_entries={}
        fields=[
            ("Volume do cilindro (L):",26),("Quantidade de cilindros:",1),("Pressão (bar):",220),
            ("Temperatura (°C):",20),("Altitude (m):",50),("Fator Z:",0.92),
            ("Massa molar (kg/mol):",0.01604),("Massa específica de referência (kg/m³):",0.76)
        ]
        for key,val in fields:
            box.add_widget(self._label(key)); e=self._entry(val); self.calc_entries[key]=e; box.add_widget(e)
            # V28.13: o cartão de destaque agora mostra o volume de GNV
            # calculado; ele é atualizado pelo método _calcular().

        # Cartão de destaque do VOLUME DE GNV em m³.
        # Ele não representa a capacidade física do cilindro; mostra os
        # volumes equivalentes calculados pelo sistema.
        self.calc_cylinder_m3_card = Label(
            text='VOLUME DE GNV — m³\nAguardando cálculo...\nANP / equivalente físico',
            size_hint_y=None, height=dp(165), font_size=sp(16), bold=True,
            halign='center', valign='middle', padding=(dp(8), dp(8))
        )
        self.calc_cylinder_m3_card.bind(size=lambda w,*a: setattr(w, 'text_size', (w.width-dp(16), w.height-dp(12))))
        with self.calc_cylinder_m3_card.canvas.before:
            self._cyl_card_color = Color(0.75,0.55,0.35,1)
            self._cyl_card_rect = Rectangle(pos=self.calc_cylinder_m3_card.pos, size=self.calc_cylinder_m3_card.size)
        self.calc_cylinder_m3_card.bind(pos=lambda w,*a:setattr(self._cyl_card_rect,'pos',w.pos), size=lambda w,*a:setattr(self._cyl_card_rect,'size',w.size))
        box.add_widget(self.calc_cylinder_m3_card)
        row=BoxLayout(size_hint_y=None,height=dp(45),spacing=dp(5))
        row.add_widget(self._button("Valores padrão",lambda *_:self._default_calc_values(),40))
        row.add_widget(self._button("Calcular",lambda *_:self._calcular(),40))
        row.add_widget(self._button("Limpar",lambda *_:self._clear_calc(),40))
        box.add_widget(row)

        # ---------------------------------------------------------------
        # Compartilhamento do resultado.
        # ---------------------------------------------------------------
        share_row = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(5))
        share_row.add_widget(
            self._button(
                "Copiar para WhatsApp",
                lambda *_: self._copiar_whatsapp(),
                40
            )
        )
        share_row.add_widget(
            self._button(
                "Abrir WhatsApp",
                lambda *_: self._abrir_whatsapp(),
                40
            )
        )
        share_row.add_widget(
            self._button(
                "Gerar JPG",
                lambda *_: self._gerar_jpg_whatsapp(),
                40
            )
        )
        box.add_widget(share_row)

        self.calc_share_status = self._label("", h=42)
        box.add_widget(self.calc_share_status)
        self.calc_result=ScrollText("", horizontal=False, size_hint_y=None, height=dp(300)); box.add_widget(self.calc_result)
        self._set_screen_content(0,box)

    def _update_gnv_m3_card(self, resultado=None, anp=None):
        """Destaca o volume físico e os quatro resultados equivalentes.

        Os resultados ANP e Físico Z são expressos na referência de 20 °C.
        "T informada" usa a temperatura digitada pelo usuário antes da
        conversão para 20 °C; o cenário "20 °C" recalcula como se o gás
        estivesse a 20 °C. Portanto, acima de 20 °C o valor informado é
        menor; abaixo de 20 °C ele é maior, mantendo pressão e volume físico.
        """
        card = getattr(self, 'calc_cylinder_m3_card', None)
        if card is None:
            return
        try:
            if resultado is None:
                card.text = (
                    'VOLUME DE GNV — m³\n'
                    'Aguardando cálculo...\n'
                    'Todos os volumes serão apresentados na referência ANP de 20 °C.'
                )
                return

            temp = float(resultado.get('temperatura_c', 20.0))
            anp_t = float(resultado.get(
                'volume_anp_ideal_m3_temperatura_informada', 0.0))
            anp20 = float(resultado.get(
                'volume_anp_ideal_m3_20c', anp or 0.0))
            fis_t = float(resultado.get(
                'volume_equivalente_m3_temperatura_informada', 0.0))
            fis20 = float(resultado.get(
                'volume_equivalente_m3_20c', 0.0))

            volume_l = float(resultado.get('volume_cilindro_l', 0.0))
            volume_m3 = float(resultado.get('volume_cilindro_m3', volume_l / 1000.0))

            # IMPORTANTE: os quatro resultados são volumes equivalentes
            # EXPRESSOS NA REFERÊNCIA DE 20 °C.
            #
            # "T informada" significa: quantidade calculada usando a
            # temperatura digitada pelo usuário e depois convertida para
            # a referência de 20 °C. Por isso, em uma temperatura acima
            # de 20 °C, esse resultado é MENOR que o cenário a 20 °C.
            card.text = (
                'VOLUME DE GNV — m³\n'
                f'Físico do cilindro: {volume_l:.3f} L = {volume_m3:.6f} m³\n'
                f'ANP — T informada ({temp:.1f} °C) → 20 °C: {anp_t:.3f} m³\n'
                f'ANP — cenário a 20 °C: {anp20:.3f} m³\n'
                f'Físico Z — T informada ({temp:.1f} °C) → 20 °C: {fis_t:.3f} m³\n'
                f'Físico Z — cenário a 20 °C: {fis20:.3f} m³'
            )
        except Exception:
            card.text = 'VOLUME DE GNV — m³\nErro ao atualizar o destaque.'

    def _style_gnv_m3_card(self, bg_hex=None, text_hex=None):
        """Aplica o fundo e o texto configurados no cartão de volume de GNV."""
        card=getattr(self,'calc_cylinder_m3_card',None)
        if card is None: return
        bg_hex=bg_hex or self.colors.get('row1','#F7ECE3')
        text_hex=text_hex or self.colors.get('result_text','#2E211B')
        card.color=self._hex_rgba(text_hex)
        try: card._cyl_card_color.rgba=self._hex_rgba(bg_hex)
        except Exception: pass

    def _default_calc_values(self):
        vals={"Volume do cilindro (L):":26,"Quantidade de cilindros:":1,"Pressão (bar):":220,"Temperatura (°C):":20,"Altitude (m):":50,"Fator Z:":0.92,"Massa molar (kg/mol):":0.01604,"Massa específica de referência (kg/m³):":0.76}
        for k,v in vals.items(): self.calc_entries[k].text=str(v)
        # O cartão de volume de GNV só é preenchido após um cálculo.
        self._update_gnv_m3_card(None, None)

    def _num(self,key): return converter_numero(self.calc_entries[key].text)

    def _calcular(self):
        try:
            volume=self._num("Volume do cilindro (L):")
            quantidade=self._num("Quantidade de cilindros:")
            pressao=self._num("Pressão (bar):")
            temp=self._num("Temperatura (°C):")
            alt=self._num("Altitude (m):")
            z=self._num("Fator Z:")
            mm=self._num("Massa molar (kg/mol):")
            dens=self._num("Massa específica de referência (kg/m³):")
            if volume<=0 or quantidade<=0 or pressao<0 or temp<=-273.15 or z<=0 or mm<=0 or dens<0:
                raise ValueError(self._t("Dados físicos inválidos."))

            volume_total=volume*quantidade
            r=calcular_quantidade_gnv(volume_total,pressao,temp,alt,z,mm,dens)
            anp_t=r["volume_anp_ideal_m3_temperatura_informada"]
            anp20=r["volume_anp_ideal_m3_20c"]
            fis_t=r["volume_equivalente_m3_temperatura_informada"]
            fis20=r["volume_equivalente_m3_20c"]
            patm = calcular_pressao_atmosferica(alt)
            z_t_calc = calcular_Z(pressao + patm, temp + 273.15)["Z"]
            z_20_calc = calcular_Z(pressao + patm, 293.15)["Z"]
            self._update_gnv_m3_card(r,anp20)
            faixas=r['volume_circuito_gnv']

            # Tabela térmica: recalcula CADA temperatura.
            # Todas as colunas representam m³ na referência ANP de 20 °C.
            # Portanto, acima de 20 °C o valor diminui; abaixo de 20 °C aumenta.
            tabela_temp = []
            patm = calcular_pressao_atmosferica(alt)
            delta_p = max(0.0, pressao)
            volume_m3_calc = volume_total / 1000.0
            for tc in range(-50, 101, 10):
                tk = tc + 273.15
                if tk <= 0:
                    continue

                # ANP/idealizado: gás ideal, padronizado para 20 °C.
                anp_tc_20 = (
                    volume_m3_calc
                    * delta_p
                    / PRESSAO_REFERENCIA_ANP_BAR
                    * (293.15 / tk)
                )

                # Físico Z: PV = ZnRT, com Z informado pelo usuário,
                # também convertido para a referência de 20 °C.
                # Quantidade ADICIONADA: ΔP = Pfinal - Pinicial.
                # Como Pinicial = 0 bar manométrico, ΔP = pressao.
                n_tc = (
                    pressao * 100000.0 * volume_m3_calc
                    / (z * R * tk)
                )
                fis_tc_20 = calcular_volume_referencia_m3(
                    n_tc, 20.0, PRESSAO_REFERENCIA_ANP_BAR, 1.0
                )

                tabela_temp.append(
                    f"{tc:>4} °C | ANP / 20 °C: {anp_tc_20:.3f} m³ | "
                    f"Físico Z / 20 °C: {fis_tc_20:.3f} m³"
                )

            linhas=[
                self._t("RELATÓRIO DOS CÁLCULOS DE GNV"),"="*70,"",
                f"Capacidade física total: {volume_total:.2f} L = {volume_total/1000:.6f} m³",
                f"Quantidade de cilindros: {self.calc_entries['Quantidade de cilindros:'].text}",
                f"Pressão: {pressao:.2f} bar",
                f"Temperatura informada: {temp:.2f} °C",
                f"Altitude: {alt:.2f} m", f"Fator Z: {z:.4f}",
                f"Massa molar: {mm:.8f} kg/mol",
                f"Massa específica de referência: {dens:.6f} kg/m³", "",
                "==============================================================",
                "VOLUME DE GNV — RESULTADO PRINCIPAL",
                "==============================================================",
                "TODOS OS VALORES ABAIXO ESTÃO EXPRESSOS NA REFERÊNCIA ANP DE 20 °C.",
                f"ANP / idealizado — usando T informada ({temp:.2f} °C): {anp_t:.6f} m³",
                f"ANP / idealizado — cenário recalculado a 20 °C: {anp20:.6f} m³",
                f"Físico Z — usando T informada ({temp:.2f} °C): {fis_t:.6f} m³",
                f"Físico Z — cenário recalculado a 20 °C: {fis20:.6f} m³",
                f"Fator Z calculado — T informada ({temp:.2f} °C): {z_t_calc:.6f}",
                f"Fator Z calculado — 20 °C: {z_20_calc:.6f}", "",
                "LEITURA PARA LEIGOS:",
                f"O cilindro tem {volume_total/1000:.6f} m³ de espaço físico.",
                f"Com T informada de {temp:.2f} °C, o ANP/idealizado resulta em {anp_t:.3f} m³ na referência de 20 °C.",
                f"No modelo Físico Z, a mesma condição resulta em {fis_t:.3f} m³ na referência de 20 °C.",
                f"Se o mesmo cilindro fosse considerado a 20 °C, os valores seriam ANP {anp20:.3f} m³ e Físico Z {fis20:.3f} m³.",
                "Portanto, acima de 20 °C a quantidade de GNV estimada para o mesmo volume e a mesma pressão é MENOR; abaixo de 20 °C é MAIOR.",
                "Os m³ acima são volumes equivalentes/padronizados e NÃO são o espaço físico ocupado dentro do cilindro.", "",
                "EFEITO DA TEMPERATURA — -50 °C a +100 °C (passo de 10 °C)",
                "TODAS AS LINHAS ESTÃO CONVERTIDAS PARA A REFERÊNCIA DE 20 °C. Isso permite comparar diretamente o efeito da temperatura.",
                *tabela_temp, "",
                f"Massa de GNV: {r['massa']:.6f} kg",
                f"Densidade calculada no cilindro: {r['densidade']:.3f} kg/m³",
                f"Quantidade de matéria: {r['mols']:.6f} mol", "",
                self._t("ESTIMATIVA GEOMÉTRICA DO CIRCUITO ACESSÓRIO DE GNV"),
                self._t("(EXCLUINDO O CILINDRO / RESERVATÓRIO)"),
                f"Tubulação alta pressão — mínimo: {faixas['minimo']['volume_tubulacao_ml']:.1f} mL",
                f"Tubulação alta pressão — médio: {faixas['medio']['volume_tubulacao_ml']:.1f} mL",
                f"Tubulação alta pressão — máximo: {faixas['maximo']['volume_tubulacao_ml']:.1f} mL",
                "Válvulas de serviço: 3 / 4 / 5 mL",
                "Redutor — câmaras de gás: 32 / 46 / 65 mL",
                "Baixa pressão/filtro/flauta: 0 / 25 / 140 mL",
                "VOLUME TOTAL DO CIRCUITO: 75 / 130 / 280 mL", "",
                self._t("Base geométrica da tubulação: V = pi × r² × L."),
                self._t("Premissa: tubo externo 6 mm, parede 1,2 mm, interno 3,6 mm, comprimento 4,0–5,5 m."),
                self._t("ATENÇÃO: estimativa de engenharia; não substitui medição real do veículo.")
            ]
            self.calc_result.set_text("\n".join(linhas))
        except Exception as exc:
            self.calc_result.set_text(self._t("Erro")+f": {exc}")

    # =====================================================================
    # WHATSAPP — COMPARTILHAMENTO DOS RESULTADOS
    # =====================================================================

    def _whatsapp_message(self):
        """
        Monta o relatório COMPLETO para WhatsApp.

        O relatório diferencia:
        - volume físico do cilindro (capacidade geométrica);
        - volume equivalente ANP/idealizado;
        - volume equivalente do modelo físico.

        Também inclui todos os dados de entrada e os principais resultados.
        """
        try:
            if not self.calc_result._text_value.strip():
                return "Execute um cálculo antes de compartilhar."

            volume_l = self._num("Volume do cilindro (L):")
            quantidade = self._num("Quantidade de cilindros:")
            pressao_bar = self._num("Pressão (bar):")
            temperatura_c = self._num("Temperatura (°C):")
            altitude_m = self._num("Altitude (m):")
            fator_z = self._num("Fator Z:")
            massa_molar = self._num("Massa molar (kg/mol):")
            densidade_ref = self._num("Massa específica de referência (kg/m³):")

            volume_total_l = volume_l * quantidade
            volume_fisico_m3 = volume_total_l / 1000.0

            r = calcular_quantidade_gnv(
                volume_total_l,
                pressao_bar,
                temperatura_c,
                altitude_m,
                fator_z,
                massa_molar,
                densidade_ref
            )

            anp_t = float(r.get("volume_anp_ideal_m3_temperatura_informada", 0.0))
            anp = float(r.get("volume_anp_ideal_m3_20c", 0.0))
            vol_t = float(r.get("volume_equivalente_m3_temperatura_informada", 0.0))
            vol_20 = float(r.get("volume_equivalente_m3_20c", 0.0))

            faixas = r.get("volume_circuito_gnv", {})
            minimo = faixas.get("minimo", {})
            medio = faixas.get("medio", {})
            maximo = faixas.get("maximo", {})

            massa = float(r.get("massa", 0.0))
            densidade = float(r.get("densidade", 0.0))
            mols = float(r.get("mols", 0.0))
            volume_especifico = float(r.get("volume_especifico", 0.0))
            volume_real = float(r.get("volume_real", 0.0))
            pressao_atm = float(r.get("pressao_atm", 0.0))
            pressao_abs = float(r.get("pressao_absoluta", 0.0))
            temperatura_k = float(r.get("temperatura_k", 0.0))

            # Tabela térmica do WhatsApp: recalcula cada temperatura e
            # apresenta TODOS os valores na referência ANP de 20 °C.
            tabela_temperatura_whatsapp = []
            for tc in range(-50, 101, 10):
                tk = tc + 273.15
                anp_tc_20 = (
                    volume_fisico_m3
                    * pressao_bar
                    / PRESSAO_REFERENCIA_ANP_BAR
                    * (293.15 / tk)
                )
                # Quantidade adicionada: usar ΔP, pois a pressão inicial
                # é 0 bar manométrico neste cálculo.
                n_tc = (
                    pressao_bar * 100000.0 * volume_fisico_m3
                    / (fator_z * R * tk)
                )
                fis_tc_20 = calcular_volume_referencia_m3(
                    n_tc, 20.0, PRESSAO_REFERENCIA_ANP_BAR, 1.0
                )
                tabela_temperatura_whatsapp.append(
                    f"{tc:>4} C | ANP / 20 C: {anp_tc_20:.3f} m3 | "
                    f"Fisico Z / 20 C: {fis_tc_20:.3f} m3"
                )

            linhas = [
                "==================================================",
                "GNV - ANALISE COMPLETA DO ABASTECIMENTO",
                "==================================================",
                "",
                "IDENTIFICACAO",
                "Sistema de Calculos e Analise da Capacidade do Cilindro de GNV - V28.26",
                "Autor: Christiano T. Gaio",
                "Perfil GitHub: https://github.com/Gaio-Christiano",
                "Repositorio oficial: https://github.com/Gaio-Christiano/GNV_m-_carro",
                "",
                "--------------------------------------------------",
                "1. DADOS DE ENTRADA",
                "--------------------------------------------------",
                f"Capacidade de cada cilindro: {volume_l:.3f} L",
                f"Quantidade de cilindros: {quantidade:.0f}",
                f"Capacidade fisica TOTAL: {volume_total_l:.3f} L",
                f"Capacidade fisica TOTAL: {volume_fisico_m3:.6f} m3",
                f"Pressao manometrica: {pressao_bar:.3f} bar",
                f"Pressao atmosferica estimada: {pressao_atm:.6f} bar",
                f"Pressao absoluta: {pressao_abs:.6f} bar",
                f"Temperatura: {temperatura_c:.3f} C",
                f"Temperatura absoluta: {temperatura_k:.3f} K",
                f"Altitude: {altitude_m:.3f} m",
                f"Fator Z: {fator_z:.6f}",
                f"Massa molar: {massa_molar:.8f} kg/mol",
                f"Massa especifica de referencia: {densidade_ref:.6f} kg/m3",
                "",
                "--------------------------------------------------",
                "2. DESTAQUE - VOLUME FISICO DO CILINDRO",
                "--------------------------------------------------",
                f"VOLUME FISICO: {volume_fisico_m3:.6f} m3",
                f"({volume_total_l:.3f} L de capacidade geometrica)",
                "",
                "IMPORTANTE: o volume fisico acima e a capacidade geometrica.",
                "Os valores em m3 abaixo sao volumes EQUIVALENTES da mesma",
                "quantidade de gas quando expressos em outra condicao.",
                "",
                "--------------------------------------------------",
                "3. VOLUMES EQUIVALENTES DE GNV — RESULTADO PRINCIPAL",
                "--------------------------------------------------",
                "TODOS OS VALORES ABAIXO ESTÃO EXPRESSOS NA REFERÊNCIA ANP DE 20 C.",
                f"ANP / IDEALIZADO - usando T informada ({temperatura_c:.3f} C): {anp_t:.6f} m3",
                f"ANP / IDEALIZADO - cenário recalculado a 20 C: {anp:.6f} m3",
                f"MODELO FISICO Z - usando T informada ({temperatura_c:.3f} C): {vol_t:.6f} m3",
                f"MODELO FISICO Z - cenário recalculado a 20 C: {vol_20:.6f} m3",
                f"Fator Z calculado - T informada ({temperatura_c:.3f} C): {calcular_Z(pressao_abs, temperatura_k)['Z']:.6f}",
                f"Fator Z calculado - 20 C: {calcular_Z(pressao_abs, 293.15)['Z']:.6f}",
                "",
                "LEITURA PARA LEIGOS",
                f"O cilindro tem {volume_fisico_m3:.6f} m3 de espaco fisico.",
                f"Com T informada de {temperatura_c:.3f} C, ANP/idealizado = {anp_t:.3f} m3 na referencia de 20 C.",
                f"No modelo Fisico Z, o resultado e {vol_t:.3f} m3 na referencia de 20 C.",
                f"Se o cilindro fosse considerado a 20 C, os resultados seriam ANP {anp:.3f} m3 e Fisico Z {vol_20:.3f} m3.",
                "Acima de 20 C, para o mesmo volume fisico e a mesma pressao, a quantidade de gas e menor; abaixo de 20 C, e maior.",
                "Os m3 sao volumes equivalentes/padronizados e nao o espaco fisico do cilindro.",
                "",
                "--------------------------------------------------",
                "4. EFEITO DA TEMPERATURA -50 C A +100 C",
                "--------------------------------------------------",
                "TODAS AS LINHAS ABAIXO ESTAO EXPRESSAS NA REFERENCIA ANP DE 20 C.",
                *tabela_temperatura_whatsapp,
                "",
                "--------------------------------------------------",
                "5. RESULTADOS FISICOS",
                "--------------------------------------------------",
                f"Massa estimada de GNV: {massa:.9f} kg",
                f"Densidade calculada no cilindro: {densidade:.6f} kg/m3",
                f"Quantidade de materia: {mols:.9f} mol",
                f"Volume especifico: {volume_especifico:.9f} m3/kg",
                f"Volume real calculado: {volume_real:.9f} m3",
                f"Densidade informada: {densidade_ref:.6f} kg/m3",
            ]

            massa_ref = r.get("massa_referencia_informada_kg")
            dif_massa = r.get("diferenca_massa_referencia_kg")
            if massa_ref is not None:
                linhas.append(f"Massa na referencia informada: {float(massa_ref):.9f} kg")
            if dif_massa is not None:
                linhas.append(f"Diferenca de massa na referencia: {float(dif_massa):.9f} kg")

            linhas += [
                "",
                "--------------------------------------------------",
                "5. VOLUME DO SISTEMA DE GNV (EXCLUINDO O CILINDRO)",
                "--------------------------------------------------",
                f"MINIMO - tubulacao: {float(minimo.get('volume_tubulacao_ml', 0)):.1f} mL",
                f"MINIMO - valvulas: {float(minimo.get('volume_valvulas_ml', 0)):.1f} mL",
                f"MINIMO - redutor: {float(minimo.get('volume_redutor_ml', 0)):.1f} mL",
                f"MINIMO - baixa pressao/filtro/flauta: {float(minimo.get('volume_baixa_ml', 0)):.1f} mL",
                f"MINIMO TOTAL: {float(minimo.get('volume_total_ml', 0)):.1f} mL",
                "",
                f"MEDIO - tubulacao: {float(medio.get('volume_tubulacao_ml', 0)):.1f} mL",
                f"MEDIO - valvulas: {float(medio.get('volume_valvulas_ml', 0)):.1f} mL",
                f"MEDIO - redutor: {float(medio.get('volume_redutor_ml', 0)):.1f} mL",
                f"MEDIO - baixa pressao/filtro/flauta: {float(medio.get('volume_baixa_ml', 0)):.1f} mL",
                f"MEDIO TOTAL: {float(medio.get('volume_total_ml', 0)):.1f} mL",
                "",
                f"MAXIMO - tubulacao: {float(maximo.get('volume_tubulacao_ml', 0)):.1f} mL",
                f"MAXIMO - valvulas: {float(maximo.get('volume_valvulas_ml', 0)):.1f} mL",
                f"MAXIMO - redutor: {float(maximo.get('volume_redutor_ml', 0)):.1f} mL",
                f"MAXIMO - baixa pressao/filtro/flauta: {float(maximo.get('volume_baixa_ml', 0)):.1f} mL",
                f"MAXIMO TOTAL: {float(maximo.get('volume_total_ml', 0)):.1f} mL",
                "",
                "--------------------------------------------------",
                "6. COMO SE CHEGA AO RESULTADO",
                "--------------------------------------------------",
                "1) litros -> m3;",
                "2) pressao manometrica + atmosfera -> pressao absoluta;",
                "3) Celsius -> Kelvin;",
                "4) modelo fisico: PV = Z*n*R*T;",
                "5) mols -> volume equivalente;",
                "6) modelo ANP/idealizado calculado separadamente;",
                "7) os resultados sao apresentados lado a lado.",
                "",
                "--------------------------------------------------",
                "7. INTERPRETACAO",
                "--------------------------------------------------",
                "O numero em m3 equivalente NAO significa que esse volume",
                "ocupa fisicamente o interior do cilindro naquela condicao.",
                "Ele representa a mesma quantidade de gas expressa na",
                "condicao de referencia do respectivo modelo.",
                "",
                "Os valores sao analiticos/estimativos e nao substituem",
                "instrumentos metrologicos oficiais.",
                "",
                "FIM DO RELATORIO",
                "=================================================="
            ]
            return "\n".join(linhas)

        except Exception as exc:
            return f"Nao foi possivel montar a mensagem: {exc}"

    def _copiar_whatsapp(self, *_args):
        """Copia a análise pronta para o clipboard do Windows/desktop."""
        try:
            from kivy.core.clipboard import Clipboard
            texto = self._whatsapp_message()
            Clipboard.copy(texto)
            if hasattr(self, "calc_share_status"):
                self.calc_share_status.text = "Resultado copiado para a área de transferência."
        except Exception as exc:
            if hasattr(self, "calc_share_status"):
                self.calc_share_status.text = f"Erro ao copiar: {exc}"

    def _abrir_whatsapp(self, *_args):
        """
        Abre o WhatsApp Web com o texto já preenchido.

        O navegador/WhatsApp ainda exige que o usuário escolha o contato e
        confirme o envio. O programa não envia mensagens sem confirmação.
        """
        try:
            texto = self._whatsapp_message()
            url = "https://wa.me/?text=" + urllib.parse.quote(texto)
            webbrowser.open(url)
            if hasattr(self, "calc_share_status"):
                self.calc_share_status.text = "WhatsApp Web aberto com o texto preparado."
        except Exception as exc:
            if hasattr(self, "calc_share_status"):
                self.calc_share_status.text = f"Erro ao abrir WhatsApp: {exc}"

    def _gerar_jpg_whatsapp(self, *_args):
        """
        Gera um JPG vertical pronto para ser anexado manualmente no WhatsApp.

        O arquivo é salvo no diretório de dados do aplicativo:
        compartilhamento_gnv.jpg
        """
        try:
            from PIL import Image, ImageDraw, ImageFont

            texto = self._whatsapp_message()
            linhas = texto.splitlines()

            largura = 1400
            margem = 70
            largura_texto = largura - 2 * margem
            tamanho_fonte = 34
            espaco_linha = 50

            # Tenta usar uma fonte comum do Windows; caso não exista,
            # utiliza a fonte padrão do Pillow.
            fontes = [
                r"C:\Windows\Fonts\segoeui.ttf",
                r"C:\Windows\Fonts\arial.ttf",
            ]
            fonte = None
            fonte_negrito = None
            for caminho in fontes:
                if os.path.exists(caminho):
                    fonte = ImageFont.truetype(caminho, tamanho_fonte)
                    caminho_bold = caminho.replace(".ttf", "b.ttf")
                    if os.path.exists(caminho_bold):
                        fonte_negrito = ImageFont.truetype(
                            caminho_bold, tamanho_fonte
                        )
                    else:
                        fonte_negrito = fonte
                    break

            if fonte is None:
                fonte = ImageFont.load_default()
                fonte_negrito = fonte

            # Quebra linhas longas para o JPG.
            draw_dummy = ImageDraw.Draw(Image.new("RGB", (10, 10), "white"))
            linhas_finais = []
            for linha in linhas:
                if not linha:
                    linhas_finais.append("")
                    continue

                palavras = linha.split()
                atual = ""
                for palavra in palavras:
                    teste = (atual + " " + palavra).strip()
                    if draw_dummy.textlength(teste, font=fonte) <= largura_texto:
                        atual = teste
                    else:
                        if atual:
                            linhas_finais.append(atual)
                        atual = palavra
                if atual:
                    linhas_finais.append(atual)

            altura = max(
                500,
                margem * 2 + len(linhas_finais) * espaco_linha
            )

            imagem = Image.new("RGB", (largura, altura), "#FFFDF8")
            draw = ImageDraw.Draw(imagem)

            y = margem
            for linha in linhas_finais:
                negrito = (
                    "VOLUME DE GNV" in linha
                    or "ANP / idealizado" in linha
                    or "Modelo físico" in linha
                )
                draw.text(
                    (margem, y),
                    linha,
                    fill="#201510",
                    font=fonte_negrito if negrito else fonte
                )
                y += espaco_linha

            caminho = Path(self.base_dir) / "compartilhamento_gnv.jpg"
            imagem.crop((0, 0, largura, min(altura, y + margem))).save(
                caminho,
                "JPEG",
                quality=95,
                optimize=True
            )

            # No Windows, abre o Explorer apontando para o arquivo.
            if hasattr(os, "startfile"):
                os.startfile(str(caminho))
            else:
                webbrowser.open(caminho.as_uri())

            if hasattr(self, "calc_share_status"):
                self.calc_share_status.text = (
                    f"JPG criado: {caminho}"
                )
        except Exception as exc:
            if hasattr(self, "calc_share_status"):
                self.calc_share_status.text = (
                    f"Erro ao gerar JPG. Verifique se Pillow está instalado: {exc}"
                )

    def _clear_calc(self):
        for e in self.calc_entries.values(): e.text=""
        self.calc_result.set_text("")
        if hasattr(self, 'calc_cylinder_m3_card'):
            self.calc_cylinder_m3_card.text = 'VOLUME FÍSICO DO CILINDRO\nInforme a capacidade e a quantidade.'

    # ---------------- ABASTECIMENTOS ----------------
    def _build_abastecimentos(self):
        box=self._content_box(); self.ab_entries={}
        ultimo=0.0
        try:
            rows=self.banco.listar_abastecimentos(); ultimo=max([float(r[4] or 0) for r in rows] or [0.0])
        except Exception: pass
        fields=[("Data / hora:",datetime.now().strftime("%Y-%m-%d %H:%M:%S")),("Posto:",""),("Cidade:",""),("Odômetro:",str(ultimo)),("Volume GNV (m³):","0"),("Preço/m³:","0"),("Capacidade do cilindro (L):","26"),("Pressão inicial (bar):","0"),("Pressão final (bar):","220"),("Temperatura (°C):","20"),("Altitude (m):","50"),("Fator Z:","0.92"),("Massa molar (kg/mol):","0.01604"),("Densidade de referência (kg/m³):","0.76"),("Observações:","")]
        for k,v in fields:
            box.add_widget(self._label(k)); e=self._entry(v); self.ab_entries[k]=e; box.add_widget(e)
        box.add_widget(self._button("Salvar Abastecimento",lambda *_:self._save_refuel()))
        self.ab_status=self._label(""); box.add_widget(self.ab_status)
        box.add_widget(self._label("Comparação / cálculo:",bold=True)); self.ab_compare=ScrollText("",horizontal=True,size_hint_y=None,height=dp(300)); box.add_widget(self.ab_compare)
        self._set_screen_content(1,box)

    def _save_refuel(self):
        try:
            posto=self.ab_entries["Posto:"].text.strip(); cidade=normalizar_cidade(self.ab_entries["Cidade:"].text.strip())
            odo=self._mobile_float(self.ab_entries["Odômetro:"].text); vol=self._mobile_float(self.ab_entries["Volume GNV (m³):"].text); preco=self._mobile_float(self.ab_entries["Preço/m³:"].text)
            cap=self._mobile_float(self.ab_entries["Capacidade do cilindro (L):"].text); pi=self._mobile_float(self.ab_entries["Pressão inicial (bar):"].text); pf=self._mobile_float(self.ab_entries["Pressão final (bar):"].text); temp=self._mobile_float(self.ab_entries["Temperatura (°C):"].text); alt=self._mobile_float(self.ab_entries["Altitude (m):"].text); z=self._mobile_float(self.ab_entries["Fator Z:"].text); mm=self._mobile_float(self.ab_entries["Massa molar (kg/mol):"].text); dens=self._mobile_float(self.ab_entries["Densidade de referência (kg/m³):"].text)
            if not posto or not cidade or odo<=0 or vol<=0 or cap<=0 or pf<pi: raise ValueError("Informe posto, cidade, odômetro, volume e pressões válidas.")
            anp=calcular_volume_anp_referencia(cap,pi,pf,temp,alt)
            cient=calcular_volume_cientifico_gas_real(cap,pi,pf,temp,alt,z,mm)
            fisico=cient["volume_referencia_m3"]
            status,d_anp,d_fis,p_anp,p_fis=self._classificar_bomba(vol,anp,fisico)
            ab=Abastecimento(self.ab_entries["Data / hora:"].text.strip(),posto,cidade,odo,vol,preco,temp,pf,alt,self.ab_entries["Observações:"].text.strip(),cap,pi,pf,dens,fisico,anp,fisico,status,d_anp,d_fis,p_anp,p_fis)
            self.banco.salvar_abastecimento(ab)
            linhas=[
                f"STATUS AUTOMÁTICO: {status}",
                f"{self._t('Volume informado pela bomba')}: {vol:.3f} m³",
                f"{self._t('Volume real calculado — ANP/idealizado')}: {anp:.3f} m³",
                f"{self._t('Diferença bomba − ANP')}: {d_anp:+.3f} m³",
                f"{self._t('Diferença percentual bomba × ANP')}: {p_anp:.2f}%",
                f"{self._t('Volume real calculado — modelo físico')} (Z={z:.3f}): {fisico:.3f} m³",
                f"{self._t('Diferença bomba − físico')}: {d_fis:+.3f} m³",
                f"{self._t('Diferença percentual bomba × físico')}: {p_fis:.2f}%",
                "",
                "INTERPRETAÇÃO: o sistema compara a leitura da bomba com dois modelos independentes.",
                "A classificação é um alerta comparativo e não constitui, isoladamente, prova metrológica de fraude."
            ]
            # Atualiza histórico e reconstrói a tela sem perder o resultado da comparação.
            self._refresh_all()
            self._build_abastecimentos()
            self.ab_status.text=f"Abastecimento salvo. Classificação automática: {status}"
            self.ab_compare.set_text("\n".join(linhas))
        except Exception as exc: self.ab_status.text=f"Erro: {exc}"


    def _classificar_bomba(self, volume_bomba, volume_anp=0.0, volume_fisico=0.0, tolerancia_percentual=5.0):
        """Classifica automaticamente o volume da bomba contra ANP e modelo físico."""
        vb=float(volume_bomba); anp=float(volume_anp); fis=float(volume_fisico)
        if vb<=0 or (anp<=0 and fis<=0):
            return "NAO_CLASSIFICADO", 0.0, 0.0, 0.0, 0.0
        d_anp=vb-anp if anp>0 else 0.0; d_fis=vb-fis if fis>0 else 0.0
        p_anp=abs(d_anp)/anp*100.0 if anp>0 else 0.0
        p_fis=abs(d_fis)/fis*100.0 if fis>0 else 0.0
        maior=max(p_anp,p_fis)
        status="CORRETO" if maior<=tolerancia_percentual else "ADULTERADO_DIVERGENTE"
        return status,d_anp,d_fis,p_anp,p_fis

    def _classicar_bomba(self, *args, **kwargs):
        """Alias de compatibilidade para chamadas antigas com o nome incorreto."""
        return self._classificar_bomba(*args, **kwargs)

    def _mobile_float(self,x): return converter_numero(x)

    # ---------------- ANP ----------------
    def _build_anp(self):
        box=self._content_box(); self.anp_entries={}
        fields=[("Capacidade física (L):","26"),("Pressão inicial manométrica (bar):","0"),("Pressão final manométrica (bar):","220"),("Temperatura ambiente (°C):","20"),("Altitude (m):","50")]
        for k,v in fields: box.add_widget(self._label(k)); self.anp_entries[k]=self._entry(v); box.add_widget(self.anp_entries[k])
        box.add_widget(self._button("Calcular condição ANP",lambda *_:self._calc_anp()))
        self.anp_result=ScrollText("", horizontal=False, size_hint_y=None, height=dp(260)); box.add_widget(self.anp_result); self._set_screen_content(2,box)
    def _calc_anp(self):
        try:
            vals=[self._mobile_float(self.anp_entries[k].text) for k in self.anp_entries]
            cap,pi,pf,temp,alt=vals
            v=calcular_volume_anp_referencia(*vals)
            patm=calcular_pressao_atmosferica(alt); p1=pi+patm; p2=pf+patm; delta=max(0.0,p2-p1)
            txt=("CONDIÇÃO DE REFERÊNCIA ANP\n"+"="*55+"\n"
                 f"Capacidade física: {cap:.3f} L\n"
                 f"Pressão inicial manométrica: {pi:.4f} bar\n"
                 f"Pressão final manométrica: {pf:.4f} bar\n"
                 f"Pressão atmosférica estimada: {patm:.5f} bar\n"
                 f"Pressão inicial absoluta: {p1:.5f} bar\n"
                 f"Pressão final absoluta: {p2:.5f} bar\n"
                 f"Temperatura ambiente: {temp:.2f} °C\n"
                 f"Altitude: {alt:.2f} m\n\n"
                 "Referência: 20 °C / 1,01325 bar\n"
                 f"ΔP absoluto: {delta:.5f} bar\n\n"
                 "Vref = Vcil × ΔP / Pref × Tref / T\n\n"
                 f"Volume adicionado equivalente ANP/idealizado: {v:.5f} m³\n"
                 f"Volume equivalente: {v*1000:.2f} L\n\n"
                 "Este cálculo é uma estimativa física/idealizada (Z=1) e não reproduz necessariamente o algoritmo interno do dispenser.")
            self.anp_result.set_text(txt)
        except Exception as e:self.anp_result.set_text(f"Erro: {e}")


    # ---------------- COMPRESSAO ----------------
    def _build_compressao(self):
        box=self._content_box(); self.comp_entries={}
        fields=[("Pressão inicial (bar):","1"),("Pressão final (bar):","220"),("Temperatura inicial (°C):","20"),("Altitude (m):","50"),("k:","1.294")]
        for k,v in fields: box.add_widget(self._label(k)); self.comp_entries[k]=self._entry(v); box.add_widget(self.comp_entries[k])
        box.add_widget(self._button("Simular aquecimento/compressão",lambda *_:self._calc_comp()))
        self.comp_result=ScrollText("",horizontal=False,size_hint_y=None,height=dp(240)); box.add_widget(self.comp_result)
        box.add_widget(self._label("Pressão × Temperatura",bold=True)); self.comp_chart_pt=GNVLineChart(size_hint_y=None,height=dp(360)); box.add_widget(self.comp_chart_pt)
        box.add_widget(self._label("Temperatura × Pressão",bold=True)); self.comp_chart_tp=GNVLineChart(size_hint_y=None,height=dp(360)); box.add_widget(self.comp_chart_tp)
        box.add_widget(self._label("Pressão × Volume Relativo",bold=True)); self.comp_chart_pv=GNVLineChart(size_hint_y=None,height=dp(360)); box.add_widget(self.comp_chart_pv)
        self.comp_chart_result=ScrollText("",horizontal=False,size_hint_y=None,height=dp(220)); box.add_widget(self.comp_chart_result)
        self._set_screen_content(3,box)

    def _calc_comp(self):
        try:
            vals=[self._mobile_float(self.comp_entries[k].text) for k in self.comp_entries]
            r=calcular_compressao_ideal_adiabatica(*vals); pontos=gerar_pontos_compressao_adiabatica(*vals)
            self.comp_result.set_text("\n".join(f"{k}: {v:.6f}" if isinstance(v,(int,float)) else f"{k}: {v}" for k,v in r.items()))
            press=[p["pressao_man_bar"] for p in pontos]; temp=[p["temperatura_c"] for p in pontos]; vol=[p["volume_relativo"] for p in pontos]
            labels=[str(i+1) for i in range(len(pontos))]
            self.comp_chart_pt.set_data([press],labels,"Pressão × Temperatura",["Pressão (bar)"],x_values=temp)
            self.comp_chart_tp.set_data([temp],labels,"Temperatura × Pressão",["Temperatura (°C)"],x_values=press)
            self.comp_chart_pv.set_data([vol],labels,"Pressão × Volume Relativo",["Volume relativo V/V₀"],x_values=press)
            self.comp_chart_result.set_text("Ponto | Pressão (bar) | Temperatura (°C) | Volume relativo\n"+"\n".join(f"{i+1:02d} | {p['pressao_man_bar']:.2f} | {p['temperatura_c']:.2f} | {p['volume_relativo']:.5f}" for i,p in enumerate(pontos)))
        except Exception as e:self.comp_result.set_text(f"Erro: {e}")


    # ---------------- HISTORICO ----------------
    def _build_historico(self):
        box=self._content_box(); box.add_widget(self._button("Atualizar histórico",lambda *_:self._refresh_history())); self.hist_result=ScrollText("", horizontal=False, size_hint_y=None, height=dp(320)); box.add_widget(self.hist_result); self._set_screen_content(4,box)
        self._refresh_history()
    def _refresh_history(self):
        try:
            rows=self.banco.listar_abastecimentos();
            lines=["STATUS | DATA | POSTO | CIDADE | ODOMETRO | VOLUME | PREÇO | TOTAL | ANP | FÍSICO"]
            status_map={"CORRETO":"🟩 CORRETO","ADULTERADO_DIVERGENTE":"🟥 DIVERGENTE","NAO_CLASSIFICADO":"🟨 NÃO CLASSIFICADO"}
            lines += [f"{status_map.get(str(r[19] or 'NAO_CLASSIFICADO'), '🟨 NÃO CLASSIFICADO')} | {r[0]} | {r[2]} | {r[3]} | {r[4]} | {r[5]} | {r[6]} | {r[7]} | {r[17]} | {r[18]}" for r in rows]
            self.hist_result.set_text("\n".join(lines))
        except Exception as e:self.hist_result.set_text(str(e))

    # ---------------- SQLITE ----------------
    def _build_sqlite(self):
        box=self._content_box(); box.add_widget(self._button("Atualizar Banco SQLite",lambda *_:self._refresh_sqlite())); self.sqlite_result=ScrollText("", horizontal=False, size_hint_y=None, height=dp(320)); box.add_widget(self.sqlite_result); self._set_screen_content(5,box); self._refresh_sqlite()
    def _refresh_sqlite(self):
        try:
            rows=self.banco.listar_abastecimentos(); lines=["ID | STATUS | DATA | POSTO | CIDADE | ODOMETRO | VOLUME | PREÇO | TOTAL | ANP | FÍSICO"]
            status_map={"CORRETO":"🟩 CORRETO","ADULTERADO_DIVERGENTE":"🟥 DIVERGENTE","NAO_CLASSIFICADO":"🟨 NÃO CLASSIFICADO"}
            lines += [f"{r[0]} | {status_map.get(str(r[19] or 'NAO_CLASSIFICADO'),'🟨 NÃO CLASSIFICADO')} | {r[1]} | {r[2]} | {r[3]} | {r[4]} | {r[5]} | {r[6]} | {r[7]} | {r[17]} | {r[18]}" for r in rows]
            self.sqlite_result.set_text("\n".join(lines))
        except Exception as e:self.sqlite_result.set_text(str(e))

    # ---------------- EXPORTACAO ----------------
    def _build_excel(self):
        box=self._content_box(); box.add_widget(self._label("Exportação e integração com Excel / arquivos de dados",bold=True)); self.export_status=self._label(""); box.add_widget(self._button("Exportar CSV",lambda *_:self._export_csv())); box.add_widget(self._button("Gerar PDF detalhado",lambda *_:self._export_pdf())); box.add_widget(self.export_status); self._set_screen_content(6,box)
    def _export_csv(self):
        try:
            rows=self.banco.listar_abastecimentos()
            csv_path=self.base_dir/"historico_abastecimentos.csv"
            with open(csv_path,"w",encoding="utf-8-sig") as f:
                f.write("id,data,posto,cidade,odometro,volume_m3,preco_m3,valor_total\n")
                for r in rows: f.write(",".join(str(x).replace(",",".") for x in r[:8])+"\n")
            xlsx_path=self.base_dir/"historico_abastecimentos.xlsx"
            wb=Workbook(); ws=wb.active; ws.title="Abastecimentos"
            ws.append(["ID","Data","Posto","Cidade","Odômetro","Volume (m³)","Preço/m³","Valor total"])
            for r in rows: ws.append(list(r[:8]))
            wb.save(xlsx_path)
            self.export_status.text=f"CSV: {csv_path}\nExcel: {xlsx_path}"
        except Exception as e:self.export_status.text=f"Erro: {e}"
    def _export_pdf(self):
        try:
            path=self.base_dir/"relatorio_gnv_android.pdf"
            pdf=FPDF(); pdf.add_page(); pdf.set_font("Arial",size=10)
            pdf.multi_cell(0,6,"RELATÓRIO TÉCNICO DE GNV")
            pdf.multi_cell(0,6,APP_TITLE)
            pdf.multi_cell(0,6,self.calc_result.label.text if hasattr(self.calc_result,'label') else "Execute um cálculo para obter os dados.")
            pdf.output(str(path)); self.export_status.text=f"PDF salvo em: {path}"
        except Exception as e:self.export_status.text=f"Erro PDF: {e}"

    # ---------------- GRAFICOS ----------------
    def _build_graficos(self):
        box=self._content_box(); box.add_widget(self._label("Gráficos de Abastecimento",bold=True))
        self.chart_spinner=Spinner(text=self._t("Bomba × Teórico"), values=(self._t("Bomba × Teórico"),self._t("Volume por abastecimento"),self._t("Consumo km/m³")), size_hint_y=None,height=dp(44))
        self.chart_spinner.bind(text=lambda *_:self._refresh_chart())
        box.add_widget(self.chart_spinner)
        box.add_widget(self._button("Atualizar Gráfico",lambda *_:self._refresh_chart()))
        self.chart=GNVLineChart(size_hint_y=None,height=dp(300)); box.add_widget(self.chart)
        self.chart_result=ScrollText("",horizontal=False,size_hint_y=None,height=dp(220)); box.add_widget(self.chart_result)
        self._set_screen_content(7,box); self._refresh_chart()

    def _refresh_chart(self):
        try:
            rows=self.banco.listar_abastecimentos(); mode=self.chart_spinner.text if hasattr(self,'chart_spinner') else self._t("Bomba × Teórico")
            labels=[str(i) for i,_ in enumerate(rows,1)]
            if self._t("Bomba × Teórico") == mode:
                pump=[float(r[5] or 0) for r in rows]; theo=[float(r[16] or 0) for r in rows]
                self.chart.set_data([pump,theo],labels,self._t("Bomba × Teórico (m³)"),[self._t("Bomba"),self._t("Teórico")])
                self.chart_result.set_text(self._t("Abastecimento | Bomba (m³) | Teórico (m³)")+"\n"+"\n".join(f"{i+1:03d} | {pump[i]:.3f} | {theo[i]:.3f}" for i in range(len(rows))))
            elif self._t("Volume por abastecimento") == mode:
                vals=[float(r[5] or 0) for r in rows]; self.chart.set_data([vals],labels,self._t("Volume por abastecimento (m³)"),[self._t("Bomba")]); self.chart_result.set_text("\n".join(f"{i+1:03d}: {v:.3f} m³" for i,v in enumerate(vals)))
            else:
                cons=[]; c_labels=[]; prev=None
                for r in rows:
                    try:
                        odo=float(r[4]); vol=float(r[5]);
                        if prev is not None and odo>prev and vol>0: cons.append((odo-prev)/vol); c_labels.append(str(len(cons)))
                        prev=odo
                    except: pass
                self.chart.set_data([cons],c_labels,self._t("Consumo: km/m³"),[self._t("Consumo")]); self.chart_result.set_text(self._t("Consumo = distância entre odômetros válidos ÷ volume abastecido pela bomba.")+"\n\n"+"\n".join(f"{i+1:03d}: {v:.2f} km/m³" for i,v in enumerate(cons)))
        except Exception as e: self.chart_result.set_text(self._t("Erro")+f": {e}")

    def _refresh_chart_text(self): self._refresh_chart()

    # ---------------- CONFIG ----------------
    def _build_config(self):
        """
        Monta a aba Configurações.

        A mudança mais importante desta versão é a remoção do antigo
        Spinner de nomes de cores. Agora cada cor é um botão visual: ao
        clicar, abre o seletor arco-íris com intensidade e luminosidade.
        """
        box = self._content_box()

        box.add_widget(self._label('Configurações do Sistema', bold=True, h=42))
        box.add_widget(self._label(
            'Tema, alinhamento, degradê e cores. Clique em uma amostra para abrir o arco-íris.',
            h=52
        ))
        # Atalho explícito: mesmo que o Spinner de abas fique fora da tela,
        # o usuário consegue abrir os Créditos diretamente daqui.
        box.add_widget(self._button('ABRIR CRÉDITOS DO PROJETO', lambda *_: self._open_credits_tab(), 42))

        # V28.13: mostra exatamente onde ficam os arquivos usados pelo aplicativo.
        box.add_widget(self._label(f'Banco SQLite: {self.db_path}', h=54))
        box.add_widget(self._label(f'Configurações: {self.config_path}', h=54))

        # ===============================================================
        # TEMA
        # ===============================================================
        box.add_widget(self._label('Tema:', h=34))
        self.theme_spinner = Spinner(
            text=self.colors.get('theme', 'claro'),
            values=('claro', 'escuro'),
            size_hint_y=None, height=dp(44),
            option_cls=ConfigSpinnerOption
        )
        self.theme_spinner.bind(text=self._theme_changed)
        box.add_widget(self.theme_spinner)

        # ===============================================================
        # ALINHAMENTO DOS RESULTADOS
        # ===============================================================
        box.add_widget(self._label('Alinhamento das respostas dos cálculos:', h=38))
        alinh = {'left': 'Esquerda', 'center': 'Centro', 'right': 'Direita'}
        self.alignment_spinner = Spinner(
            text=alinh.get(self.colors.get('alignment', 'center'), 'Centro'),
            values=('Esquerda', 'Centro', 'Direita'),
            size_hint_y=None, height=dp(44),
            option_cls=ConfigSpinnerOption
        )
        self.alignment_spinner.bind(text=self._alignment_changed)
        box.add_widget(self.alignment_spinner)

        # ===============================================================
        # DEGRADÊ DOS CAMPOS
        # ===============================================================
        box.add_widget(self._label('Degradê nos campos de preenchimento:', h=38))
        self.gradient_spinner = Spinner(
            text='Ativado' if self.colors.get('gradient_enabled', True) else 'Desativado',
            values=('Ativado', 'Desativado'),
            size_hint_y=None, height=dp(44),
            option_cls=ConfigSpinnerOption
        )
        self.gradient_spinner.bind(text=self._gradient_changed)
        box.add_widget(self.gradient_spinner)

        # ===============================================================
        # PALETA DE REFERÊNCIA
        # ===============================================================
        self.color_controls = {}

        color_fields = [
            ('Cor de fundo', 'bg'),
            ('Cor das letras gerais', 'text'),
            ('Cor dos botões', 'button'),
            ('Cor base dos campos', 'field'),
            ('Início do degradê dos campos', 'field_start'),
            ('Fim do degradê dos campos', 'field_end'),
            ('Fundo geral dos resultados', 'result'),
            # V28.12: duas caixas independentes pedidas pelo usuário.
            # Uma controla os números/textos digitados nos TextInput; a outra
            # controla os textos produzidos pelo sistema.
            ('Cor das letras dos campos de entrada', 'input_text'),
            ('Cor das letras dos resultados do sistema', 'result_text'),
            ('Fundo da linha 1', 'row1'),
            ('Fundo da linha 2', 'row2')
        ]

        # Nomes conhecidos apenas para exibição amigável. O usuário não fica
        # limitado a eles: o ColorPickerPopup permite qualquer #RRGGBB.
        self.palette_map = {
            'Preto': '#000000', 'Branco': '#FFFFFF',
            'Cinza muito claro': '#F7F7F7', 'Cinza claro': '#E0E0E0',
            'Cinza médio': '#9E9E9E', 'Cinza escuro': '#616161',
            'Cinza muito escuro': '#212121',
            'Vermelho muito claro': '#FFEBEE', 'Vermelho claro': '#FFCDD2',
            'Vermelho médio': '#EF9A9A', 'Vermelho escuro': '#E57373',
            'Vermelho muito escuro': '#C62828',
            'Laranja muito claro': '#FFF3E0', 'Laranja claro': '#FFE0B2',
            'Laranja médio': '#FFCC80', 'Laranja escuro': '#FF9800',
            'Laranja muito escuro': '#E65100',
            'Amarelo muito claro': '#FFFDE7', 'Amarelo claro': '#FFF9C4',
            'Amarelo médio': '#FFF59D', 'Amarelo escuro': '#FDD835',
            'Amarelo muito escuro': '#F9A825',
            'Verde muito claro': '#E8F5E9', 'Verde claro': '#C8E6C9',
            'Verde médio': '#A5D6A7', 'Verde escuro': '#43A047',
            'Verde muito escuro': '#1B5E20',
            'Turquesa muito claro': '#E0F2F1', 'Turquesa claro': '#B2DFDB',
            'Turquesa médio': '#80CBC4', 'Turquesa escuro': '#00897B',
            'Turquesa muito escuro': '#004D40',
            'Ciano muito claro': '#E0F7FA', 'Ciano claro': '#B2EBF2',
            'Ciano médio': '#80DEEA', 'Ciano escuro': '#00ACC1',
            'Ciano muito escuro': '#006064',
            'Azul muito claro': '#E3F2FD', 'Azul claro': '#BBDEFB',
            'Azul médio': '#64B5F6', 'Azul escuro': '#1976D2',
            'Azul muito escuro': '#0D47A1',
            'Índigo muito claro': '#E8EAF6', 'Índigo claro': '#C5CAE9',
            'Índigo médio': '#7986CB', 'Índigo escuro': '#3949AB',
            'Índigo muito escuro': '#1A237E',
            'Roxo muito claro': '#F3E5F5', 'Roxo claro': '#E1BEE7',
            'Roxo médio': '#BA68C8', 'Roxo escuro': '#8E24AA',
            'Roxo muito escuro': '#4A148C',
            'Magenta muito claro': '#FCE4EC', 'Magenta claro': '#F8BBD0',
            'Magenta médio': '#F06292', 'Magenta escuro': '#D81B60',
            'Magenta muito escuro': '#880E4F',
            'Marrom muito claro': '#EFEBE9', 'Marrom claro': '#D7B899',
            'Marrom médio': '#A67B5B', 'Marrom escuro': '#5D4037',
            'Marrom muito escuro': '#3E2723'
        }

        box.add_widget(self._label(
            'Paleta arco-íris: clique na amostra e escolha matiz, intensidade e luminosidade.',
            h=52
        ))

        for label, key in color_fields:
            box.add_widget(self._label(label, h=32))

            row = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(6))
            hx = self.colors.get(key, '#FFFFFF')
            name = self._color_name(hx, self.palette_map)

            control = Button(
                text=f'{name}\n{hx.upper()}',
                size_hint_x=1,
                font_size=sp(11),
                bold=True,
                background_normal='',
                background_down=''
            )
            control.background_color = self._hex_rgba(hx)
            control.color = self._contrast_color(hx)
            control.bind(on_release=lambda _btn, kk=key, title=label:
                         self._open_color_picker(kk, title))

            self.color_controls[key] = control
            row.add_widget(control)
            box.add_widget(row)

        box.add_widget(self._button('Salvar Configurações', lambda *_: self._save_config(), 46))
        box.add_widget(self._button('Restaurar Padrão', lambda *_: self._restore_config(), 46))
        self.config_status = self._label('', h=44)
        box.add_widget(self.config_status)

        self._set_screen_content(8, box)

    def _update_config_diagnostic(self):
        """Exibe as propriedades visuais atualmente armazenadas e aplicadas."""
        if not hasattr(self, 'config_diagnostic'):
            return
        c = self.colors
        self.config_diagnostic.text = (
            f"ATIVOS — entrada: {c.get('input_text')} | resultados: {c.get('result_text')}\n"
            f"Linha 1: {c.get('row1')} | Linha 2: {c.get('row2')} | alinhamento: {c.get('alignment')}\n"
            f"Degradê: {('ATIVADO' if c.get('gradient_enabled', True) else 'DESATIVADO')} "
            f"| início: {c.get('field_start')} | fim: {c.get('field_end')}\n"
            f"Botões: {c.get('button')} | tema: {c.get('theme')} | personalizadas: {self._colors_personalizadas}"
        )

    def _open_color_picker(self, key, title):
        """Abre o seletor arco-íris para uma propriedade da interface."""
        atual = self.colors.get(key, '#FFFFFF')

        def apply_color(hex_value):
            # O valor é guardado em hexadecimal porque é simples de estudar,
            # editar manualmente e transportar entre Windows e Android.
            self.colors[key] = hex_value.upper()
            self._colors_personalizadas = True
            # Atualiza imediatamente o botão visual correspondente antes de
            # redesenhar os demais componentes.
            control = getattr(self, 'color_controls', {}).get(key)
            if control is not None:
                control.background_normal = ''
                control.background_down = ''
                control.background_color = self._hex_rgba(self.colors[key])
                control.color = self._contrast_color(self.colors[key])
                control.text = f'{self._color_name(self.colors[key], self.palette_map)}\n{self.colors[key]}'
            self._apply_colors()
            Clock.schedule_once(lambda _dt: self._apply_colors(), 0)

        popup = ColorPickerPopup(title, atual, apply_color)
        popup.open()

    def _color_name(self, value, palette):
        """Retorna um nome conhecido ou 'Cor personalizada'."""
        for name, hx in palette.items():
            if str(value).lower() == hx.lower():
                return name
        return 'Cor personalizada'

    def _alignment_changed(self, spinner, *_args):
        """Salva imediatamente a preferência de alinhamento dos resultados."""
        self.colors['alignment'] = {
            'Esquerda': 'left', 'Centro': 'center', 'Direita': 'right'
        }.get(spinner.text, 'center')
        self._apply_all_visual_now()
        Clock.schedule_once(lambda _dt: self._apply_all_visual_now(), 0)

    def _gradient_changed(self, spinner, *_args):
        """Liga/desliga o degradê dos campos sem alterar os cálculos."""
        self.colors['gradient_enabled'] = spinner.text == 'Ativado'
        self._apply_all_visual_now()
        Clock.schedule_once(lambda _dt: self._apply_colors(), 0)

    def _load_config(self):
        """
        V28.13 - Carrega configuração com controle explícito de personalização.

        cores_personalizadas=False -> usa a paleta oficial do tema.
        cores_personalizadas=True  -> usa as cores que o usuário salvou.

        A primeira leitura de arquivos antigos migra para a paleta padrão e
        evita que cores de versões anteriores contaminem o tema atual.
        """
        try:
            if self.config_path.exists():
                cfg=json.loads(self.config_path.read_text(encoding='utf-8'))
                self.idioma=cfg.get('idioma','pt-BR')
                old_colors=cfg.get('cores',{})
                saved_theme=str(cfg.get('tema',old_colors.get('theme','claro'))).lower()
                self._colors_personalizadas=bool(cfg.get('cores_personalizadas',False))
                self.colors=self._theme_defaults(saved_theme)
                if self._colors_personalizadas:
                    self.colors.update(old_colors)
        except Exception:
            self.idioma='pt-BR'
            self._colors_personalizadas=False
            self.colors=self._theme_defaults('claro')

        self.colors.setdefault('alignment','center')
        self.colors.setdefault('gradient_enabled',True)
        self.colors.setdefault('field_start',self.colors.get('field','#FFFFFF'))
        self.colors.setdefault('field_end',self.colors.get('field','#FFFFFF'))
        theme=str(self.colors.get('theme','claro')).lower()
        self.colors.setdefault('input_text','#FFFDF8' if theme=='escuro' else '#241713')
        self.colors.setdefault('result_text','#FFF4EC' if theme=='escuro' else '#2E211B')

    def _persist_config(self):
        """
        Grava configuracoes.json de forma atômica. O .tmp é substituído pelo
        arquivo final somente depois de toda a escrita ser concluída.
        """
        payload={
            'versao_visual':6,
            'idioma':self.idioma,
            'tema':self.colors.get('theme','claro'),
            'cores_personalizadas':bool(getattr(self,'_colors_personalizadas',False)),
            'cores':self.colors
        }
        tmp=self.config_path.with_suffix('.json.tmp')
        tmp.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding='utf-8')
        tmp.replace(self.config_path)

    def _save_config(self):
        """Salva todas as configurações sem sobrescrever cores escolhidas."""
        try:
            self.colors['theme']=self.theme_spinner.text
            self.colors['alignment']={'Esquerda':'left','Centro':'center','Direita':'right'}.get(self.alignment_spinner.text,'center')
            self.colors['gradient_enabled']=self.gradient_spinner.text=='Ativado'
            self._persist_config()
            # IMPORTANTE: não chamamos _refresh_all() aqui.
            # Recriar as telas destruía a aparência recém-configurada.
            # A atualização é direta nas instâncias existentes de TODAS as abas.
            self._apply_all_visual_now()
            self._update_config_controls_from_colors()
            self._save_config_status()
        except Exception as exc:
            if hasattr(self,'config_status'):
                self.config_status.text=f'Erro ao salvar configurações: {exc}'

    def _save_config_status(self):
        if hasattr(self,'config_status'):
            self.config_status.text=self._t('Configurações salvas com sucesso.')

    def _restore_config(self):
        """Restaura o tema claro oficial sem reconstruir nenhuma aba."""
        self.colors=self._theme_defaults('claro')
        self.idioma='pt-BR'
        self._colors_personalizadas=False
        if hasattr(self,'theme_spinner'): self.theme_spinner.text='claro'
        if hasattr(self,'alignment_spinner'): self.alignment_spinner.text='Centro'
        if hasattr(self,'gradient_spinner'): self.gradient_spinner.text='Ativado'
        # Atualiza controles de paleta imediatamente.
        self._update_config_controls_from_colors()
        self._apply_all_visual_now()
        if hasattr(self,'config_status'):
            self.config_status.text='Configurações restauradas. Aplicadas em todas as abas.'
        try:
            self._persist_config()
        except Exception as exc:
            if hasattr(self,'config_status'): self.config_status.text=f'Erro ao salvar padrão: {exc}'

    def _theme_defaults(self, theme):
        """Retorna uma paleta completa e coerente para o tema selecionado."""
        if str(theme).lower()=='escuro':
            return {'theme':'escuro','bg':'#17110F','text':'#FFF4EC','input_text':'#FFFDF8','result_text':'#FFF4EC','button':'#6D4C41','field':'#39271F','field_start':'#5A4034','field_end':'#241713','gradient_enabled':True,'result':'#493229','row1':'#5A4035','row2':'#32221D','alignment':'center'}
        return {'theme':'claro','bg':'#F5EFEA','text':'#2E211B','input_text':'#241713','result_text':'#2E211B','button':'#B88962','field':'#FFF8F0','field_start':'#FFFFFF','field_end':'#E6CDB8','gradient_enabled':True,'result':'#F0E2D7','row1':'#F7ECE3','row2':'#E1C8B5','alignment':'center'}

    def _theme_changed(self, spinner, *_args):
        """Troca de tema em tempo real, sem reconstruir as telas."""
        self.colors=self._theme_defaults(spinner.text)
        self._colors_personalizadas=False
        if hasattr(self,'alignment_spinner'): self.alignment_spinner.text='Centro'
        if hasattr(self,'gradient_spinner'): self.gradient_spinner.text='Ativado'
        self._refresh_color_controls()
        self._apply_all_visual_now()
        try: self._persist_config()
        except Exception: pass
        if hasattr(self,'config_status'):
            self.config_status.text=f'Tema {spinner.text} aplicado em todas as abas.'

    def _apply_theme(self):
        # Apenas aplica as cores atuais. Não sobrescreve as cores personalizadas
        # quando o idioma é alterado ou quando o aplicativo é iniciado.
        self._apply_colors()

    # ---------------- FORMULAS ----------------
    def _open_credits_tab(self):
        """Abre a aba de Créditos diretamente a partir das Configurações."""
        try:
            self.sm.current = self.screen_names[11]
            self.tab_spinner.text = self.screen_names_for_language()[11]
        except Exception:
            pass

    def _build_formulas(self):
        scr=self.sm.get_screen(self.screen_names[9]); scr.clear_widgets(); self.formula_view=ScrollText(self.formula_pt, horizontal=False); scr.clear_widgets(); scr.add_widget(self.formula_view)
    # ---------------- TOTAL ----------------
    def _build_creditos(self):
        """Aba de créditos: história e propósito do projeto."""
        box=self._content_box(padding=14)
        box.add_widget(self._label('CRÉDITOS — POR QUE ESTE SISTEMA FOI CRIADO', h=48, bold=True))
        texto = """Este projeto nasceu de uma situação que deveria ser simples: abastecer um veículo com GNV, conferir o volume informado pela bomba e seguir viagem.

Mas, repetidas vezes, surgia a mesma situação: um cilindro com determinada capacidade física recebia na bomba uma quantidade em m³ acima do que parecia fisicamente possível. Quando questionado, o consumidor frequentemente ouve explicações como “é a pressão da bomba”, “o gás aqueceu”, “é a temperatura”, “é a tubulação”, “é o sistema do carro” ou outras justificativas difíceis de verificar no momento do abastecimento.

Depois de tantas paradas em postos, dúvidas, discussões e a sensação de não possuir ferramentas suficientes para conferir os números, surgiu a decisão de transformar a frustração em estudo. Este sistema foi desenvolvido para registrar abastecimentos, comparar volumes, estudar as condições de pressão e temperatura, calcular referências físicas e organizar os dados para que uma pergunta simples possa ser analisada com números.

A proposta não é substituir instrumentos metrológicos oficiais nem afirmar, sozinho, que um posto está errado. A proposta é dar ao usuário uma ferramenta de análise, histórico e comparação para que diferenças possam ser investigadas com mais organização e menos achismo.

O projeto também nasceu de muito trabalho de programação, testes, erros, reconstruções e aprendizado. Cada versão representa uma tentativa de tornar os cálculos, o banco de dados e a apresentação das informações mais úteis para o uso real.

AUTOR E DESENVOLVEDOR
Christiano T. Gaio
Analista de Sistemas e Pesquisador

OBJETIVO PROFISSIONAL
Apresentar o projeto como estudo técnico e software experimental, buscando colaboração, revisão técnica, parcerias e oportunidades com empresas interessadas em GNV, instrumentação, software, análise de dados e telemetria.

GITHUB
Repositório oficial do projeto: https://github.com/Gaio-Christiano/GNV_m-_carro
GitHub: https://github.com/Gaio-Christiano

Aviso: os resultados do programa são análises e estimativas computacionais. Para uma conclusão metrológica sobre um abastecimento específico devem ser utilizados os dados e instrumentos adequados."""
        txt=ScrollText(texto,size_hint_y=None,height=dp(680))
        box.add_widget(txt)
        box.add_widget(self._label('Créditos e documentação do projeto — V28.26',h=40))
        box.add_widget(self._button(
            'ABRIR PERFIL GITHUB — Gaio-Christiano',
            lambda *_: self._abrir_github_perfil(),
            44
        ))
        box.add_widget(self._button(
            'ABRIR REPOSITÓRIO OFICIAL — GNV_m-_carro',
            lambda *_: self._abrir_github_repositorio(),
            44
        ))
        self._set_screen_content(11,box)

    def _abrir_github_perfil(self, *_args):
        """Abre o perfil oficial do autor no navegador."""
        webbrowser.open("https://github.com/Gaio-Christiano")

    def _abrir_github_repositorio(self, *_args):
        """Abre o repositorio oficial do projeto no navegador."""
        webbrowser.open("https://github.com/Gaio-Christiano/GNV_m-_carro")

    def _build_total(self):
        box=self._content_box(); box.add_widget(self._label("Total de Abastecimentos",bold=True)); box.add_widget(self._label("Classificação automática: 🟩 correto | 🟥 divergente | 🟨 não classificado. O status é um alerta comparativo e não prova isolada de fraude.")); self.total_result=ScrollText("", horizontal=False, size_hint_y=None, height=dp(360)); box.add_widget(self._button("Atualizar total",lambda *_:self._refresh_total())); box.add_widget(self.total_result); self._set_screen_content(10,box); self._refresh_total()
    def _refresh_total(self):
        try:
            rows=self.banco.listar_abastecimentos()
            vols=[float(r[5] or 0) for r in rows]; vals=[float(r[7] or 0) for r in rows]
            unique=[]; seen=set()
            for r in rows:
                odo=float(r[4] or 0)
                key=(round(odo,3),round(float(r[5] or 0),6))
                if key not in seen: unique.append(r); seen.add(key)
            dist=0.0; cons_b=[]; cons_anp=[]; cons_fis=[]
            prev=None
            for r in sorted(unique,key=lambda x:(float(x[4] or 0),str(x[1]))):
                odo=float(r[4] or 0)
                if prev is not None and odo>prev:
                    d=odo-prev; dist+=d; pump=float(r[5] or 0); anp=float(r[17] or 0); fis=float(r[18] or 0)
                    if pump>0: cons_b.append(d/pump)
                    if anp>0: cons_anp.append(d/anp)
                    if fis>0: cons_fis.append(d/fis)
                prev=odo
            media_b=dist/sum([float(r[5] or 0) for r in unique[1:]]) if len(unique)>1 and sum(float(r[5] or 0) for r in unique[1:])>0 else 0
            status_counts={"CORRETO":0,"ADULTERADO_DIVERGENTE":0,"NAO_CLASSIFICADO":0}
            status_vol={k:0.0 for k in status_counts}
            for r in rows:
                st=str(r[19] or "NAO_CLASSIFICADO").upper() if len(r)>18 else "NAO_CLASSIFICADO"
                if st not in status_counts: st="NAO_CLASSIFICADO"
                status_counts[st]+=1; status_vol[st]+=float(r[5] or 0)
            lines=["RESUMO TOTAL DE ABASTECIMENTOS","="*60,f"Registros: {len(rows)}",f"Registros únicos considerados: {len(unique)}",f"Volume total da bomba: {sum(vols):.3f} m³",f"Valor total: R$ {sum(vals):.2f}",f"Distância total considerada: {dist:.1f} km",f"MÉDIA GLOBAL — BOMBA: {media_b:.2f} km/m³",f"MÉDIA — ANP: {(sum(cons_anp)/len(cons_anp) if cons_anp else 0):.2f} km/m³",f"MÉDIA — FÍSICA: {(sum(cons_fis)/len(cons_fis) if cons_fis else 0):.2f} km/m³","",f"CORRETOS: {status_counts['CORRETO']} registros | {status_vol['CORRETO']:.3f} m³",f"ADULTERADO / DIVERGENTE: {status_counts['ADULTERADO_DIVERGENTE']} registros | {status_vol['ADULTERADO_DIVERGENTE']:.3f} m³",f"NÃO CLASSIFICADOS: {status_counts['NAO_CLASSIFICADO']} registros | {status_vol['NAO_CLASSIFICADO']:.3f} m³"]
            self.total_result.set_text("\n".join(lines))
        except Exception as e:self.total_result.set_text(str(e))

    def change_language(self, spinner, idioma):
        self.idioma=idioma
        current=self.sm.current if hasattr(self,'sm') else 'screen_0'
        self._build_all_screens()
        self.sm.current=current
        self._apply_language()
        # A reconstrução criou novos widgets; reaplica no próximo ciclo.
        Clock.schedule_once(lambda _dt: self._apply_colors(), 0)
        try: self._save_config()
        except Exception: pass

    def _apply_language(self):
        names=self.screen_names_for_language(); self.tab_spinner.values=tuple(names); self.tab_spinner.text=names[self.screen_names.index(self.sm.current)] if self.sm.current in self.screen_names else names[0]
        title={"pt-BR": APP_TITLE, "English": f"CNG Cylinder Capacity Calculation and Analysis System - {APP_VERSION}", "Español": f"Sistema de Cálculos y Análisis de la Capacidad del Cilindro GNV - {APP_VERSION}", "Français": f"Système de Calcul et d’Analyse de la Capacité du Cylindre GNV - {APP_VERSION}", "Italiano": f"Sistema di Calcolo e Analisi della Capacità del Cilindro GNV - {APP_VERSION}", "Deutsch": f"Berechnungs- und Analysesystem für CNG-Zylinderkapazität - {APP_VERSION}", "日本語": f"CNGシリンダー容量計算・分析システム - {APP_VERSION}", "中文": f"CNG气瓶容量计算与分析系统 - {APP_VERSION}"}.get(self.idioma, APP_TITLE)
        self.header.text=title
        self.footer.text={"pt-BR":"Analista de Sistemas e Pesquisador - Christiano T.Gaio - Desenvolvedor | Projeto iniciado o Desenvolvimento em 06/2026","English":"Systems Analyst and Researcher - Christiano T.Gaio - Developer | Project Development Started in 06/2026","Español":"Analista de Sistemas e Investigador - Christiano T.Gaio - Desarrollador | Proyecto iniciado en 06/2026","Français":"Analyste Systèmes et Chercheur - Christiano T.Gaio - Développeur | Développement du projet commencé en 06/2026","Italiano":"Analista di Sistemi e Ricercatore - Christiano T.Gaio - Sviluppatore | Sviluppo del progetto iniziato nel 06/2026","Deutsch":"Systemanalyst und Forscher - Christiano T.Gaio - Entwickler | Projektentwicklung begonnen im 06/2026","日本語":"システムアナリスト・研究者 - Christiano T.Gaio - 開発者 | プロジェクト開発開始: 06/2026","中文":"系统分析师兼研究员 - Christiano T.Gaio - 开发者 | 项目开发启动于 06/2026"}.get(self.idioma,self.footer.text)
        if hasattr(self,'formula_view'): self.formula_view.set_text(FORMULAS_I18N_COMPLETAS.get(self.idioma,self.formula_pt))
        self._apply_theme()

    def _go_from_spinner(self, spinner, text):
        names=self.screen_names_for_language()
        try: self.sm.current=self.screen_names[names.index(text)]
        except ValueError: pass

    def _refresh_all(self):
        self._refresh_history(); self._refresh_sqlite(); self._refresh_total(); self._refresh_chart_text()


if __name__ == "__main__":
    MobileGNVApp().run()
